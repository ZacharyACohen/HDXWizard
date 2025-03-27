import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
from tkinter import messagebox
from tkinter import PhotoImage
from tkinter import Toplevel


import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment

import numpy as np
import pandas as pd

import matplotlib.pyplot as plt
import matplotlib.patches as patches
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.lines import Line2D
from matplotlib.widgets import Button, RectangleSelector
import matplotlib.ticker as ticker



import xlwings as xw
import fitz  # PyMuPDF
from PIL import Image, ImageTk


import torch
import torch.nn as nn
import torch.nn.functional as Fu
import torch.optim as optim
from torch.utils.data import DataLoader, TensorDataset
from torchvision.transforms import ToTensor

from Bio.PDB import PDBParser
from Bio import PDB
from Bio.SeqUtils import seq1
from Bio import Align
from Bio.PDB.PDBExceptions import PDBConstructionWarning



import warnings
import shutil
import re
import os
import subprocess
import platform
import csv
import tempfile
import atexit
import json
from statistics import mode
import traceback

delete_faulty_sheets = True   #### Should always be true except for debugging


def open_file(filepath):
    if platform.system() == 'Windows':
        os.startfile(filepath)
    elif platform.system() == 'Darwin':  # macOS
        subprocess.call(('open', filepath))
    else:  # Linux
        subprocess.call(('xdg-open', filepath))



os.makedirs('RecentLegends', exist_ok=True)
os.makedirs("englander_jsons", exist_ok=True)
os.makedirs("baserate_jsons", exist_ok=True)

if not os.path.isdir("Alterations"):
    os.makedirs("Alterations")

window = tk.Tk()
window.geometry("1500x900")
if os.name == 'nt':
    window.state('zoomed')
window.title("HDXWizard")
canvas = tk.Canvas(window, width=1500, height=900)
canvas.place(x=0, y=0)


def reset_preferences():
    user_choice = tk.messagebox.askyesno("Reset Preferences", "Are you sure you want to reset preferences? This will reset all formatting and data processing preferences to their default values. This will not change any saved color schemes.")
    if user_choice:
        try:
            if os.path.exists("Alterations"):
                for filename in os.listdir("Alterations"):
                    if filename == "New Text Document.txt":
                        continue
                    file_path = os.path.join("Alterations", filename)
                    if os.path.isfile(file_path):
                        os.remove(file_path)
            tk.messagebox.showinfo("Preferences Reset", "Preferences Successfully Reset")
        except:
            tk.messagebox.showerror("Preference Reset Error", f"There has been an error resetting preferences. This can be done manually by navigating to {os.getcwd()}\Alterations and deleting all files except 'New Text Document'.")
    else:
        return
    

def open_settings():
    settings_popup = tk.Toplevel(window)
    settings_popup.geometry("300x200")
    settings_popup.title("Settings")
    settings_canvas = tk.Canvas(settings_popup, width=300, height=200)
    settings_canvas.place(x=0, y=0)
    tk.Label(settings_popup, text="Settings").place(x=120, y=10)
    x1 = 5
    y1 = 5
    x2 = 295
    y2 = 195
    settings_canvas.create_rectangle(x1, y1, x2, y2)
    y = 35
    settings_canvas.create_line(x1, y, x2, y)
    
    reset_preferences_bt = tk.Button(settings_popup, text="Reset Preferences", command=reset_preferences)
    reset_preferences_bt.place(x=40, y=40)
    
    info_bt = tk.Button(settings_popup, text="INFO", width=5, command=open_info)
    info_bt.place(x=210, y=40)
    

#settings_image = PhotoImage(file="settings-icon-2048x2046-cw28eevx.png")
settings_image = Image.open("settings-icon-2048x2046-cw28eevx.png")
resized_settings_image = settings_image.resize((45, 45), Image.LANCZOS)
new_settings_image = ImageTk.PhotoImage(resized_settings_image)

#settings_bt = tk.Button(window, text='\u2699', bg="gray", fg="black", command=open_settings)
settings_bt = tk.Button(window, image=new_settings_image, command=open_settings)
settings_bt.place(x=1486, y=30)

initial_protein_set = set()
sdbt_clicked = False
cdbt_clicked = False
seqbt_txt_clicked = False
seqbt_fasta_clicked = False
skip_bt_clicked = False
txt_h_bt_clicked = False

data = []
seq = None

courier_new_style = Font(name='Courier New')
size_5_courier_new_style = Font(size=5, name='Courier New')



prot_seq_dic = {}
def open_sd_file():
    global sdbt_clicked, cdbt_clicked, temp_file_path_excel, examiner_clicked, cd_dfs, examiner_dfs, workbench_clicked, workbench_dfs, sd_dfs, sd_bt
    
    sd_bt.config(state="disabled")
    sd_bt.config(relief="sunken", bg="white", fg="black")
    sd_file_paths = filedialog.askopenfilenames(filetypes=[("Excel and CSV files", "*.xlsx *.csv"), ("All files", "*.*")])
    if not sd_file_paths:
        if sdbt_clicked == True:
            sd_bt.config(state="normal")
            sd_bt.config(relief="raised", bg="green", fg="white")
        else:
            sd_bt.config(state="normal")
            sd_bt.config(relief="raised", bg="orange", fg="black")
        return
    for sd_file_path in sd_file_paths:
        not_state_data = False
        try:
            if sd_file_path.endswith(".xlsx"):
                df = pd.read_excel(sd_file_path)
            if sd_file_path.endswith(".csv"):
                df = pd.read_csv(sd_file_path)
            if 'Center SD' not in df.columns:
                if len(sd_file_paths) == 1:
                    tk.messagebox.showerror("DynamX State Data Error", "This File is not DynamX State Data. Please try again.")
                else:
                    tk.messagebox.showerror("DynamX State Data Error", f"The File [{sd_file_path}] is not DynamX State Data and has been excluded.")
                if sdbt_clicked is True:
                    sd_bt.config(state="normal")
                    sd_bt.config(relief="raised", bg="green", fg="white")
                else:
                    sd_bt.config(state="normal")
                    sd_bt.config(relief="raised", bg="orange", fg="black")
                not_state_data = True
                
        except PermissionError as e:
            if len(sd_file_paths) == 1:
                tk.messagebox.showerror("Permission Error", "File is open. Please close the file and try again")
            else:
                tk.messagebox.showerror("Permission Error", f"The File [{sd_file_path}] is open and has been excluded.")
            if sdbt_clicked is True:
                sd_bt.config(state="normal")
                sd_bt.config(relief="raised", bg="green", fg="white")
            else:
                sd_bt.config(state="normal")
                sd_bt.config(relief="raised", bg="orange", fg="black")
            not_state_data = True
            
        except Exception as e:
            if len(sd_file_paths) == 1:
                tk.messagebox.showerror("DynamX State Data Error", "This File is not DynamX State Data. Please try again.")
            else:
                tk.messagebox.showerror("DynamX State Data Error", f"The File [{sd_file_path}] is not Dynamx State Data and has been excluded.")
            if sdbt_clicked is True:
                sd_bt.config(state="normal")
                sd_bt.config(relief="raised", bg="green", fg="white")
            else:
                sd_bt.config(state="normal")
                sd_bt.config(relief="raised", bg="orange", fg="black")
            not_state_data = True
            
        if not_state_data == True:
            continue
        
        sd_dfs.append(df)
        for each_df in sd_dfs:
            for index, row in each_df.iterrows():
                protein = row["Protein"]
                initial_protein_set.add(protein)


        sd_bt = tk.Button(window, text="State Data", bg="green", fg="white",  width=12,  command=open_sd_file)
        sd_bt.place(x=120, y=30)
        sdbt_clicked = True


        cdbt_clicked = False
        examiner_clicked = False
        workbench_clicked = False
        cd_dfs = []
        examiner_dfs = []
        workbench_dfs = []

        examiner_bt = tk.Button(window, text="HDExaminer", bg="orange", fg="black", width=12, command=open_examiner_file)
        examiner_bt.place(x=120, y=60)
        workbench_bt = tk.Button(window, text="HDXWorkbench", bg="orange", fg="black", width=12, command=open_workbench_file)
        workbench_bt.place(x=220, y=60)
        cd_bt = tk.Button(window, text="Cluster Data", bg="orange", fg="black", width=12, command=open_cd_file)
        cd_bt.place(x=220, y=30)

        
        check_button_clicks()
    else:
        if sdbt_clicked == True:
            sd_bt.config(state="normal")
            sd_bt.config(relief="raised", bg="green", fg="white")
        else:
            sd_bt.config(state="normal")
            sd_bt.config(relief="raised", bg="orange", fg="black")


cd_dfs = []
examiner_dfs = []
workbench_dfs = []
sd_dfs = []
def open_cd_file():
    global cdbt_clicked, cd_dfs, sdbt_clicked, examiner_clicked, examiner_dfs, workbench_clicked, workbench_dfs, sd_dfs, cd_bt
    cd_bt.config(state="disabled")
    cd_bt.config(relief="sunken", bg="white", fg="black")
    
    cd_file_paths = filedialog.askopenfilenames(filetypes=[("Excel and CSV files", "*.xlsx *.csv"), ("All files", "*.*")])

    if not cd_file_paths:
        if cdbt_clicked is True:
            cd_bt.config(state="normal")
            cd_bt.config(relief="raised", bg="green", fg="white")
        else:
            cd_bt.config(state="normal")
            cd_bt.config(relief="raised", bg="orange", fg="black")
        return
    
    
    for cd_file_path in cd_file_paths:
        not_cluster_data = False
        
        try:
            if cd_file_path.endswith(".xlsx"):
                df = pd.read_excel(cd_file_path)
            if cd_file_path.endswith(".csv"):
                df = pd.read_csv(cd_file_path)
            if 'File' not in df.columns or 'Modification' not in df.columns:
                if len(cd_file_paths) == 1:
                    tk.messagebox.showerror("DynamX Cluster Data Error", "This File is not DynamX Cluster Data. Please try again.")
                else:
                    tk.messagebox.showerror("DynamX Cluster Data Error", f"The File [{cd_file_path}] is not DynamX Cluster Data and has been exluded.")

                if cdbt_clicked is True:
                    cd_bt.config(state="normal")
                    cd_bt.config(relief="raised", bg="green", fg="white")
                else:
                    cd_bt.config(state="normal")
                    cd_bt.config(relief="raised", bg="orange", fg="black")

                not_cluster_data = True
                
        except PermissionError as e:
            if len(cd_file_paths) == 1:
                tk.messagebox.showerror("Permission Error", "File is open. Please close the file and try again")
            else:
                tk.messagebox.showerror("Permission Error", f"The File [{cd_file_path}] is open and has been excluded.")
            if cdbt_clicked is True:
                cd_bt.config(state="normal")
                cd_bt.config(relief="raised", bg="green", fg="white")
            else:
                cd_bt.config(state="normal")
                cd_bt.config(relief="raised", bg="orange", fg="black")
            not_cluster_data = True
            
        except Exception as e:
            if len(cd_file_paths) == 1:
                tk.messagebox.showerror("DynamX Cluster Data Error", "This File is not DynamX Cluster Data. Please try again.")
            else:
                tk.messagebox.showerror("DynamX Cluster Data Error", f"The File [{cd_file_path}] is not DynamX Cluster Data and has been exluded.")

            if cdbt_clicked is True:
                cd_bt.config(state="normal")
                cd_bt.config(relief="raised", bg="green", fg="white")
            else:
                cd_bt.config(state="normal")
                cd_bt.config(relief="raised", bg="orange", fg="black")
            
            not_cluster_data = True
        
       
        
        if not_cluster_data == True:
            continue
        
        cd_dfs.append(df)
        for each_df in cd_dfs:
            for index, row in each_df.iterrows():
                protein = row["Protein"]
                initial_protein_set.add(protein)

        

        cd_bt = tk.Button(window, text="Cluster Data", bg="green", fg="white",  width=12,  command=open_cd_file)
        cd_bt.place(x=220, y=30)
        cdbt_clicked = True
        sdbt_clicked = False
        examiner_clicked = False
        workbench_clicked = False
        examiner_dfs = []
        workbench_dfs = []
        sd_dfs = []
        
        sd_bt = tk.Button(window, text="State Data", bg="orange", fg="black",  width=12, command=open_sd_file)
        sd_bt.place(x=120, y=30)
        
        
        examiner_bt = tk.Button(window, text="HDExaminer", bg="orange", fg="black", width=12, command=open_examiner_file)
        examiner_bt.place(x=120, y=60)
        workbench_bt = tk.Button(window, text="HDXWorkbench", bg="orange", fg="black", width=12, command=open_workbench_file)
        workbench_bt.place(x=220, y=60)
        
        check_button_clicks()
    else:
        if cdbt_clicked is True:
            cd_bt.config(state="normal")
            cd_bt.config(relief="raised", bg="green", fg="white")
        else:
            cd_bt.config(state="normal")
            cd_bt.config(relief="raised", bg="orange", fg="black")

            
            
examiner_clicked = False
workbench_clicked = False
def open_examiner_file():
    global cdbt_clicked, examiner_dfs, sdbt_clicked, examiner_clicked, cd_dfs, examiner_bt, workbench_clicked, workbench_dfs, sd_dfs
    examiner_bt.config(state="disabled")
    examiner_bt.config(relief="sunken", bg="white", fg="black")
    
    examiner_file_paths = filedialog.askopenfilenames(filetypes=[("Excel and CSV files", "*.xlsx *.csv"), ("All files", "*.*")])
    if not examiner_file_paths:
        if examiner_clicked is True:
            examiner_bt.config(state="normal")
            examiner_bt.config(relief="raised", bg="green", fg="white")
        else:
            examiner_bt.config(state="normal")
            examiner_bt.config(relief="raised", bg="orange", fg="black")
        return
    for examiner_file_path in examiner_file_paths:
        not_examiner_data = False
        try:
            if examiner_file_path.endswith(".xlsx"):
                df = pd.read_excel(examiner_file_path)
            if examiner_file_path.endswith(".csv"):
                df = pd.read_csv(examiner_file_path)
            if 'Conf Interval (#D)' not in df.columns:
                if len(examiner_file_paths) == 1:
                    tk.messagebox.showerror("HDExaminer Data Error", "This File is not HDExaminer Data. Please try again.")
                else:
                    tk.messagebox.showerror("HDExaminer Data Error", f"The File [{examiner_file_path}] is not HDExaminer Data and has been excluded.")
                if examiner_clicked is True:
                    examiner_bt.config(state="normal")
                    examiner_bt.config(relief="raised", bg="green", fg="white")
                else:
                    examiner_bt.config(state="normal")
                    examiner_bt.config(relief="raised", bg="orange", fg="black")
                not_examiner_data = True
                
        except PermissionError as e:
            if len(examiner_file_paths) == 1:
                tk.messagebox.showerror("Permission Error", "File is open. Please close the file and try again")
            else:
                tk.messagebox.showerror("Permission Error", f"The File [{examiner_file_path}] is open and has been excluded.")
            if examiner_clicked is True:
                examiner_bt.config(state="normal")
                examiner_bt.config(relief="raised", bg="green", fg="white")
            else:
                examiner_bt.config(state="normal")
                examiner_bt.config(relief="raised", bg="orange", fg="black")
            not_examiner_data = True
            
        except Exception as e:
            if len(examiner_file_paths) == 1:
                tk.messagebox.showerror("HDExaminer Data Error", "This File is not HDExaminer Data. Please try again.")
            else:
                tk.messagebox.showerror("HDExaminer Data Error", f"The File [{examiner_file_path}] is not HDExaminer Data and has been excluded.")
            if examiner_clicked is True:
                examiner_bt.config(state="normal")
                examiner_bt.config(relief="raised", bg="green", fg="white")
            else:
                examiner_bt.config(state="normal")
                examiner_bt.config(relief="raised", bg="orange", fg="black")
            not_examiner_data = True
            
        if not_examiner_data == True:
            continue
            
        examiner_dfs.append(df)
        for each_df in examiner_dfs:
            for index, row in each_df.iterrows():
                protein = row["Protein"]
                initial_protein_set.add(protein)


        examiner_bt = tk.Button(window, text="HDExaminer", bg="green", fg="white", width=12, command=open_examiner_file)
        examiner_bt.place(x=120, y=60)
        examiner_clicked = True

        cd_bt = tk.Button(window, text="Cluster Data", bg="orange", fg="black",  width=12,  command=open_cd_file)
        cd_bt.place(x=220, y=30)
        cdbt_clicked = False
        sdbt_clicked = False
        cd_dfs = []
        workbench_clicked = False
        workbench_dfs = []
        sd_dfs = []
        
        workbench_bt = tk.Button(window, text="HDXWorkbench", bg="orange", fg="black", width=12, command=open_workbench_file)
        workbench_bt.place(x=220, y=60)
        
        
        sd_bt = tk.Button(window, text="State Data", bg="orange", fg="black",  width=12,  command=open_sd_file)
        sd_bt.place(x=120, y=30)
        
        check_button_clicks()
    else:
        if examiner_clicked is True:
            examiner_bt.config(state="normal")
            examiner_bt.config(relief="raised", bg="green", fg="white")
        else:
            examiner_bt.config(state="normal")
            examiner_bt.config(relief="raised", bg="orange", fg="black")
            

def open_workbench_file():
    global cdbt_clicked, examiner_dfs, sdbt_clicked, examiner_clicked, cd_dfs, examiner_bt, workbench_clicked, workbench_dfs, workbench_bt, sd_dfs
    workbench_bt.config(state="disabled")
    workbench_bt.config(relief="sunken", bg="white", fg="black")
    
    workbench_file_paths = filedialog.askopenfilenames(filetypes=[("Excel and CSV files", "*.xlsx *.csv"), ("All files", "*.*")])
    if not workbench_file_paths:
        if workbench_clicked is True:
            workbench_bt.config(state="normal")
            workbench_bt.config(relief="raised", bg="green", fg="white")
        else:
            workbench_bt.config(state="normal")
            workbench_bt.config(relief="raised", bg="orange", fg="black")
        return
    for workbench_file_path in workbench_file_paths:
        try:
            if workbench_file_path.endswith(".xlsx"):
                df = pd.read_excel(workbench_file_path, skiprows=29)
            if workbench_file_path.endswith(".csv"):
                df = pd.read_csv(workbench_file_path, skiprows=29)
            not_workbench_data = False
            if 'features' not in df.columns:
                if len(workbench_file_paths) == 1:
                    tk.messagebox.showerror("HDXWorkbench Data Error", "This File is not HDXWorkbench Data. Please try again.")
                else:
                    tk.messagebox.showerror("HDXWorkbench Data Error", f"The File [{workbench_file_path}] is not HDXWorkbench Data and has been excluded.")
                if workbench_clicked is True:
                    workbench_bt.config(state="normal")
                    workbench_bt.config(relief="raised", bg="green", fg="white")
                else:
                    workbench_bt.config(state="normal")
                    workbench_bt.config(relief="raised", bg="orange", fg="black")
                not_workbench_data = True
        except PermissionError as e:
            if len(workbench_file_paths) == 1:
                tk.messagebox.showerror("Permission Error", "File is open. Please close the file and try again")
            else:
                tk.messagebox.showerror("Permission Error", f"The File [{workbench_file_path}] is open and has been excluded.")
            if workbench_clicked is True:
                workbench_bt.config(state="normal")
                workbench_bt.config(relief="raised", bg="green", fg="white")
            else:
                workbench_bt.config(state="normal")
                workbench_bt.config(relief="raised", bg="orange", fg="black")
            not_workbench_data = True
            
        except Exception as e:
            if len(workbench_file_paths) == 1:
                tk.messagebox.showerror("HDXWorkbench Data Error", "This File is not HDXWorkbench Data. Please try again.")
            else:
                tk.messagebox.showerror("HDXWorkbench Data Error", f"The File [{workbench_file_path}] is not HDXWorkbench Data and has been excluded.")
            if workbench_clicked is True:
                workbench_bt.config(state="normal")
                workbench_bt.config(relief="raised", bg="green", fg="white")
            else:
                workbench_bt.config(state="normal")
                workbench_bt.config(relief="raised", bg="orange", fg="black")
            not_workbench_data = True
            
        if not_workbench_data == True:
            continue
            
        workbench_dfs.append(df)
        for each_df in workbench_dfs:
            for index, row in each_df.iterrows():
                protein = 'protein' 
                initial_protein_set.add(protein)


        workbench_bt = tk.Button(window, text="HDXWorkbench", bg="green", fg="white", width=12, command=open_workbench_file)
        workbench_bt.place(x=220, y=60)
        workbench_clicked = True

        cd_bt = tk.Button(window, text="Cluster Data", bg="orange", fg="black",  width=12,  command=open_cd_file)
        cd_bt.place(x=220, y=30)
        
        cdbt_clicked = False
        sdbt_clicked = False
        cd_dfs = []
        examiner_clicked = False
        examiner_dfs = []
        sd_dfs = []
        
        examiner_bt = tk.Button(window, text="HDExaminer", bg="orange", fg="black", width=12, command=open_examiner_file)
        examiner_bt.place(x=120, y=60)
        
        
        sd_bt = tk.Button(window, text="State Data", bg="orange", fg="black",  width=12,  command=open_sd_file)
        sd_bt.place(x=120, y=30)
        
        check_button_clicks()
    else:
        if workbench_clicked is True:
            workbench_bt.config(state="normal")
            workbench_bt.config(relief="raised", bg="green", fg="white")
        else:
            workbench_bt.config(state="normal")
            workbench_bt.config(relief="raised", bg="orange", fg="black")
        
def clear_data_sdcd():
    global cd_dfs, examiner_dfs, sdbt_clicked, cdbt_clicked, examiner_clicked, workbench_clicked, workbench_dfs, sd_dfs
    sdbt_clicked = False
    cdbt_clicked = False
    examiner_clicked = False
    workbench_clicked = False
    cd_dfs = []
    examiner_dfs = []
    workbench_dfs = []
    sd_dfs = []
    initial_protein_set = set()
    
    cd_bt = tk.Button(window, text="Cluster Data", bg="orange", fg="black",  width=12,  command=open_cd_file)
    cd_bt.place(x=220, y=30)
    
    sd_bt = tk.Button(window, text="State Data", bg="orange", fg="black",  width=12,  command=open_sd_file)
    sd_bt.place(x=120, y=30)
    
    examiner_bt = tk.Button(window, text="HDExaminer", bg="orange", fg="black", width=12, command=open_examiner_file)
    examiner_bt.place(x=120, y=60)
    
    workbench_bt = tk.Button(window, text="HDX Workbench", bg="orange", fg="black", width=12, command=open_workbench_file)
    workbench_bt.place(x=220, y=60)
    
    for widget in window.winfo_children():
        if widget.winfo_x() > 370 and widget != settings_bt:
            widget.destroy()
    for item in canvas.find_all():
        coords = canvas.coords(item)
        # For lines and shapes, coords are a list of x, y pairs. We check the first x-coordinate.
        if coords and coords[0] > 370:
            canvas.delete(item)
                
    for widget in window.winfo_children():
        if widget.winfo_y() > 180 and widget != settings_bt:
            widget.destroy()
    for item in canvas.find_all():
        coords = canvas.coords(item)
        # For lines and shapes, coords are a list of x, y pairs. We check the first x-coordinate.
        if coords and coords[1] > 180:
            canvas.delete(item)

            

def get_max_theo(peptide):
    if n_min_one == True or n_min_two == True:
        length = len(peptide)
        prolinecount=0
        for letter in peptide:
            if letter == 'P':
                prolinecount = prolinecount+1
        if peptide[0] == 'P':
            max_theo = length-prolinecount
        else:
            max_theo = (length-1)-prolinecount
        if n_min_two == True:
            if peptide[1] == 'P':
                pass
            else:
                max_theo = max_theo - 1
    elif calculate_theoretical_back_exchange == True:
        englander_time_in_H2O = englander_time_in_H2O_trap + (englander_time_in_H2O_constant * peptide_RT_dic[peptide])
        max_theo = calculate_peptide_deuterium_remaining(peptide, englander_pH, englander_tempK, englander_time_in_H2O)
    else:
        tk.messagebox.showerror("Max Uptake Error", "No deuterium incorporation limit specified")
    return max_theo
    
    
        
    



class sd_peptide:
    __slots__ = ("Sequence", "Startvalue", "Endvalue", "State", "Protein", "Timepoint", "Uptake", "Stdev", "RT")
    Sequence: str
    Startvalue: int
    Endvalue: int
    State: str
    Protein: str
    Timepoint: float
    Stdev: float
    RT: float
    
    def __init__(self, Sequence, Startvalue, Endvalue, State, Protein, Timepoint, Uptake, Stdev, RT):
        self.Sequence = Sequence
        self.Startvalue = Startvalue
        self.Endvalue = Endvalue
        self.State = State
        self.Protein = Protein
        self.Timepoint = Timepoint
        self.Uptake = Uptake
        self.Stdev = Stdev
        self.RT = RT

class cd_peptide:
    __slots__ = ("Sequence", "Startvalue", "Endvalue", "State", "Protein", "File", "Timepoint", "Charge", "Intensity", "Center", "MHP", "Uptake", "Stdev", "RT", "und_std")
    Sequence: str
    Startvalue: int
    Endvalue: int
    State: str
    Protein: str
    File: str
    Timepoint: float
    Charge: int
    Intensity: float
    Center: float
    MHP: float
    Uptake: float
    Stdev: float
    RT: float
    und_std: float
    
    
    def __init__(self, Sequence, Startvalue, Endvalue, State, Protein, File, Timepoint, Charge, Intensity, Center, MHP, Uptake, Stdev, RT, und_std):
        self.Sequence = Sequence
        self.Startvalue = Startvalue
        self.Endvalue = Endvalue
        self.State = State
        self.Protein = Protein
        self.File = File
        self.Timepoint = Timepoint
        self.Charge = Charge
        self.Intensity = Intensity
        self.Center = Center
        self.MHP = MHP
        self.Uptake = Uptake
        self.Stdev = Stdev
        self.RT = RT
        self.und_std = und_std
        
        
class examiner_peptide:
    __slots__ = ("Sequence", "Startvalue", "Endvalue", "State", "Protein", "Timepoint", "Uptake", "Stdev", "RT")
    Sequence: str
    Startvalue: int
    Endvalue: int
    State: str
    Protein: str
    Timepoint: float
    Uptake: float
    Stdev: float
    RT: float
    
    
    def __init__(self, Sequence, Startvalue, Endvalue, State, Protein, Timepoint, Uptake, Stdev, RT):
        self.Sequence = Sequence
        self.Startvalue = Startvalue
        self.Endvalue = Endvalue
        self.State = State
        self.Protein = Protein
        self.Timepoint = Timepoint
        self.Uptake = Uptake
        self.Stdev = Stdev
        self.RT = RT
        
class workbench_peptide:
    __slots__ = ("Sequence", "Startvalue", "Endvalue", "State", "Protein", "Timepoint", "Charge", "Center", "Uptake", "Stdev", "Discarded", "RT", "und_std", "File")
    Sequence: str
    Startvalue: int
    Endvalue: int
    State: str
    Protein: str
    Timepoint: float
    Charge: int
    Center: float
    Uptake: float
    Stdev: float
    Discarded: str
    RT: float
    und_std: float
    File: str
    
    def __init__(self, Sequence, Startvalue, Endvalue, State, Protein, Timepoint, Charge, Center, Uptake, Stdev, Discarded, RT, und_std, File):
        self.Sequence = Sequence
        self.Startvalue = Startvalue
        self.Endvalue = Endvalue
        self.State = State
        self.Protein = Protein
        self.Timepoint = Timepoint
        self.Charge = Charge
        self.Center = Center
        self.Uptake = Uptake
        self.Stdev = Stdev
        self.Discarded = Discarded
        self.RT = RT
        self.und_std = und_std
        self.File = File
        

def open_sequence_txt():
    global seqbt_txt_clicked, seq
    file_path = filedialog.askopenfilename(filetypes=[("Text Files", "*.txt")])
    if file_path:
        seq = open(file_path, 'r')
        for line in seq:
            if line.startswith(">"):
                tk.messagebox.showerror("Sequence File Error", "This button is for inputting one sequence only, NOT in FASTA format")
                return
        seqbt2 = tk.Button(window, text=".txt (p)", bg="green", fg="white",  width=5, command=lambda: [open_sequence_txt(), skip_sequence_off(), open_sequence_fasta_off(), txt_h_off()])
        seqbt2.place(x=220, y=105)
        seqbt_txt_clicked = True
        check_button_clicks()
def seq_txt_off():
    global seqbt_txt_clicked
    seqbt_txt_clicked = False
    seqbt_txt = tk.Button(window, text=".txt (p)", bg="orange", fg="black", width=5,  command=lambda: [open_sequence_txt(), skip_sequence_off(), open_sequence_fasta_off(), txt_h_off()])
    seqbt_txt.place(x=220, y=105)

def open_sequence_fasta():
    global seqbt_fasta_clicked, fasta_file_path, prot_seq_dic
    
    if os.path.exists("Alterations/NoMessage.json"):
        pass
    else:
        user_choice = tk.messagebox.askyesno("Fasta Format", "This program will read .fasta file protein names only until the first space or |, i.e. >BSA BSA will be read as BSA. This name must match the name of the protein from the data file exactly to be matched.  \n\nDo you want to see this message again?", default='no')
        if user_choice:
            pass
        else:
            dont_show_path = "Alterations/NoMessage.json"
            with open(dont_show_path, 'w') as file:
                json.dump({"Empty": 0}, file)
                
    fasta_file_paths = filedialog.askopenfilenames(filetypes=[("Fasta Files", "*.fasta")])
    if not fasta_file_paths:
        return
    for fasta_file_path in fasta_file_paths:
        seq_headers = open(fasta_file_path, 'r')
        for i, line in enumerate(seq_headers):
            if i == 0:
                if not line.startswith(">"):
                    tk.messagebox.showerror("Sequence File Error", "Fasta file format is incorrect")
                    return
        seq_headers.seek(0)
        seqbt_fasta_clicked = True
#        for line in seq_headers:
#            if line.startswith(">"):
#                new_line = line.lstrip(">")
#                protein_name = new_line.split("|")[0].split()[0]
#                if protein_name not in prot_seq_dic:
#                    next_line = next(seq_headers, None)  # Read the next line
#                    if next_line is not None and next_line.strip() != "":
#                        prot_seq_dic[protein_name] = next_line.strip()
#                    else:
#                        next_next_line = next(seq_headers, None)
#                        if next_next_line is not None and next_next_line.strip() != "":
#                            prot_seq_dic[protein_name] = next_next_line.strip()
        current_protein = None
        current_sequence = []
        for line in seq_headers:
            line = line.strip()
            if line.startswith(">"):
                if current_protein:
                    prot_seq_dic[current_protein] = "".join(current_sequence)
                current_protein = line.split("|")[0].split()[0].lstrip(">").strip()
                current_sequence = []
            else:
                current_sequence.append(line)
        if current_protein:
            prot_seq_dic[current_protein] = "".join(current_sequence)
                    
                    
            

        seqbt_fasta = tk.Button(window, text=".fasta",bg="green",fg="white", width=5, command=lambda: [open_sequence_fasta(), seq_txt_off, skip_sequence_off(), txt_h_off()])
        seqbt_fasta.place(x=120, y=105)
        check_button_clicks()
def open_sequence_fasta_off():
    global seqbt_fasta_clicked
    seqbt_fasta_clicked = False
    seqbt_fasta = tk.Button(window, text=".fasta",bg="orange",fg="black", width=5, command=lambda: [open_sequence_fasta(), seq_txt_off, skip_sequence_off(), txt_h_off()])
    seqbt_fasta.place(x=120, y=105)

def txt_h_on():
    global txt_h_bt_clicked
    global prot_seq_dic
    
    if os.path.exists("Alterations/NoMessage.json"):
        pass
    else:
        user_choice = tk.messagebox.askyesno("Fasta Format", "This program will read .fasta file protein names only until the first space or |, i.e. >BSA BSA will be read as BSA. This name must match the name of the protein from the data file exactly to be matched. \n\nDo you want to see this message again?", default='no')
        if user_choice:
            pass
        else:
            dont_show_path = "Alterations/NoMessage.json"
            with open(dont_show_path, 'w') as file:
                json.dump({"Empty": 0}, file)
            
    txt_h_file_paths = filedialog.askopenfilenames(filetypes=[("Text Files", "*.txt")])
    if not txt_h_file_paths:
        return
    for txt_h_file_path in txt_h_file_paths:
        seq_headers = open(txt_h_file_path, 'r')
        
        for i, line in enumerate(seq_headers):
            if i == 0:
                if not line.startswith(">"):
                    tk.messagebox.showerror("Sequence File Error", "Fasta format is incorrect. Please insert .txt file in fasta format.")
                    return
        
        seq_headers.seek(0)
        txt_h_bt_clicked = True
        current_protein = None
        current_sequence = []
        for line in seq_headers:
            line = line.strip()
            if line.startswith(">"):
                if current_protein:
                    prot_seq_dic[current_protein] = "".join(current_sequence)
                current_protein = line.split("|")[0].split()[0].lstrip(">").strip()
                current_sequence = []
            else:
                current_sequence.append(line)
        if current_protein:
            prot_seq_dic[current_protein] = "".join(current_sequence)
#        for line in seq_headers:
#            if line.startswith(">"):
#                new_line = line.lstrip(">")
#                protein_name = new_line.split("|")[0].split()[0]
#                if protein_name not in prot_seq_dic:
#                    next_line = next(seq_headers, None)  # Read the next line
#                    if next_line is not None and next_line.strip() != "":
#                        prot_seq_dic[protein_name] = next_line.strip()
#                    else:
#                        next_next_line = next(seq_headers, None)
#                        if next_next_line is not None and next_next_line.strip() != "":
#                            prot_seq_dic[protein_name] = next_next_line.strip()
            
        txt_h_bt = tk.Button(window, text=".txt (>)",bg="green",fg="white", width=5, command=lambda: [seq_txt_off(), skip_sequence_off(), open_sequence_fasta_off(), txt_h_on()])
        txt_h_bt.place(x=170, y=105)

        check_button_clicks()
        

def txt_h_off():
    global txt_h_bt_clicked
    txt_h_bt_clicked = False
    txt_h_bt = tk.Button(window, text=".txt (>)",bg="orange",fg="black", width=5, command=lambda: [seq_txt_off(), skip_sequence_off(), open_sequence_fasta_off(), txt_h_on()])
    txt_h_bt.place(x=170, y=105)


def skip_sequence():
    global skip_bt_clicked
    skip_bt_clicked = True
    skip_bt2 = tk.Button(window, text="Skip",bg="green",fg="white", width=5, command=skip_sequence)
    skip_bt2.place(x=270, y=105)
    check_button_clicks()

def skip_sequence_off():
    global skip_bt_clicked
    skip_bt_clicked = False
    skip_bt = tk.Button(window, text="Skip",bg="orange",fg="black", width=5, command=lambda: [skip_sequence(), seq_txt_off(), open_sequence_fasta_off(), txt_h_off()])
    skip_bt.place(x=270, y=105)

def open_info():
    try:
#        os.startfile("HDXWizard_Operating_Instructions_1.0.pdf")
        open_file("HDXWizard_Operating_Instructions_1.0.pdf")
    except:
        tk.messagebox.showerror("Error", "Cannot find operating instructions file")


#def go_to_git():
#    webbrowser.open("https://github.com/ZacharyACohen/HDXWizard.git")
#
#if program_needs_update is True:
#    popup_window_update = tk.Toplevel(window)  # Create a new window for the popup menu
#    popup_window_update.geometry("500x100")
#    popup_window_update.title("Update Available")
#    tk.Label(popup_window_update, text=f"Current Version: {version_number}").place(x=10, y=10)
#    tk.Label(popup_window_update, text=f"Newest Version: {newest_version}").place(x=10, y=40)
#    update_label = tk.Label(popup_window_update, text="Please go to https://github.com/ZacharyACohen/HDXWizard.git to update program")
#    update_label.place(x=10, y=70)
#    go_bt = tk.Button(popup_window_update, text="GO", command=go_to_git).place(x=460, y=68)
#    popup_window_update.attributes("-topmost", True)



file_enter_lab = tk.Label(window, text="File Entry")
file_enter_lab.place(x=40, y=5)


sdlab = tk.Label(window, text="DynamX Data: ")
sdlab.place(x=13, y=30)

tk.Label(window, text="Other Data:").place(x=13, y=60)


sd_bt = tk.Button(window, text="State Data",bg="orange",fg="black", width=12, command=open_sd_file)
sd_bt.place(x=120, y=30)

cd_bt = tk.Button(window, text="Cluster Data", bg="orange", fg="black", width=12, command=open_cd_file)
cd_bt.place(x=220, y=30)

clear_bt = tk.Button(window, text="Clear", width=5, command=clear_data_sdcd)
clear_bt.place(x=320, y=30)

examiner_bt = tk.Button(window, text="HDExaminer", bg="orange", fg="black", width=12, command=open_examiner_file)
examiner_bt.place(x=120, y=60)

workbench_bt = tk.Button(window, text="HDXWorkbench", bg="orange", fg="black", width=12, command=open_workbench_file)
workbench_bt.place(x=220, y=60)


seqlab = tk.Label(window, text="Insert Sequence: ")
seqlab.place(x=15, y=105)

seqbt_txt = tk.Button(window, text=".txt (p)",bg="orange",fg="black", width=5, command=lambda: [open_sequence_txt(), skip_sequence_off(), open_sequence_fasta_off(), txt_h_off()])
seqbt_txt.place(x=220, y=105)

seqbt_fasta = tk.Button(window, text=".fasta",bg="orange",fg="black", width=5, command=lambda: [open_sequence_fasta(), seq_txt_off, skip_sequence_off(), txt_h_off()])
seqbt_fasta.place(x=120, y=105)

skip_bt = tk.Button(window, text="Skip",bg="orange",fg="black", width=5, command=lambda: [skip_sequence(), seq_txt_off(), open_sequence_fasta_off(), txt_h_off()])
skip_bt.place(x=270, y=105)

txt_h_bt = tk.Button(window, text=".txt (>)",bg="orange",fg="black", width=5, command=lambda: [seq_txt_off(), skip_sequence_off(), open_sequence_fasta_off(), txt_h_on()])
txt_h_bt.place(x=170, y=105)

x1 = 10
x2 = 370
y=95
canvas.create_line(x1, y, x2, y)
seq_explain_lb = tk.Label(window, text="For fasta and .txt (>) files (.txt with fasta format): add unlimited")
seq_explain_lb.place(x=15, y=135)
seq_explain_lb2 = tk.Label(window, text="For .txt (p), add one file containing only one sequence (no header)")
seq_explain_lb2.place(x=15, y=150)


def clear_data_seqs():
    global prot_seq_dic, seq
    prot_seq_dic = {}
    seq = None
    
    seq_txt_off()
    open_sequence_fasta_off()
    txt_h_off()
    skip_sequence_off()
    
    for widget in window.winfo_children():
        if widget.winfo_x() > 370 and widget != settings_bt:
            widget.destroy()
    for item in canvas.find_all():
        coords = canvas.coords(item)
        # For lines and shapes, coords are a list of x, y pairs. We check the first x-coordinate.
        if coords and coords[0] > 370:
            canvas.delete(item)
                
    for widget in window.winfo_children():
        if widget.winfo_y() > 180 and widget != settings_bt:
            widget.destroy()
    for item in canvas.find_all():
        coords = canvas.coords(item)
        # For lines and shapes, coords are a list of x, y pairs. We check the first x-coordinate.
        if coords and coords[1] > 180:
            canvas.delete(item)

clear_seqs_bt = tk.Button(window, text="Clear", width=5, command=clear_data_seqs)
clear_seqs_bt.place(x=320, y=105)


def n_min_one_on():
    global n_min_one, n_min_two, calculate_theoretical_back_exchange
    n_min_one_bt = tk.Button(window, text="N-1", bg="green", fg="white", command=lambda: [n_min_one_off(), n_min_two_on()])
    n_min_one_bt.place(x=185, y=250)
    n_min_one = True

def n_min_one_off():
    global n_min_one, n_min_two, calculate_theoretical_back_exchange
    n_min_one_bt = tk.Button(window, text="N-1", bg="orange", fg="black", command=lambda: [n_min_one_on(), n_min_two_off(), theo_calculation_off()])
    n_min_one_bt.place(x=185, y=250)
    n_min_one = False

def n_min_two_on():
    global n_min_one, n_min_two, calculate_theoretical_back_exchange
    n_min_two_bt = tk.Button(window, text="N-2", bg="green", fg="white", command=lambda: [n_min_one_on(), n_min_two_off(), theo_calculation_off()])
    n_min_two_bt.place(x=225, y=250)
    n_min_two = True
    if os.path.exists("Alterations/Min2.json"):
        pass
    else:
        user_choice = tk.messagebox.askyesno("Default Changes", "Do you want to make this the default value?", default='no')
        dont_show_path = "Alterations/Min2.json"
        if user_choice:
            with open(dont_show_path, 'w') as file:
                json.dump({"Min2Json": 2}, file)
        else:
            with open(dont_show_path, 'w') as file:
                json.dump({"Min2Json": 1}, file)

def n_min_two_off():
    global n_min_one, n_min_two, calculate_theoretical_back_exchange
    n_min_two_bt = tk.Button(window, text="N-2", bg="orange", fg="black", command=lambda: [n_min_one_off(), n_min_two_on(), theo_calculation_off()])
    n_min_two_bt.place(x=225, y=250)
    n_min_two = False
    
def theo_calculation_on():
    global n_min_one, n_min_two, calculate_theoretical_back_exchange, englander_window
                
    
    englander_window = Toplevel(window)
    englander_window.title("Englander Deuterium Incorporation Calculation")
    englander_window.geometry("600x300")
    

    englander_window.lift()
    englander_window.attributes('-topmost', True)
    
    tk.Label(englander_window, text="Input Back Exchange Parameters").place(x=10, y=10)
    
    tk.Label(englander_window, text="pH:").place(x=10, y=40)
    englander_pH_enter = tk.Entry(englander_window, width=6)
    englander_pH_enter.place(x=50, y=40)
    englander_pH_enter.insert(0, "2.3")
    
    tk.Label(englander_window, text="T(K):").place(x=10, y=70)
    englander_tempk_enter = tk.Entry(englander_window, width=6)
    englander_tempk_enter.place(x=50, y=70)
    englander_tempk_enter.insert(0, "273.15")
    
    tk.Label(englander_window, text="Time in H2O (m) = ").place(x=10, y=100)
    
    tk.Label(englander_window, text="Trapping:").place(x=130, y=100)
    englander_trap_enter = tk.Entry(englander_window, width=3)
    englander_trap_enter.place(x=190, y=100)
    englander_trap_enter.insert(0, "3")
    
    tk.Label(englander_window, text="+ (Constant").place(x=210, y=100)
    englander_constant_enter = tk.Entry(englander_window, width=3)
    englander_constant_enter.place(x=280, y=100)
    englander_constant_enter.insert(0, "0")
    tk.Label(englander_window, text="* Peptide RT)").place(x=300, y=100)
        
    save_englander_bt = tk.Button(englander_window, text="Save Values", command=lambda: save_englander_rates(englander_pH_enter.get(), englander_tempk_enter.get(), englander_trap_enter.get(), englander_constant_enter.get(), englander_json_dropdown))
    save_englander_bt.place(x=20, y=130)
    
    
    tk.Label(englander_window, text="Input Base Rates for ND/H2O").place(x=410, y=10)
    
    tk.Label(englander_window, text="log10_k_acid_ref:").place(x=410, y=40)
    log10_k_acid_ref_enter = tk.Entry(englander_window, width=6)
    log10_k_acid_ref_enter.place(x=510, y=40)
    log10_k_acid_ref_enter.insert(0, "1.4")  
    
    tk.Label(englander_window, text="log10_k_base_ref:").place(x=410, y=70)
    log10_k_base_ref_enter = tk.Entry(englander_window, width=6)
    log10_k_base_ref_enter.place(x=510, y=70)
    log10_k_base_ref_enter.insert(0, "10") 
    
    tk.Label(englander_window, text="log10_k_water_ref:").place(x=410, y=100)
    log10_k_water_ref_enter = tk.Entry(englander_window, width=6)
    log10_k_water_ref_enter.place(x=510, y=100)
    log10_k_water_ref_enter.insert(0, "-1.6") 
    

    save_baserate_bt = tk.Button(englander_window, text="Save Values", command=lambda: save_baserate_rates(log10_k_acid_ref_enter.get(), log10_k_base_ref_enter.get(), log10_k_water_ref_enter.get(), baserate_json_dropdown))
    save_baserate_bt.place(x=410, y=130)
    
    tk.Label(englander_window, text="Englander Param File:").place(x=100, y=170)
    englander_jsons_list = os.listdir("englander_jsons")
    englander_json_dropdown = ttk.Combobox(englander_window, values=englander_jsons_list, width=32)
    if os.path.exists("Alterations/englander_default_json.json"):
        with open("Alterations/englander_default_json.json", "r") as file:
            data = json.load(file)
            file_path = data.get("FilePath", "")  # Get the value, default to an empty string if missing
            englander_json_dropdown.set(file_path)
    englander_json_dropdown.place(x=250, y=170)
    
    set_as_default_bt = tk.Button(englander_window, text="Set as default", command=lambda: set_englander_default(englander_json_dropdown.get()))
    set_as_default_bt.place(x=470, y=165)
    
    tk.Label(englander_window, text="Base Rate File:").place(x=100, y=200)
    baserate_jsons_list = os.listdir("baserate_jsons")
    baserate_json_dropdown = ttk.Combobox(englander_window, values=baserate_jsons_list, width=32)
    if os.path.exists("Alterations/baserate_default_json.json"):
        with open("Alterations/baserate_default_json.json", "r") as file:
            data = json.load(file)
            file_path = data.get("FilePath", "")  # Get the value, default to an empty string if missing
            baserate_json_dropdown.set(file_path)
    baserate_json_dropdown.place(x=250, y=200)
    
    set_br_as_default_bt = tk.Button(englander_window, text="Set as default", command=lambda: set_baserate_default(baserate_json_dropdown.get()))
    set_br_as_default_bt.place(x=470, y=195)
    

    englander_out_bt = tk.Button(englander_window, text="OK", command=lambda: close_englander_window(englander_json_dropdown.get(), baserate_json_dropdown.get()))
    englander_out_bt.place(x=350, y=230)
    
    
def set_englander_default(small_file_path):
    if os.path.exists("Alterations/englander_default_json.json"):
        os.remove("Alterations/englander_default_json.json")
    with open("Alterations/englander_default_json.json", 'w') as file:
        json.dump({"FilePath": small_file_path}, file)   
        
def set_baserate_default(small_file_path):
    if os.path.exists("Alterations/baserate_default_json.json"):
        os.remove("Alterations/baserate_default_json.json")
    with open("Alterations/baserate_default_json.json", 'w') as file:
        json.dump({"FilePath": small_file_path}, file)  
    
def save_englander_rates(pH, tempK, trap, constant, englander_json_dropdown):
    try:
        a, b, c, d = float(pH), float(tempK), float(trap), float(constant)
    except:
        tk.messagebox.showerror("Value Error", "Please Make sure all inputs are numbers and try again")
        return
    data = {
        "englander_pH": pH,
        "englander_tempK": tempK,
        "englander_time_in_H2O_trap": trap,
        "englander_time_in_H2O_constant": constant
    }
    
    filename = f"englander_jsons/pH{pH}_T{tempK}_t{trap}plus({constant})(RT).json"
    with open(filename, "w") as file:
        json.dump(data, file, indent=4)
    tk.messagebox.showinfo("File Saved", f"File Saved as {filename}")
    
    englander_jsons_list = os.listdir("englander_jsons")
    englander_json_dropdown["values"] = englander_jsons_list
    englander_json_dropdown.set(filename.split("/")[1])
    

def save_baserate_rates(log10_k_acid_ref, log10_k_base_ref, log10_k_water_ref, baserate_json_dropdown):
    try:
        a, b, c = float(log10_k_acid_ref), float(log10_k_base_ref), float(log10_k_water_ref)
    except:
        tk.messagebox.showerror("Value Error", "Please Make sure all inputs are numbers and try again")
        return
    data = {
        "log10_k_acid_ref": log10_k_acid_ref,
        "log10_k_base_ref": log10_k_base_ref,
        "log10_k_water_ref": log10_k_water_ref,
    }
    
    filename = f"baserate_jsons/a{log10_k_acid_ref} b{log10_k_base_ref} w{log10_k_water_ref}.json"
    with open(filename, "w") as file:
        json.dump(data, file, indent=4)
    tk.messagebox.showinfo("File Saved", f"File Saved as {filename}")
    
    baserate_jsons_list = os.listdir("baserate_jsons")
    baserate_json_dropdown["values"] = baserate_jsons_list
    baserate_json_dropdown.set(filename.split("/")[1])

    
def close_englander_window(small_file_path, br_file_path): 
    global calculate_theoretical_back_exchange, englander_time_in_H2O_trap, englander_time_in_H2O_constant, englander_pH, englander_tempK, log10_k_acid_ref, log10_k_base_ref, log10_k_water_ref
    
    englander_path = "englander_jsons/" + small_file_path
    br_path = "baserate_jsons/" + br_file_path
    try:
        with open(englander_path, "r") as file:
            pass
    except:
        tk.messagebox.showerror("Error", "Error, please make sure to providea paramter file before exiting")
        return
    try:
        with open(br_path, "r") as file:
            pass
    except:
        tk.messagebox.showerror("Error", "Error, please make sure to providea baserate file before exiting")
        return
    
    with open(englander_path, "r") as file:
        data = json.load(file)
        englander_time_in_H2O_trap = float(data["englander_time_in_H2O_trap"])
        englander_time_in_H2O_constant = float(data["englander_time_in_H2O_constant"])
        englander_pH = float(data["englander_pH"])
        englander_tempK = float(data["englander_tempK"])
        
    with open(br_path, "r") as file:
        data = json.load(file)
        log10_k_acid_ref = float(data["log10_k_acid_ref"])
        log10_k_base_ref = float(data["log10_k_base_ref"])
        log10_k_water_ref = float(data["log10_k_water_ref"])
    
    englander_window.destroy()  # Close the window
    if os.path.exists("Alterations/BoolEnglander.json"):
        pass
    else:
        user_choice = tk.messagebox.askyesno("Default Changes", "Do you want to make this the default value?", default='no')
        dont_show_path = "Alterations/BoolEnglander.json"
        if user_choice:
            with open(dont_show_path, 'w') as file:
                json.dump({"BoolEnglanderJson": 3}, file)
        else:
            with open(dont_show_path, 'w') as file:
                json.dump({"BoolEnglanderJson": 1}, file) 

    

        
        
    theo_calc_bt = tk.Button(window, text="Englander Rates", bg="green", fg="white", command=lambda: [theo_calculation_off(), n_min_two_off(), n_min_one_on()])
    theo_calc_bt.place(x=265, y=250)
    calculate_theoretical_back_exchange = True
    
    
    
def theo_calculation_off():
    global n_min_one, n_min_two, calculate_theoretical_back_exchange, englander_window
    theo_calc_bt = tk.Button(window, text="Englander Rates", bg="orange", fg="black", command=lambda: [n_min_one_off(), n_min_two_off(), theo_calculation_on()])
    theo_calc_bt.place(x=265, y=250)
    
    englander_window.destroy()
    calculate_theoretical_back_exchange = False
    

def check_button_clicks():
    global difmap_bt_on, pepmap_bt_on, chic_bt_on, cdif_bt_on, condpeps_bt_on, difcond_bt_on, uptake_plot_bt_on, heatmap_bt_on    
    if (sdbt_clicked or cdbt_clicked or examiner_clicked or workbench_clicked) and (seqbt_txt_clicked or seqbt_fasta_clicked or skip_bt_clicked or txt_h_bt_clicked):
        exp_bt_off()
        theo_bt_off()
        msg1 = tk.Label(window, text="RFU Calculation and Correction")
        msg1.place(x=15, y=190)
        exp_bt = tk.Button(window, text="MaxD Corrected",bg="orange",fg="black",command=lambda: [theo_bt_off(), exp_bt_on()])
        exp_bt.place(x=150, y=220)
        theo_bt = tk.Button(window, text="No maxD",bg="orange",fg="black",command=lambda: [exp_bt_off(), theo_bt_on()])
        theo_bt.place(x=50, y=220)

        x1 = 10
        y1 = 182
        x2 = 370
        y2 = 880
        canvas.create_rectangle(x1, y1, x2, y2, outline="black", fill="")
        
        global n_min_one, n_min_two, calculate_theoretical_back_exchange
        n_min_one = True
        n_min_two = False
        calculate_theoretical_back_exchange = False
        n_one_vs_n_two_lab = tk.Label(window, text="Deuterium Incorporation Limit:")
        n_one_vs_n_two_lab.place(x=15, y=250)
        n_min_one_bt = tk.Button(window, text="N-1", bg="green", fg="white", command=lambda: [n_min_one_off(), n_min_two_on(), theo_calculation_off()])
        n_min_one_bt.place(x=185, y=250)
        n_min_two_bt = tk.Button(window, text="N-2", bg="orange", fg="black", command=lambda: [n_min_one_off(), n_min_two_on(), theo_calculation_off()])
        n_min_two_bt.place(x=225, y=250)
        theo_calc_bt = tk.Button(window, text="Englander Rates", bg="orange", fg="black", command=lambda: [n_min_one_off(), n_min_two_off(), theo_calculation_on()])
        theo_calc_bt.place(x=265, y=250)
        
        
        if os.path.exists("Alterations/Min2.json"):
            with open("Alterations/Min2.json", 'r') as file:
                json_data = json.load(file)
                Min2Json = int(float(json_data["Min2Json"]))
                if Min2Json == 2:
                    n_min_two_on()
                    n_min_one_off()
                    theo_calculation_off()
        if os.path.exists("Alterations/BoolEnglander.json"):
            with open("Alterations/BoolEnglander.json", 'r') as file:
                json_data = json.load(file)
                BoolEnglander = int(float(json_data["BoolEnglanderJson"]))
                if BoolEnglander == 3:
                    n_min_one_off()
                    n_min_two_off()
                    theo_calculation_on()
                    
        if os.path.exists("Alterations/Min2.json") and os.path.exists("Alterations/BoolEnglander.json"):
            mtime1 = os.path.getmtime("Alterations/Min2.json")
            mtime2 = os.path.getmtime("Alterations/BoolEnglander.json")
            
            if mtime1 > mtime2:
                with open("Alterations/Min2.json", 'r') as file:
                    json_data = json.load(file)
                    Min2Json = int(float(json_data["Min2Json"]))
                    if Min2Json == 2:
                        n_min_two_on()
                        n_min_one_off()
                        theo_calculation_off()
            else:
                pass #BoolEnglander ran most recently anyways
                
                
                
        
        difmap_bt_on = False
        pepmap_bt_on = False
        chic_bt_on = False
        cdif_bt_on = False
        condpeps_bt_on = False
        difcond_bt_on = False
        uptake_plot_bt_on = False
        heatmap_bt_on = False
        
        for widget in window.winfo_children():
            if widget.winfo_x() > 370 and widget != settings_bt:
                widget.destroy()
        for item in canvas.find_all():
            coords = canvas.coords(item)
            # For lines and shapes, coords are a list of x, y pairs. We check the first x-coordinate.
            if coords and coords[0] > 370:
                canvas.delete(item)
        
        if txt_h_bt_clicked is True or seqbt_fasta_clicked is True:
            protein_count = 0
            missing_protein_count = 0
            missing_protein_list = []
            for protein in initial_protein_set:
                protein_count += 1
                if protein not in prot_seq_dic.keys():
                    missing_protein_count += 1
                    missing_protein_list.append(protein)
            if missing_protein_count > 0:
                tk.messagebox.showinfo("Protein Sequences Not Found", f"For {protein_count} sequences from your data files, {missing_protein_count} still do not have sequences assigned. They are: \n\n {missing_protein_list}. \n\n Your provided protein titles are: \n\n {list(prot_seq_dic.keys())} \n\n You can move on without providing a sequence or add more sequences. Make sure the protein names match.")
                    
                    



x1, y1 = 10, 10  # Top-left coordinates of the rectangle
x2, y2 = 370, 180  # Bottom-right coordinates of the rectangle
canvas.create_rectangle(x1, y1, x2, y2, outline="black", fill="")

def increase_progress(x):
    progress['value'] += x
    window.update()

def start_progress():
    global progress
    if reduce_states_var.get() == 0:
        states_to_look_in = statedic_of_pepdic_cor
    if reduce_states_var.get() == 1:
        states_to_look_in = order_state_dic.values()
        states_to_look_in = [x for x in states_to_look_in if x != False]
    pmax = 1
    if pepmap_bt_on:
        pmax = pmax + len(states_to_look_in)
    if difmap_bt_on:
        pmax = pmax + 1.5*len(new_dic_of_dif_list) + 1
    if chic_bt_on:
        pmax = pmax+0.33
    if cdif_bt_on:
        pmax=pmax + 0.33
    if condpeps_bt_on:
        pmax = pmax + len(states_to_look_in)
    if difcond_bt_on:
        pmax = pmax + len(new_dic_of_dif_list)
    if uptake_plot_bt_on:
        pmax = pmax + 2
    style = ttk.Style()
    style.theme_use('clam')
    style.configure("blue.Horizontal.TProgressbar", foreground='blue', background='blue')
    progress = ttk.Progressbar(window, style='blue.Horizontal.TProgressbar', orient='horizontal', mode='determinate', length=200, maximum=pmax)
    progress.place(x=1270, y=200, width=200, height=25)  # Position the progress bar at the bottom left
    window.update()

difmap_bt_on = False
pepmap_bt_on = False
chic_bt_on = False
cdif_bt_on = False
condpeps_bt_on = False
difcond_bt_on = False
uptake_plot_bt_on = False
heatmap_bt_on = False
def difmap_on():
    global difmap_bt_on
    difmap_bt_2 = tk.Button(window, text="Peptide Difference",bg="green",fg="white",width=17, command=lambda: (difmap_off(), heatmap_off()))
    difmap_bt_2.place(x=1340,y=80)
    difmap_bt_on = True
def difmap_off():
    global difmap_bt_on
    difmap_bt = tk.Button(window, text="Peptide Difference",bg="orange",fg="black",width=17, command=difmap_on)
    difmap_bt.place(x=1340,y=80)
    difmap_bt_on = False
def pepmap_on():
    global pepmap_bt_on
    pepmap_bt_2 = tk.Button(window, text="Peptide Plot",bg="green",fg="white",width=17, command=pepmap_off)
    pepmap_bt_2.place(x=1190,y=80)
    pepmap_bt_on = True
def pepmap_off():
    global pepmap_bt_on
    pepmap_bt = tk.Button(window, text="Peptide Plot",bg="orange",fg="black",width=17, command=pepmap_on)
    pepmap_bt.place(x=1190,y=80)
    pepmap_bt_on = False
def chiclet_on():
    global chic_bt_on
    chiclet_bt_2 = tk.Button(window, text="Chiclet Plot",bg="green",fg="white",width=17, command=chiclet_off)
    chiclet_bt_2.place(x=1190,y=40)
    chic_bt_on = True
def chiclet_off():
    global chic_bt_on
    chic_bt = tk.Button(window, text="Chiclet Plot",bg="orange",fg="black", width=17, command=chiclet_on)
    chic_bt.place(x=1190,y=40)
    chic_bt_on = False
def cdif_on():
    global cdif_bt_on
    cdif_bt_2 = tk.Button(window, text="Chiclet Difference",bg="green",fg="white",width=17, command=cdif_off)
    cdif_bt_2.place(x=1340,y=40)
    cdif_bt_on = True
def cdif_off():
    global cdif_bt_on
    cdif_bt = tk.Button(window, text="Chiclet Difference",bg="orange",fg="black",width=17, command=cdif_on)
    cdif_bt.place(x=1340,y=40)
    cdif_bt_on = False
def condpeps_on():
    global condpeps_bt_on
    condpeps_bt_2 = tk.Button(window, text="Condensed Peptide",bg="green",fg="white",width=17, command=condpeps_off)
    condpeps_bt_2.place(x=1190,y=120)
    condpeps_bt_on = True
def condpeps_off():
    global condpeps_bt_on
    condpeps_bt = tk.Button(window, text="Condensed Peptide",bg="orange",fg="black",width=17, command=condpeps_on)
    condpeps_bt.place(x=1190,y=120)
    condpeps_bt_on = False
def difcond_on():
    global difcond_bt_on
    difcond_bt_2 = tk.Button(window, text="Condensed Difference",bg="green",fg="white",width=17, command=lambda: (difcond_off(), heatmap_off()))
    difcond_bt_2.place(x=1340,y=120)
    difcond_bt_on = True
def difcond_off():
    global difcond_bt_on
    difcond_bt = tk.Button(window, text="Condensed Difference",bg="orange",fg="black",width=17, command=difcond_on)
    difcond_bt.place(x=1340,y=120)
    difcond_bt_on = False
def uptake_plot_on():
    global uptake_plot_bt_on
    uptake_plot_bt = tk.Button(window, text="Uptake Plots",bg="green", fg="white", width=17, command=uptake_plot_off)
    uptake_plot_bt.place(x=1190, y=160)
    uptake_plot_bt_on = True
def uptake_plot_off():
    global uptake_plot_bt_on
    uptake_plot_bt = tk.Button(window, text="Uptake Plots",bg="orange", fg="black", width=17, command=uptake_plot_on)
    uptake_plot_bt.place(x=1190, y=160)
    uptake_plot_bt_on = False
def heatmap_on():
    global heatmap_bt_on
    heatmap_bt = tk.Button(window, text="Localized Differences",bg="green", fg="white", width=17, command=heatmap_off)
    heatmap_bt.place(x=1340, y=160)
    heatmap_bt_on = True
def heatmap_off():
    global heatmap_bt_on
    heatmap_bt = tk.Button(window, text="Localized Differences", bg="orange", fg="black", width=17, command=lambda: (heatmap_on(), difcond_on(), difmap_on()))
    heatmap_bt.place(x=1340, y=160)
    heatmap_bt_on = False

    
    
    
def on_closing_custom_colors():
    global custom_colors_open
    custom_colors_open = False
    popup_window_uptake.destroy()
    
custom_colors_open = False
def create_custom_colors():
    global custom_colors_open, popup_window_uptake
    if custom_colors_open:
        user_choice = tk.messagebox.askyesno("Custom Colors", "Create Custom Colors may already be open. Do you want to close and open a new window?", default='no')
        if user_choice:
            custom_colors_open = False
            popup_window_uptake.destroy()
        else:
            popup_window_uptake.lift()
            return
    
        
    def show_examples():
        try:
            #os.startfile("Creating Custom Color Schemes.pdf")
            open_file("Creating Custom Color Schemes.pdf")
        except:
            tk.messagebox.showerror("Error", "Cannot find example file")
            
    popup_window_uptake = tk.Toplevel(window)  # Create a new window for the popup menu
    popup_window_uptake.geometry("1050x500")
    custom_colors_open = True
    popup_window_uptake.protocol("WM_DELETE_WINDOW", on_closing_custom_colors)
    canvas = tk.Canvas(popup_window_uptake, width=1050, height=500)
    canvas.place(x=0, y=0)
    x1 = 5
    y1 = 5
    x2 = 345
    y2 = 495
    canvas.create_rectangle(x1, y1, x2, y2, outline="black", fill="")
    tk.Label(popup_window_uptake, text="Create Custom Colors for All Uptake Maps").place(x=20, y=1)
    tk.Label(popup_window_uptake, text="Enter RFU as a decimal, with the highest exchanging color first").place(x=10, y=25)
    tk.Label(popup_window_uptake, text="Hexadecimal Color").place(x=115, y=50)
    tk.Label(popup_window_uptake, text="White Text").place(x=235, y=50)


    lab1 = tk.Label(popup_window_uptake, text="If RFU >           :")
    lab1.place(x=20, y=70)
    lab2 = tk.Label(popup_window_uptake, text="If RFU >           :")
    lab2.place(x=20, y=100)
    lab3 = tk.Label(popup_window_uptake, text="If RFU >           :")
    lab3.place(x=20, y=130)
    lab4 = tk.Label(popup_window_uptake, text="If RFU >           :")
    lab4.place(x=20, y=160)
    lab5 = tk.Label(popup_window_uptake, text="If RFU >           :")
    lab5.place(x=20, y=190)
    lab6 = tk.Label(popup_window_uptake, text="If RFU >           :")
    lab6.place(x=20, y=220)
    lab7 = tk.Label(popup_window_uptake, text="If RFU >           :")
    lab7.place(x=20, y=250)
    lab8 = tk.Label(popup_window_uptake, text="If RFU >           :")
    lab8.place(x=20, y=280)
    lab9 = tk.Label(popup_window_uptake, text="If RFU >           :")
    lab9.place(x=20, y=310)
    lab10 = tk.Label(popup_window_uptake, text="If RFU >  0       :")
    lab10.place(x=20, y=340)
    lab11 = tk.Label(popup_window_uptake, text="If RFU <  0       :")
    lab11.place(x=20, y=370)
    lab12 = tk.Label(popup_window_uptake, text="If RFU =  0       :")
    lab12.place(x=20, y=400)
    lab13 = tk.Label(popup_window_uptake, text="If Peptide Absent :")
    lab13.place(x=20, y=430)

    val1 = tk.Entry(popup_window_uptake, width=4)
    val1.place(x=67, y=70)
    val2 = tk.Entry(popup_window_uptake, width=4)
    val2.place(x=67, y=100)
    val3 = tk.Entry(popup_window_uptake, width=4)
    val3.place(x=67, y=130)
    val4 = tk.Entry(popup_window_uptake, width=4)
    val4.place(x=67, y=160)
    val5 = tk.Entry(popup_window_uptake, width=4)
    val5.place(x=67, y=190)
    val6 = tk.Entry(popup_window_uptake, width=4)
    val6.place(x=67, y=220)
    val7 = tk.Entry(popup_window_uptake, width=4)
    val7.place(x=67, y=250)
    val8 = tk.Entry(popup_window_uptake, width=4)
    val8.place(x=67, y=280)
    val9 = tk.Entry(popup_window_uptake, width=4)
    val9.place(x=67, y=310)


    col1 = tk.Entry(popup_window_uptake, width=8)
    col1.place(x=140, y=70)
    col2 = tk.Entry(popup_window_uptake, width=8)
    col2.place(x=140, y=100)
    col3 = tk.Entry(popup_window_uptake, width=8)
    col3.place(x=140, y=130)
    col4 = tk.Entry(popup_window_uptake, width=8)
    col4.place(x=140, y=160)
    col5 = tk.Entry(popup_window_uptake, width=8)
    col5.place(x=140, y=190)
    col6 = tk.Entry(popup_window_uptake, width=8)
    col6.place(x=140, y=220)
    col7 = tk.Entry(popup_window_uptake, width=8)
    col7.place(x=140, y=250)
    col8 = tk.Entry(popup_window_uptake, width=8)
    col8.place(x=140, y=280)
    col9 = tk.Entry(popup_window_uptake, width=8)
    col9.place(x=140, y=310)
    col10 = tk.Entry(popup_window_uptake, width=8)
    col10.place(x=140, y=340)
    col11 = tk.Entry(popup_window_uptake, width=8)
    col11.place(x=140, y=370)
    col12 = tk.Entry(popup_window_uptake, width=8)
    col12.place(x=140, y=400)
    col12.insert(0, "F2F2F2")
    col13 = tk.Entry(popup_window_uptake, width=8)
    col13.place(x=140, y=430)
    col13.insert(0, "FAE8D7")

    global chkval_1, chkval_2, chkval_3, chkval_4, chkval_5, chkval_6, chkval_7, chkval_8, chkval_9, chkval_10, chkval_11, chkval_12, chkval_13
    chkval_1 = tk.IntVar(value=1)
    txtchk1 = tk.Checkbutton(popup_window_uptake, text='', variable=chkval_1)
    txtchk1.place(x=255, y=65)

    chkval_2 = tk.IntVar(value=1)
    txtchk2 = tk.Checkbutton(popup_window_uptake, text='', variable=chkval_2)
    txtchk2.place(x=255, y=95)

    chkval_3 = tk.IntVar(value=1)
    txtchk3 = tk.Checkbutton(popup_window_uptake, text='', variable=chkval_3)
    txtchk3.place(x=255, y=125)

    chkval_4 = tk.IntVar(value=1)
    txtchk4 = tk.Checkbutton(popup_window_uptake, text='', variable=chkval_4)
    txtchk4.place(x=255, y=155)

    chkval_5 = tk.IntVar(value=1)
    txtchk5 = tk.Checkbutton(popup_window_uptake, text='', variable=chkval_5)
    txtchk5.place(x=255, y=185)

    chkval_6 = tk.IntVar(value=1)
    txtchk6 = tk.Checkbutton(popup_window_uptake, text='', variable=chkval_6)
    txtchk6.place(x=255, y=215)

    chkval_7 = tk.IntVar(value=1)
    txtchk7 = tk.Checkbutton(popup_window_uptake, text='', variable=chkval_7)
    txtchk7.place(x=255, y=245)

    chkval_8 = tk.IntVar(value=1)
    txtchk8 = tk.Checkbutton(popup_window_uptake, text='', variable=chkval_8)
    txtchk8.place(x=255, y=275)

    chkval_9 = tk.IntVar(value=1)
    txtchk9 = tk.Checkbutton(popup_window_uptake, text='', variable=chkval_9)
    txtchk9.place(x=255, y=305)

    chkval_10 = tk.IntVar(value=1)
    txtchk10 = tk.Checkbutton(popup_window_uptake, text='', variable=chkval_10)
    txtchk10.place(x=255, y=335)

    chkval_11 = tk.IntVar(value=1)
    txtchk11 = tk.Checkbutton(popup_window_uptake, text='', variable=chkval_11)
    txtchk11.place(x=255, y=365)

    chkval_12 = tk.IntVar(value=1)
    txtchk12 = tk.Checkbutton(popup_window_uptake, text='', variable=chkval_12)
    txtchk12.place(x=255, y=395)

    chkval_13 = tk.IntVar(value=1)
    txtchk13 = tk.Checkbutton(popup_window_uptake, text='', variable=chkval_13)
    txtchk13.place(x=255, y=425)

    def save_colors():
        vals = []
        if val1.get() != "":
            vals.append(val1.get())
        if val2.get() != "":
            vals.append(val2.get())
        if val3.get() != "":
            vals.append(val3.get())
        if val4.get() != "":
            vals.append(val4.get())
        if val5.get() != "":
            vals.append(val5.get())
        if val6.get() != "":
            vals.append(val6.get())
        if val7.get() != "":
            vals.append(val7.get())
        if val8.get() != "":
            vals.append(val8.get())
        if val9.get() != "":
            vals.append(val9.get())
        potential_val_error_1 = False
        potential_val_error_2 = False
        for i, val in enumerate(vals):
            if float(val) > 1:
                potential_val_error_2 = True
                break
            try:
                next_val = vals[i + 1]
            except:
                continue
            if next_val > val:
                potential_val_error_1 = True
                break


        if potential_val_error_1 == True:
            tk.messagebox.showwarning("Potential Error Found", f"{next_val} > {val}. This may cause an error when running the program. Please list RFU criticial points in descending order")
        if potential_val_error_2 == True:
            tk.messagebox.showwarning("Potential Error Found", f"{val} > 1. This may cause an error when running the program. Please list RFUs as a decimal.")
            tk.Label(popup_window_uptake, text=f"Potential Error found. {val} > 1")


        uptake_color_dic = {}
        if val1.get() != "" and col1.get() != "":
            uptake_color_dic[val1.get()] = col1.get()
        if val2.get() != "" and col2.get() != "":
            uptake_color_dic[val2.get()] = col2.get()
        if val3.get() != "" and col3.get() != "":
            uptake_color_dic[val3.get()] = col3.get()
        if val4.get() != "" and col4.get() != "":
            uptake_color_dic[val4.get()] = col4.get()
        if val5.get() != "" and col5.get() != "":
            uptake_color_dic[val5.get()] = col5.get()
        if val6.get() != "" and col6.get() != "":
            uptake_color_dic[val6.get()] = col6.get()
        if val7.get() != "" and col7.get() != "":
            uptake_color_dic[val7.get()] = col7.get()
        if val8.get() != "" and col8.get() != "":
            uptake_color_dic[val8.get()] = col8.get()
        if val9.get() != "" and col9.get() != "":
            uptake_color_dic[val9.get()] = col9.get()
        if col10.get() != "":
            uptake_color_dic[">0"] = col10.get()
        else:
            uptake_color_dic[">0"] = "000000"
        if col11.get() != "":
            uptake_color_dic["<0"] = col11.get()
        else:
            uptake_color_dic["<0"] = "000000"
        if col12.get() != "":
            uptake_color_dic["=0"] = col12.get()
        else:
            uptake_color_dic["=0"] = "F2F2F2"
        if col13.get() != "":
            uptake_color_dic[-99999] = col13.get()
        else:
            uptake_color_dic[-99999] = "FAE8D7"


        uptake_text_dic = {}
        if val1.get() != "" and col1.get != "":
            uptake_text_dic[val1.get()] = chkval_1.get()
        if val2.get() != "" and col2.get != "":
            uptake_text_dic[val2.get()] = chkval_2.get()
        if val3.get() != "" and col3.get != "":
            uptake_text_dic[val3.get()] = chkval_3.get()
        if val4.get() != "" and col4.get != "":
            uptake_text_dic[val4.get()] = chkval_4.get()
        if val5.get() != "" and col5.get() != "":
            uptake_text_dic[val5.get()] = chkval_5.get()
        if val6.get() != "" and col6.get() != "":
            uptake_text_dic[val6.get()] = chkval_6.get()
        if val7.get() != "" and col7.get() != "":
            uptake_text_dic[val7.get()] = chkval_7.get()
        if val8.get() != "" and col8.get() != "":
            uptake_text_dic[val8.get()] = chkval_8.get()
        if val9.get() != "" and col9.get() != "":
            uptake_text_dic[val9.get()] = chkval_9.get()
        uptake_text_dic[">0"] = chkval_10.get()
        uptake_text_dic["<0"] = chkval_11.get()
        uptake_text_dic["=0"] = chkval_12.get()
        uptake_text_dic[-99999] = chkval_13.get()
        json_data = {
            "header": "Uptake Colors",
            "uptake_color_dic": uptake_color_dic,
            "uptake_text_dic": uptake_text_dic
        }
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".json",
            initialdir="./Colors",
            title="Save File",
            initialfile="uptake_new_colors",
            filetypes=[("JSON files", "*.json")]
        )
        if file_path:
            with open(file_path, 'w') as f:
                json.dump(json_data, f, indent=4)
                title = os.path.basename(file_path)
                tk.messagebox.showinfo("File Saved", f"File Saved as {title}")
                
            uptake_file_names, dif_file_names, local_file_names = update_dir_lists()
            update_color_comboboxes(uptake_file_names, dif_file_names, local_file_names)
            popup_window_uptake.focus_set()
            
        else:
            tk.messagebox.showerror("File Not Saved", "File was not saved")
            popup_window_uptake.focus_set()
        




    save_bt_uptake = tk.Button(popup_window_uptake, text = "Save Colors", command=save_colors)
    save_bt_uptake.place(x=20, y=460)


    x1 = 347
    y1 = 5
    x2 = 915 + 130
    y2 = 495
    canvas.create_rectangle(x1, y1, x2, y2, outline="black", fill="")
    tk.Label(popup_window_uptake, text="Create Custom Colors for All Difference Maps").place(x=370, y=1)
    tk.Label(popup_window_uptake, text="Enter Difference with the highest absolute value differences first. For RFU enter as a decimal").place(x=360, y=40)
    tk.Label(popup_window_uptake, text="Protection").place(x=440, y=65)
    x1=352
    y=89
    x2=625
    canvas.create_line(x1, y, x2, y)
    tk.Label(popup_window_uptake, text="Deprotection").place(x=720, y=65)
    x1 = 635
    y=89
    x2=895
    canvas.create_line(x1, y, x2, y)
    tk.Label(popup_window_uptake, text="Hexadecimal Color").place(x=445, y=90)
    tk.Label(popup_window_uptake, text="White Text").place(x=565, y=90)
    tk.Label(popup_window_uptake, text="Hexadecimal Color").place(x=727, y=90)
    tk.Label(popup_window_uptake, text="White Text").place(x=847, y=90)
    examples_bt = tk.Button(popup_window_uptake, text="See examples", bg='green', fg='white', command=show_examples)
    examples_bt.place(x=925, y=15)


    plab1 = tk.Label(popup_window_uptake, text="If dif >           :")
    plab1.place(x=360, y=120)
    plab2 = tk.Label(popup_window_uptake, text="If dif >           :")
    plab2.place(x=360, y=150)
    plab3 = tk.Label(popup_window_uptake, text="If dif >           :")
    plab3.place(x=360, y=180)
    plab4 = tk.Label(popup_window_uptake, text="If dif >           :")
    plab4.place(x=360, y=210)
    plab5 = tk.Label(popup_window_uptake, text="If dif >           :")
    plab5.place(x=360, y=240)
    plab6 = tk.Label(popup_window_uptake, text="If dif >  0        :")
    plab6.place(x=360, y=270)

    dlab1 = tk.Label(popup_window_uptake, text="If dif >           :")
    dlab1.place(x=640, y=120)
    dlab2 = tk.Label(popup_window_uptake, text="If dif >           :")
    dlab2.place(x=640, y=150)
    dlab3 = tk.Label(popup_window_uptake, text="If dif >           :")
    dlab3.place(x=640, y=180)
    dlab4 = tk.Label(popup_window_uptake, text="If dif >           :")
    dlab4.place(x=640, y=210)
    dlab5 = tk.Label(popup_window_uptake, text="If dif >           :")
    dlab5.place(x=640, y=240)
    dlab6 = tk.Label(popup_window_uptake, text="If dif >  0        :")
    dlab6.place(x=640, y=270)

    blab1 = tk.Label(popup_window_uptake, text="If dif =  0       :")
    blab1.place(x=360, y=300)
    blab2 = tk.Label(popup_window_uptake, text="If Peptide Absent :")
    blab2.place(x=360, y=330)

    pval1 = tk.Entry(popup_window_uptake, width=4)
    pval1.place(x=401, y=120)
    pval2 = tk.Entry(popup_window_uptake, width=4)
    pval2.place(x=401, y=150)
    pval3 = tk.Entry(popup_window_uptake, width=4)
    pval3.place(x=401, y=180)
    pval4 = tk.Entry(popup_window_uptake, width=4)
    pval4.place(x=401, y=210)
    pval5 = tk.Entry(popup_window_uptake, width=4)
    pval5.place(x=401, y=240)

    dval1 = tk.Entry(popup_window_uptake, width=4)
    dval1.place(x=681, y=120)
    dval2 = tk.Entry(popup_window_uptake, width=4)
    dval2.place(x=681, y=150)
    dval3 = tk.Entry(popup_window_uptake, width=4)
    dval3.place(x=681, y=180)
    dval4 = tk.Entry(popup_window_uptake, width=4)
    dval4.place(x=681, y=210)
    dval5 = tk.Entry(popup_window_uptake, width=4)
    dval5.place(x=681, y=240)

    pcol1 = tk.Entry(popup_window_uptake, width=8)
    pcol1.place(x=474, y=120)
    pcol2 = tk.Entry(popup_window_uptake, width=8)
    pcol2.place(x=474, y=150)
    pcol3 = tk.Entry(popup_window_uptake, width=8)
    pcol3.place(x=474, y=180)
    pcol4 = tk.Entry(popup_window_uptake, width=8)
    pcol4.place(x=474, y=210)
    pcol5 = tk.Entry(popup_window_uptake, width=8)
    pcol5.place(x=474, y=240)
    pcol6 = tk.Entry(popup_window_uptake, width=8)
    pcol6.place(x=474, y=270)
    pcol6.insert(0, "F2F2F2")

    bcol1 = tk.Entry(popup_window_uptake, width=8)
    bcol1.place(x=474, y=300)
    bcol1.insert(0, "F2F2F2")
    bcol2 = tk.Entry(popup_window_uptake, width=8)
    bcol2.place(x=474, y=330)
    bcol2.insert(0, "FAE8D7")

    dcol1 = tk.Entry(popup_window_uptake, width=8)
    dcol1.place(x=754, y=120)
    dcol2 = tk.Entry(popup_window_uptake, width=8)
    dcol2.place(x=754, y=150)
    dcol3 = tk.Entry(popup_window_uptake, width=8)
    dcol3.place(x=754, y=180)
    dcol4 = tk.Entry(popup_window_uptake, width=8)
    dcol4.place(x=754, y=210)
    dcol5 = tk.Entry(popup_window_uptake, width=8)
    dcol5.place(x=754, y=240)
    dcol6 = tk.Entry(popup_window_uptake, width=8)
    dcol6.place(x=754, y=270)
    dcol6.insert(0, "F2F2F2")

    global pchkval_1, pchkval_2, pchkval_3, pchkval_4, pchkval_5, pchkval_6, bchkval_1, bchkval_2, dchkval_1, dchkval_2, dchkval_3, dchkval_4, dchkval_5, dchkval_6
    pchkval_1 = tk.IntVar(value=1)
    ptxtchk1 = tk.Checkbutton(popup_window_uptake, text='', variable=pchkval_1)
    ptxtchk1.place(x=589, y=115)

    pchkval_2 = tk.IntVar(value=1)
    ptxtchk2 = tk.Checkbutton(popup_window_uptake, text='', variable=pchkval_2)
    ptxtchk2.place(x=589, y=145)

    pchkval_3 = tk.IntVar(value=1)
    ptxtchk3 = tk.Checkbutton(popup_window_uptake, text='', variable=pchkval_3)
    ptxtchk3.place(x=589, y=175)

    pchkval_4 = tk.IntVar(value=1)
    ptxtchk4 = tk.Checkbutton(popup_window_uptake, text='', variable=pchkval_4)
    ptxtchk4.place(x=589, y=205)

    pchkval_5 = tk.IntVar(value=1)
    ptxtchk5 = tk.Checkbutton(popup_window_uptake, text='', variable=pchkval_5)
    ptxtchk5.place(x=589, y=235)

    pchkval_6 = tk.IntVar(value=0)
    ptxtchk6 = tk.Checkbutton(popup_window_uptake, text='', variable=pchkval_6)
    ptxtchk6.place(x=589, y=265)

    dchkval_1 = tk.IntVar(value=1)
    dtxtchk1 = tk.Checkbutton(popup_window_uptake, text='', variable=dchkval_1)
    dtxtchk1.place(x=869, y=115)

    dchkval_2 = tk.IntVar(value=1)
    dtxtchk2 = tk.Checkbutton(popup_window_uptake, text='', variable=dchkval_2)
    dtxtchk2.place(x=869, y=145)

    dchkval_3 = tk.IntVar(value=1)
    dtxtchk3 = tk.Checkbutton(popup_window_uptake, text='', variable=dchkval_3)
    dtxtchk3.place(x=869, y=175)

    dchkval_4 = tk.IntVar(value=1)
    dtxtchk4 = tk.Checkbutton(popup_window_uptake, text='', variable=dchkval_4)
    dtxtchk4.place(x=869, y=205)

    dchkval_5 = tk.IntVar(value=1)
    dtxtchk5 = tk.Checkbutton(popup_window_uptake, text='', variable=dchkval_5)
    dtxtchk5.place(x=869, y=235)

    dchkval_6 = tk.IntVar(value=0)
    dtxtchk6 = tk.Checkbutton(popup_window_uptake, text='', variable=dchkval_6)
    dtxtchk6.place(x=869, y=265)

    bchkval_1 = tk.IntVar(value=0)
    btxtchk1 = tk.Checkbutton(popup_window_uptake, text='', variable=bchkval_1)
    btxtchk1.place(x=589, y=295)

    bchkval_2 = tk.IntVar(value=1)
    btxtchk2 = tk.Checkbutton(popup_window_uptake, text='', variable=bchkval_2)
    btxtchk2.place(x=589, y=325)

    
    def save_colors2():
        pvals = []
        dvals = []
        if pval1.get() != "":
            pvals.append(pval1.get())
        if pval2.get() != "":
            pvals.append(pval2.get())
        if pval3.get() != "":
            pvals.append(pval3.get())
        if pval4.get() != "":
            pvals.append(pval4.get())
        if pval5.get() != "":
            pvals.append(pval5.get())
        if dval1.get() != "":
            dvals.append(dval1.get())
        if dval2.get() != "":
            dvals.append(dval2.get())
        if dval3.get() != "":
            dvals.append(dval3.get())
        if dval4.get() != "":
            dvals.append(dval4.get())
        if dval5.get() != "":
            dvals.append(dval5.get())
        potential_val_error_1 = False
        for i, val in enumerate(pvals):
            try:
                next_val = pvals[i + 1]
            except:
                continue
            if next_val > val:
                potential_val_error_1 = True
                break
        if potential_val_error_1 == False:
            for i, val in enumerate(dvals):
                try:
                    next_val = dvals[i + 1]
                except:
                    continue
                if next_val > val:
                    potential_val_error_1 = True
                    break
        if potential_val_error_1 == True:
            tk.messagebox.showwarning("Potential Error Found", f"{next_val} > {val}. This may cause an error when running the program. Please make sure differences are listed in descending differences, and not that both protection and deprotection values should be written as a positive number")
           

        protection_color_dic = {}
        if pval1.get() != "" and pcol1.get() != "":
            protection_color_dic[pval1.get()] = pcol1.get()
        if pval2.get() != "" and pcol2.get() != "":
            protection_color_dic[pval2.get()] = pcol2.get()
        if pval3.get() != "" and pcol3.get() != "":
            protection_color_dic[pval3.get()] = pcol3.get()
        if pval4.get() != "" and pcol4.get() != "":
            protection_color_dic[pval4.get()] = pcol4.get()
        if pval5.get() != "" and pcol5.get() != "":
            protection_color_dic[pval5.get()] = pcol5.get()
        if pcol6.get() != "":
            protection_color_dic[">0"] = pcol6.get()
        else:
            protection_color_dic[">0"] = "F2F2F2"
        deprotection_color_dic = {}
        if dval1.get() != "" and dcol1.get() != "":
            deprotection_color_dic[dval1.get()] = dcol1.get()
        if dval2.get() != "" and dcol2.get() != "":
            deprotection_color_dic[dval2.get()] = dcol2.get()
        if dval3.get() != "" and dcol3.get() != "":
            deprotection_color_dic[dval3.get()] = dcol3.get()
        if dval4.get() != "" and dcol4.get() != "":
            deprotection_color_dic[dval4.get()] = dcol4.get()
        if dval5.get() != "" and dcol5.get() != "":
            deprotection_color_dic[dval5.get()] = dcol5.get()
        if dcol6.get() != "":
            deprotection_color_dic[">0"] = dcol6.get()
        else:
            deprotection_color_dic[">0"] = "F2F2F2"
        both_color_dic = {}
        if bcol1.get() != "":
            both_color_dic["=0"] = bcol1.get()
        else:
            both_color_dic["=0"] = "F2F2F2"
        if bcol2.get() != "":
            both_color_dic[-99999] = bcol2.get()
        else:
            both_color_dic[-99999] = "FAE8D7"


        protection_text_dic = {}
        if pval1.get() != "" and pcol1.get != "":
            protection_text_dic[pval1.get()] = pchkval_1.get()
        if pval2.get() != "" and pcol2.get != "":
            protection_text_dic[pval2.get()] = pchkval_2.get()
        if pval3.get() != "" and pcol3.get != "":
            protection_text_dic[pval3.get()] = pchkval_3.get()
        if pval4.get() != "" and pcol4.get != "":
            protection_text_dic[pval4.get()] = pchkval_4.get()
        if pval5.get() != "" and pcol5.get() != "":
            protection_text_dic[pval5.get()] = pchkval_5.get()
        protection_text_dic[">0"] = pchkval_6.get()

        deprotection_text_dic = {}
        if dval1.get() != "" and dcol1.get != "":
            deprotection_text_dic[dval1.get()] = dchkval_1.get()
        if dval2.get() != "" and dcol2.get != "":
            deprotection_text_dic[dval2.get()] = dchkval_2.get()
        if dval3.get() != "" and dcol3.get != "":
            deprotection_text_dic[dval3.get()] = dchkval_3.get()
        if dval4.get() != "" and dcol4.get != "":
            deprotection_text_dic[dval4.get()] = dchkval_4.get()
        if dval5.get() != "" and dcol5.get() != "":
            deprotection_text_dic[dval5.get()] = dchkval_5.get()
        deprotection_text_dic[">0"] = dchkval_6.get()

        both_text_dic = {}
        both_text_dic["=0"] = bchkval_1.get()
        both_text_dic[-99999] = bchkval_2.get()
        json_data = {
            "header": "Difference Colors",
            "protection_color_dic": protection_color_dic,
            "protection_text_dic": protection_text_dic,
            "deprotection_color_dic": deprotection_color_dic,
            "deprotection_text_dic": deprotection_text_dic,
            "both_color_dic": both_color_dic,
            "both_text_dic": both_text_dic
        }
        
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".json",
            initialdir="./Colors",
            title="Save File",
            initialfile="dif_new_colors",
            filetypes=[("JSON files", "*.json")]
        )
        if file_path:
            with open(file_path, 'w') as f:
                json.dump(json_data, f, indent=4)
                title = os.path.basename(file_path)
                tk.messagebox.showinfo("File Saved", f"File Saved as {title}")
                
            uptake_file_names, dif_file_names, local_file_names = update_dir_lists()
            update_color_comboboxes(uptake_file_names, dif_file_names, local_file_names)
            popup_window_uptake.focus_set()
            
        else:
            tk.messagebox.showerror("File Not Saved", "File was not saved")
            popup_window_uptake.focus_set()
    


    save_bt_dif = tk.Button(popup_window_uptake, text = "Save Colors", command=save_colors2)
    save_bt_dif.place(x=725, y=330)
    
    x1 = 347
    y1 = 360
    x2 = 915 + 130
    y2 = 495
    
    canvas.create_rectangle(x1, y1, x2, y2, outline="black", fill="")

    
    
    
    
    def save_colors3():
        try:
            significance_cutoff = significance_entry.get()
            significance_cutoff = float(significance_cutoff)
        except:
            tk.messagebox.showerror("Significance Cut-Off Error", "Please make sure significance cut-off is a retrievable value and try again")
            popup_window_uptake.focus_set()
            return
        lcols = []
        lcol_0 = lcol_entry_0.get()
        lcols.append(lcol_0)
        lcol_1 = lcol_entry_1.get()
        lcols.append(lcol_1)
        lcol_2 = lcol_entry_2.get()
        lcols.append(lcol_2)
        lcol_3 = lcol_entry_3.get()
        lcols.append(lcol_3)
        lcol_4 = lcol_entry_4.get()
        lcols.append(lcol_4)
        lcol_5 = lcol_entry_5.get()
        lcols.append(lcol_5)
        for lcol in lcols:
            if is_valid_hexadecimal(lcol) == False:
                tk.messagebox.showerror("Hex Color Error", "Please make sure at least every color except manual options are a valid hex color and try again")
                popup_window_uptake.focus_set()
                return
        
        lcol_6 = lcol_entry_6.get()
        if is_valid_hexadecimal(lcol_6) == False:
            lcol_6 = False
        lcol_7 = lcol_entry_7.get()
        if is_valid_hexadecimal(lcol_7) == False:
            lcol_7 = False
        lcol_8 = lcol_entry_8.get()
        if is_valid_hexadecimal(lcol_8) == False:
            lcol_8 = False
        lcol_9 = lcol_entry_9.get()
        if is_valid_hexadecimal(lcol_9) == False:
            lcol_9 = False
        
        
        json_data = {
            "header": "Localized Difference Plot Colors",
            "lcols": lcols,
            "significance_cutoff": significance_cutoff,
            "lcol_6": lcol_6,
            "lcol_7": lcol_7,
            "lcol_8": lcol_8,
            "lcol_9": lcol_9
        }
        
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".json",
            initialdir="./Colors",
            title="Save File",
            initialfile="local_new_colors",
            filetypes=[("JSON files", "*.json")]
        )
        if file_path:
            with open(file_path, 'w') as f:
                json.dump(json_data, f, indent=4)
                title = os.path.basename(file_path)
                tk.messagebox.showinfo("File Saved", f"File Saved as {title}")
                
                
            uptake_file_names, dif_file_names, local_file_names = update_dir_lists()
            update_color_comboboxes(uptake_file_names, dif_file_names, local_file_names)
            popup_window_uptake.focus_set()
    
        else:
            tk.messagebox.showerror("File Not Saved", "File was not saved")
            popup_window_uptake.focus_set()
    

    tk.Label(popup_window_uptake, text="Create Custom Colors for Localized Difference Plots - Manual Options are Optional").place(x=370, y=362)
    
    
    x1 = 493
    y1 = 383
    x2 = 860
    y2 = 451
    canvas.create_rectangle(x1, y1, x2, y2, outline="black", fill="", dash=(4,3))
    
    x1 = 986
    y1 = 383
    x2 = 1044
    y2 = 451
    canvas.create_rectangle(x1, y1, x2, y2, outline="black", fill="", dash=(4,3))
    
    tk.Label(popup_window_uptake, text="Manual").place(x=365, y=385)
    tk.Label(popup_window_uptake, text="Option").place(x=366, y=405)
    lcol_entry_7 = tk.Entry(popup_window_uptake, width=8)
    lcol_entry_7.place(x=362, y=425)
    
    tk.Label(popup_window_uptake, text="Manual").place(x=430, y=385)
    tk.Label(popup_window_uptake, text="Option").place(x=431, y=405)
    lcol_entry_6 = tk.Entry(popup_window_uptake, width=8)
    lcol_entry_6.place(x=427, y=425)
    
    tk.Label(popup_window_uptake, text="Significant").place(x=430+65, y=385)
    tk.Label(popup_window_uptake, text="Protection").place(x=430+65, y=405)
    lcol_entry_2 = tk.Entry(popup_window_uptake, width=8)
    lcol_entry_2.place(x=433+65, y=425)
    
    tk.Label(popup_window_uptake, text="Questionable").place(x=500+65, y=385)
    tk.Label(popup_window_uptake, text="Protection").place(x=505+65, y=405)
    lcol_entry_1 = tk.Entry(popup_window_uptake, width=8)
    lcol_entry_1.place(x=507+65, y=425)
    
    tk.Label(popup_window_uptake, text="No").place(x=595+65, y=385)
    tk.Label(popup_window_uptake, text="Difference").place(x=575+65, y=405)
    lcol_entry_0 = tk.Entry(popup_window_uptake, width=8)
    lcol_entry_0.place(x=577+65, y=425)
    lcol_entry_0.insert(0, "F2F2F2")
    
    tk.Label(popup_window_uptake, text="Questionable").place(x=640+65, y=385)
    tk.Label(popup_window_uptake, text="Deprotection").place(x=641+65, y=405)
    lcol_entry_4 = tk.Entry(popup_window_uptake, width=8)
    lcol_entry_4.place(x=650+65, y=425)
    
    tk.Label(popup_window_uptake, text="Significant").place(x=725+65, y=385)
    tk.Label(popup_window_uptake, text="Deprotection").place(x=720+65, y=405)
    lcol_entry_5 = tk.Entry(popup_window_uptake, width=8)
    lcol_entry_5.place(x=730+65, y=425)
    
    tk.Label(popup_window_uptake, text="Manual").place(x=801+65, y=385)
    tk.Label(popup_window_uptake, text="Option").place(x=802+65, y=405)
    lcol_entry_8 = tk.Entry(popup_window_uptake, width=8)
    lcol_entry_8.place(x=798+65, y=425)
    
    tk.Label(popup_window_uptake, text="Manual").place(x=801+130, y=385)
    tk.Label(popup_window_uptake, text="Option").place(x=802+130, y=405)
    lcol_entry_9 = tk.Entry(popup_window_uptake, width=8)
    lcol_entry_9.place(x=798+130, y=425)
    
    tk.Label(popup_window_uptake, text="No").place(x=874+130, y=385)
    tk.Label(popup_window_uptake, text="Coverage").place(x=858+130, y=405)
    lcol_entry_3 = tk.Entry(popup_window_uptake, width=8)
    lcol_entry_3.place(x=859+130, y=425)
    lcol_entry_3.insert(0, "FAE8D7")
    
    tk.Label(popup_window_uptake, text="Significance Cut-off (Da ~ 0.5 or RFU ~ 0.05):").place(x=360, y=455)
    significance_entry = tk.Entry(popup_window_uptake, width=5)
    significance_entry.place(x=630, y=456)
    
    save_bt_loc = tk.Button(popup_window_uptake, text = "Save Colors", command=save_colors3)
    save_bt_loc.place(x=725, y=456)
    
    
   




def update_dir_lists():
    folder_path = "./Colors"  # Path to the "Color Patterns" folder
    file_names = os.listdir(folder_path)  # Get a list of file names in the folder
    uptake_file_names = []
    dif_file_names = []
    local_file_names = []
    for file_name in file_names:
        with open("./Colors/" + file_name, 'r') as f:
            json_data = json.load(f)
            if json_data.get("header") == "Uptake Colors":
                uptake_file_names.append(file_name)
            elif json_data.get("header") == "Difference Colors": 
                dif_file_names.append(file_name)
            elif json_data.get("header") == "Localized Difference Plot Colors":
                local_file_names.append(file_name)
    return uptake_file_names, dif_file_names, local_file_names

def update_color_comboboxes(uptake_file_names, dif_file_names, local_file_names):
    global uptake_color_scheme_dropdown, difference_color_scheme_dropdown, localized_color_scheme_dropdown
    uptake_color_scheme_dropdown['values'] = uptake_file_names
    difference_color_scheme_dropdown['values'] = dif_file_names
    localized_color_scheme_dropdown['values'] = local_file_names

    
    
def white_variable_changed(*args): 
    if os.path.exists("Alterations/WhiteVar.json"):
        pass
    else:
        user_choice = tk.messagebox.askyesno("Default Changes", "Do you want to make this the default value?", default='no')
        dont_show_path = "Alterations/WhiteVar.json"
        if user_choice:
            with open(dont_show_path, 'w') as file:
                json.dump({"WhiteVarJson": 0}, file)
        else:
            with open(dont_show_path, 'w') as file:
                json.dump({"WhiteVarJson": 1}, file)
                
def insig_variable_changed(*args): 
    if os.path.exists("Alterations/Insig.json"):
        pass
    else:
        user_choice = tk.messagebox.askyesno("Default Changes", "Do you want to make this the default value?", default='no')
        dont_show_path = "Alterations/Insig.json"
        if user_choice:
            with open(dont_show_path, 'w') as file:
                json.dump({"InsigJson": 0}, file)
        else:
            with open(dont_show_path, 'w') as file:
                json.dump({"InsigJson": 1}, file)
                
def sd_variable_changed(*args): 
    if os.path.exists("Alterations/ShowSD.json"):
        pass
    else:
        user_choice = tk.messagebox.askyesno("Default Changes", "Do you want to make this the default value?", default='no')
        dont_show_path = "Alterations/ShowSD.json"
        if user_choice:
            with open(dont_show_path, 'w') as file:
                json.dump({"ShowSDJson": 1}, file)
        else:
            with open(dont_show_path, 'w') as file:
                json.dump({"ShowSDJson": 0}, file)
                
def proline_variable_changed(*args): 
    if os.path.exists("Alterations/Proline.json"):
        pass
    else:
        user_choice = tk.messagebox.askyesno("Default Changes", "Do you want to make this the default value?", default='no')
        dont_show_path = "Alterations/Proline.json"
        if user_choice:
            with open(dont_show_path, 'w') as file:
                json.dump({"ProlineJson": 0}, file)
        else:
            with open(dont_show_path, 'w') as file:
                json.dump({"ProlineJson": 1}, file)
                
def ResEnum_variable_changed(*args): 
    if os.path.exists("Alterations/ResEnum.json"):
        pass
    else:
        user_choice = tk.messagebox.askyesno("Default Changes", "Do you want to make this the default value?", default='no')
        dont_show_path = "Alterations/ResEnum.json"
        if user_choice:
            with open(dont_show_path, 'w') as file:
                json.dump({"ResEnumJson": 1}, file)
        else:
            with open(dont_show_path, 'w') as file:
                json.dump({"ResEnumJson": 0}, file)
    
    
def create_format_box():
    format_title = tk.Label(window, text="Formatting Options")
    format_title.place(x=960, y=5)

    x1,y1 = 922, 10
    x2, y2 = 1170, 450
    canvas.create_rectangle(x1, y1, x2, y2, outline="black", fill="")

    uptake_file_names, dif_file_names, local_file_names = update_dir_lists()
        
        
    global uptake_color_scheme_dropdown, difference_color_scheme_dropdown, localized_color_scheme_dropdown, colors_already_formatted
    uptake_color_scheme_dropdown = ttk.Combobox(window, values=uptake_file_names, width=19)
    if exp_bt_on_c:
        uptake_color_scheme_dropdown.set("corrected_df.json")
    if theo_bt_on_c:
        if float(be_entry.get()) != 0:
            uptake_color_scheme_dropdown.set("corrected_df.json")
        else:
            uptake_color_scheme_dropdown.set("uncorrected_df.json")
    uptake_color_scheme_dropdown.bind("<<ComboboxSelected>>")
    uptake_color_scheme_dropdown.place(x=1030, y=30)
    colors_already_formatted = True
    tk.Label(window, text="Uptake Colors: ").place(x=928, y=30)
    tk.Label(window, text="Difference Colors: ").place(x=928, y=60)
    tk.Label(window, text="Localized Colors: ").place(x=928, y=90)
    difference_color_scheme_dropdown = ttk.Combobox(window, values=dif_file_names, width=19)
    if exp_bt_on_c == True:
        difference_color_scheme_dropdown.set("Da_green_blue.json")
    if theo_bt_on_c == True:
        difference_color_scheme_dropdown.set("Da_green_blue.json")
    difference_color_scheme_dropdown.bind("<<ComboboxSelected>>")
    difference_color_scheme_dropdown.place(x=1030, y=60)
    
    localized_color_scheme_dropdown = ttk.Combobox(window, values=local_file_names, width=19)
    if exp_bt_on_c:
        localized_color_scheme_dropdown.set("9_Da_green_blue.json")
    if theo_bt_on_c:
        localized_color_scheme_dropdown.set("9_Da_green_blue.json")
    localized_color_scheme_dropdown.bind("<<ComboboxSelected>>")
    localized_color_scheme_dropdown.place(x=1030, y=90)
    
    
    create_colors = tk.Button(window, text="Create Custom Colors", command=create_custom_colors)
    create_colors.place(x=980, y=118)
    chiclet_options_title = tk.Label(window, text="Chiclet Options")
    chiclet_options_title.place(x=1000, y=152)
    x1 = 930
    x2 = 1162
    y1 = 176
    y2 = 150
    canvas.create_rectangle(x1, y1, x2, y2, outline="black", fill="")
    pepgap_lb = tk.Label(window, text="Add Gaps if Pep in Only One State:")
    pepgap_lb.place(x=925, y=180)
    global white_var
    white_var = tk.IntVar(value=1)
    if os.path.exists("Alterations/WhiteVar.json"):
        with open("Alterations/WhiteVar.json", 'r') as file:
            json_data = json.load(file)
            WhiteVarJson = int(float(json_data["WhiteVarJson"]))
            white_var = tk.IntVar(value=WhiteVarJson)
    white_var.trace_add('write', white_variable_changed) 
    chk1 = tk.Checkbutton(window, text='', variable=white_var)
    chk1.place(x=1140, y=178)
    
    reorder_states_bt = tk.Button(window, text="Reorder States", command=ReorderableListbox) 
    reorder_states_bt.place(x=1002, y=203)

    global con_pep_height_enter, con_pep_width_enter, full_pep_height_enter, full_pep_width_enter
    full_pepmap_title = tk.Label(window, text="Peptide Plot Options")
    full_pepmap_title.place(x=982, y=240)
    x1 = 930
    x2 = 1162
    y1 = 264
    y2 = 238
    canvas.create_rectangle(x1, y1, x2, y2, outline="black", fill="")
#    full_pep_width_lb = tk.Label(window, text = "Cell Width:")
#    full_pep_width_lb.place(x=925, y=235)
#    full_pep_width_enter = tk.Entry(window, width=5)
#    full_pep_width_enter.insert(0, "4")
#    full_pep_width_enter.place(x=1000, y=235)
    reduce_states_label = tk.Label(window, text = "Only Show States From Uptake Plot Box:")
    reduce_states_label.place(x=925, y=268)
    global reduce_states_var
    reduce_states_var = tk.IntVar(value=0)
    reduce_states_chk = tk.Checkbutton(window, text='', variable=reduce_states_var)
    reduce_states_chk.place(x=1140, y=266)
    
    enumerate_residues_label = tk.Label(window, text = "Enumerate All Residues:")
    enumerate_residues_label.place(x=925, y=292)
    global enumerate_residues_var
    enumerate_residues_var = tk.IntVar(value=0)
    if os.path.exists("Alterations/ResEnum.json"):
        with open("Alterations/ResEnum.json", 'r') as file:
            json_data = json.load(file)
            ResEnumJson = int(float(json_data["ResEnumJson"]))
            enumerate_residues_var = tk.IntVar(value=ResEnumJson)
    enumerate_residues_var.trace_add('write', ResEnum_variable_changed) 
    enumerate_residues_chk = tk.Checkbutton(window, text='', variable=enumerate_residues_var)
    enumerate_residues_chk.place(x=1140, y=290)



    con_pepmap_title = tk.Label(window, text="Condensed Peptide Plot Options")
    con_pepmap_title.place(x=960, y=322)
    x1 = 930
    x2 = 1162
    y1 = 346
    y2 = 320
    canvas.create_rectangle(x1, y1, x2, y2, outline="black", fill="")
#    con_pep_width_lb = tk.Label(window, text = "Cell Width:")
#    con_pep_width_lb.place(x=925, y=331)
#    con_pep_width_enter = tk.Entry(window, width=5)
#    con_pep_width_enter.insert(0, "2.5")
#    con_pep_width_enter.place(x=1000, y=331)



    insig_dif_lb = tk.Label(window, text="Show Insignificant Values:")
    insig_dif_lb.place(x=925, y=352)
    global insig_dif_chk, SE_enter
    insig_dif_chk = tk.IntVar(value=1)
    if os.path.exists("Alterations/Insig.json"):
        with open("Alterations/Insig.json", 'r') as file:
            json_data = json.load(file)
            InsigJson = int(float(json_data["InsigJson"]))
            insig_dif_chk = tk.IntVar(value=InsigJson)
    insig_dif_chk.trace_add('write', insig_variable_changed) 
    insig_check = tk.Checkbutton(window, text='', variable=insig_dif_chk)
    insig_check.place(x=1140, y=350)

    tk.Label(window, text="Show Standard Error:").place(x=925, y=373)
    tk.Label(window, text="n = ").place(x=1055, y=373)
    SE_enter = tk.Entry(window, width=4)
    SE_enter.place(x=1080, y=373)
    SE_enter.insert(0, "3")
    global sd_checkvar
    sd_checkvar = tk.IntVar(value=0)
    if os.path.exists("Alterations/ShowSD.json"):
        with open("Alterations/ShowSD.json", 'r') as file:
            json_data = json.load(file)
            ShowSDJson = int(float(json_data["ShowSDJson"]))
            sd_checkvar = tk.IntVar(value=ShowSDJson)
    sd_checkvar.trace_add('write', sd_variable_changed) 
    sd_check = tk.Checkbutton(window, text='', variable=sd_checkvar)
    sd_check.place(x=1140, y=370)
    
    loc_pepmap_title = tk.Label(window, text="Localized Difference Plot Options")
    loc_pepmap_title.place(x=960, y=398)
    x1 = 930
    x2 = 1162
    y1 = 396
    y2 = 422
    canvas.create_rectangle(x1, y1, x2, y2, outline="black", fill="")
    
    global proline_checkvar
    tk.Label(window, text="Exclude Exchange at Prolines:").place(x=925, y=426)
    proline_checkvar = tk.IntVar(value=1)
    if os.path.exists("Alterations/Proline.json"):
        with open("Alterations/Proline.json", 'r') as file:
            json_data = json.load(file)
            ProlineJson = int(float(json_data["ProlineJson"]))
            proline_checkvar = tk.IntVar(value=ProlineJson)
    proline_checkvar.trace_add('write', proline_variable_changed) 
    proline_check = tk.Checkbutton(window, text='', variable=proline_checkvar)
    proline_check.place(x=1140, y=424)
    
    



class ReorderableListbox(tk.Tk): 
    def __init__(self):
        global states
        super().__init__()

        self.title("Reorder List")
        self.geometry("300x400")

        # Frame to hold Listbox and Scrollbar
        frame = tk.Frame(self)
        frame.pack(fill=tk.BOTH, expand=True)

        # Scrollbar
        scrollbar = tk.Scrollbar(frame, orient=tk.VERTICAL)

        # Listbox with scrollbar
        self.listbox = tk.Listbox(frame, selectmode=tk.SINGLE, yscrollcommand=scrollbar.set)
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar.config(command=self.listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Insert items into Listbox
        for state in states:
            self.listbox.insert(tk.END, state)

        # Bind mouse events for drag-and-drop reordering
        self.listbox.bind("<Button-1>", self.on_click)
        self.listbox.bind("<B1-Motion>", self.on_drag)

        self.dragged_item_index = None

        # Save Button
        save_button = tk.Button(self, text="Save Order", command=self.save_order)
        save_button.pack(pady=10)

    def on_click(self, event):
        widget = event.widget
        self.dragged_item_index = widget.nearest(event.y)

    def on_drag(self, event):
        widget = event.widget
        new_index = widget.nearest(event.y)

        if self.dragged_item_index is not None and new_index != self.dragged_item_index:
            item_text = widget.get(self.dragged_item_index)
            widget.delete(self.dragged_item_index)
            widget.insert(new_index, item_text)
            self.dragged_item_index = new_index

    def save_order(self):
        global states
        ordered_items = self.listbox.get(0, tk.END)
        states = list(ordered_items)
        self.destroy()


def create_run_box():
    global run_bt
    run_box_title =tk.Label(window, text="Choose Scripts")
    run_box_title.place(x=1210, y=5)
    x1,y1 = 1172,10
    x2,y2= 1485, 450
    canvas.create_rectangle(x1, y1, x2, y2, outline="black", fill="")

    run_bt = tk.Button(window, text="\u23F5",bg="blue",fg="white",width=7, command=r_initialize)
    run_bt.place(x=1190,y=200)
    chic_bt = tk.Button(window, text="Chiclet Plot",bg="orange",fg="black", width=17, command=chiclet_on)
    chic_bt.place(x=1190,y=40)
    cdif_bt = tk.Button(window, text="Chiclet Difference",bg="orange",fg="black", width=17, command=cdif_on)
    cdif_bt.place(x=1340,y=40)
    pepmap_bt = tk.Button(window, text="Peptide Plot",bg="orange",fg="black", width=17, command=pepmap_on)
    pepmap_bt.place(x=1190,y=80)
    difmap_bt = tk.Button(window, text="Peptide Difference",bg="orange",fg="black",width=17, command=difmap_on)
    difmap_bt.place(x=1340,y=80)
    condpep_bt = tk.Button(window, text="Condensed Peptide",bg="orange",fg="black",width=17, command=condpeps_on)
    condpep_bt.place(x=1190,y=120)
    difcond_bt = tk.Button(window, text="Condensed Difference",bg="orange",fg="black",width=17, command=difcond_on)
    difcond_bt.place(x=1340,y=120)
    uptake_plot_bt = tk.Button(window, text="Uptake Plots",bg="orange", fg="black", width=17, command=uptake_plot_on)
    uptake_plot_bt.place(x=1190, y=160)
    heatmap_bt = tk.Button(window, text="Localized Differences", bg="orange", fg="black", width=17, command=lambda: (heatmap_on(), difcond_on(), difmap_on()))
    heatmap_bt.place(x=1340, y=160)
    
    
def create_uptakeplot_box():
    global correction, uptake_plot_colors, uptake_plot_symbols, show_last, maxD_dash, state_selects, col_entries, sym_entries, size_entries, x_enter, y_enter, linewidth_enter, pep_search_enter, a_horizontal, a_vertical, title_entries, legend_size_entry, leg_ur, leg_ul, leg_bl, leg_br, leg_pos, legend_linewidth_entry, dot_chkval, cplt_chkval
    x1, y1 = 372, 452
    x2, y2 = 1485, 880
    canvas.create_rectangle(x1, y1, x2, y2, outline="black", fill="")
    tk.Label(window, text="Format Uptake Plots").place(x=380, y=455)
    tk.Label(window, text="Y Axis:").place(x=500, y=455)
    y_enter = tk.Entry(window, width=15)
    y_enter.place(x=540, y=455)
    y_enter.insert(0, "Uptake (Da)")
    tk.Label(window, text="X Axis:").place(x=630, y=455)
    x_enter = tk.Entry(window, width=15)
    x_enter.place(x=670, y=455)
    x_enter.insert(0, "Exposure (min)")
    Uncorrected_plot_bt = tk.Button(window, text="Uncorrected", bg="green", fg="white", command=lambda: [r_unc_bt_off(), r_c_bt_on()])
    Uncorrected_plot_bt.place(x=375, y=480)
    corrected_plot_bt = tk.Button(window, text="Corrected", bg="orange", fg="black", command=lambda: [r_unc_bt_off(), r_c_bt_on()])
    corrected_plot_bt.place(x=461, y=480)
    show_last_bt = tk.Button(window, text="Show Last Timepoint", bg="green", fg="white", command=show_last_bt_off)
    show_last_bt.place(x=535, y=480)
#    maxD_dash_bt = tk.Button(window, text="maxD Dashed Line", bg = 'orange', fg='black', command=maxd_dash_on)
#    maxD_dash_bt.place(x=880, y=835)
    show_bt = tk.Button(window, text="Show", command=create_example_plot)
    show_bt.place(x=663, y=480)
    tk.Label(window, text="Linewidth:").place(x=375, y=510)
    linewidth_enter = tk.Entry(window, width=3)
    linewidth_enter.place(x=435, y=510)
    linewidth_enter.insert(0, "0.6")
    dashed_label = tk.Label(window, text="Dash")
    dashed_label.place(x=460, y=510)
    dot_chkval = tk.IntVar(value=0)
    dot_chk = tk.Checkbutton(window, text='', variable=dot_chkval)
    dot_chk.place(x=490, y=510)
    tk.Label(window, text="Search for residue:").place(x=518, y=510)
    pep_search_enter = tk.Entry(window, width=5)
    pep_search_enter.place(x=620, y=510)
    next_bt = tk.Button(window, text="Next Peptide", command=next_peptide)
    next_bt.place(x=708, y=480)
    previous_bt = tk.Button(window, text="Last Peptide", command=previous_peptide)
    previous_bt.place(x=710, y=510)
    search_bt = tk.Button(window, text="Search", command=searching)
    search_bt.place(x=660, y=510)
    horizontal_bt = tk.Button(window, text="Horizontal", bg="green", fg="white", command=lambda: [horizontal_bt_off(), vertical_bt_on()], width=8)
    horizontal_bt.place(x=787, y=480)
    vertical_bt = tk.Button(window, text="Vertical", bg="orange", fg="black", command=lambda: [horizontal_bt_off(), vertical_bt_on()], width=8)
    vertical_bt.place(x=787, y=510)
    a_horizontal = True
    a_vertical = False
    correction = False
    show_last = True
    maxD_dash = False
    
#    complete_states_label = tk.Label(window, text="Complete States Only:")
#    complete_states_label.place(x=1152, y=830)
#    cplt_chkval = tk.IntVar(value=0)
#    cplt_chk = tk.Checkbutton(window, text='', variable=cplt_chkval)
#    cplt_chk.place(x=1282, y=830)
    
    tk.Label(window, text="State").place(x=425, y=537)
    tk.Label(window, text="Hex Color").place(x=573, y=537)
    tk.Label(window, text="Unicode Symbol").place(x=633, y=537)
    tk.Label(window, text="Size").place(x=738, y=537)
    tk.Label(window, text="Legend Title").place(x=780, y=537)
    
    uptake_plot_colors = ["1F77B4", "FF7F0E", "2CA02C", "D62728", "9467BD", "8C564B", "E377C2", "7F7F7F", "BCBD22", "17BECF", "140003", "32CD32", "E47565", "C7703D", "D31C80", "800000"]
    uptake_plot_symbols = ["U+2B24", "U+25A0", "U+25B2", "U+25C6", "U+25BC", "U+25CB", "U+25A1", "U+025C1", "U+025C7", "U+025AE", "U+025B1", "U+025B3", "U+025B0", "U+025C8", "U+025C9", "U+25A3"]
    
    
    def on_change_sym(sv, identifier):
        if is_valid_unicode(sv.get()) == True:
            if "+" in sv.get():
                new_symbol = sv.get().split("+")[1]
            else:
                sym_unicode_labels[identifier-1].config(text="")
            try:
                unicode_char = chr(int(new_symbol, 16))
                sym_unicode_labels[identifier-1].config(text=unicode_char)
            except:
                sym_unicode_labels[identifier-1].config(text="")
        else:
            sym_unicode_labels[identifier-1].config(text="")
            
    def on_change_col(sv, identifier):
        if is_valid_hexadecimal(sv.get()) == True:
            sym_unicode_labels[identifier-1].config(foreground=f"#{sv.get()}")
        else:
            sym_unicode_labels[identifier-1].config(foreground=f"#F0F0F0")
        
        
    
    state_selects = []
    col_entries = []
    sym_entries = []
    size_entries = []
    title_entries = []
    text_vars = []
    text_vars_cols = []
    sym_unicode_labels = []
    
    x = 0
    states_o = list(states.keys())
    for state in states_o:
        if x <= 15:
            x += 1
            state_var = tk.StringVar(value=state)
            na_var = tk.StringVar(value="N/A")
            state_select = ttk.Combobox(window, values=(states_o + ["N/A"]), width=28)
            state_select.place(x=375, y=(537 + (20*x)))
        
        
            if x <= 2:
                state_select.set(state_var.get())
            else:
                state_select.set(na_var.get())
            title_entry = tk.Entry(window, width=13)
            state_select.bind("<<ComboboxSelected>>", lambda event, title_entry=title_entry, state_select=state_select: on_state_selected(event, title_entry, state_select))

            text_var = tk.StringVar()
            sym_unicode_label = tk.Label(window, text=text_var, foreground=f"#{uptake_plot_colors[x-1]}")
            sym_unicode_label.place(x=710, y=(537 + (20*x)))
            sym_unicode_labels.append(sym_unicode_label)
            
            text_var2 = tk.StringVar()
            text_var2.trace_add("write", lambda name, index, mode, sv=text_var2, identifier=x: on_change_col(sv, identifier))
            text_vars_cols.append(text_var2)
            col_entry = tk.Entry(window, textvariable=text_var2, width=9)
            col_entry.place(x=580, y=(537 + (20*x)))
            col_entry.insert(0, uptake_plot_colors[x-1])
            
            
            text_var.trace_add("write", lambda name, index, mode, sv=text_var, identifier=x: on_change_sym(sv, identifier))
            text_vars.append(text_var)  # Store the StringVar to maintain scope
            sym_entry = tk.Entry(window, textvariable=text_var, width=9)
            sym_entry.place(x=648, y=(537 + (20*x)))
            sym_entry.insert(0, uptake_plot_symbols[x-1])
            
            
            
            size_entry = tk.Entry(window, width=3)
            size_entry.place(x=740, y=(537 + (20*x)))
            size_entry.insert(0, "3")
#            if x <= 2:
#                title_entry.insert(0, state.split("~")[1])
            title_entry.insert(0, state)
            title_entry.place(x=770, y=(537 + (20*x)))
            state_selects.append(state_select)
            col_entries.append(col_entry)
            sym_entries.append(sym_entry)
            size_entries.append(size_entry)
            title_entries.append(title_entry)
            
    tk.Label(window, text="Legend Options").place(x=910, y=455)
    x1 = 870
    y = 480
    x2 = 1055
    canvas.create_line(x1, y, x2, y)
    tk.Label(window, text="Legend Size:").place(x=870, y=485)
    legend_size_entry = tk.Entry(window, width = 6)
    legend_size_entry.insert(0, "10")
    legend_size_entry.place(x=945, y=485)
    tk.Label(window, text="Legend Linewidth:").place(x=870, y=515)
    legend_linewidth_entry = tk.Entry(window, width=6)
    legend_linewidth_entry.insert(0, "1")
    legend_linewidth_entry.place(x=975, y=515)
    
    
    
    tk.Label(window, text="Errorbar Options").place(x=910, y=615)
    x1 = 870
    y = 640
    x2 = 1055
    canvas.create_line(x1, y, x2, y)
    
    
    
    global chkval_errorbars, show_error_bars, SE_num_entry
    show_error_bars = False
    chkval_errorbars = tk.IntVar(value=0)
    errorbar_checkbox = tk.Checkbutton(window, text='Show Errorbars (SE)', variable=chkval_errorbars, command=errorbar_toggle)
    errorbar_checkbox.place(x=860, y=645)
    
    tk.Label(window, text="n=").place(x=990, y=645)
    SE_num_entry = tk.Entry(window, width=6)
    SE_num_entry.place(x=1010, y=645)
    SE_num_entry.insert(0, "3")

    
    global capsize_entry, errorbar_linewidth_entry, capthick_entry
    tk.Label(window, text="Capsize:").place(x=890, y=675)
    capsize_entry = tk.Entry(window, width=6)
    capsize_entry.place(x=960, y=675)
    capsize_entry.insert(0, "5")
    tk.Label(window, text="Linewdith:").place(x=890, y=705)
    errorbar_linewidth_entry = tk.Entry(window, width=6)
    errorbar_linewidth_entry.place(x=960, y=705)
    errorbar_linewidth_entry.insert(0, "0.5")
    tk.Label(window, text="CapThick:").place(x=890, y=735)
    capthick_entry = tk.Entry(window, width=6)
    capthick_entry.place(x=960, y=735)
    capthick_entry.insert(0, "0.5")
    
    
    x1 = 870
    y = 765
    x2 = 1055
    canvas.create_line(x1, y, x2, y)
    
    global temp_max_plot_height_entry, temp_max_plot_chkval, change_scale
    change_scale = False
    tk.Label(window, text="Set max plot height (for .png only)").place(x=870, y=770)
    temp_max_plot_height_entry = tk.Entry(window, width=6)
    temp_max_plot_height_entry.place(x=870, y=800)
    
    temp_max_plot_chkval = tk.IntVar(value=0)
    temp_max_plot_checkbox = tk.Checkbutton(window, variable=temp_max_plot_chkval, command=temp_max_plot_toggle)
    temp_max_plot_checkbox.place(x=920, y=800)
    
    x1 = 870
    y = 830
    x2 = 1055
    canvas.create_line(x1, y, x2, y)
    
    
    
    
    leg_ur = False
    leg_ul = False
    leg_bl = False
    leg_br = False
    ur_bt = tk.Button(window, text="Upper Right", bg="orange", fg="black", width = 10, command = lambda: [ur_bt_on(), ul_bt_off(), bl_bt_off(), br_bt_off()])
    ur_bt.place(x=955, y=545)
    ul_bt = tk.Button(window, text="Upper Left",  bg="orange", fg="black" ,width = 10, command = lambda: [ur_bt_off(), ul_bt_on(), bl_bt_off(), br_bt_off()])
    ul_bt.place(x=870, y=545)
    br_bt = tk.Button(window, text="Bottom Right", bg="orange", fg="black", width = 10, command = lambda: [ur_bt_off(), ul_bt_off(), bl_bt_off(), br_bt_on()])
    br_bt.place(x=955, y=575)
    bl_bt = tk.Button(window, text="Bottom Left",  bg="orange", fg="black" ,width = 10, command = lambda: [ur_bt_off(), ul_bt_off(), bl_bt_on(), br_bt_off()])
    bl_bt.place(x=870, y=575)
    create_example_plot()
    
def errorbar_toggle():
    global chkval_errorbars, show_error_bars
    if chkval_errorbars.get() == 1:
        show_error_bars = True
    else:
        show_error_bars = False
    create_example_plot()
    
def temp_max_plot_toggle():
    global change_scale, temp_max_plot_chkval
    if temp_max_plot_chkval.get() == 1:
        change_scale = True
    else:
        change_scale = False
    create_example_plot()
    
def ur_bt_on():
    global leg_ur, leg_ul, leg_bl, leg_br, leg_pos
    leg_ur = True
    leg_ul = False
    leg_bl = False
    leg_br = False
    leg_pos = 'upper right'
    ur_bt = tk.Button(window, text="Upper Right", bg="green", fg="white", width = 10, command = ur_bt_off)
    ur_bt.place(x=955, y=545)
    create_example_plot()
    
def ur_bt_off():
    global leg_ur
    leg_ur = False
    ur_bt = tk.Button(window, text="Upper Right", bg="orange", fg="black", width = 10, command = lambda: [ur_bt_on(), ul_bt_off(), bl_bt_off(), br_bt_off()])
    ur_bt.place(x=955, y=545)
    create_example_plot()
    
def ul_bt_on():
    global leg_ur, leg_ul, leg_bl, leg_br, leg_pos
    leg_ur = False
    leg_ul = True
    leg_bl = False
    leg_br = False
    leg_pos = 'upper left'
    ul_bt = tk.Button(window, text="Upper Left",  bg="green", fg="white" ,width = 10, command = ul_bt_off)
    ul_bt.place(x=870, y=545)
    create_example_plot()

def ul_bt_off():
    global leg_ul
    leg_ul = False
    ul_bt = tk.Button(window, text="Upper Left",  bg="orange", fg="black" ,width = 10, command = lambda: [ur_bt_off(), ul_bt_on(), bl_bt_off(), br_bt_off()])
    ul_bt.place(x=870, y=545)
    create_example_plot()
    
def bl_bt_on():
    global leg_ur, leg_ul, leg_bl, leg_br, leg_pos
    leg_ur = False
    leg_ul = False
    leg_bl = True
    leg_br = False
    leg_pos = 'lower left'
    bl_bt = tk.Button(window, text="Bottom Left",  bg="green", fg="white" ,width = 10, command = bl_bt_off)
    bl_bt.place(x=870, y=575)
    create_example_plot()
    
def bl_bt_off():
    global leg_bl
    leg_bl = False
    bl_bt = tk.Button(window, text="Bottom Left",  bg="orange", fg="black" ,width = 10, command = lambda: [ur_bt_off(), ul_bt_off(), bl_bt_on(), br_bt_off()])
    bl_bt.place(x=870, y=575)
    create_example_plot()
    
def br_bt_on():
    global leg_ur, leg_ul, leg_bl, leg_br, leg_pos
    leg_ur = False
    leg_ul = False
    leg_bl = False
    leg_br = True
    leg_pos = 'lower right'
    br_bt = tk.Button(window, text="Bottom Right", bg="green", fg="white", width = 10, command = br_bt_on)
    br_bt.place(x=955, y=575)
    create_example_plot()

def br_bt_off():
    global leg_br
    leg_br = False
    br_bt = tk.Button(window, text="Bottom Right", bg="orange", fg="black", width = 10, command = lambda: [ur_bt_off(), ul_bt_off(), bl_bt_off(), br_bt_on()])
    br_bt.place(x=955, y=575)
    create_example_plot()
    

    
    
    

    
def on_state_selected(event, title_entry, state_select):
    selected_value = state_select.get()
    if "~" in selected_value:
        title_entry.delete(0, tk.END)
        title_entry.insert(0, selected_value.split("~")[1])
    elif selected_value == "N/A":
        title_entry.delete(0, tk.END)
    else:
        title_entry.delete(0, tk.END)
        title_entry.insert(0, selected_value)

def next_peptide():
    global current_peptide_index
    current_peptide_index += 1
    create_example_plot()

def previous_peptide():
    global current_peptide_index
    current_peptide_index -= 1
    create_example_plot()

def r_unc_bt_on():
    global correction
    correction = False
    Uncorrected_plot_bt = tk.Button(window, text="Uncorrected", bg="green", fg="white", command=lambda: [r_unc_bt_off(), r_c_bt_on()])
    Uncorrected_plot_bt.place(x=375, y=480)
    create_example_plot()
    
def r_unc_bt_off():
    global correction
    correction = True
    Uncorrected_plot_bt = tk.Button(window, text="Uncorrected", bg="orange", fg="black", command=lambda: [r_unc_bt_on(), r_c_bt_off()])
    Uncorrected_plot_bt.place(x=375, y=480)
    create_example_plot()

def r_c_bt_on():
    global correction
    correction = True
    corrected_plot_bt = tk.Button(window, text="Corrected", bg="green", fg="white", command=lambda: [r_unc_bt_on(), r_c_bt_off()])
    corrected_plot_bt.place(x=461, y=480)
    create_example_plot()

def r_c_bt_off():
    global correction
    correction = False
    corrected_plot_bt = tk.Button(window, text="Corrected", bg="orange", fg="black", command=lambda: [r_unc_bt_off(), r_c_bt_on()])
    corrected_plot_bt.place(x=461, y=480)
    create_example_plot()

def show_last_bt_on():
    global show_last
    show_last = True
    show_last_bt = tk.Button(window, text="Show Last Timepoint", bg="green", fg="white", command=show_last_bt_off)
    show_last_bt.place(x=535, y=480)
    create_example_plot()
    
def show_last_bt_off():
    global show_last
    show_last = False
    show_last_bt = tk.Button(window, text="Show Last Timepoint", bg="orange", fg="black", command=show_last_bt_on)
    show_last_bt.place(x=535, y=480)
    create_example_plot()
    
def horizontal_bt_on():
    global a_horizontal, a_vertical
    horizontal_bt = tk.Button(window, text="Horizontal", bg="green", fg="white", command=lambda: [horizontal_bt_off(), vertical_bt_on()], width=8)
    horizontal_bt.place(x=787, y=480)
    a_horizontal = True
    a_vertical = False

def horizontal_bt_off():
    global a_horizontal, a_vertical
    horizontal_bt = tk.Button(window, text="Horizontal", bg="orange", fg="black", command=lambda: [horizontal_bt_on(), vertical_bt_off()], width=8)
    horizontal_bt.place(x=787, y=480)
    a_horizontal = False
    a_vertical = True
    
def vertical_bt_on():
    global a_horizontal, a_vertical
    vertical_bt = tk.Button(window, text="Vertical", bg="green", fg="white", command=lambda: [horizontal_bt_on(), vertical_bt_off()], width=8)
    vertical_bt.place(x=787, y=510)
    a_vertical = True
    a_horizontal = False
    
def vertical_bt_off():
    global a_horizontal, a_vertical
    vertical_bt = tk.Button(window, text="Vertical", bg="orange", fg="black", command=lambda: [horizontal_bt_off(), vertical_bt_on()], width=8)
    vertical_bt.place(x=787, y=510)
    a_vertical = False
    a_horizontal = True
    
#def maxd_dash_on():
#    global maxD_dash
#    maxD_dash_bt = tk.Button(window, text="maxD Dashed Line", bg = 'green', fg='white', command=maxd_dash_off)
#    maxD_dash_bt.place(x=880, y=835)
#    maxD_dash = True
#    create_example_plot()
#    
#def maxd_dash_off():
#    global maxD_dash
#    maxD_dash_bt = tk.Button(window, text="maxD Dashed Line", bg = 'orange', fg='black', command=maxd_dash_on)
#    maxD_dash_bt.place(x=880, y=835)
#    maxD_dash = False
#    create_example_plot()

global search_on
search_on = False
def searching():
    global search_on
    search_on = True
    create_example_plot()
    
def search_for_startvalue():
    if pep_search_enter.get() == None:
        return False
    try: 
        pep_search_num = int(pep_search_enter.get())
        return pep_search_num
    except:
        return False
    
def search_for_protein():
    current_protein_list = list()
    for state in order_state_dic.values():
        if state != False and state != "False":
            protein = state.split("~")[0]
            if protein not in current_protein_list:
                current_protein_list.append(protein)
    return current_protein_list
            
    
        
    
    
def is_valid_hexadecimal(s):
    return bool(re.fullmatch(r'#[0-9a-fA-F]{6}|[0-9a-fA-F]{6}|#[0-9a-fA-F]{8}|[0-9a-fA-F]{8}', s))

def is_valid_unicode(input_str):
    try:
        input_str.encode('utf-8').decode('utf-8')
        return True
    except UnicodeError:
        return False

def parse_data():
    global up_plot_colors, up_plot_symbols, up_plot_sizes, up_plot_titles
    global order_color_dic, order_symbol_dic, order_size_dic, order_state_dic, orders, linewidth_in_use, order_title_dic
    up_plot_states = []
    up_plot_colors = []
    up_plot_symbols = []
    up_plot_sizes = []
    up_plot_titles = []
    for i, state in enumerate(state_selects):
        if is_valid_hexadecimal(col_entries[i].get()) is True:
            col = col_entries[i].get()
        else:
            col = False
        if is_valid_unicode(sym_entries[i].get()) is True:
            sym = sym_entries[i].get()
        else:
            sym = False
        if not size_entries[i].get() is None:
            try: 
                size_ent = int(size_entries[i].get())
            except:
                size_ent = 10
        else:
            size_ent = 10
        if not title_entries[i].get() == None:
            tit_ent = title_entries[i].get()
        else:
            tit_ent = "No title"
        if state.get() == "N/A":
            size_ent = False
            col = False
            sym = False
            tit_ent = False
        up_plot_colors.append((state.get(), col))
        up_plot_symbols.append((state.get(), sym))
        up_plot_sizes.append((state.get(), size_ent))
        up_plot_titles.append((state.get(), tit_ent))
    
    try:
        linewidth_in_use = float(linewidth_enter.get())
    except:
        linewidth_in_use = 1
    
    orders_initial = ["first", "second", "third", "fourth", "fifth", "sixth", "seventh", "eighth", "ninth", "tenth", "eleventh", "twelfth", "thirteenth", "fourteenth", "fifteenth", "sixteenth"]
    orders = []
    for i in range(min(len(states), 16)):
        orders.append(orders_initial[i])
        
    order_color_dic = {}
    order_symbol_dic = {}
    order_size_dic = {}
    order_state_dic = {}
    order_title_dic = {}
    for i, order in enumerate(orders):
        key, color = up_plot_colors[i]
        if color is not False:
            if not color.startswith("#"):
                color = '#' + color
        order_color_dic[order] = color
        if key != "N/A":
            order_state_dic[order] = key
        else:
            order_state_dic[order] = False
        
        key, symbol = up_plot_symbols[i]
        if symbol is not False:
            symbol = chr(int(symbol.removeprefix('U+'), 16))
        order_symbol_dic[order] = symbol
        
        key, size = up_plot_sizes[i]
        order_size_dic[order] = size
        
        key, title = up_plot_titles[i]
        order_title_dic[order] = title
        
        if color is False or symbol is False or size is False or title is False:
            order_state_dic[order] = False
            
    global errorbar_capsize, errorbar_linewidth, errorbar_capthick
    errorbar_capsize = float(capsize_entry.get())
    errorbar_linewidth = float(errorbar_linewidth_entry.get())
    errorbar_capthick = float(capthick_entry.get())
        
global current_peptide_index
current_peptide_index = 0

    
class ProteinSelectDialog:
    def __init__(self, parent, protein_list):
        self.search_popup = tk.Toplevel(parent)
        self.search_popup.title("Protein Select")
        self.search_popup.geometry("300x100+600+400")
        self.search_popup.transient(parent)  # Makes this window a transient window of 'parent'
        self.search_popup.grab_set()  # Directs all events to this window until it is closed

        self.selected_value = tk.StringVar()
        self.dropdown = ttk.Combobox(self.search_popup, textvariable=self.selected_value, values=protein_list)
        self.dropdown.pack(pady=20)

        # Binding selection event to call on_protein_value_selected
        self.dropdown.bind("<<ComboboxSelected>>", self.on_protein_value_selected)

        # Wait for the window to close before continuing in the main program
        parent.wait_window(self.search_popup)

    def on_protein_value_selected(self, event):
        # When a selection is made, close the popup
        self.search_popup.destroy()

    def get_selected_value(self):
        # Return the selected value from the dropdown
        return self.selected_value.get()
    
def get_protein_selection(window, protein_list):
    if len(protein_list) > 1:
        dialog = ProteinSelectDialog(window, protein_list)
        return dialog.get_selected_value()
    return protein_list[0]



copy_statedic = {}
def create_example_plot():
    global legend_size_entry, leg_pos, legend_linewidth_entry, line_legend_entries
    try:
        legend_size = float(legend_size_entry.get())
    except:
        legend_size = 10
    try:
        legend_linewidth = float(legend_linewidth_entry.get())
    except:
        legend_linewidth = 1
        
    global current_peptide_index, search_on
    parse_data()
    try: 
        picture_widget.destroy()
    except:
        pass
    

    r_process_data()
    
    global linestyle_in_use
    if dot_chkval.get() == 0:
        linestyle_in_use = "-"
    if dot_chkval.get() == 1:
        linestyle_in_use = "--"
            
    
    
    
    all_peptides = []      
    for state in order_state_dic.values():
        if state != False and state != "False":
            protein = state.split("~")[0]
            for peptide in peplist[state]:
                if (protein, peptide) in all_peptides:
                    continue
                all_peptides.append((protein, peptide))  
                
    
    
    
    seg_proteins = True
    if seg_proteins == False:             
        sorted_all_peptides = sorted(all_peptides, key=lambda x: (int(pro_peptide_starts.get((x[0], x[1]), [0])[0]), len(x[1])))
    if seg_proteins == True:
        sorted_all_peptides = sorted(
            all_peptides,
            key=lambda x: (
                x[0],  # Sort primarily by protein name
                int(pro_peptide_starts.get((x[0], x[1]), [0])[0]),  # Secondary sort by start position
                len(x[1])  # Tertiary sort by peptide length
            )
        )
        
    
    max_protein_tps = {}
    for this_protein in protein_states.keys():
        max_state_tps = []
        for state in statedic_of_pepdic_raw2.keys():
            if not state.split("~")[0] == this_protein:
                continue
            small_tuplist = []
            for any_peptide, tuplist in statedic_of_pepdic_raw2[state].items():
                for each_tuple in tuplist:
                    small_tuplist.append(each_tuple[1])
            
            if show_last == False:
                max_tp_state = sorted(set(small_tuplist), reverse=True)[1]
            else:
                max_tp_state = max(small_tuplist)
            max_state_tps.append(max_tp_state)
        
        max_protein_tp = max(max_state_tps)
        max_protein_tps[this_protein] = max_protein_tp
        
    
    min_protein_tps = {}
    for this_protein in protein_states.keys():
        min_state_tps = []
        for state in statedic_of_pepdic_raw2.keys():
            if not state.split("~")[0] == this_protein:
                continue
            small_tuplist = []
            for any_peptide, tuplist in statedic_of_pepdic_raw2[state].items():
                for each_tuple in tuplist:
                    if each_tuple[1] != 0:
                        small_tuplist.append(each_tuple[1])
            
           
            min_tp_state = min(small_tuplist)
            min_state_tps.append(min_tp_state)
        
        min_protein_tp = min(min_state_tps)
        min_protein_tps[this_protein] = min_protein_tp
        
    
    if search_on is True:
        pep_search_num = search_for_startvalue()
        protein_list = search_for_protein()
        selected_protein = get_protein_selection(window, protein_list)
        peptide_search_found = False
        if not pep_search_num is False:
            for (protein, pep) in sorted_all_peptides:
                if protein == selected_protein:
                    if (pro_peptide_ends[(protein, pep)])[0] >= pep_search_num:
                        peptide_search_found = True
                        current_peptide_index = sorted_all_peptides.index((protein, pep))
                        break
            if peptide_search_found == False:
                current_peptide_index = len(sorted_all_peptides) - 1
                        
    search_on = False
    

    if current_peptide_index < 0:
        current_peptide_index = 0
        
    if current_peptide_index > (len(sorted_all_peptides) - 1):
        current_peptide_index = len(sorted_all_peptides) - 1
        
    
    current_peptide = sorted_all_peptides[current_peptide_index][1]
    current_protein = sorted_all_peptides[current_peptide_index][0]
    

        
    max_theo = get_max_theo(current_peptide)
    startvalues = pro_peptide_starts.get((current_protein, current_peptide), None)
    startvalue= int(startvalues[0])
    endvalues = pro_peptide_ends.get((current_protein, current_peptide), None)
    endvalue = int(endvalues[0])
    
    fig = Figure()
    ax = fig.add_subplot(111)

    ax.set_ylabel(y_enter.get())
    
    
    line_legend_entries = []
    legend_entries = []
    for state in statedic_of_pepdic_raw2:
        if not state.split("~")[0] == current_protein:
                continue
                
        if show_last is False:
            small_tuplist = []
            for key, value in statedic_of_pepdic_raw2.items():
                if not key.split("~")[0] == current_protein:
                    continue
                for any_peptide, tuplist in value.items():
                    for each_tuple in tuplist:
                        small_tuplist.append(each_tuple[1])
            max_tp_here = max(small_tuplist)
            
        up_list = []
        tp_list = []
        sd_list = []
        if correction is False:
            ax.set_ylim(0, max_theo)
            if change_scale is True:
                try:
                    ax.set_ylim(0, int(temp_max_plot_height_entry.get()))
                except:
                    pass
            if current_peptide in statedic_of_pepdic_raw2[state]:
                for up, tp in statedic_of_pepdic_raw2[state][current_peptide]:
                    up_list.append(up)
                    tp_list.append(tp)
                for sd, tp in statedic_of_sddic_raw2[state][current_peptide]:
                    sd_list.append(sd)
        if correction is True:
            ax.set_ylim(0, (max_theo + 2))
            if change_scale is True:
                try:
                    ax.set_ylim(0, int(temp_max_plot_height_entry.get()))
                except:
                    pass
            if current_peptide in statedic_of_pepdic_cor[state]:
                for up, tp in statedic_of_pepdic_cor[state][current_peptide]:
                    up_list.append(up)
                    tp_list.append(tp)
                for sd, tp in statedic_of_sddic_cor[state][current_peptide]:
                    sd_list.append(sd)
                
        if tp_list != []:
            max_timepoint = max(tp_list)


    
        if (correction is False and current_peptide in statedic_of_pepdic_raw2[state]) or (correction is True and current_peptide in statedic_of_pepdic_cor[state]):
            if tp_list != []:
                if show_last is True:
                    pass
                if show_last is False:
                    if tp_list[-1] == max_tp_here:
                        up_list = up_list[0:-1]
                        tp_list = tp_list[0:-1]
                        sd_list = sd_list[0:-1]

                if tp_list[0] == 0:
                    tp_list = tp_list[1:]
                    up_list = up_list[1:]
                    sd_list = sd_list[1:]
                
                
                filtered_pairs = [(up, tp) for up, tp in zip(up_list, tp_list) if up != -99999]
                if filtered_pairs:
                    up_list, tp_list = list(zip(*filtered_pairs))
                    up_list = list(up_list)
                    tp_list = list(tp_list)
                else:
                    continue
                    
                sd_list = [z if z != -99999 else np.nan for z in sd_list]
                
                if correction is True:
                    up_list = [z * max_theo for z in up_list]
                    sd_list = [z * max_theo if z != np.nan else np.nan for z in sd_list]
                    
                SE_list = [z / np.sqrt(float(SE_num_entry.get())) if z != np.nan else np.nan for z in sd_list]                      
                
                        
                for order, st in order_state_dic.items():
                    if st == state:
                        if show_error_bars == False:
                            ax.plot(tp_list, up_list, color=order_color_dic[order], linestyle=linestyle_in_use, linewidth = linewidth_in_use, label=order_title_dic[order])
                        else:
                            try:
                                ax.errorbar(tp_list, up_list, yerr=SE_list, color=order_color_dic[order], linestyle=linestyle_in_use, linewidth=linewidth_in_use, label=order_title_dic[order], capsize=errorbar_capsize, elinewidth=errorbar_linewidth, capthick=errorbar_capthick)
                            except Exception as e:
                                ax.plot(tp_list, up_list, color=order_color_dic[order], linestyle=linestyle_in_use, linewidth = linewidth_in_use, label=order_title_dic[order])
                            
#                        line = Line2D([0], [0], color=order_color_dic[order], linestyle='-', linewidth=legend_linewidth, label=order_title_dic[order])
#                        line_legend_entries.append(line)
                        for x, y in zip(tp_list, up_list):
                            ax.text(x, y, order_symbol_dic[order], color=order_color_dic[order], ha='center', va='center', fontsize=order_size_dic[order])
#        else:
#            for order, st in order_state_dic.items():
#                if st == state:
#                    line = Line2D([0], [0], color=order_color_dic[order], linestyle='-', linewidth=legend_linewidth, label=order_title_dic[order])
#                    line_legend_entries.append(line)
                    
                 
    for state in statedic_of_pepdic_raw2:
#        if state.split("~")[0] == current_protein:
#                continue
        for order, st in order_state_dic.items():
            if st == state:
                line = Line2D([0], [0], color=order_color_dic[order], linestyle='-', linewidth=legend_linewidth, label=order_title_dic[order])
                line_legend_entries.append(line)
            
    if change_scale is True or calculate_theoretical_back_exchange is True:
        old_max_theo = max_theo
        try:
            max_theo = int(np.ceil(temp_max_plot_height_entry.get()))
        except:
            max_theo = int(np.ceil(max_theo))
                        
    if correction is True:
        max_theo += 2
        
    if max_theo <= 7:
        step = 1
    elif max_theo == 8 or max_theo == 10:
        step = 2
    elif max_theo == 9:
        step = 3
    elif max_theo in [11, 13, 14]:
        step = 2
    elif max_theo in [12, 15]:
        step = 3
    elif max_theo in [16, 17, 19, 20]:
        step = 4
    elif max_theo == 18:
        step = 6
    elif max_theo == 21:
        step = 7
    elif max_theo in [22, 23, 24]:
        step = 4
    elif max_theo >= 25:
        step = 5
        
    y_ticks = list(range(0, max_theo + 1, step))

    if max_theo % step > 1 and max_theo < 25:
        y_ticks.append(max_theo)
    if max_theo % step > 2 and max_theo >= 25:
        y_ticks.append(max_theo)
    ax.set_yticks(y_ticks)
    
    if correction is True:
        max_theo -= 2
    
#    if maxD_dash is True:
#        print(statedic_of_pepdic_raw2[state])
#        upt_tp_tups = statedic_of_pepdic_raw2[state][peptide]
#        maxfile_up_tp_tups = statedic_of_pepdic_raw2[maxfile][peptide]
#        max_tp = max(maxfile_up_tp_tups, key=lambda x: x[1])[1]
#        maxD_uptake = next(x[0] for x in maxfile_up_tp_tups if x[1] == max_tp)
#        ax.axhline(y=maxD_uptake, color='black', linestyle='--', linewidth=1)
    
    if change_scale is True or calculate_theoretical_back_exchange is True:
        max_theo = old_max_theo
    
    ax.set_xlabel(x_enter.get())
    ax.set_xscale('log')
    #ax.xaxis.set_major_locator(ticker.LogLocator(base=10.0, numticks=10))
    ax.xaxis.set_major_locator(ticker.LogLocator(base=10.0, subs=[1.0, 10.0], numticks=10))
    ax.xaxis.set_minor_locator(ticker.LogLocator(base=10.0, subs='auto', numticks=10))

    fig.subplots_adjust(left=0.15)
    
    if len(current_peptide) < 20:
        ax.set_title(f'$^{{{startvalue}}} {current_peptide} ^{{{endvalue}}}$')
    else:
        new_title = current_peptide[:9] + "..." + current_peptide[-9:]
        ax.set_title(f'$^{{{startvalue}}} {new_title} ^{{{endvalue}}}$')
    fig.subplots_adjust(bottom=0.2)
    if leg_ur or leg_ul or leg_bl or leg_br is True:
        legend = ax.legend(handles=line_legend_entries, loc=leg_pos, prop={'size': legend_size})
        
    picture = FigureCanvasTkAgg(fig, master=window)
    picture_widget = picture.get_tk_widget()
    picture_widget.place(x=1065, y=455, width=390, height = 425)
    
    
    save_uptakeplot_button = tk.Button(window, text="Save as PNG", command=lambda: save_figure(fig, startvalue, endvalue))
    save_uptakeplot_button.place(x=970, y=800)
    
def save_figure(fig, startvalue, endvalue):
    figure_title = f"{startvalue}-{endvalue}"
    png_tit = filedialog.asksaveasfilename(initialfile=figure_title,
                                            defaultextension=".png",
                                            filetypes=[("PNG files", "*.png")])
    if not png_tit:
        tk.messagebox.showinfo("Save PNG", f"The PNG was not saved.")
        return
    else:
        fig.savefig(png_tit, dpi=1000)
        tk.messagebox.showinfo("Save PDF", f"The PNG has been saved as '{png_tit}'.")
    
def check_button_clicks2():
        
    for widget in window.winfo_children():
        if widget.winfo_x() > 370 and widget != settings_bt:
            widget.destroy()
    for item in canvas.find_all():
        coords = canvas.coords(item)
        # For lines and shapes, coords are a list of x, y pairs. We check the first x-coordinate.
        if coords and coords[0] > 370:
            canvas.delete(item)
                
    global states, peplist, startvallist, endvallist, state_options, protein_states, sdbt_clicked, cdbt_clicked, cd_dfs, examiner_dfs, class_peptides, workbench_dfs, sd_dfs
    class_peptides = []
    
    states = {}

    peplist = {}

    startvallist = {}

    endvallist = {}

    state_options =[]

    protein_states = {}


    msg2 = tk.Label(window, text="Add Differences")
    msg2.place(x=395, y=5)
    dif_bt = tk.Button(window, text="+Dif",bg="white",fg="black",command=dif_bt_done)
    dif_bt.place(x=495, y=10)


    x1, y1 = 372, 10  # Top-left coordinates of the rectangle
    x2, y2 = 920, 450  # Bottom-right coordinates of the rectangle
    canvas.create_rectangle(x1, y1, x2, y2, outline="black", fill="")

    create_format_box()

    create_run_box()
    

    if sdbt_clicked is True:
        for df in sd_dfs:
            for index, row in df.iterrows():
                protein = row["Protein"]
                if pd.notna(row["Modification"]):
                    protein = row["Protein"] + "_PTM"
                if protein not in protein_states:
                    protein_states[protein] = []
                state = protein + "~" + row["State"].replace("~", "_").replace("/", "_").replace(":", ";").replace("|", "_")
                if state not in states:
                    protein_states[protein].append(state)
                    states[state] = True
                    peplist[state] = list()
                    startvallist[state] = list()
                    endvallist[state] = list()
                if row["Sequence"] not in peplist[state]:
                    peplist[state].append(row["Sequence"])
                    startvallist[state].append(row["Start"])
                    endvallist[state].append(row["End"])
        
                
            
                uptake = row["Uptake"]
                if pd.isna(row["Uptake"]):
                    uptake = -99999
                    sd = -99999
                    
                sd = row["Uptake SD"]
                if pd.isna(row["Uptake SD"]):
                    sd = -99999
                if sd == 0:
                    sd = -99999
        
                peptide_instance = sd_peptide(Sequence = row["Sequence"], Startvalue = row["Start"], Endvalue = row["End"], State = state, Protein = protein, Timepoint = row["Exposure"], Uptake = uptake, Stdev = sd, RT = row["RT"])
                
                class_peptides.append(peptide_instance)
                
    if cdbt_clicked is True:
        for df in cd_dfs:
            for index, row in df.iterrows():
                protein = row["Protein"]
                if pd.notna(row["Modification"]):
                    protein = row["Protein"] + "_PTM"
                if protein not in protein_states:
                    protein_states[protein] = []
                state = protein + "~" + row["State"].replace("~", "_").replace("/", "_").replace(":", ";").replace("|", "_")
                if state not in states:
                    protein_states[protein].append(state)
                    states[state] = True
                    peplist[state] = list()
                    startvallist[state] = list()
                    endvallist[state] = list()
                if row["Sequence"] not in peplist[state]:
                    peplist[state].append(row["Sequence"])
                    startvallist[state].append(row["Start"])
                    endvallist[state].append(row["End"])
                file = row["File"]
                
                center = row["Center"]
                if pd.isna(row["Center"]):
                    center = -99999
     


                peptide_instance = cd_peptide(Sequence = row["Sequence"], Startvalue = row["Start"], Endvalue = row["End"], State = state, Protein = protein, File = file, Timepoint = row["Exposure"], Charge = row["z"], Intensity = row["Inten"], Center = center, MHP = row["MHP"], Uptake = -99999, Stdev = -99999, RT = row["RT"], und_std = -99999)

                class_peptides.append(peptide_instance)
                
    
    if examiner_clicked is True:
        for df in examiner_dfs:
            for index, row in df.iterrows():
                protein = row["Protein"]
                if protein not in protein_states:
                    protein_states[protein] = []
                state = protein + "~" + row["Protein State"].replace("~", "_").replace("/", "_").replace(":", ";").replace("|", "_")
                if state not in states:
                    protein_states[protein].append(state)
                    states[state] = True
                    peplist[state] = list()
                    startvallist[state] = list()
                    endvallist[state] = list()
                if row["Sequence"] not in peplist[state]:
                    peplist[state].append(row["Sequence"])
                    startvallist[state].append(row["Start"])
                    endvallist[state].append(row["End"])
                    
                timepoint = row["Deut Time (sec)"]
                if timepoint == "MAX":
                    timepoint = 3245908.2345029
                    
                uptake = row["#D"]
                if pd.isna(row["#D"]):
                    uptake = -99999
                    sd = -99999
                    
                sd = row["Stddev"]
                if pd.isna(row["Stddev"]):
                    sd = -99999
                if sd == 0:
                    sd = -99999
                
                    
                peptide_instance = examiner_peptide(Sequence = row["Sequence"], Startvalue = row["Start"], Endvalue = row["End"], State = state, Protein = protein, Timepoint = float(timepoint), Uptake = uptake, Stdev = sd, RT = row["RT (min)"])
                
                class_peptides.append(peptide_instance)
                    
    if workbench_clicked is True:
        for df in workbench_dfs:
            for index, row in df.iterrows():
                protein = 'protein'
                if protein not in protein_states:
                    protein_states[protein] = []
                if pd.isna(row["sample"]):
                    continue
                state = protein + "~" + row["sample"].replace("~", "_").replace("/", "_").replace(":", ";").replace("|", "_")
                if state not in states:
                    protein_states[protein].append(state)
                    states[state] = True
                    peplist[state] = list()
                    startvallist[state] = list()
                    endvallist[state] = list()
                if row["peptide"] not in peplist[state]:
                    peplist[state].append(row["peptide"])
                    startvallist[state].append(row["start"])
                    endvallist[state].append(row["end"])
                    
                timepoint = row["timepoint"]
                while type(timepoint) == str:
                    try:
                        timepoint = float(timepoint)
                    except:
                        timepoint = timepoint[:-1]
                        
                center = row["centroid"]
                if pd.isna(row["centroid"]):
                    center = -99999
                    
                file = row["replicate"]
                    
                rt_range = row["rt_aggregate"][1:-1].split("-")
                RT = np.mean([float(x) for x in rt_range])
                    
                peptide_instance = workbench_peptide(Sequence = row["peptide"], Startvalue = row["start"], Endvalue = row["end"], State = state, Protein = protein, Timepoint = timepoint, Charge = row["charge"], Center = center, Uptake = -99999, Stdev = -99999, Discarded = row["discarded_replicate"], RT=RT, und_std = -99999, File = file)
                
            
                if peptide_instance.Discarded == False:
                    class_peptides.append(peptide_instance)
                
    global peptide_RT_dic     
    unique_sequences = {peptide_instance.Sequence for peptide_instance in class_peptides}
    peptide_RT_dic = {sequence: np.mean([peptide_instance.RT for peptide_instance in class_peptides if peptide_instance.Sequence == sequence])for sequence in unique_sequences}

    for state in states:
        if state not in state_options:
            state_options.append(state)

            
    if exp_bt_on_c is True:
        make_maxdic_dropdowns()
    
    
    r_process_data()
    create_uptakeplot_box()
    dif_bt_done()



def make_maxdic_dropdowns():
    global vsb, maxdic_canvas, maxdic_frame
    def onFrameConfigure(maxdic_canvas):
        maxdic_canvas.configure(scrollregion=maxdic_canvas.bbox("all"))
        
    def remove_focus(event=None):
        window.focus_set()  # Set focus to another widget
        
    try:
        maxdic_canvas.destroy()
    except:
        pass
    try:
        maxdic_frame.destroy()
    except:
        pass
    
    global maxdic, dropdowns, snum, dropdown_widgets, label_widgets
    maxdic_canvas = tk.Canvas(window, borderwidth=0)
    vsb = ttk.Scrollbar(window, orient="vertical", command=maxdic_canvas.yview)
    maxdic_canvas.configure(yscrollcommand=vsb.set)
    vsb.place(x=350, y=340, width=20, height=530)
    maxdic_canvas.place(x=15, y=340, width=330, height=530)
    
    maxdic_frame = ttk.Frame(maxdic_canvas)
    maxdic_canvas.create_window((0, 0), window=maxdic_frame, anchor="nw")
    maxdic_frame.bind("<Configure>", lambda event, canvas=maxdic_canvas: onFrameConfigure(canvas))

    
    snum = 0
    dropdown_widgets = []  # List to store dropdown widgets
    label_widgets = []  # List to store label widgets
    for state in states:
        state_label = tk.Label(maxdic_frame, text=state + ":")
        font_size = 12
        while state_label.winfo_reqwidth() < 138:
            font_size = font_size+1
            state_label.config(font=("Arial", font_size))
        while state_label.winfo_reqwidth() > 138:
            font_size = font_size-1
            state_label.config(font=("Arial", font_size))
        state_label.grid(row=snum, column=0)
        label_widgets.append(state_label)

        dropdown_var = tk.StringVar(value=state)  # Create a unique StringVar for each dropdown
        dropdown = ttk.Combobox(maxdic_frame, values=state_options, width=28)
        dropdown.set(dropdown_var.get())
        dropdown.grid(row=snum, column=1)
        dropdown.bind("<<ComboboxSelected>>", remove_focus)
        dropdown_widgets.append(dropdown)

        dropdowns[state] = dropdown  # Map the dropdown variable to the state

        snum += 1






def create_custom_state():
    global state_options, avg1, avg2, popup_window

    popup_window = tk.Toplevel(window)  # Create a new window for the popup menu
    popup_window.geometry("200x250")

     # Calculate the desired position for the popup window
    x = window.winfo_x() + 100  # Adjust the value as needed
    y = window.winfo_y() + 200  # Adjust the value as needed

    # Set the position of the popup window
    popup_window.geometry(f"+{x}+{y}")

    average_lb = tk.Label(popup_window, text="Average two states:")
    average_lb.grid(column=0, row=0)


    options = state_options


    avg1 = ttk.Combobox(popup_window, values=options, width=28)
    avg1.grid(column=0, row=1, padx=5)

    avg2 = ttk.Combobox(popup_window, values=options, width=28)
    avg2.grid(column=0, row=2, padx=5)

    state_save_bt = tk.Button(popup_window, text="Save State", bg="white", fg="black", command=state_save)
    state_save_bt.grid(column=0, row=3)


    # You can further customize the popup window properties here if needed

    popup_window.transient(window)  # Set the main window as the parent of the popup window
    popup_window.grab_set()  # Grab the focus to the popup window
    popup_window.mainloop()  # Start the main loop for the popup window

def state_save():
    global state_options, popup_window
    state_options.append("pyAVG|"+f'{avg1.get()}'+'|' f'{avg2.get()}')
    state_save_lb = tk.Label(popup_window, text="State Saved")
    state_save_lb.grid(column=0, row=4)
    state_sav_lb2 = tk.Label(popup_window, text="(Only for maxD)")
    state_sav_lb2.grid(column=0, row=5)
    make_maxdic_dropdowns()


    global new_states_dic
    new_states_dic = {}
    for state_o in state_options:
        if state_o.startswith("pyAVG"):
            toavg_list = list()
            first_split = state_o.split("|")
            toavg_list.extend(first_split)
            toavg_list_con = [toavg_list[1], toavg_list[2]]
            new_states_dic[state_o] = toavg_list_con







def maxD_Da_dif_bt_on():
    global maxD_Da_dif_on_c, maxD_dif_bt_list
    maxD_Da_dif_on_c = True
    maxD_Da_dif_bt = tk.Button(window, text="Corrected (Da)", fg="white", bg="green", width=13, command=lambda: [maxD_Da_dif_bt_off(), maxD_rfu_dif_bt_on()])
    maxD_Da_dif_bt.place(x=250, y=283)
    difference_color_scheme_dropdown.set("Da_green_blue.json")
    localized_color_scheme_dropdown.set("9_Da_green_blue.json")
    maxD_dif_bt_list.append(maxD_Da_dif_bt)
    
def maxD_Da_dif_bt_off():
    global maxD_Da_dif_on_c, maxD_dif_bt_list
    maxD_Da_dif_on_c = False
    maxD_Da_dif_bt = tk.Button(window, text="Corrected (Da)", fg="black", bg="orange", width=13, command=lambda: [maxD_Da_dif_bt_on(), maxD_rfu_dif_bt_off()])
    maxD_Da_dif_bt.place(x=250, y=283)
    maxD_dif_bt_list.append(maxD_Da_dif_bt)
    
def maxD_rfu_dif_bt_on():
    global maxD_rfu_dif_on_c, maxD_Da_dif_bt
    maxD_rfu_dif_on_c = True
    maxD_rfu_dif_bt = tk.Button(window, text="Corrected (RFU)", fg="white", bg="green", width=13, command=lambda: [maxD_Da_dif_bt_on(), maxD_rfu_dif_bt_off()])
    maxD_rfu_dif_bt.place(x=130, y=283)
    difference_color_scheme_dropdown.set("RFU_green_blue.json")
    localized_color_scheme_dropdown.set("9_RFU_green_blue.json")
    maxD_dif_bt_list.append(maxD_rfu_dif_bt)
    
def maxD_rfu_dif_bt_off():
    global maxD_rfu_dif_on_c, maxD_rfu_dif_bt, maxD_dif_bt_list
    maxD_rfu_dif_on_c = False
    maxD_rfu_dif_bt = tk.Button(window, text="Corrected (RFU)", fg="black", bg="orange", width=13, command=lambda: [maxD_Da_dif_bt_off(), maxD_rfu_dif_bt_on()])
    maxD_rfu_dif_bt.place(x=130, y=283)
    maxD_dif_bt_list.append(maxD_rfu_dif_bt)
    

exp_bt_on_c = False
theo_bt_on_c = False
def exp_bt_on():
    global exp_bt_on_c, maxD_label, custom_state_bt, exp_st_lb, maxD_peptides_lb, maxD_label_line, state_label_line, choose_rfu_or_da_label, maxD_dif_bt_list, maxD_rfu_dif_on_c, maxD_Da_dif_on_c
    
    try:
        for item in maxD_dif_bt_list:
            item.destroy()
    except:
        pass
    
    exp_bt2 = tk.Button(window, text="MaxD Corrected",bg="green",fg="white",command=lambda: [exp_bt_off(), theo_bt_on()])
    exp_bt2.place(x=150, y=220)
    exp_bt_on_c = True
    maxD_dif_bt_list = []

    
    global maxdic, dropdowns
    maxdic = {}  # Initialize an empty dictionary
    dropdowns = {}  # Initialize an empty dictionary to store dropdown variables

    exp_st_lb = tk.Label(window, text="Protein~State")
    exp_st_lb.place(x=45, y=313)
    maxD_peptides_lb = tk.Label(window, text="maxD Peptide Extraction")
    maxD_peptides_lb.place(x=175, y=313)
    
#    custom_state_bt = tk.Button(window, text="Custom State", bg="white", fg="black", command=create_custom_state)
#    custom_state_bt.place(x=275, y=220)
    
    
    choose_rfu_or_da_label = tk.Label(window, text="Show Differences as:")
    choose_rfu_or_da_label.place(x=15, y=283)
    
    maxD_Da_dif_on_c = True
    maxD_Da_dif_bt = tk.Button(window, text="Corrected (Da)", fg="white", bg="green", width=13, command=lambda: [maxD_Da_dif_bt_off(), maxD_rfu_dif_bt_on()])
    maxD_Da_dif_bt.place(x=250, y=283)
    maxD_dif_bt_list.append(maxD_Da_dif_bt)
    
    maxD_rfu_dif_on_c = False
    maxD_rfu_dif_bt = tk.Button(window, text="Corrected (RFU)", fg="black", bg="orange", width=13, command=lambda: [maxD_Da_dif_bt_off(), maxD_rfu_dif_bt_on()])
    maxD_rfu_dif_bt.place(x=130, y=283)
    maxD_dif_bt_list.append(maxD_rfu_dif_bt)
    

    x1 = 17
    y = 337
    x2 = 148
    state_label_line = canvas.create_line(x1, y, x2, y)

    x1 = 160
    y=337
    x2 = 340
    maxD_label_line = canvas.create_line(x1, y, x2, y)
    
    check_button_clicks2()

def set_default_be():
    if os.path.exists("Alterations/DefaultBE.json"):
        os.remove("Alterations/DefaultBE.json")
        
    default_be = be_entry.get()
    with open("Alterations/DefaultBE.json", 'w') as file:
        json.dump({"Default Back Exchange": default_be}, file)
    tk.messagebox.showinfo("Back Exchange Saved", f"New Default Back Exchange Saved as {default_be}%")

colors_already_formatted = False
def theo_bt_on():
    def on_change_be(sv):
        try_worked = False
        try:
            float_be = float(sv.get())
            try_worked = True
        except:
            pass
        if try_worked == True:
            if float_be == 0:
                try:
                    uptake_color_scheme_dropdown.set("uncorrected_df.json")
                except:
                    pass
            else:
                try:
                    uptake_color_scheme_dropdown.set("corrected_df.json")
                except:
                    pass
        if try_worked == False:
            try:
                uptake_color_scheme_dropdown.set("uncorrected_df.json")
            except:
                pass

        
    global theo_bt_on_c, be_entry, per_label, back_exchange_label, back_exchange_label, be_default_bt
    theo_bt2 = tk.Button(window, text="No maxD",bg="green",fg="white",command=lambda: [theo_bt_off(), exp_bt_on()])
    theo_bt2.place(x=50, y=220)
    theo_bt_on_c = True
    back_exchange_label = tk.Label(window, text="Correct for Back Exchange:")
    back_exchange_label.place(x=20, y=285)
    be_default_bt = tk.Button(window, text="Set as default", command=set_default_be)
    be_default_bt.place(x=220, y=285)
    global be_entry
    
    text_var_be = tk.StringVar()
    text_var_be.trace_add("write", lambda name, index, mode, sv=text_var_be: on_change_be(sv)) 
    be_entry = tk.Entry(window, textvariable=text_var_be, width=5)
    if os.path.exists("Alterations/DefaultBE.json"):
        with open("Alterations/DefaultBE.json", 'r') as file:
            json_data = json.load(file)
            default_be = float(json_data["Default Back Exchange"])
            be_entry.insert(0, default_be)
    else:
        be_entry.insert(0, "0")
        if colors_already_formatted is True:
            try:
                uptake_color_scheme_dropdown.set("uncorrected_df.json")
            except:
                pass
    be_entry.place(x=170, y=285)
    
    
    per_label = tk.Label(window, text="%")
    per_label.place(x=200, y=285)

    check_button_clicks2()


def exp_bt_off():
    global exp_bt_on_c
    exp_bt1 = tk.Button(window, text="MaxD Corrected",bg="orange",fg="black",command=lambda: [theo_bt_off(), exp_bt_on()])
    exp_bt1.place(x=150, y=220)
    exp_bt_on_c = False
    global dropdown_widgets, label_widgets, custom_state_bt
    try:
        for dropdown in dropdown_widgets:
            dropdown.destroy()
        for label in label_widgets:
            label.destroy()
    except:
        pass
    try:
        custom_state_bt.destroy()
        exp_st_lb.destroy()
        maxD_peptides_lb.destroy()
    except:
        pass
    try:
        canvas.delete(maxD_label_line)
        canvas.delete(state_label_line)
    except:
        pass
    try:
        choose_rfu_or_da_label.destroy()
        for item in maxD_dif_bt_list:
            item.destroy()
    except:
        pass
    try:
        vsb.destroy()
    except:
        pass
    try:
        maxdic_canvas.destroy()
    except:
        pass
    try:
        maxdic_frame.destroy()
    except:
        pass
        


def theo_bt_off():
    global theo_bt_on_c, be_entry, per_label, back_exchange_label, be_default_bt
    theo_bt1 = tk.Button(window, text="No maxD",bg="orange",fg="black",command=lambda: [exp_bt_off(), theo_bt_on()])
    theo_bt1.place(x=50, y=220)
    theo_bt_on_c = False
    try:
        back_exchange_label.destroy()
        per_label.destroy()
        be_entry.destroy()
        be_default_bt.destroy()
    except:
        pass

onedif_state = tk.StringVar()
twodif_state = tk.StringVar()

def update_dropdown_options(event):
    if onedif_state.get() != "":
        onedif_dropdown["values"] = state_options
    if twodif_state.get() != "":
        twodif_dropdown["values"] = state_options


def dif_bt_done():
    minus_label = tk.Label(window, text="-")
    minus_label.place(x=575, y=60)

    filtered_options = []
    for state in state_options:
        if not state.startswith("pyAVG"):
            filtered_options.append(state)

    global onedif_dropdown
    onedif_dropdown = ttk.Combobox(window, values=filtered_options, width=28)
    onedif_dropdown.place(x=380, y=60)
    onedif_dropdown.bind("<<ComboboxSelected>>", update_dropdown_options)

    global twodif_dropdown
    twodif_dropdown = ttk.Combobox(window, values=filtered_options, width=28)
    twodif_dropdown.place(x=590, y=60)
    twodif_dropdown.bind("<<ComboboxSelected>>", update_dropdown_options)

    global s_entry1
    s_entry1 = tk.Entry(window)
    s_entry1.place(x=790, y=60)
    d_label = tk.Label(window, text="Difference Title:")
    d_label.place(x=805, y=35)
    onedif_lb = tk.Label(window, text="Protein~State One")
    onedif_lb.place(x=415, y=35)
    twodif_lb = tk.Label(window, text="Protein~State Two")
    twodif_lb.place(x=620, y=35)

    dif_bt2 = tk.Button(window, text="+Dif", bg="white", fg="black", command=dif_bt2_done)
    dif_bt2.place(x=495, y=10)

def dif_bt2_done():
    minus_label2 = tk.Label(window, text="-")
    minus_label2.place(x=575, y=90)

    filtered_options = []
    for state in state_options:
        if not state.startswith("pyAVG"):
            filtered_options.append(state)

    global onedif2_dropdown
    onedif2_dropdown = ttk.Combobox(window, values=filtered_options, width=28)
    onedif2_dropdown.place(x=380, y=90)
    onedif2_dropdown.bind("<<ComboboxSelected>>", update_dropdown_options)

    global twodif2_dropdown
    twodif2_dropdown = ttk.Combobox(window, values=filtered_options, width=28)
    twodif2_dropdown.place(x=590, y=90)
    twodif2_dropdown.bind("<<ComboboxSelected>>", update_dropdown_options)

    global s_entry2
    s_entry2 = tk.Entry(window)
    s_entry2.place(x=790, y=90)

    dif_bt3 = tk.Button(window, text="+Dif", bg="white", fg="black", command=dif_bt3_done)
    dif_bt3.place(x=495, y=10)

def dif_bt3_done():
    minus_label3 = tk.Label(window, text="-")
    minus_label3.place(x=575, y=120)

    filtered_options = []
    for state in state_options:
        if not state.startswith("pyAVG"):
            filtered_options.append(state)

    global onedif3_dropdown
    onedif3_dropdown = ttk.Combobox(window, values=filtered_options, width=28)
    onedif3_dropdown.place(x=380, y=120)
    onedif3_dropdown.bind("<<ComboboxSelected>>", update_dropdown_options)

    global twodif3_dropdown
    twodif3_dropdown = ttk.Combobox(window, values=filtered_options, width=28)
    twodif3_dropdown.place(x=590, y=120)
    twodif3_dropdown.bind("<<ComboboxSelected>>", update_dropdown_options)

    global s_entry3
    s_entry3 = tk.Entry(window)
    s_entry3.place(x=790, y=120)

    dif_bt4 = tk.Button(window, text="+Dif", bg="white", fg="black", command=dif_bt4_done)
    dif_bt4.place(x=495, y=10)

def dif_bt4_done():
    minus_label4 = tk.Label(window, text="-")
    minus_label4.place(x=575, y=150)

    filtered_options = []
    for state in state_options:
        if not state.startswith("pyAVG"):
            filtered_options.append(state)

    global onedif4_dropdown
    onedif4_dropdown = ttk.Combobox(window, values=filtered_options, width=28)
    onedif4_dropdown.place(x=380, y=150)
    onedif4_dropdown.bind("<<ComboboxSelected>>", update_dropdown_options)

    global twodif4_dropdown
    twodif4_dropdown = ttk.Combobox(window, values=filtered_options, width=28)
    twodif4_dropdown.place(x=590, y=150)
    twodif4_dropdown.bind("<<ComboboxSelected>>", update_dropdown_options)

    global s_entry4
    s_entry4 = tk.Entry(window)
    s_entry4.place(x=790, y=150)

    dif_bt5 = tk.Button(window, text="+Dif", bg="white", fg="black", command=dif_bt5_done)
    dif_bt5.place(x=495, y=10)

def dif_bt5_done():
    minus_label5 = tk.Label(window, text="-")
    minus_label5.place(x=575, y=180)

    filtered_options = []
    for state in state_options:
        if not state.startswith("pyAVG"):
            filtered_options.append(state)

    global onedif5_dropdown
    onedif5_dropdown = ttk.Combobox(window, values=filtered_options, width=28)
    onedif5_dropdown.place(x=380, y=180)
    onedif5_dropdown.bind("<<ComboboxSelected>>", update_dropdown_options)

    global twodif5_dropdown
    twodif5_dropdown = ttk.Combobox(window, values=filtered_options, width=28)
    twodif5_dropdown.place(x=590, y=180)
    twodif5_dropdown.bind("<<ComboboxSelected>>", update_dropdown_options)

    global s_entry5
    s_entry5 = tk.Entry(window)
    s_entry5.place(x=790, y=180)

    dif_bt6 = tk.Button(window, text="+Dif", bg="white", fg="black", command=dif_bt6_done)
    dif_bt6.place(x=495, y=10)

def dif_bt6_done():
    minus_label6 = tk.Label(window, text="-")
    minus_label6.place(x=575, y=210)

    filtered_options = []
    for state in state_options:
        if not state.startswith("pyAVG"):
            filtered_options.append(state)

    global onedif6_dropdown
    onedif6_dropdown = ttk.Combobox(window, values=filtered_options, width=28)
    onedif6_dropdown.place(x=380, y=210)
    onedif6_dropdown.bind("<<ComboboxSelected>>", update_dropdown_options)

    global twodif6_dropdown
    twodif6_dropdown = ttk.Combobox(window, values=filtered_options, width=28)
    twodif6_dropdown.place(x=590, y=210)
    twodif6_dropdown.bind("<<ComboboxSelected>>", update_dropdown_options)

    global s_entry6
    s_entry6 = tk.Entry(window)
    s_entry6.place(x=790, y=210)

    dif_bt7 = tk.Button(window, text="+Dif", bg="white", fg="black", command=dif_bt7_done)
    dif_bt7.place(x=495, y=10)

def dif_bt7_done():
    minus_label7 = tk.Label(window, text="-")
    minus_label7.place(x=575, y=240)

    filtered_options = []
    for state in state_options:
        if not state.startswith("pyAVG"):
            filtered_options.append(state)

    global onedif7_dropdown
    onedif7_dropdown = ttk.Combobox(window, values=filtered_options, width=28)
    onedif7_dropdown.place(x=380, y=240)
    onedif7_dropdown.bind("<<ComboboxSelected>>", update_dropdown_options)

    global twodif7_dropdown
    twodif7_dropdown = ttk.Combobox(window, values=filtered_options, width=28)
    twodif7_dropdown.place(x=590, y=240)
    twodif7_dropdown.bind("<<ComboboxSelected>>", update_dropdown_options)

    global s_entry7
    s_entry7 = tk.Entry(window)
    s_entry7.place(x=790, y=240)

    dif_bt8 = tk.Button(window, text="+Dif", bg="white", fg="black", command=dif_bt8_done)
    dif_bt8.place(x=495, y=10)

def dif_bt8_done():
    minus_label8 = tk.Label(window, text="-")
    minus_label8.place(x=575, y=270)

    filtered_options = []
    for state in state_options:
        if not state.startswith("pyAVG"):
            filtered_options.append(state)

    global onedif8_dropdown
    onedif8_dropdown = ttk.Combobox(window, values=filtered_options, width=28)
    onedif8_dropdown.place(x=380, y=270)
    onedif8_dropdown.bind("<<ComboboxSelected>>", update_dropdown_options)

    global twodif8_dropdown
    twodif8_dropdown = ttk.Combobox(window, values=filtered_options, width=28)
    twodif8_dropdown.place(x=590, y=270)
    twodif8_dropdown.bind("<<ComboboxSelected>>", update_dropdown_options)

    global s_entry8
    s_entry8 = tk.Entry(window)
    s_entry8.place(x=790, y=270)

    dif_bt9 = tk.Button(window, text="+Dif", bg="white", fg="black", command=dif_bt9_done)
    dif_bt9.place(x=495, y=10)
def dif_bt9_done():
    minus_label9 = tk.Label(window, text="-")
    minus_label9.place(x=575, y=300)

    filtered_options = []
    for state in state_options:
        if not state.startswith("pyAVG"):
            filtered_options.append(state)

    global onedif9_dropdown
    onedif9_dropdown = ttk.Combobox(window, values=filtered_options, width=28)
    onedif9_dropdown.place(x=380, y=300)
    onedif9_dropdown.bind("<<ComboboxSelected>>", update_dropdown_options)

    global twodif9_dropdown
    twodif9_dropdown = ttk.Combobox(window, values=filtered_options, width=28)
    twodif9_dropdown.place(x=590, y=300)
    twodif9_dropdown.bind("<<ComboboxSelected>>", update_dropdown_options)

    global s_entry9
    s_entry9 = tk.Entry(window)
    s_entry9.place(x=790, y=300)

    dif_bt10 = tk.Button(window, text="+Dif", bg="white", fg="black", command=dif_bt10_done)
    dif_bt10.place(x=495, y=10)

def dif_bt10_done():
    minus_label10 = tk.Label(window, text="-")
    minus_label10.place(x=575, y=330)

    filtered_options = []
    for state in state_options:
        if not state.startswith("pyAVG"):
            filtered_options.append(state)

    global onedif10_dropdown
    onedif10_dropdown = ttk.Combobox(window, values=filtered_options, width=28)
    onedif10_dropdown.place(x=380, y=330)
    onedif10_dropdown.bind("<<ComboboxSelected>>", update_dropdown_options)

    global twodif10_dropdown
    twodif10_dropdown = ttk.Combobox(window, values=filtered_options, width=28)
    twodif10_dropdown.place(x=590, y=330)
    twodif10_dropdown.bind("<<ComboboxSelected>>", update_dropdown_options)

    global s_entry10
    s_entry10 = tk.Entry(window)
    s_entry10.place(x=790, y=330)

    dif_bt11 = tk.Button(window, text="+Dif", bg="white", fg="black", command=dif_bt11_done)
    dif_bt11.place(x=495, y=10)
def dif_bt11_done():
    minus_label11 = tk.Label(window, text="-")
    minus_label11.place(x=575, y=360)

    filtered_options = []
    for state in state_options:
        if not state.startswith("pyAVG"):
            filtered_options.append(state)

    global onedif11_dropdown
    onedif11_dropdown = ttk.Combobox(window, values=filtered_options, width=28)
    onedif11_dropdown.place(x=380, y=360)
    onedif11_dropdown.bind("<<ComboboxSelected>>", update_dropdown_options)

    global twodif11_dropdown
    twodif11_dropdown = ttk.Combobox(window, values=filtered_options, width=28)
    twodif11_dropdown.place(x=590, y=360)
    twodif11_dropdown.bind("<<ComboboxSelected>>", update_dropdown_options)

    global s_entry11
    s_entry11 = tk.Entry(window)
    s_entry11.place(x=790, y=360)

    dif_bt12 = tk.Button(window, text="+Dif", bg="white", fg="black", command=dif_bt12_done)
    dif_bt12.place(x=495, y=10)

def dif_bt12_done():
    minus_label12 = tk.Label(window, text="-")
    minus_label12.place(x=575, y=390)

    filtered_options = []
    for state in state_options:
        if not state.startswith("pyAVG"):
            filtered_options.append(state)

    global onedif12_dropdown
    onedif12_dropdown = ttk.Combobox(window, values=filtered_options, width=28)
    onedif12_dropdown.place(x=380, y=390)
    onedif12_dropdown.bind("<<ComboboxSelected>>", update_dropdown_options)

    global twodif12_dropdown
    twodif12_dropdown = ttk.Combobox(window, values=filtered_options, width=28)
    twodif12_dropdown.place(x=590, y=390)
    twodif12_dropdown.bind("<<ComboboxSelected>>", update_dropdown_options)

    global s_entry12
    s_entry12 = tk.Entry(window)
    s_entry12.place(x=790, y=390)

    dif_bt13 = tk.Button(window, text="+Dif", bg="white", fg="black", command=dif_bt13_done)
    dif_bt13.place(x=495, y=10)

def dif_bt13_done():
    maxstates_error1 = tk.Label(window, text="Sorry, a maximum of 12 differences is supported")
    maxstates_error1.place(x=480, y=420)



def check_dif_reqs():
    global new_dic_of_dif_list
    dif_list = []
    dic_of_dif_list = {}
    new_dic_of_dif_list = {}
    pairlist = []
    title_list = []
    try:
        title = s_entry1.get()[:20]
        if title == "":
            title = "untitled"
        x = 2
        while title in dic_of_dif_list.keys():
            if title.startswith(f"{x-1}_"):
                title = title.removeprefix(f"{x-1}_")
            title = f"{x}_" + title 
            x += 1
        dic_of_dif_list[title] = [onedif_dropdown.get(), twodif_dropdown.get()]
    except:
        pass
    try:
        title = s_entry2.get()[:20]
        if title == "":
            title = "untitled"
        x = 2
        while title in dic_of_dif_list.keys():
            if title.startswith(f"{x-1}_"):
                title = title.removeprefix(f"{x-1}_")
            title = f"{x}_" + title 
            x += 1
        dic_of_dif_list[title] = [onedif2_dropdown.get(), twodif2_dropdown.get()]
    except:
        pass
    try:
        title = s_entry3.get()[:20]
        if title == "":
            title = "untitled"
        x = 2
        while title in dic_of_dif_list.keys():
            if title.startswith(f"{x-1}_"):
                title = title.removeprefix(f"{x-1}_")
            title = f"{x}_" + title 
            x += 1
        dic_of_dif_list[title] =[onedif3_dropdown.get(), twodif3_dropdown.get()]
    except:
        pass
    try:
        title = s_entry4.get()[:20]
        if title == "":
            title = "untitled"
        x = 2
        while title in dic_of_dif_list.keys():
            if title.startswith(f"{x-1}_"):
                title = title.removeprefix(f"{x-1}_")
            title = f"{x}_" + title 
            x += 1
        dic_of_dif_list[title] = [onedif4_dropdown.get(), twodif4_dropdown.get()]
    except:
        pass
    try:
        title = s_entry5.get()[:20]
        if title == "":
            title = "untitled"
        x = 2
        while title in dic_of_dif_list.keys():
            if title.startswith(f"{x-1}_"):
                title = title.removeprefix(f"{x-1}_")
            title = f"{x}_" + title 
            x += 1
        dic_of_dif_list[title] = [onedif5_dropdown.get(), twodif5_dropdown.get()]
    except:
        pass
    try:
        title = s_entry6.get()[:20]
        if title == "":
            title = "untitled"
        x = 2
        while title in dic_of_dif_list.keys():
            if title.startswith(f"{x-1}_"):
                title = title.removeprefix(f"{x-1}_")
            title = f"{x}_" + title 
            x += 1
        dic_of_dif_list[title] = [onedif6_dropdown.get(), twodif6_dropdown.get()]
    except:
        pass
    try:
        title = s_entry7.get()[:20]
        if title == "":
            title = "untitled"
        x = 2
        while title in dic_of_dif_list.keys():
            if title.startswith(f"{x-1}_"):
                title = title.removeprefix(f"{x-1}_")
            title = f"{x}_" + title 
            x += 1
        dic_of_dif_list[title] = [onedif7_dropdown.get(), twodif7_dropdown.get()]
    except:
        pass
    try:
        title = s_entry8.get()[:20]
        if title == "":
            title = "untitled"
        x = 2
        while title in dic_of_dif_list.keys():
            if title.startswith(f"{x-1}_"):
                title = title.removeprefix(f"{x-1}_")
            title = f"{x}_" + title 
            x += 1
        dic_of_dif_list[title] =[onedif8_dropdown.get(), twodif8_dropdown.get()]
    except:
        pass
    try:
        title = s_entry9.get()[:20]
        if title == "":
            title = "untitled"
        x = 2
        while title in dic_of_dif_list.keys():
            if title.startswith(f"{x-1}_"):
                title = title.removeprefix(f"{x-1}_")
            title = f"{x}_" + title 
            x += 1
        dic_of_dif_list[title] = [onedif9_dropdown.get(), twodif9_dropdown.get()]
    except:
        pass
    try:
        title = s_entry10.get()[:20]
        if title == "":
            title = "untitled"
        x = 2
        while title in dic_of_dif_list.keys():
            if title.startswith(f"{x-1}_"):
                title = title.removeprefix(f"{x-1}_")
            title = f"{x}_" + title 
            x += 1
        dic_of_dif_list[title] = [onedif10_dropdown.get(), twodif10_dropdown.get()]
    except:
        pass
    try:
        title = s_entry11.get()[:20]
        if title == "":
            title = "untitled"
        x = 2
        while title in dic_of_dif_list.keys():
            if title.startswith(f"{x-1}_"):
                title = title.removeprefix(f"{x-1}_")
            title = f"{x}_" + title 
            x += 1
        dic_of_dif_list[title] = [onedif11_dropdown.get(), twodif11_dropdown.get()]
    except:
        pass
    try:
        title = s_entry12.get()[:20]
        if title == "":
            title = "untitled"
        x = 2
        while title in dic_of_dif_list.keys():
            if title.startswith(f"{x-1}_"):
                title = title.removeprefix(f"{x-1}_")
            title = f"{x}_" + title 
            x += 1
        dic_of_dif_list[title] = [onedif12_dropdown.get(), twodif12_dropdown.get()]
    except:
        pass
    
    for stt, pair in  dic_of_dif_list.items():
        if pair[0] == "" or pair[1] == "":
            continue
        pairlist.append(pair)
        new_stt = stt.replace("/", "_")
        new_stt = new_stt.replace("~", "_")
        new_stt = new_stt.replace("|", "_")
        title_list.append(new_stt.replace(":", ";"))
    x=0
    for title in title_list:
        new_dic_of_dif_list[title[:20]] = pairlist[x]
        x=x+1
    


global comp_error_lab
comp_error_lab = None

def r_initialize():
    global comp_error_lab
    try:
        tit_bt.destroy()
    except:
        pass
    try:
        pdf_bt.destroy()
    except:
        pass
    try:
        mapviewer_bt.destroy()
    except:
        pass
    try:
        be_label.config(text="                                                                 ")
    except:
        pass
    if comp_error_lab is not None:
        comp_error_lab.destroy()
        
    try:
        on_closing_mapviewer()
    except:
        pass
    
    new_dic_of_dif_lists = {}
    check_dif_reqs()
    
    for key, value in new_dic_of_dif_list.items():
        if value[0].split("~")[0] != value[1].split("~")[0]:
            user_choice_diflist = tk.messagebox.askyesno("Potential Error in Difference States", "Potential Error in Difference States: One or more difference contains different proteins. This may cause issues. Do you wish to continue anyways?", default='no')
            if user_choice_diflist:
                continue
            else:
                return
            
    
    if heatmap_bt_on:
        if new_dic_of_dif_list == {}:
            tk.messagebox.showerror("Difference Error", "There are no differences selected, you cannot run [Localized Differences]")
            return
    if cdif_bt_on:
        if new_dic_of_dif_list == {}:
            tk.messagebox.showerror("Difference Error", "There are no differences selected, you cannot run [Chiclet Difference]")
            return
    if difmap_bt_on:
        if new_dic_of_dif_list == {}:
            tk.messagebox.showerror("Difference Error", "There are no differences selected, you cannot run [Peptide Difference]")
            return
    if difcond_bt_on:
        if new_dic_of_dif_list == {}:
            tk.messagebox.showerror("Difference Error", "There are no differences selected, you cannot run [Condensed Difference]")
            return
    



    p = r_extract_uptake_colors_from_JSON()
    if p is False:
        tk.messagebox.showerror("Color Error", "Could not extract uptake colors")
        return
    
    p = r_extract_difference_colors_from_JSON()
    if p is False:
        tk.messagebox.showerror("Color Error", "Could not extract difference colors")
        return
    
    p = r_extract_localized_colors_from_JSON()
    if p is False:
        tk.messagebox.showerror("Color Error", "Could not extract localized colors")
        return
    
    if difmap_bt_on == False and pepmap_bt_on == False and chic_bt_on == False and cdif_bt_on == False and condpeps_bt_on == False and difcond_bt_on == False and uptake_plot_bt_on == False and heatmap_bt_on == False:
        tk.messagebox.showerror("Run Error", "Please make sure you have selected what visualizations to produce and try again")
        return
    
    r_process_data()
    
    run_bt.config(state="disabled")
    run_bt.config(relief="sunken", bg="white", fg="black")
    
    start_progress()


    global wb
    wb = openpyxl.Workbook()

    r_make_legend1()
    r_make_legend2(True)
    r_make_legend3()
    create_example_plot()
    coincident_peptides()
    if uptake_plot_bt_on == True:
        create_example_plot()
        r_uptake_plots()
    if chic_bt_on == True:
        r_chiclet()
    if cdif_bt_on == True:
        r_chicdif()
    if pepmap_bt_on == True:
        r_pepmaps()
    if difmap_bt_on == True:
        r_difmaps()
    if condpeps_bt_on == True:
        r_condpeps()
    if difcond_bt_on == True:
        r_difcond()
    if heatmap_bt_on == True:
        r_heat_map()
    if uptake_plot_bt_on == True:
        save_pdf()
    if chic_bt_on or cdif_bt_on or pepmap_bt_on or difmap_bt_on or condpeps_bt_on or difcond_bt_on or heatmap_bt_on:
        save_wb()
    if heatmap_bt_on:
        create_mapviewer_bt()
        
def coincident_peptides():  
    global coincident_protein_peptides
    coincident_protein_peptides = {} 
    small_protein_list = []
    for state in states:
        protein = state.split("~")[0]
        small_protein_list.append(protein)
    for protein in small_protein_list:
        peptides_for_protein = [peptide for (p, peptide) in pro_peptide_starts.keys() if p == protein]
        banned_peptide_set = set()
        for peptide in peptides_for_protein:
            for state_name in states:
                current_protein = state_name.split("~")[0]
                if current_protein != protein:
                    continue
                if peptide not in statedic_of_pepdic_cor[state_name].keys():
                    banned_peptide_set.add(peptide)

        coincident_peptides = [peptide for peptide in peptides_for_protein if peptide not in banned_peptide_set]
        coincident_protein_peptides[protein] = coincident_peptides
    


def r_extract_uptake_colors_from_JSON():
    global uptake_color_length, uptake_val_1, uptake_col_1, uptake_val_2, uptake_col_2, uptake_val_3, uptake_col_3, uptake_val_4, uptake_col_4, uptake_val_5, uptake_col_5, uptake_val_6, uptake_col_6, uptake_val_7, uptake_col_7, uptake_val_8, uptake_col_8, uptake_val_9, uptake_col_9, uptake_eqz_key, uptake_abs_key, uptake_ltz_key, uptake_gtz_key, uptake_text_1, uptake_text_2, uptake_text_3, uptake_text_4, uptake_text_5, uptake_text_6, uptake_text_7, uptake_text_8, uptake_text_9, uptake_gtz_text, uptake_eqz_text, uptake_ltz_text, uptake_abs_text, comp_error_lab
    uptake_val_1 = 0
    uptake_val_2 = 0
    uptake_val_3 = 0
    uptake_val_4 = 0
    uptake_val_5 = 0
    uptake_val_6 = 0
    uptake_val_7 = 0
    uptake_val_8 = 0
    uptake_val_9 = 0
    uptake_color_length = 0
    with open("./Colors/" + uptake_color_scheme_dropdown.get(), 'r') as f:
        json_data = json.load(f)
        if json_data.get("header") == "Uptake Colors":
            uptake_color_dic = json_data.get("uptake_color_dic", {})
            uptake_text_dic = json_data.get("uptake_text_dic", {})
            x = 1
            for key, color in uptake_color_dic.items():
                if key == "=0":
                    uptake_eqz_key = color
                    continue
                if key == "-99999":
                    uptake_abs_key = color
                    continue
                if key == ">0":
                    uptake_gtz_key = color
                    continue
                if key == "<0":
                    uptake_ltz_key = color
                    continue
                if x == 1:
                    uptake_val_1 = float(key)
                    uptake_col_1 = color
                    uptake_color_length = 1
                elif x == 2:
                    uptake_val_2 = float(key)
                    uptake_col_2 = color
                    uptake_color_length = 2
                elif x == 3:
                    uptake_val_3 = float(key)
                    uptake_col_3 = color
                    uptake_color_length = 3
                elif x == 4:
                    uptake_val_4 = float(key)
                    uptake_col_4 = color
                    uptake_color_length = 4
                elif x == 5:
                    uptake_val_5 = float(key)
                    uptake_col_5 = color
                    uptake_color_length = 5
                elif x == 6:
                    uptake_val_6 = float(key)
                    uptake_col_6 = color
                    uptake_color_length = 6
                elif x == 7:
                    uptake_val_7 = float(key)
                    uptake_col_7 = color
                    uptake_color_length = 7
                elif x == 8:
                    uptake_val_8 = float(key)
                    uptake_col_8 = color
                    uptake_color_length = 8
                elif x == 9:
                    uptake_val_9 = float(key)
                    uptake_col_9 = color
                    uptake_color_length = 9
                x += 1
            for key, text in uptake_text_dic.items():
                if key not in ["-99999", "=0", ">0", "<0"]:
                    key = float(key)
                if text == 1:
                    text = 'FFFFFFFF'
                if text == 0:
                    text = 'FF000000'
                if key == "=0":
                    uptake_eqz_text = text
                if key == "-99999":
                    uptake_abs_text = text
                if key == ">0":
                    uptake_gtz_text = text
                if key == "<0":
                    uptake_ltz_text = text
                if key == uptake_val_1:
                    uptake_text_1 = text
                if key == uptake_val_2:
                    uptake_text_2 = text
                if key == uptake_val_3:
                    uptake_text_3 = text
                if key == uptake_val_4:
                    uptake_text_4 = text
                if key == uptake_val_5:
                    uptake_text_5 = text
                if key == uptake_val_6:
                    uptake_text_6 = text
                if key == uptake_val_7:
                    uptake_text_7 = text
                if key == uptake_val_8:
                    uptake_text_8 = text
                if key == uptake_val_9:
                    uptake_text_9 = text
            return True
        else:
            comp_error_lab = tk.Label(window, text="Uptake color selection is not compatible")
            comp_error_lab.place(x=1190, y=230)
            run_bt.config(state="normal")
            run_bt.config(relief="raised")
            return False
        
def r_extract_difference_colors_from_JSON():
    global p_val_1, p_val_2, p_val_3, p_val_4, p_val_5, d_val_1, d_val_2, d_val_3, d_val_4, d_val_5, p_col_1, p_col_2, p_col_3, p_col_4, p_col_5, d_col_1, d_col_2, d_col_3, d_col_4, d_col_5, p_col_gtz, p_col_length, p_text_1, p_text_2, p_text_3, p_text_4, p_text_5, d_text_1, d_text_2, d_text_3, d_text_4, d_text_5, p_text_gtz, d_col_gtz, d_text_gtz, d_col_length, b_col_eqz, b_col_abs, b_text_eqz, b_text_abs, comp_error_lab
    p_val_1 = 0
    p_val_2 = 0
    p_val_3 = 0
    p_val_4 = 0
    p_val_5 = 0
    d_val_1 = 0
    d_val_2 = 0
    d_val_3 = 0
    d_val_4 = 0
    d_val_5 = 0
    p_col_length = 0
    d_col_length = 0
    with open("./Colors/" + difference_color_scheme_dropdown.get(), 'r') as f:
        json_data = json.load(f)
        if json_data.get("header") == "Difference Colors":
            pval_color_dic = json_data.get("protection_color_dic", {})
            pval_text_dic = json_data.get("protection_text_dic", {})
            dval_color_dic = json_data.get("deprotection_color_dic", {})
            dval_text_dic = json_data.get("deprotection_text_dic", {})
            bval_color_dic = json_data.get("both_color_dic", {})
            bval_text_dic = json_data.get("both_text_dic", {})
            x = 1
            for key, color in pval_color_dic.items():
                if key == ">0":
                    p_col_gtz = color
                    continue
                if x == 1:
                    p_val_1 = float(key)
                    p_col_1 = color
                    p_col_length = 1
                elif x == 2:
                    p_val_2 = float(key)
                    p_col_2 = color
                    p_col_length = 2
                elif x == 3:
                    p_val_3 = float(key)
                    p_col_3 = color
                    p_col_length = 3
                elif x == 4:
                    p_val_4 = float(key)
                    p_col_4 = color
                    p_col_length = 4
                elif x == 5:
                    p_val_5 = float(key)
                    p_col_5 = color
                    p_col_length = 5
                x += 1
            for key, text in pval_text_dic.items():
                try:
                    key = float(key)
                except:
                    pass
                if text == 1:
                    text = 'FFFFFFFF'
                if text == 0:
                    text = 'FF000000'
                if key == ">0":
                    p_text_gtz = text
                if key == p_val_1:
                    p_text_1 = text
                elif key == p_val_2:
                    p_text_2 = text
                elif key == p_val_3:
                    p_text_3 = text
                elif key == p_val_4:
                    p_text_4 = text
                elif key == p_val_5:
                    p_text_5 = text

            x = 1
            for key, color in dval_color_dic.items():
                if key == ">0":
                    d_col_gtz = color
                    continue
                if x == 1:
                    d_val_1 = float(key)
                    d_col_1 = color
                    d_col_length = 1
                elif x == 2:
                    d_val_2 = float(key)
                    d_col_2 = color
                    d_col_length = 2
                elif x == 3:
                    d_val_3 = float(key)
                    d_col_3 = color
                    d_col_length = 3
                elif x == 4:
                    d_val_4 = float(key)
                    d_col_4 = color
                    d_col_length = 4
                elif x == 5:
                    d_val_5 = float(key)
                    d_col_5 = color
                    d_col_length = 5
                x += 1
            for key, text in dval_text_dic.items():
                try:
                    key = float(key)
                except:
                    pass
                if text == 1:
                    text = 'FFFFFFFF'
                if text == 0:
                    text = 'FF000000'
                if key == ">0":
                    d_text_gtz = text
                if key == d_val_1:
                    d_text_1 = text
                elif key == d_val_2:
                    d_text_2 = text
                elif key == d_val_3:
                    d_text_3 = text
                elif key == d_val_4:
                    d_text_4 = text
                elif key == d_val_5:
                    d_text_5 = text

            for key, color in bval_color_dic.items():
                if key == "=0":
                    b_col_eqz = color
                if key == "-99999":
                    b_col_abs = color
            for key, text in bval_text_dic.items():
                if text == 1:
                    text = 'FFFFFFFF'
                if text == 0:
                    text = 'FF000000'
                if key == "=0":
                    b_text_eqz = text
                if key == "-99999":
                    b_text_abs = text
            return True
        else:
            comp_error_lab = tk.Label(window, text="Difference color selection is not compatible")
            comp_error_lab.place(x=1190, y=230)
            run_bt.config(state="normal")
            run_bt.config(relief="raised")
            return False
        
def r_extract_localized_colors_from_JSON():
    global comp_error_lab, lcol0, lcol1, lcol2, lcol3, lcol4, lcol5, lcol6, lcol7, lcol8, lcol9, future_linear_map_multiplier
    with open("./Colors/" + localized_color_scheme_dropdown.get(), 'r') as f:
        json_data = json.load(f)
        if json_data.get("header") == "Localized Difference Plot Colors":
            lcol_list = json_data.get("lcols", [])
            future_linear_map_multiplier = json_data.get("significance_cutoff", 0)
            lcol6 = json_data.get("lcol_6", False)
            lcol7 = json_data.get("lcol_7", False)
            lcol8 = json_data.get("lcol_8", False)
            lcol9 = json_data.get("lcol_9", False)
            if future_linear_map_multiplier == 0:
                tk.messagebox.showerror("Invalid Color Scheme", f"Significance Cut-Off Value in {localized_color_scheme_dropdown.get()} is invalid")
                return
            
            
            lcol0 = lcol_list[0]
            lcol1 = lcol_list[1]
            lcol2 = lcol_list[2]
            lcol3 = lcol_list[3]
            lcol4 = lcol_list[4]
            lcol5 = lcol_list[5]
            if lcol6 != False:
                if lcol6.lower() == "false":
                    lcol6 = False
            if lcol7 != False:
                if lcol7.lower() == "false":
                    lcol7 = False
            if lcol8 != False:
                if lcol8.lower() == "false":
                    lcol8 = False
            if lcol9 != False:
                if lcol9.lower() == "false":
                    lcol9 = False
        
        else:
            comp_error_lab = tk.Label(window, text="Difference color selection is not compatible")
            comp_error_lab.place(x=1190, y=230)
            run_bt.config(state="normal")
            run_bt.config(relief="raised")
            return False
                
                
def r_process_data():  
    global statedic_of_pepdic_cor, new_dic_of_dif_list, s_timepoints, sdbt_clicked, cdbt_clicked
    statedic_of_pepdic_cor = {}
    s_timepoints = {}
    for state in states:
        timepoints = list()
        if sdbt_clicked is True:
            for peptide_instance in class_peptides:
                if peptide_instance.State == state:
                    timepoint = peptide_instance.Timepoint
                    if timepoint not in timepoints:
                        timepoints.append(timepoint)            
        if cdbt_clicked is True:
            for peptide_instance in class_peptides:
                if peptide_instance.State == state:
                    timepoint = peptide_instance.Timepoint
                    if timepoint not in timepoints:
                        timepoints.append(timepoint)
        if examiner_clicked is True:
            for peptide_instance in class_peptides:
                if peptide_instance.State == state:
                    timepoint = peptide_instance.Timepoint
                    if timepoint not in timepoints:
                        timepoints.append(timepoint)
            if 3245908.2345029 in timepoints:
                timepoints.remove(3245908.2345029)
                timepoints.append((max(timepoints))*10)
            for peptide_instance in class_peptides:
                if peptide_instance.State == state and peptide_instance.Timepoint == 3245908.2345029:
                    peptide_instance.Timepoint = timepoints[-1]
                    if peptide_instance.Uptake == 0:
                        peptide_instance.Uptake = -99999
                        peptide_instance.Stdev = -99999
        if workbench_clicked is True:
            for peptide_instance in class_peptides:
                if peptide_instance.State == state:
                    timepoint = peptide_instance.Timepoint
                    if timepoint not in timepoints:
                        timepoints.append(timepoint)
        timepoints.sort()
        s_timepoints[state] = timepoints
        





    



    #here we get uptake, sd, timepoint, do out averaging if replicates, makes statedic_of_pepdic_raw and statedic_of_sddic_raw
    global statedic_of_pepdic_raw, statedic_of_sddic_raw, peplist, multiple_repeated_peptides
    statedic_of_pepdic_raw = {}
    statedic_of_sddic_raw = {}
    
    multiple_repeated_peptides = False
    gt3_replicates = False
    
    if sdbt_clicked == True or examiner_clicked == True:
        for state in states:
            if gt3_replicates == True:
                break
            for peptide in peplist[state]:
                if gt3_replicates == True:
                    break
                filtered_peptides = [p for p in class_peptides if p.Sequence == peptide and p.State == state]
                nonset_timepoints = []
                for p in filtered_peptides:
                    nonset_timepoints.append(p.Timepoint)
                    num_replicates = sum([1 for x in nonset_timepoints if x == p.Timepoint])
                    if num_replicates >= 3:
                        gt3_replicates = True
                        break
                    if num_replicates >= 2:
                        multiple_repeated_peptides = True
    
    elif cdbt_clicked == True or workbench_clicked == True:
        multiple_repeated_peptides = True
        for state in states:
            if gt3_replicates == True:
                break
            for peptide in peplist[state]:
                if gt3_replicates == True:
                    break
                filtered_peptides = [p for p in class_peptides if p.Sequence == peptide and p.State == state]
                set_timepoint_tups = set()
                for p in filtered_peptides:
                    set_timepoint_tups.add((p.Timepoint, p.File))
                    num_replicates = sum([1 for x in set_timepoint_tups if x[0] == p.Timepoint])
                    if num_replicates >= 3:
                        gt3_replicates = True
                        break
                    if num_replicates >= 2:
                        multiple_repeated_peptides = True
            
                
                
                ###if multiple_repeated_peptides is True, every single sd value from the source worksheet will be ignored, and it will be calculated in HDXWizard

    for state in states:
        if state not in statedic_of_pepdic_raw:
            statedic_of_pepdic_raw[state] = True
            pepdic_raw = {}
            sddic_raw = {}
        for peptide in peplist[state]:
            upt_tp_tup_list = list()
            sd_tp_tup_list = list()
            upt_tp_tup_list_initial = list()
            upt_tp_tup_dic = {}
            
            
            if sdbt_clicked is True or examiner_clicked is True:
                pass
            if cdbt_clicked is True or workbench_clicked is True:
                filtered_peptides = [p for p in class_peptides if p.Sequence == peptide and p.State == state]
                und_instances = []
                for peptide_instance in filtered_peptides:
                    if peptide_instance.Timepoint == 0:
                        und_instances.append(peptide_instance)
                und_mass_replicate_tups = []
                for instance in und_instances:
                    if instance.Center == -99999:
                        pass
                    else:
                        instance_mass = (instance.Center * instance.Charge) - instance.Charge
                        instance.Uptake = 0
                        instance.Stdev = -99999
                        und_mass_replicate_tups.append((instance_mass, instance.File))
                if len(und_mass_replicate_tups) == 0:
                    und_mass = -99999
                    und_std = -99999
                else:
                    mass_by_file = {}
                    for mass, file in und_mass_replicate_tups:
                        if file not in mass_by_file:
                            mass_by_file[file] = []
                        mass_by_file[file].append(mass)
                    first_average_list = [np.mean(masses) for (file, masses) in mass_by_file.items()]
                    und_mass = np.mean(first_average_list)
                    if gt3_replicates == True:
                        num_there = sum([1 for mass in first_average_list if mass != -99999])
                        if num_there >= 3:
                            und_std = np.std(first_average_list, ddof=1)
                        else:
                            und_std = -99999
                    else:
                        und_std = -99999

                    
                for peptide_instance in filtered_peptides:
                    peptide_instance.und_std = und_std
                    if peptide_instance.Uptake == 0:
                        continue
                    if peptide_instance.Center != -99999:
                        instance_mass = (peptide_instance.Center * peptide_instance.Charge) - peptide_instance.Charge
                    else:
                        instance_mass = -99999
                    if und_mass != -99999 and instance_mass != -99999:
                        peptide_instance.Uptake = instance_mass - und_mass
                    else:
                        peptide_instance.Uptake = -99999
                    peptide_instance.Stdev = -99999  
                    
            
            filtered_peptides = [p for p in class_peptides if p.Sequence == peptide and p.State == state]
            for peptide_instance in filtered_peptides:
                if peptide_instance.Uptake == -99999:
                    peptide_instance.Stdev = -99999
                if peptide_instance.Timepoint == 0:
                    peptide_instance.Stdev = -99999
                if multiple_repeated_peptides == False:
                    upt_tp_tup_list.append((peptide_instance.Uptake, peptide_instance.Timepoint))
                    sd_tp_tup_list.append((peptide_instance.Stdev, peptide_instance.Timepoint))
                if sdbt_clicked == True or examiner_clicked == True:
                    upt_tp_tup_list_initial.append((peptide_instance.Uptake, peptide_instance.Timepoint))
                if cdbt_clicked == True or workbench_clicked == True:
                    if peptide_instance.File not in upt_tp_tup_dic.keys():
                        upt_tp_tup_dic[peptide_instance.File] = []
                    upt_tp_tup_dic[peptide_instance.File].append((peptide_instance.Uptake, peptide_instance.Timepoint))


            if multiple_repeated_peptides == True:
                if cdbt_clicked == True or workbench_clicked == True:
                    for file, list_of_tups in upt_tp_tup_dic.items():
                        timepoint_dic = {}
                        for uptake, timepoint in list_of_tups:
                            if timepoint not in timepoint_dic:
                                timepoint_dic[timepoint] = []
                            timepoint_dic[timepoint].append(uptake)
                        
                        for timepoint, uptakes in timepoint_dic.items():
                            filtered_uptakes = [uptake for uptake in uptakes if uptake != -99999]
                            if len(filtered_uptakes) == 0:
                                mean_uptake = -99999
                            else:
                                mean_uptake = np.mean(filtered_uptakes)
                            upt_tp_tup_list_initial.append((mean_uptake, timepoint))
                            
                        
                upt_tp_tup_list_dic = {}
                for uptake, timep in upt_tp_tup_list_initial:
                    if timep not in upt_tp_tup_list_dic.keys():
                        upt_tp_tup_list_dic[timep] = []
                    upt_tp_tup_list_dic[timep].append(uptake)
                for timep, up_list in upt_tp_tup_list_dic.items():
                    up_list = [x for x in up_list if x != -99999]
                    if len(up_list) <= 2:
                        sd_tp_tup_list.append((-99999, timep))
                    else:
                        if timep != 0:
                            if sdbt_clicked or examiner_clicked:
                                sd_tp_tup_list.append((np.std(up_list, ddof=1), timep))
                            if workbench_clicked or cdbt_clicked:
                                deuterated_std = np.std(up_list, ddof=1)
                                uptake_std = [peptide_instance.und_std for peptide_instance in filtered_peptides][0]
                                if uptake_std == -99999:
                                    sd_tp_tup_list.append((-99999, timep))
                                else:
                                    combined_std = np.sqrt((deuterated_std **2)+(uptake_std ** 2))
                                    sd_tp_tup_list.append((combined_std, timep))

                        else:
                            sd_tp_tup_list.append((-99999, timep))
                    if up_list != []:
                        upt_tp_tup_list.append((np.mean(up_list), timep))
                    else:
                        upt_tp_tup_list.append((-99999, timep))
                            
                            
                
                    
                    



            upt_tp_tup_list = sorted(upt_tp_tup_list, key=lambda x: x[1])
            sd_tp_tup_list = sorted(sd_tp_tup_list, key=lambda x: x[1])

            pepdic_raw[peptide] = upt_tp_tup_list
            sddic_raw[peptide] = sd_tp_tup_list
            
            
                
                
                
        statedic_of_pepdic_raw[state] = pepdic_raw
        statedic_of_sddic_raw[state] = sddic_raw
        

    
    
    
        

    #statedic of pepdic raw and of ssdic_raw add -99999 as placeholders for missing values
    global statedic_of_pepdic_raw2
    statedic_of_pepdic_raw2 = {}
    for state, pepdic_raw in statedic_of_pepdic_raw.items():
        pepdic_raw2 = {}
        for peptide, upt_tp_tups in pepdic_raw.items():
            upt_tp_tups2 = list()
            for timepoint in s_timepoints[state]:
                timepoint_found = False
                for uptake, tp in upt_tp_tups:
                    if timepoint == tp:
                        timepoint_found = True
                        break
                if timepoint_found == False:
                    new_tup = (-99999, timepoint)
                    upt_tp_tups2.append(new_tup)
                    pepdic_raw2[peptide] = upt_tp_tups2
                if timepoint_found == True:
                    upt_tp_tups2.append((uptake, tp))
            upt_tp_tups2 = sorted(upt_tp_tups2, key=lambda x: x[1])
            pepdic_raw2[peptide] = upt_tp_tups2

        statedic_of_pepdic_raw2[state] = pepdic_raw2

    global statedic_of_sddic_raw2
    statedic_of_sddic_raw2 = {}
    for state, sddic_raw in statedic_of_sddic_raw.items():
        sddic_raw2 = {}
        for peptide, sd_tp_tups in sddic_raw.items():
            sd_tp_tups2 = list()
            for timepoint in s_timepoints[state]:
                timepoint_found = False
                for sd, tp in sd_tp_tups:
                    if timepoint == tp:
                        timepoint_found = True
                        break
                if timepoint_found == False:
                    new_sd = (-99999, timepoint)
                    sd_tp_tups2.append(new_sd)
                    sddic_raw2[peptide] = upt_tp_tups2
                if timepoint_found == True:
                    sd_tp_tups2.append((sd, tp))
            sd_tp_tups2 = sorted(sd_tp_tups2, key=lambda x: x[1])
            sddic_raw2[peptide] = sd_tp_tups2
        statedic_of_sddic_raw2[state] = sddic_raw2
        
        
        
    #now we are dealing with rfu
    global noD_dic_states, statedic_of_sddic_cor
    noD_dic_states = {}
    statedic_of_pepdic_cor = {}
    statedic_of_sddic_cor = {}
    if exp_bt_on_c == True:
        for state, dropdown in dropdowns.items():
            selected_value = dropdown.get()
            maxdic[state] = selected_value
            

        for state, pepdic_raw2 in statedic_of_pepdic_raw2.items():
            if not maxdic[state].startswith("pyAVG"):
                maxfile = maxdic[state]

                noD_dic_peptides = {}
                maxd_list = list()
                maxSD_list = list()
                maxtheo_list = list()
                for peptide, upt_tp_tups in pepdic_raw2.items():
                    try:
                        maxfile_up_tp_tups = statedic_of_pepdic_raw2[maxfile][peptide]
                        max_tp = max(maxfile_up_tp_tups, key=lambda x: x[1])[1]
                        maxD = next(x[0] for x in maxfile_up_tp_tups if x[1] == max_tp)
                    except:
                        maxD = -99999
                    try:
                        maxfile_sd_tp_tups = statedic_of_sddic_raw2[maxfile][peptide]
                        max_tp = max(maxfile_sd_tp_tups, key=lambda x: x[1])[1]
                        maxSD = next(x[0] for x in maxfile_sd_tp_tups if x[1] == max_tp)
                    except:
                        maxSD = -99999
                    if maxD != -99999:
                        maxd_list.append(maxD)
                        if maxSD != -99999:
                            maxSD_list.append(maxSD)

                        max_theo = get_max_theo(peptide)
                        maxtheo_list.append(max_theo)
                    else:
                        noD_dic_peptides[peptide] = True
                noD_dic_states[state] = noD_dic_peptides
                if len(maxd_list) != 0:
                    total_uptake = sum(maxd_list)
                    total_theo = sum(maxtheo_list)
                    average_rfu = (total_uptake / total_theo)
                    #average_rfu_sd_relerr = np.std(maxd_list, ddof=1)/np.mean(maxd_list)
                    #average_rfu_sd = average_rfu_sd_relerr*average_rfu
                    average_rfu_sd = -99999
#                    if len(maxSD_list) != 0:
#                        sd_array_squared = np.asarray(maxSD_list) ** 2
#                        sd_comb = (np.sqrt(np.sum(sd_array_squared)))/len(maxSD_list)
#                        average_rfu_sd_percent = sd_comb / (total_uptake/len(maxd_list))
#                        average_rfu_sd = average_rfu_sd_percent * average_rfu
#                    else:
#                        average_rfu_sd = 0
                    
                else:
                    average_rfu = 1
                    average_rfu_sd = -99999 



                pepdic_cor = {}
                sddic_cor = {}
                for peptide, upt_tp_tups in pepdic_raw2.items():
                    try:
                        maxfile_up_tp_tups = statedic_of_pepdic_raw2[maxfile][peptide]
                        max_tp = max(maxfile_up_tp_tups, key=lambda x: x[1])[1]
                        maxD = next(x[0] for x in maxfile_up_tp_tups if x[1] == max_tp)
                    except:
                        maxD = -99999
                    try:
                        maxfile_sd_tp_tups = statedic_of_sddic_raw2[maxfile][peptide]
                        max_tp = max(maxfile_sd_tp_tups, key=lambda x: x[1])[1]
                        maxSD = next(x[0] for x in maxfile_sd_tp_tups if x[1] == max_tp)
                    except:
                        maxSD = -99999
                    if maxD == -99999:

                        maxD_i = get_max_theo(peptide)
                        maxD = maxD_i * average_rfu
                        #maxSD = (float(average_rfu_sd) / float(average_rfu)) * maxD
                        maxSD = -99999
                    
                    if maxSD != -99999:
                        maxSD = maxSD 
                        
                    Dcorrected_tups = list()
                    Dcorrected_sd_tups = list()
                    sd_tp_tups = statedic_of_sddic_raw2[state][peptide]
                    for uptake, timepoint in upt_tp_tups:
                        if uptake != -99999:
                            Dcorrected = (float(uptake) / float(maxD))
                            for sd, tp in sd_tp_tups:
                                if tp == timepoint:
                                    uptake_SD = float(sd)
                                    break
                            if uptake_SD != -99999 and uptake != 0 and maxSD != -99999 and uptake_SD != maxSD:
                                Dcorrected_SD_percent = np.sqrt((((uptake_SD)/(float(uptake))) ** 2) + (((maxSD)/(float(maxD))) ** 2))
                                Dcorrected_SD = Dcorrected_SD_percent * Dcorrected
                            elif uptake_SD == maxSD:
                                Dcorrected_SD_percent = maxSD/float(maxD)
                                Dcorrected_SD = Dcorrected_SD_percent * Dcorrected                               
                            else:
                                Dcorrected_SD = -99999

                        else:
                            Dcorrected = -99999
                            Dcorrected_SD = -99999
                        Dcorrected_tups.append((Dcorrected, timepoint))
                        Dcorrected_sd_tups.append((Dcorrected_SD, timepoint))
                    pepdic_cor[peptide] = Dcorrected_tups
                    sddic_cor[peptide] = Dcorrected_sd_tups
                statedic_of_pepdic_cor[state] = pepdic_cor
                statedic_of_sddic_cor[state] = sddic_cor
                
            
    

#            if maxdic[state].startswith("pyAVG"):
#                pepdic_cor = {}
#                sddic_cor = {}
#                maxd_list = list()
#                maxtheo_list = list()
#                maxSD_list = list()
#                maxfiles = new_states_dic[maxdic[state]]
#                st1 = maxfiles[0]
#                st2 = maxfiles[1]
#
#                noD_dic_peptides = {}
#                for peptide, upt_tp_tups in pepdic_raw2.items():
#                    try:
#                        up_tp_tups1 = statedic_of_pepdic_raw2[st1][peptide]
#                        max_tp1 = max(up_tp_tups1, key=lambda x: x[1])[1]
#                        maxD1 = next(x[0] for x in up_tp_tups1 if x[1] == max_tp1)
#                    except:
#                        maxD1 = -99999
#                    try:
#                        up_tp_tups2 = statedic_of_pepdic_raw2[st2][peptide]
#                        max_tp2 = max(up_tp_tups2, key=lambda x: x[1])[1]
#                        maxD2 = next(x[0] for x in up_tp_tups2 if x[1] == max_tp2)
#                    except:
#                        maxD2 = -99999
#                    try:
#                        sd_tp_tups1 = statedic_of_sddic_raw2[st1][peptide]
#                        max_tp1 = max(sd_tp_tups1, key=lambda x: x[1])[1]
#                        maxSD1 = next(x[0] for x in sd_tp_tups1 if x[1] == max_tp1)
#                    except:
#                        maxSD1 = -99999
#                    try:
#                        sd_tp_tups2 = statedic_of_sddic_raw2[st2][peptide]
#                        max_tp2 = max(sd_tp_tups2, key=lambda x: x[1])[1]
#                        maxSD2 = next(x[0] for x in sd_tp_tups2 if x[1] == max_tp2)
#                    except:
#                        maxSD2 = -99999
#                    if maxD1 != -99999:
#                        maxd_list.append(maxD1)
#                        if maxSD1 != -99999:
#                            maxSD_list.append(maxSD1)
#
#                        max_theo = get_max_theo(peptide)
#                        maxtheo_list.append(max_theo)
#
#                    if maxD2 != -99999:
#                        maxd_list.append(maxD2)
#                        if maxSD1 != -99999:
#                            maxSD_list.append(maxSD2)
#
#                        max_theo = get_max_theo(peptide)
#                        maxtheo_list.append(max_theo)
#                    if maxD1 == -99999 and maxD2 == -99999:
#                        noD_dic_peptides[peptide] = True
#                noD_dic_states[state] = noD_dic_peptides
#                if len(maxd_list) != 0:
#                    total_uptake = sum(maxd_list)
#                    total_theo = sum(maxtheo_list)
#                    average_rfu = (total_uptake / total_theo)
#                    if len(maxSD_list) != 0:
#                        sd_array_squared = np.asarray(maxSD_list) ** 2
#                        sd_comb = (np.sqrt(np.sum(sd_array_squared)))/len(maxSD_list)
#                        average_rfu_sd_percent = sd_comb / (total_uptake/len(maxd_list))
#                        average_rfu_sd = average_rfu_sd_percent * average_rfu
#                    else:
#                        average_rfu_sd = 0
#                else:
#                    average_rfu = 1
#                    average_rfu_sd = 0
#
#
#                for peptide, upt_tp_tups in pepdic_raw2.items():
#                    if peptide in peplist[st1]:
#                        if peptide in peplist[st2]:
#                            up_tp_tups1 = statedic_of_pepdic_raw2[st1][peptide]
#                            up_tp_tups2 = statedic_of_pepdic_raw2[st2][peptide]
#                            sd_tp_tups1 = statedic_of_sddic_raw2[st1][peptide]
#                            sd_tp_tups2 = statedic_of_sddic_raw2[st2][peptide]
#                            max_tp_1 = max(up_tp_tups1, key=lambda x: x[1])[1]
#                            maxD_1 = next(x[0] for x in up_tp_tups1 if x[1] == max_tp_1)
#                            max_tp_2 = max(up_tp_tups2, key=lambda x: x[1])[1]
#                            maxD_2 = next(x[0] for x in up_tp_tups2 if x[1] == max_tp_2)
#                            max_tp_sd1 = max(sd_tp_tups1, key=lambda x: x[1])[1]
#                            maxSD_1 = next(x[0] for x in sd_tp_tups1 if x[1] == max_tp_sd1)
#                            max_tp_sd2 = max(sd_tp_tups2, key=lambda x: x[1])[1]
#                            maxSD_2 = next(x[0] for x in sd_tp_tups2 if x[1] == max_tp_sd2)
#
#                            if maxD_1 != -99999 and maxD_2 != -99999:
#                                maxD = (maxD_1 + maxD_2)/2
#                                if maxSD_1 != -99999 and maxSD_2 != -99999:
#                                    maxSD = (np.sqrt(((maxSD_1) ** 2) + ((maxSD_2) ** 2)))/2
#                                elif maxSD_1 != -99999:
#                                    maxSD = maxSD_1
#                                elif maxSD_2 != -99999:
#                                    maxSD = maxSD_2
#                                else:
#                                    maxSD = -99999
#                            elif maxD_1 != -99999:
#                                maxD = maxD_1
#                                if maxSD_1 != -99999:
#                                    maxSD = maxSD_1
#                                else:
#                                    maxSD = -99999
#                            elif maxD_2 != -99999:
#                                maxD = maxD_2
#                                if maxSD_2 != -99999:
#                                    maxSD = maxSD_2
#                                else:
#                                    maxSD = -99999
#                            else:
#
#                                maxD_i = get_max_theo(peptide)
#                                maxD = maxD_i * average_rfu
#                                maxSD = (float(average_rfu_sd) / float(average_rfu)) * maxD
#
#                        else:
#                            up_tp_tups1 = statedic_of_pepdic_raw2[st1][peptide]
#                            max_tp_1 = max(up_tp_tups1, key=lambda x: x[1])[1]
#                            maxD = next(x[0] for x in up_tp_tups1 if x[1] == max_tp_1)
#                            sd_tp_tups1 = statedic_of_sddic_raw2[st1][peptide]
#                            max_tp_sd1 = max(sd_tp_tups1, key=lambda x: x[1])[1]
#                            maxSD_1 = next(x[0] for x in sd_tp_tups1 if x[1] == max_tp_sd1)
#                            if maxD == -99999:
#
#                                maxD_i = get_max_theo(peptide)
#                                maxD = maxD_i * average_rfu
#                                maxSD = (float(average_rfu_sd) / float(average_rfu)) * maxD
#                            else:
#                                maxSD = maxSD_1
#                        Dcorrected_tups = list()
#                        Dcorrected_sd_tups = list()
#                        sd_tp_tups = statedic_of_sddic_raw2[state][peptide]
#                        for uptake, timepoint in upt_tp_tups:
#                            if uptake != -99999:
#                                Dcorrected = (float(uptake) / float(maxD))
#                                for sd, tp in sd_tp_tups:
#                                    if tp == timepoint:
#                                        uptake_SD = float(sd)
#                                        break
#                                if uptake_SD != -99999 and uptake != 0 and maxSD != -99999:
#                                    Dcorrected_SD_percent = np.sqrt((((uptake_SD)/(float(uptake))) ** 2) + (((maxSD)/(float(maxD))) ** 2))
#                                    Dcorrected_SD = Dcorrected_SD_percent * Dcorrected
#                                else:
#                                    Dcorrected_SD = -99999
#                            else:
#                                Dcorrected = -99999
#                                Dcorrected_SD = -99999
#                            Dcorrected_tups.append((Dcorrected, timepoint))
#                            Dcorrected_sd_tups.append((Dcorrected_SD, timepoint))
#                        pepdic_cor[peptide] = Dcorrected_tups
#                        sddic_cor[peptide] = Dcorrected_sd_tups
#
#                    if peptide in peplist[st2]:
#                        if peptide not in peplist[st1]:
#                            up_tp_tups2 = statedic_of_pepdic_raw2[st2][peptide]
#                            max_tp_2 = max(up_tp_tups2, key=lambda x: x[1])[1]
#                            maxD = next(x[0] for x in up_tp_tups2 if x[1] == max_tp_2)
#                            sd_tp_tups2 = statedic_of_sddic_raw2[st2][peptide]
#                            max_tp_sd2 = max(sd_tp_tups2, key=lambda x: x[1])[1]
#                            maxSD_2 = next(x[0] for x in sd_tp_tups2 if x[1] == max_tp_sd2)
#                            if maxD == -99999:
#
#                                maxD_i = get_max_theo(peptide)
#                                maxD = maxD_i * average_rfu
#                                maxSD = (float(average_rfu_sd) / float(average_rfu)) * maxD
#                            else:
#                                maxSD = maxSD_2
#                            Dcorrected_tups = list()
#                            Dcorrected_sd_tups = list()
#                            sd_tp_tups = statedic_of_sddic_raw2[state][peptide]
#                            for uptake, timepoint in upt_tp_tups:
#                                if uptake != -99999:
#                                    Dcorrected = (float(uptake) / float(maxD))
#                                    for sd, tp in sd_tp_tups:
#                                        if tp == timepoint:
#                                            uptake_SD = float(sd)
#                                            break
#                                    if uptake_SD != -99999 and uptake != 0 and maxSD != -99999:
#                                        Dcorrected_SD_percent = np.sqrt((((uptake_SD)/(float(uptake))) ** 2) + (((maxSD)/(float(maxD))) ** 2))
#                                        Dcorrected_SD = Dcorrected_SD_percent * Dcorrected
#                                    else:
#                                        Dcorrected_SD = -99999
#                                else:
#                                    Dcorrected = -99999
#                                    Dcorrected_SD = -99999
#                                Dcorrected_tups.append((Dcorrected, timepoint))
#                                Dcorrected_sd_tups.append((Dcorrected_SD, timepoint))
#                            pepdic_cor[peptide] = Dcorrected_tups
#                            sddic_cor[peptide] = Dcorrected_sd_tups
#                    if peptide not in peplist[st1] and peptide not in peplist[st2]:
#
#                        maxD_i = get_max_theo(peptide)
#                        maxD = maxD_i * average_rfu
#                        maxSD = (float(average_rfu_sd) / float(average_rfu)) * maxD
#                        Dcorrected_tups = list()
#                        Dcorrected_sd_tups = list()
#                        sd_tp_tups = statedic_of_sddic_raw2[state][peptide]
#                        for uptake, timepoint in upt_tp_tups:
#                            if uptake != -99999:
#                                Dcorrected = (float(uptake) / float(maxD))
#                                for sd, tp in sd_tp_tups:
#                                    if tp == timepoint:
#                                        uptake_SD = float(sd)
#                                        break
#                                if uptake_SD != -99999 and uptake != 0 and maxSD != -99999:
#                                    Dcorrected_SD_percent = np.sqrt((((uptake_SD)/(float(uptake))) ** 2) + (((maxSD)/(float(maxD))) ** 2))
#                                    Dcorrected_SD = Dcorrected_SD_percent * Dcorrected
#                                else:
#                                    Dcorrected_SD = -99999
#                            else:
#                                Dcorrected = -99999
#                                Dcorrected_SD = -99999
#                            Dcorrected_tups.append((Dcorrected, timepoint))
#                            Dcorrected_sd_tups.append((Dcorrected_SD, timepoint))
#                        pepdic_cor[peptide] = Dcorrected_tups
#                        sddic_cor[peptide] = Dcorrected_sd_tups
#
#
#                statedic_of_pepdic_cor[state] = pepdic_cor
#                statedic_of_sddic_cor[state] = sddic_cor

    if theo_bt_on_c == True:
        global back_exchange, be_label
        be = be_entry.get()
        try:
            back_exchange = float(be)
        except:
            back_exchange = 0
            be_label = tk.Label(window, text="Invalid Back Exchange. Defaulted to 0")
            be_label.place(x=30, y=400)
        for state, pepdic_raw2 in statedic_of_pepdic_raw2.items():
            pepdic_cor = {}
            sddic_cor = {}
            for peptide, upt_tp_tups in pepdic_raw2.items():

                maxD = get_max_theo(peptide)
                be_maxD = maxD * ((100-back_exchange)/100)
                Dcorrected_tups = list()
                Dcorrected_SD_tups = list()
                sd_tp_tups = statedic_of_sddic_raw2[state][peptide]
                for uptake, timepoint in upt_tp_tups:
                    if uptake != -99999:
                        Dcorrected = (float(uptake) / float(be_maxD))
                        for sd, tp in sd_tp_tups:
                            if tp == timepoint:
                                uptake_SD = float(sd)
                                break
                        if uptake_SD != -99999 and uptake != 0:
                            Dcorrected_SD = (uptake_SD/float(uptake)) * Dcorrected
                        else:
                            Dcorrected_SD = -99999
                    else:
                        Dcorrected = -99999
                        Dcorrected_SD = -99999
                    Dcorrected_tups.append((Dcorrected, timepoint))
                    Dcorrected_SD_tups.append((Dcorrected_SD, timepoint))
                pepdic_cor[peptide] = Dcorrected_tups
                sddic_cor[peptide] = Dcorrected_SD_tups
            statedic_of_pepdic_cor[state] = pepdic_cor
            statedic_of_sddic_cor[state] = sddic_cor
            


    global pro_peptide_starts
    global pro_peptide_ends
    pro_peptide_starts = {}
    pro_peptide_ends = {}
    
    if sdbt_clicked:
        for peptide_instance in class_peptides:
            if (peptide_instance.Protein, peptide_instance.Sequence) not in pro_peptide_starts:
                pro_peptide_starts[(peptide_instance.Protein, peptide_instance.Sequence)] = [peptide_instance.Startvalue]
            if (peptide_instance.Protein, peptide_instance.Sequence) not in pro_peptide_ends:
                pro_peptide_ends[(peptide_instance.Protein, peptide_instance.Sequence)] = [peptide_instance.Endvalue]
                
    if cdbt_clicked:
        for peptide_instance in class_peptides:
            if (peptide_instance.Protein, peptide_instance.Sequence) not in pro_peptide_starts:
                pro_peptide_starts[(peptide_instance.Protein, peptide_instance.Sequence)] = [peptide_instance.Startvalue]
            if (peptide_instance.Protein, peptide_instance.Sequence) not in pro_peptide_ends:
                pro_peptide_ends[(peptide_instance.Protein, peptide_instance.Sequence)] = [peptide_instance.Endvalue]
                
    if examiner_clicked:
        for peptide_instance in class_peptides:
            if (peptide_instance.Protein, peptide_instance.Sequence) not in pro_peptide_starts:
                pro_peptide_starts[(peptide_instance.Protein, peptide_instance.Sequence)] = [peptide_instance.Startvalue]
            if (peptide_instance.Protein, peptide_instance.Sequence) not in pro_peptide_ends:
                pro_peptide_ends[(peptide_instance.Protein, peptide_instance.Sequence)] = [peptide_instance.Endvalue]
                
    if workbench_clicked:
        for peptide_instance in class_peptides:
            if (peptide_instance.Protein, peptide_instance.Sequence) not in pro_peptide_starts:
                pro_peptide_starts[(peptide_instance.Protein, peptide_instance.Sequence)] = [peptide_instance.Startvalue]
            if (peptide_instance.Protein, peptide_instance.Sequence) not in pro_peptide_ends:
                pro_peptide_ends[(peptide_instance.Protein, peptide_instance.Sequence)] = [peptide_instance.Endvalue]


    global seqlist_dic
    beginnings = {}
    seqlist_dic = {}
    seqlist_dic_proteins = {}

    if seqbt_txt_clicked == True:
        seq.seek(0)
        seqlist = list()
        for line in seq:
            line = line.rstrip()
            for r in line:
                seqlist.append(r)
                
        for state in states:
            this_protein = state.split("~")[0]
            peptide_start_list = list()
            peptide_end_list = list()
            for (pro, peptide), start in pro_peptide_starts.items():
                if pro == this_protein:
                    peptide_start_list.append(start)
            beginning_l = min(peptide_start_list)
            beginnings[state] = beginning_l[0]
            for (pro, peptide), end in pro_peptide_ends.items():
                if pro == this_protein:
                    peptide_end_list.append(end)
            ending_l = max(peptide_end_list)
            ending = ending_l[0]
            b_e_range = (ending - beginnings[state])
            
            
            pep_in_pro = False
            for (pro, peptide), start in pro_peptide_starts.items():
                if pro == this_protein:
                    if peptide in ''.join(seqlist):
                        pep_in_pro = True
                        break
            
            if pep_in_pro == False:
                seqlist_A = list()
                x = 0
                while x <= b_e_range:
                    seqlist_A.append("A")
                    x += 1
                seqlist_dic[state] = seqlist_A 
                
            if pep_in_pro == True:
                if len(seqlist) >= b_e_range:
                    seqlist_dic[state] = seqlist

                if len(seqlist) < b_e_range:
                    seqlist_A = list()
                    x = 0
                    while x <= b_e_range:
                        seqlist_A.append("A")
                        x += 1
                    seqlist_dic[state] = seqlist_A 
                
            peptide_start_list = list()
            peptide_end_list = list()
            for (the_protein, peptide), start in pro_peptide_starts.items(): 
                if the_protein == this_protein:
                    peptide_start_list.append(start)
            beginning_l = min(peptide_start_list)
            beginnings[state] = beginning_l[0]
            
            


    if seqbt_fasta_clicked or txt_h_bt_clicked:
        for protein, s in list(prot_seq_dic.items()):
            if protein + "_PTM" in protein_states.keys():
                prot_seq_dic[protein + "_PTM"] = s
        for protein, s in prot_seq_dic.items():
            seqlist = list()
            for line in s:
                line = line.strip()
                for r in line:
                    seqlist.append(r)
            seqlist_dic_proteins[protein] = seqlist
        seqlist_dic = {}
        for protein, sequence in seqlist_dic_proteins.items():
            for state in states:
                this_protein = state.split("~")[0]
                if state.startswith(protein):
                    peptide_start_list = list()
                    peptide_end_list = list()
                    for (pro, peptide), start in pro_peptide_starts.items():
                        if pro == this_protein:
                            peptide_start_list.append(start)
                    beginning_l = min(peptide_start_list)
                    beginnings[state] = beginning_l[0]
                    for (pro, peptide), end in pro_peptide_ends.items():
                        if pro == this_protein:
                            peptide_end_list.append(end)
                    ending_l = max(peptide_end_list)
                    ending = ending_l[0]
                    b_e_range = ending - beginnings[state]

                    if len(sequence) < b_e_range:
                        continue

                    seqlist_dic[state] = sequence

        for state in states:
            this_protein = state.split("~")[0]
            if state not in seqlist_dic:
                peptide_start_list = list()
                peptide_end_list = list()
                for (pro, peptide), start in pro_peptide_starts.items():
                    if pro == this_protein:
                        peptide_start_list.append(start)
                beginning_l = min(peptide_start_list)
                beginnings[state] = beginning_l[0]
                for (pro, peptide), end in pro_peptide_ends.items():
                    if pro == this_protein:
                        peptide_end_list.append(end)
                ending_l = max(peptide_end_list)
                ending = ending_l[0]
                b_e_range = ending - beginnings[state]
                seqlist = list()
                x = 0
                while x <= b_e_range:
                    seqlist.append("A")
                    x += 1
                seqlist_dic[state] = seqlist
                    



            
    if skip_bt_clicked == True:
        for state in states:
            this_protein = state.split("~")[0]
            peptide_start_list = list()
            peptide_end_list = list()
            for (pro, peptide), start in pro_peptide_starts.items():
                if pro == this_protein:
                    peptide_start_list.append(start)
            beginning_l = min(peptide_start_list)
            beginnings[state] = beginning_l[0]
            for (pro, peptide), end in pro_peptide_ends.items():
                if pro == this_protein:
                    peptide_end_list.append(end)
            ending_l = max(peptide_end_list)
            ending = ending_l[0]
            b_e_range = (ending - beginnings[state])
            seqlist = list()
            x = 0
            while x <= b_e_range:
                seqlist.append("A")
                x += 1
            seqlist_dic[state] = seqlist
            
    #here we are labelling where in the sequence the protein starts
    global seq_start
    seq_start = {}
    if seqbt_txt_clicked == True:
        for state in states:
            sequence = ""
            for res in seqlist_dic[state]:
                sequence = sequence + res
            A_count = sequence.count("A")
            if A_count == len(sequence):
                seq_start[state] = beginnings[state]
            protein = state.split("~")[0]
            for peptide in peplist[state]:
                one_peptide_sequence = peptide
                one_peptide_starts = pro_peptide_starts.get((protein, peptide), None)
                one_peptide_start = int(one_peptide_starts[0])

                if one_peptide_sequence not in sequence:
                    continue
                else:
                    split_sequence = sequence.split(peptide)
                    before_space = split_sequence[0]
                    seq_start[state] = one_peptide_start - len(before_space)
                    break
                    

                    
                    

    if skip_bt_clicked:
        for state in states:
            seq_start[state] = beginnings[state]
    if seqbt_fasta_clicked or txt_h_bt_clicked:
        for state in states:
            protein = state.split("~")[0]
            sequence = ""
            for res in seqlist_dic[state]:
                sequence = sequence + res
            for peptide in peplist[state]:
                one_peptide_sequence = peptide
                one_peptide_starts = pro_peptide_starts.get((protein, peptide), None)
                one_peptide_start = int(one_peptide_starts[0])

                if one_peptide_sequence not in sequence:
                    continue
                else:
                    split_sequence = sequence.split(peptide)
                    before_space = split_sequence[0]
                    seq_start[state] = one_peptide_start - len(before_space)
                    break
    for state in states:
        if state not in seq_start:
            try:
                seq_start[state] = beginnings[state]
            except:
                seq_start[state] = 0




def r_make_legend1():
    ws = wb.create_sheet(title="Figure Legends")
    fig, ax = plt.subplots()
    xpos = uptake_color_length + 1
    if uptake_color_length >= 1:
        color = assign_hex(uptake_col_1)
        square_1 = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
        ax.plot([xpos, xpos], [1, 1.3], color='black', linewidth=1)
        ax.text(xpos, 1.35, round(uptake_val_1 * 100), ha='center', va='bottom', fontsize=12)
        ax.plot([xpos + 1, xpos + 1], [1, 1.3], color='black', linewidth=1)
        ax.text(xpos + 1, 1.35, "100", ha='center', va='bottom', fontsize=12)
        xpos -= 1
        ax.add_patch(square_1)
    if uptake_color_length >= 2:
        color = assign_hex(uptake_col_2)
        square_2 = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
        ax.plot([xpos, xpos], [1, 1.3], color='black', linewidth=1)
        ax.text(xpos, 1.35, round(uptake_val_2 * 100), ha='center', va='bottom', fontsize=12)
        xpos -= 1
        ax.add_patch(square_2)
    if uptake_color_length >= 3:
        color = assign_hex(uptake_col_3)
        square_3 = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
        ax.plot([xpos, xpos], [1, 1.3], color='black', linewidth=1)
        ax.text(xpos, 1.35, round(uptake_val_3 * 100), ha='center', va='bottom', fontsize=12)
        xpos -= 1
        ax.add_patch(square_3)
    if uptake_color_length >= 4:
        color = assign_hex(uptake_col_4)
        square_4 = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
        ax.plot([xpos, xpos], [1, 1.3], color='black', linewidth=1)
        ax.text(xpos, 1.35, round(uptake_val_4 * 100), ha='center', va='bottom', fontsize=12)
        xpos -= 1
        ax.add_patch(square_4)
    if uptake_color_length >= 5:
        color = assign_hex(uptake_col_5)
        square_5 = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
        ax.plot([xpos, xpos], [1, 1.3], color='black', linewidth=1)
        ax.text(xpos, 1.35, round(uptake_val_5 * 100), ha='center', va='bottom', fontsize=12)
        xpos -= 1
        ax.add_patch(square_5)
    if uptake_color_length >= 6:
        color = assign_hex(uptake_col_6)
        square_6 = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
        ax.plot([xpos, xpos], [1, 1.3], color='black', linewidth=1)
        ax.text(xpos, 1.35, round(uptake_val_6 * 100), ha='center', va='bottom', fontsize=12)
        xpos -= 1
        ax.add_patch(square_6)
    if uptake_color_length >= 7:
        color = assign_hex(uptake_col_7)
        square_7 = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
        ax.plot([xpos, xpos], [1, 1.3], color='black', linewidth=1)
        ax.text(xpos, 1.35, round(uptake_val_7 * 100), ha='center', va='bottom', fontsize=12)
        xpos -= 1
        ax.add_patch(square_7)
    if uptake_color_length >= 8:
        color = assign_hex(uptake_col_8)
        square_8 = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
        ax.plot([xpos, xpos], [1, 1.3], color='black', linewidth=1)
        ax.text(xpos, 1.35, round(uptake_val_8 * 100), ha='center', va='bottom', fontsize=12)
        xpos -= 1
        ax.add_patch(square_8)
    if uptake_color_length >= 9:
        color = assign_hex(uptake_col_9)
        square_9 = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
        ax.plot([xpos, xpos], [1, 1.3], color='black', linewidth=1)
        ax.text(xpos, 1.35, round(uptake_val_9 * 100), ha='center', va='bottom', fontsize=12)
        xpos -= 1
        ax.add_patch(square_9)
    color = assign_hex(uptake_gtz_key)
    square_10 = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
    ax.plot([xpos, xpos], [1, 1.3], color='black', linewidth=1)
    ax.text(xpos, 1.35, "0", ha='center', va='bottom', fontsize=12)
    ax.add_patch(square_10)

    color = assign_hex(uptake_abs_key)
    square_11 = patches.Rectangle((xpos, -1.5), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
    ax.add_patch(square_11)
    ax.text(xpos + 2, -1.25, " No Data", ha='center', va='bottom', fontsize=14)

    ax.spines['top'].set_visible(False)
    ax.spines['bottom'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.spines['right'].set_visible(False)

    ax.set_aspect('equal')
    ax.set_xlim(-0.5, uptake_color_length + 3.5)
    ax.set_ylim(-3, 2.5)
    ax.set_xticks([])
    ax.set_yticks([])
    try:
        fig.savefig('./RecentLegends/uptakelegend.png', dpi=300) 
    except:
        tk.messagebox.showerror("File Access Error", f"An error has occured writing files to this computer. Please navigate to {os.getcwd()} and right click on the RecentLegends folder. Select Properties then Security, and make sure users have permission to modify the contents of this folder and rerun the program. If you cannot see the AppData folder, got to view>show>hidden items.")
        quit()
    plt.close()
    
    
    
    if theo_bt_on_c == True:
        if back_exchange == 0:
            ws.cell(row=1, column=8, value="Values are not corrected for Back Exchange")
        else:
            ws.cell(row=1, column=8, value=f"Values Corrected for {back_exchange}% Back Exchange")
    elif exp_bt_on_c == True:
        ws.cell(row=1, column=8, value=f"Values Corrected for maximally deuterated control")
        
    if n_min_two == True:
        ws.cell(row=2, column=8, value="All calculations assuming no measured exchange for N-2")
    if n_min_one == True:
        ws.cell(row=2, column=8, value="All calculations assuming no measured exchange for N-1")
    if calculate_theoretical_back_exchange == True:
        ws.cell(row=2, column=8, value=f"All calculations use Englander Rate constants with time in H2O equal to {englander_time_in_H2O_trap} + {englander_time_in_H2O_constant}*peptide RT, pH = {englander_pH}, and temperateure in kelvin equal to {englander_tempK}")
  
    img = openpyxl.drawing.image.Image('./RecentLegends/uptakelegend.png')
    img.anchor = 'A7'
    ws.add_image(img)

def r_make_legend2(save_to_wb):
    if save_to_wb == True:
        fig, ax = plt.subplots()
    else:
        fig, ax = plt.subplots(figsize=(6.4, 2.0))
    xpos = p_col_length + d_col_length + 1
    if save_to_wb == True:
        text_fontsize = 12
    else:
        text_fontsize = 5
    if d_col_length >= 1:
        color = assign_hex(d_col_1)
        square_1 = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
        ax.plot([xpos, xpos], [1, 1.3], color='black', linewidth=1)
        if exp_bt_on_c == True and maxD_rfu_dif_on_c == True:
            ax.text(xpos, 1.35, round(d_val_1 * 100), ha='center', va='bottom', fontsize=text_fontsize)
        else:
            ax.text(xpos, 1.35, d_val_1, ha='center', va='bottom', fontsize=text_fontsize)
        xpos -= 1
        ax.add_patch(square_1)
    if d_col_length >= 2:
        color = assign_hex(d_col_2)
        square_2 = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
        ax.plot([xpos, xpos], [1, 1.3], color='black', linewidth=1)
        if exp_bt_on_c == True and maxD_rfu_dif_on_c == True:
            ax.text(xpos, 1.35, round(d_val_2 * 100), ha='center', va='bottom', fontsize=text_fontsize)
        else:
            ax.text(xpos, 1.35, d_val_2, ha='center', va='bottom', fontsize=text_fontsize)
        xpos -= 1
        ax.add_patch(square_2)
    if d_col_length >= 3:
        color = assign_hex(d_col_3)
        square_3 = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
        ax.plot([xpos, xpos], [1, 1.3], color='black', linewidth=1)
        if exp_bt_on_c == True and maxD_rfu_dif_on_c == True:
            ax.text(xpos, 1.35, round(d_val_3 * 100), ha='center', va='bottom', fontsize=text_fontsize)
        else:
            ax.text(xpos, 1.35, d_val_3, ha='center', va='bottom', fontsize=text_fontsize)
        xpos -= 1
        ax.add_patch(square_3)
    if d_col_length >= 4:
        color = assign_hex(d_col_4)
        square_4 = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
        ax.plot([xpos, xpos], [1, 1.3], color='black', linewidth=1)
        if exp_bt_on_c == True and maxD_rfu_dif_on_c == True:
            ax.text(xpos, 1.35, round(d_val_4 * 100), ha='center', va='bottom', fontsize=text_fontsize)
        else:
            ax.text(xpos, 1.35, d_val_4, ha='center', va='bottom', fontsize=text_fontsize)
        xpos -= 1
        ax.add_patch(square_4)
    if d_col_length >= 5:
        color = assign_hex(d_col_5)
        square_5 = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
        ax.plot([xpos, xpos], [1, 1.3], color='black', linewidth=1)
        if exp_bt_on_c == True and maxD_rfu_dif_on_c == True:
            ax.text(xpos, 1.35, round(d_val_5 * 100), ha='center', va='bottom', fontsize=text_fontsize)
        else:
            ax.text(xpos, 1.35, d_val_5, ha='center', va='bottom', fontsize=text_fontsize)
        xpos -= 1
        ax.add_patch(square_5)
    color = assign_hex(d_col_gtz)
    square_6 = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
    ax.plot([xpos, xpos], [1, 1.3], color='black', linewidth=1)
    ax.text(xpos, 1.35, "0", ha='center', va='bottom', fontsize=text_fontsize)
    xpos -= 1
    ax.add_patch(square_6)
    color = assign_hex(p_col_gtz)
    square_7 = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
    ax.plot([xpos, xpos], [1, 1.3], color='black', linewidth=1)
    xpos -= 1
    ax.add_patch(square_7)
    if p_col_length >= 5:
        if exp_bt_on_c == True and maxD_rfu_dif_on_c == True:
            ax.text(xpos + 1, 1.35, round(p_val_5 * 100), ha='center', va='bottom', fontsize=text_fontsize)
        else:
            ax.text(xpos + 1, 1.35, p_val_5, ha='center', va='bottom', fontsize=text_fontsize)
        color = assign_hex(p_col_5)
        square_8 = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
        ax.plot([xpos, xpos], [1, 1.3], color='black', linewidth=1)
        xpos -= 1
        ax.add_patch(square_8)
    if p_col_length >= 4:
        if exp_bt_on_c == True and maxD_rfu_dif_on_c == True:
            ax.text(xpos + 1, 1.35, round(p_val_4 * 100), ha='center', va='bottom', fontsize=text_fontsize)
        else:
            ax.text(xpos + 1, 1.35, p_val_4, ha='center', va='bottom', fontsize=text_fontsize)
        color = assign_hex(p_col_4)
        square_9 = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
        ax.plot([xpos, xpos], [1, 1.3], color='black', linewidth=1)
        xpos -= 1
        ax.add_patch(square_9)
    if p_col_length >= 3:
        if exp_bt_on_c == True and maxD_rfu_dif_on_c == True:
            ax.text(xpos + 1, 1.35, round(p_val_3 * 100), ha='center', va='bottom', fontsize=text_fontsize)
        else:
            ax.text(xpos + 1, 1.35, p_val_3, ha='center', va='bottom', fontsize=text_fontsize)
        color = assign_hex(p_col_3)
        square_10 = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
        ax.plot([xpos, xpos], [1, 1.3], color='black', linewidth=1)
        xpos -= 1
        ax.add_patch(square_10)
    if p_col_length >= 2:
        if exp_bt_on_c == True and maxD_rfu_dif_on_c == True:
            ax.text(xpos + 1, 1.35, round(p_val_2 * 100), ha='center', va='bottom', fontsize=text_fontsize)
        else:
            ax.text(xpos + 1, 1.35, p_val_2, ha='center', va='bottom', fontsize=text_fontsize)
        color = assign_hex(p_col_2)
        square_11 = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
        ax.plot([xpos, xpos], [1, 1.3], color='black', linewidth=1)
        xpos -= 1
        ax.add_patch(square_11)
    if p_col_length >= 1:
        if exp_bt_on_c == True and maxD_rfu_dif_on_c == True:
            ax.text(xpos + 1, 1.35, round(p_val_1 * 100), ha='center', va='bottom', fontsize=text_fontsize)
        else:
            ax.text(xpos + 1, 1.35, p_val_1, ha='center', va='bottom', fontsize=text_fontsize)
        color = assign_hex(p_col_1)
        square_12 = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
        ax.add_patch(square_12)
    
    if save_to_wb == True:
        color = assign_hex(b_col_abs)
        square_13 = patches.Rectangle((xpos, -1.5), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
        ax.add_patch(square_13)
        ax.text(xpos + 3, -1.25, " No Data", ha='center', va='bottom', fontsize=14)
    if save_to_wb == False:
        color = assign_hex(b_col_abs)
        square_13 = patches.Rectangle((p_col_length + d_col_length + 3, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
        ax.add_patch(square_13)
        

    ax.spines['top'].set_visible(False)
    ax.spines['bottom'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.spines['right'].set_visible(False)

    if save_to_wb == True:
        ax.set_aspect('equal')
        ax.set_xlim(-0.5, p_col_length + d_col_length + 3.5)
        ax.set_ylim(-3, 2.5)
        ax.set_xticks([])
        ax.set_yticks([])
    if save_to_wb == False:
        ax.set_aspect('equal')
        ax.set_xticks([])
        ax.set_yticks([])
    
    if save_to_wb == True:
        fig.savefig('./RecentLegends/differencelegend.png', dpi=300)
        plt.close()
        ws = wb['Figure Legends']
        img = openpyxl.drawing.image.Image('./RecentLegends/differencelegend.png')
        img.anchor = 'A87'
        ws.add_image(img)
    else:
        return fig, ax
    


def r_make_legend3():
    new_items_list = [lcol0, lcol1, lcol2, lcol3, lcol4, lcol5, lcol6, lcol7, lcol8, lcol9]
    false_item_index_list = []
    for i, item in enumerate(new_items_list):
        if item == False:
            false_item_index_list.append(i)

    color_mapping = {}
    for i, item in enumerate(new_items_list):
        if i in false_item_index_list:
            color_mapping[i] = "#000000"
            continue
        color_mapping[i] = "#" + str(item)
    


    color_indexes = []
    color_indexes_possible = [7, 6, 2, 1, 0, 4, 5, 8, 9, 3]
    for i, item in enumerate(color_indexes_possible):
        if item in false_item_index_list:
            continue
        if color_indexes_possible[i] == 3:
            continue
        color_indexes.append(color_indexes_possible[i])
    
    
    fig, ax = plt.subplots(figsize=(6, 2))
    
    
        
    xpos = 0
    for n in color_indexes:
        color = color_mapping[n]
        square = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
        ax.plot([xpos+0.5, xpos+0.5], [1, 1.3], color='black', linewidth=1)
        xpos += 1
        ax.add_patch(square)

    xpos += 1
    color = color_mapping[3]
    square = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
    ax.plot([xpos+0.5, xpos+0.5], [1, 1.3], color='black', linewidth=1)
    ax.add_patch(square)
    
    
    ax.spines['top'].set_visible(False)
    ax.spines['bottom'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.spines['right'].set_visible(False)

    ax.set_aspect('equal')
    ax.set_xticks([])
    ax.set_yticks([])
    

    
    fig.savefig('./RecentLegends/linear_map_scale.png', dpi=300)
    
    ws = wb['Figure Legends']
    img = openpyxl.drawing.image.Image('./RecentLegends/linear_map_scale.png')
    img.anchor = 'A158'
    ws.add_image(img)

def assign_hex(col):
    color = "#" + col
    return color


def r_pepmaps():
    function_worksheet_names = []
    try:
        if reduce_states_var.get() == 0:
            states_to_look_in = statedic_of_pepdic_cor
        if reduce_states_var.get() == 1:
            states_to_look_in = order_state_dic.values()
            states_to_look_in = [x for x in states_to_look_in if x != False]

        for state in states_to_look_in:
            highest_row = 1
            asterisk_in_state = False
            protein = state.split("~")[0]
            sorted_peptides = sorted(peplist[state], key=lambda x: (int(pro_peptide_starts.get((x[0], x[1]), [0])[0]), len(x[1])))
            ws_title = (f"{state}")[-30:]
            ws = wb.create_sheet(title=ws_title)
            function_worksheet_names.append(ws_title)
            ws.append(["Timepoint"])
            ws.append([" "] + seqlist_dic[state] + [" "])
            timepoint_number = 0
            for timepoint in s_timepoints[state]:
                if timepoint_number == 0:
                    startrow = 3
                    endrow = 250
                if timepoint_number != 0:
                    startrow = ws.max_row +5
                    endrow = ws.max_row + 250
                for peptide in sorted_peptides:
                    startvalues = pro_peptide_starts.get((protein, peptide), None)
                    startvalue= int(startvalues[0]) - seq_start[state]
                    endvalues = pro_peptide_ends.get((protein, peptide), None)
                    endvalue = int(endvalues[0]) - seq_start[state]
                    peptide_length = len(peptide)
                    Cuptake = None
                    try:
                        for up, tp in statedic_of_pepdic_cor[state][peptide]:
                            if tp == timepoint:
                                Cuptake = up
                    except:
                        pass
                    if Cuptake is not None:
                        for row in ws.iter_rows(min_row=startrow, max_row=startrow):
                            row[0].value = timepoint
                        for i, row in enumerate(ws.iter_rows(min_row=startrow,max_row=endrow), start=startrow):
                            cells = row[startvalue + 1:endvalue + 2]
                            if all(cell.value is None for cell in cells):
                                if i > highest_row:
                                    highest_row = i
                                row[startvalue + 1].value = Cuptake
                                row[startvalue + 1].border = Border(top=Side(border_style='thin', color='FF000000'),
                                         bottom=Side(border_style='thin', color='FF000000'),
                                         left=Side(border_style='thin', color='FF000000'))
                                row[endvalue+1].value = Cuptake
                                row[endvalue+1].border = Border(top=Side(border_style='thin', color='FF000000'),
                                         bottom=Side(border_style='thin', color='FF000000'),
                                         right=Side(border_style='thin', color='FF000000'))
                                for cell in row[startvalue + 2:endvalue+1]:
                                    cell.value = Cuptake
                                    cell.border = Border(top=Side(border_style='thin', color='FF000000'),
                                         bottom=Side(border_style='thin', color='FF000000'))

                                try:
                                    if peptide in noD_dic_states[state]:
                                        if Cuptake != 0 and Cuptake != -99999:
                                            row[startvalue+1].value = "*"
                                            asterisk_in_state = True
                                            #row[startvalue+1].alignment = Alignment(textRotation=90, vertical='center')
                                            if uptake_color_length >= 1 and Cuptake > uptake_val_1:
                                                fill = PatternFill(start_color=uptake_col_1, end_color=uptake_col_1, fill_type='solid')
                                                font = Font(color=uptake_text_1, size=12)
                                            elif uptake_color_length >= 2 and Cuptake > uptake_val_2:
                                                fill = PatternFill(start_color=uptake_col_2, end_color=uptake_col_2, fill_type='solid')
                                                font = Font(color=uptake_text_2, size=12)
                                            elif uptake_color_length >= 3 and Cuptake > uptake_val_3:
                                                fill = PatternFill(start_color=uptake_col_3, end_color=uptake_col_3, fill_type='solid')
                                                font = Font(color=uptake_text_3, size=12)
                                            elif uptake_color_length >= 4 and Cuptake > uptake_val_4:
                                                fill = PatternFill(start_color=uptake_col_4, end_color=uptake_col_4, fill_type='solid')
                                                font = Font(color=uptake_text_4, size=12)
                                            elif uptake_color_length >= 5 and Cuptake > uptake_val_5:
                                                fill = PatternFill(start_color=uptake_col_5, end_color=uptake_col_5, fill_type='solid')
                                                font = Font(color=uptake_text_5, size=12)
                                            elif uptake_color_length >= 6 and Cuptake > uptake_val_6:
                                                fill = PatternFill(start_color=uptake_col_6, end_color=uptake_col_6, fill_type='solid')
                                                font = Font(color=uptake_text_6, size=12)
                                            elif uptake_color_length >= 7 and Cuptake > uptake_val_7:
                                                fill = PatternFill(start_color=uptake_col_7, end_color=uptake_col_7, fill_type='solid')
                                                font = Font(color=uptake_text_7, size=12)
                                            elif uptake_color_length >= 8 and Cuptake > uptake_val_8:
                                                fill = PatternFill(start_color=uptake_col_8, end_color=uptake_col_8, fill_type='solid')
                                                font = Font(color=uptake_text_8, size=12)
                                            elif uptake_color_length >= 9 and Cuptake > uptake_val_9:
                                                fill = PatternFill(start_color=uptake_col_9, end_color=uptake_col_9, fill_type='solid')
                                                font = Font(color=uptake_text_9, size=12)
                                            elif Cuptake > 0.0:
                                                fill = PatternFill(start_color=uptake_gtz_key, end_color=uptake_gtz_key, fill_type='solid')
                                                font = Font(color=uptake_gtz_text, size=12)
                                            elif Cuptake == 0:
                                                fill = PatternFill(start_color=uptake_eqz_key, end_color=uptake_eqz_key, fill_type='solid')
                                                font = Font(color=uptake_eqz_text, size=12)
                                                cell.number_format = ';;;'
                                            elif Cuptake == -99999:
                                                fill = PatternFill(start_color=uptake_abs_key, end_color=uptake_abs_key, fill_type='solid')
                                                font = Font(color=uptake_abs_text, size=12)
                                                cell.number_format = ';;;'
                                            elif Cuptake < 0.0:
                                                fill = PatternFill(start_color=uptake_ltz_key, end_color=uptake_ltz_key, fill_type='solid')
                                                font = Font(color=uptake_ltz_text, size=12)
                                            row[startvalue+1].fill = fill
                                            row[startvalue+1].font = font
                                except:
                                    pass







                                break
                            else:
                                continue



                timepoint_number = timepoint_number + 1


            for row in ws.iter_rows(min_row=2, max_row=2):
                for cell in row:
                    cell.font = courier_new_style

            if asterisk_in_state == True:
                for row in ws.iter_rows(min_row=highest_row+2, max_row=highest_row+2):
                    row[1].value = "* = no maxD for peptide, average back exchange used"


            white_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type='solid')
            for row in ws.rows:
                for cell in row:
                    if cell.value != "*":
                        cell.fill = white_fill



            for i, column in enumerate(ws.columns):
                if i == 0:
                    continue
                #ws.column_dimensions[column[0].column_letter].width = full_pep_width_enter.get()
                ws.column_dimensions[column[0].column_letter].width = "4"

            for row in ws.rows:
                ws.row_dimensions[row[0].row].height = 14.4



            for row in ws.iter_rows(min_row=3):
                for cell in row[1:]:
                    cell_v = cell.value
                    if cell_v != "*" and cell_v != "* = no maxD for peptide, average back exchange used":
                        if uptake_color_length >= 1 and cell_v is not None and cell_v > uptake_val_1:
                            fill = PatternFill(start_color=uptake_col_1, end_color=uptake_col_1, fill_type='solid')
                            cell.number_format = ';;;'
                        elif uptake_color_length >= 2 and cell_v is not None and cell_v > uptake_val_2:
                            fill = PatternFill(start_color=uptake_col_2, end_color=uptake_col_2, fill_type='solid')
                            cell.number_format = ';;;'
                        elif uptake_color_length >= 3 and cell_v is not None and cell_v > uptake_val_3:
                            fill = PatternFill(start_color=uptake_col_3, end_color=uptake_col_3, fill_type='solid')
                            cell.number_format = ';;;'
                        elif uptake_color_length >= 4 and cell_v is not None and cell_v > uptake_val_4:
                            fill = PatternFill(start_color=uptake_col_4, end_color=uptake_col_4, fill_type='solid')
                            cell.number_format = ';;;'
                        elif uptake_color_length >= 5 and cell_v is not None and cell_v > uptake_val_5:
                            fill = PatternFill(start_color=uptake_col_5, end_color=uptake_col_5, fill_type='solid')
                            cell.number_format = ';;;'
                        elif uptake_color_length >= 6 and cell_v is not None and cell_v > uptake_val_6:
                            fill = PatternFill(start_color=uptake_col_6, end_color=uptake_col_6, fill_type='solid')
                            cell.number_format = ';;;'
                        elif uptake_color_length >= 7 and cell_v is not None and cell_v > uptake_val_7:
                            fill = PatternFill(start_color=uptake_col_7, end_color=uptake_col_7, fill_type='solid')
                            cell.number_format = ';;;'
                        elif uptake_color_length >= 8 and cell_v is not None and cell_v > uptake_val_8:
                            fill = PatternFill(start_color=uptake_col_8, end_color=uptake_col_8, fill_type='solid')
                            cell.number_format = ';;;'
                        elif uptake_color_length >= 9 and cell_v is not None and cell_v > uptake_val_9:
                            fill = PatternFill(start_color=uptake_col_9, end_color=uptake_col_9, fill_type='solid')
                            cell.number_format = ';;;'
                        elif cell_v is not None and cell_v > 0.0:
                            fill = PatternFill(start_color=uptake_gtz_key, end_color=uptake_gtz_key, fill_type='solid')
                            cell.number_format = ';;;'
                        elif cell_v == 0:
                            fill = PatternFill(start_color=uptake_eqz_key, end_color=uptake_eqz_key, fill_type='solid')
                            cell.number_format = ';;;'
                        elif cell_v == -99999:
                            fill = PatternFill(start_color=uptake_abs_key, end_color=uptake_abs_key, fill_type='solid')
                            cell.number_format = ';;;'
                        elif cell_v is not None and cell_v < 0.0:
                            fill = PatternFill(start_color=uptake_ltz_key, end_color=uptake_ltz_key, fill_type='solid')
                            cell.number_format = ';;;'
                        if cell_v is not None:
                            cell.fill = fill


            increase_progress(1)

            for row in ws.iter_rows(min_row=1, max_row=1):
                num = seq_start[state]
                for cell in row:
                    if cell.column >= 2 and cell.column < ws.max_column:
                        cell.value = num
                        num = num+1

            for row in ws.iter_rows(max_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')

            if enumerate_residues_var.get() == 0:
                for row in ws.iter_rows(min_row=2, max_row=2):
                    ws.row_dimensions[2].height = 20
                    for cell in row:
                        cell.font = Font(name="Courier New", size=20)
                for row in ws.iter_rows(min_row=1, max_row=1):
                    ws.row_dimensions[1].height = 20
                    row[1].font = Font(name="Courier New", size=14)
                    for i, cell in enumerate(row[2:], start=2):
                        if cell.value == None:
                            continue
                        elif cell.value % 10 == 0:
                            try:
                                row[i-1].value = cell.value
                                ws.merge_cells(start_row=cell.row, start_column=row[i-1].column, end_row=cell.row, end_column=row[i+1].column)
                                row[i-1].alignment = Alignment(horizontal="center")
                                row[i-1].font = Font(name="Courier New", size=14)  
                            except:
                                print("Excepting during cell numbering cell merge")
                        else:
                            cell.value = None
                for cell in ws["A"]:
                    cell.font = Font(name="Courier New", size=14, bold=True)
                    cell.alignment = Alignment(horizontal="center") 
                    ws.column_dimensions["A"].width = 15  

            
    except Exception as e:
        error_message = traceback.format_exc() 
        print(error_message)
        tk.messagebox.showerror("Unknown Error in Peptide Plot", error_message)
        if delete_faulty_sheets == True:
            for sheet_name in function_worksheet_names:
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]

def r_difmaps():
    function_worksheet_names = []
    try:
        global difference_titles, difference_states, seqlist_dic
        difference_titles = []
        difference_states = []
        for stt, pair in  new_dic_of_dif_list.items():
            highest_row = 1
            asterisk_in_state = False
            first = pair[0]
            second = pair[1]
            protein_one = first.split("~")[0]
            protein_two = second.split("~")[0]
            sorted_peptides_first = sorted(peplist[first], key=lambda x: (int(pro_peptide_starts.get((x[0], x[1]), [0])[0]), len(x[1])))
            sorted_peptides_second = sorted(peplist[second], key=lambda x: (int(pro_peptide_starts.get((x[0], x[1]), [0])[0]), len(x[1])))
            difname = f"{stt}"
            wtit = (f"{difname}" + "_dif")
            difference_titles.append(difname)
            difference_states.append((first, second))
            ws = wb.create_sheet(title=wtit)
            function_worksheet_names.append(wtit)
            ws.append(["Timepoint"])
            ws.append([" "] + seqlist_dic[first] + [" "])
            timepoint_number = 0
            for timepoint in s_timepoints[first]:
                if timepoint in s_timepoints[second]:
                    if timepoint_number == 0:
                        startrow = 3
                        endrow = 250
                    if timepoint_number != 0:
                        startrow = ws.max_row +5
                        endrow = ws.max_row + 250
                    for peptide in sorted_peptides_first:
                        startvalues = pro_peptide_starts.get((protein_one, peptide), None)
                        startvalue= int(startvalues[0]) - seq_start[first]
                        endvalues = pro_peptide_ends.get((protein_one, peptide), None)
                        endvalue = int(endvalues[0]) - seq_start[first]
                        peptide_length = len(peptide)

                        if exp_bt_on_c == True and maxD_rfu_dif_on_c == True:
                            up1 = None
                            up2 = None
                            diftake = None

                            try:
                                for up, tp in statedic_of_pepdic_cor[first][peptide]:
                                    if tp == timepoint:
                                        up1 = up
                                for up, tp in statedic_of_pepdic_cor[second][peptide]:
                                    if tp == timepoint:
                                        up2 = up
                                if up1 is not None and up2 is not None and up1 != -99999 and up2 != -99999:
                                    diftake = up1 - up2
                                elif up1 is not None and up2 is not None:
                                    diftake = -99999

                            except:
                                pass

                        if exp_bt_on_c == True and maxD_Da_dif_on_c == True:
                            up1 = None
                            up2 = None
                            diftake = None

                            try:
                                for up, tp in statedic_of_pepdic_cor[first][peptide]:
                                    if tp == timepoint:
                                        up1 = up
                                for up, tp in statedic_of_pepdic_cor[second][peptide]:
                                    if tp == timepoint:
                                        up2 = up
                                if up1 is not None and up2 is not None and up1 != -99999 and up2 != -99999:
                                    max_theo = get_max_theo(peptide)
                                    diftake = max_theo * (up1 - up2)
                                elif up1 is not None and up2 is not None:
                                    diftake = -99999

                            except:
                                pass

                        if theo_bt_on_c == True and back_exchange == 0:
                            up1 = None
                            up2 = None
                            diftake = None
                            try:
                                for up, tp in statedic_of_pepdic_raw2[first][peptide]:
                                    if tp == timepoint:
                                        up1 = up
                                for up, tp in statedic_of_pepdic_raw2[second][peptide]:
                                    if tp == timepoint:
                                        up2 = up
                                if up1 is not None and up2 is not None and up1 != -99999 and up2 != -99999:
                                    diftake = up1 - up2
                                elif up1 is not None and up2 is not None:
                                    diftake = -99999
                            except:
                                pass

                        if theo_bt_on_c == True and back_exchange != 0:
                            up1 = None
                            up2 = None
                            diftake = None
                            try:
                                for up, tp in statedic_of_pepdic_cor[first][peptide]:
                                    if tp == timepoint:
                                        up1 = up
                                for up, tp in statedic_of_pepdic_cor[second][peptide]:
                                    if tp == timepoint:
                                        up2 = up
                                max_theo = get_max_theo(peptide)
                                if up1 is not None and up2 is not None and up1 != -99999 and up2 != -99999:
                                    diftake = max_theo * (up1 - up2)
                                elif up1 is not None and up2 is not None:
                                    diftake = -99999
                            except:
                                pass

                        if diftake is not None:
                            for row in ws.iter_rows(min_row=startrow, max_row=startrow):
                                row[0].value = timepoint
                            for i, row in enumerate(ws.iter_rows(min_row=startrow,max_row=endrow), start=startrow):
                                if i > highest_row:
                                    highest_row = i
                                cells = row[startvalue + 1:endvalue + 2]
                                if all(cell.value is None for cell in cells):
                                    row[startvalue + 1].value = diftake
                                    row[startvalue + 1].border = Border(top=Side(border_style='thin', color='FF000000'),
                                             bottom=Side(border_style='thin', color='FF000000'),
                                             left=Side(border_style='thin', color='FF000000'))
                                    row[endvalue+1].value = diftake
                                    row[endvalue+1].border = Border(top=Side(border_style='thin', color='FF000000'),
                                             bottom=Side(border_style='thin', color='FF000000'),
                                             right=Side(border_style='thin', color='FF000000'))
                                    for cell in row[startvalue + 2:endvalue+1]:
                                        cell.value = diftake
                                        cell.border = Border(top=Side(border_style='thin', color='FF000000'),
                                             bottom=Side(border_style='thin', color='FF000000'))
                                    ch1 = False
                                    ch2 = False
                                    try:
                                        if noD_dic_states[first][peptide] == True:
                                            ch1 = True
                                    except:
                                        pass
                                    try:
                                        if noD_dic_states[second][peptide] == True:
                                            ch2 = True
                                    except:
                                        pass

                                    if ch1 == True or ch2 == True:
                                        if diftake != 0 and diftake != -99999:
                                            row[startvalue+1].value = "*"
                                            asterisk_in_state = True
                                            row[startvalue+1].alignment = Alignment(textRotation=90, vertical='center')

                                            if d_col_length >= 1 and diftake >= d_val_1:
                                                fill = PatternFill(start_color=d_col_1, end_color=d_col_1, fill_type='solid')
                                                font = Font(color=d_text_1, size=12)
                                            elif d_col_length >= 2 and diftake >= d_val_2:
                                                fill = PatternFill(start_color=d_col_2, end_color=d_col_2, fill_type='solid')
                                                font = Font(color=d_text_2, size=12)
                                            elif d_col_length >= 3 and diftake >= d_val_3:
                                                fill = PatternFill(start_color=d_col_3, end_color=d_col_3, fill_type='solid')
                                                font = Font(color=d_text_3, size=12)
                                            elif d_col_length >= 4 and diftake >= d_val_4:
                                                fill = PatternFill(start_color=d_col_4, end_color=d_col_4, fill_type='solid')
                                                font = Font(color=d_text_4, size=12)
                                            elif d_col_length >= 5 and diftake >= d_val_5:
                                                fill = PatternFill(start_color=d_col_5, end_color=d_col_5, fill_type='solid')
                                                font = Font(color=d_text_5, size=12)
                                            elif diftake > 0:
                                                fill = PatternFill(start_color=d_col_gtz, end_color=d_col_gtz, fill_type='solid')
                                                font = Font(color=d_text_gtz, size=12)
                                            elif p_col_length >= 1 and diftake <= (-1) * p_val_1:
                                                fill = PatternFill(start_color=p_col_1, end_color=p_col_1, fill_type='solid')
                                                font = Font(color=p_text_1, size=12)
                                            elif p_col_length >= 2 and diftake <= (-1) * p_val_2:
                                                fill = PatternFill(start_color=p_col_2, end_color=p_col_2, fill_type='solid')
                                                font = Font(color=p_text_2, size=12)
                                            elif p_col_length >= 3 and diftake <= (-1) * p_val_3:
                                                fill = PatternFill(start_color=p_col_3, end_color=p_col_3, fill_type='solid')
                                                font = Font(color=p_text_3, size=12)
                                            elif p_col_length >= 4 and diftake <= (-1) * p_val_4:
                                                fill = PatternFill(start_color=p_col_4, end_color=p_col_4, fill_type='solid')
                                                font = Font(color=p_text_4, size=12)
                                            elif p_col_length >= 5 and diftake <= (-1) * p_val_5:
                                                fill = PatternFill(start_color=p_col_5, end_color=p_col_5, fill_type='solid')
                                                font = Font(color=p_text_5, size=12)
                                            elif diftake < 0:
                                                fill = PatternFill(start_color=p_col_gtz, end_color=p_col_gtz, fill_type='solid')
                                                font = Font(color=p_text_gtz, size=12)

                                            row[startvalue+1].fill = fill
                                            row[startvalue+1].font = font




                                    break
                                else:
                                    continue





                    timepoint_number = timepoint_number + 1
            increase_progress(1.5)


            if asterisk_in_state == True:
                for row in ws.iter_rows(min_row=highest_row+2, max_row=highest_row+2):
                    row[1].value = "* = no maxD for peptide, average back exchange used"


            white_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type='solid')
            for row in ws.rows:
                for cell in row:
                    if cell.value != "*":
                        cell.fill = white_fill


            for row in ws.iter_rows(min_row=2, max_row=2):
                for cell in row:
                    cell.font = courier_new_style

            for i, column in enumerate(ws.columns):
                if i == 0:
                    continue
                #ws.column_dimensions[column[0].column_letter].width = full_pep_width_enter.get()
                ws.column_dimensions[column[0].column_letter].width = "4"

            for row in ws.rows:
                ws.row_dimensions[row[0].row].height = 14.4



            for row in ws.iter_rows(min_row=1, max_row=1):
                num = seq_start[first]
                for cell in row:
                    if cell.column >= 2 and cell.column < ws.max_column:
                        cell.value = num
                        num = num+1

            for row in ws.iter_rows(min_row=3):
                for cell in row[1:]:
                    cell_v = cell.value
                    if cell_v != "*" and cell_v != "* = no maxD for peptide, average back exchange used":
                        if cell_v == -99999:
                            fill = PatternFill(start_color=b_col_abs, end_color=b_col_abs, fill_type='solid')
                            cell.number_format = ';;;'
                        elif d_col_length >= 1 and cell_v is not None and cell_v >= d_val_1:
                            fill = PatternFill(start_color=d_col_1, end_color=d_col_1, fill_type='solid')
                            cell.number_format = ';;;'
                        elif d_col_length >= 2 and cell_v is not None and cell_v >= d_val_2:
                            fill = PatternFill(start_color=d_col_2, end_color=d_col_2, fill_type='solid')
                            cell.number_format = ';;;'
                        elif d_col_length >= 3 and cell_v is not None and cell_v >= d_val_3:
                            fill = PatternFill(start_color=d_col_3, end_color=d_col_3, fill_type='solid')
                            cell.number_format = ';;;'
                        elif d_col_length >= 4 and cell_v is not None and cell_v >= d_val_4:
                            fill = PatternFill(start_color=d_col_4, end_color=d_col_4, fill_type='solid')
                            cell.number_format = ';;;'
                        elif d_col_length >= 5 and cell_v is not None and cell_v >= d_val_5:
                            fill = PatternFill(start_color=d_col_5, end_color=d_col_5, fill_type='solid')
                            cell.number_format = ';;;'
                        elif cell_v is not None and cell_v > 0:
                            fill = PatternFill(start_color=d_col_gtz, end_color=d_col_gtz, fill_type='solid')
                            cell.number_format = ';;;'
                        elif p_col_length >= 1 and cell_v is not None and cell_v <= (-1) * p_val_1:
                            fill = PatternFill(start_color=p_col_1, end_color=p_col_1, fill_type='solid')
                            cell.number_format = ';;;'
                        elif p_col_length >= 2 and cell_v is not None and cell_v <= (-1) * p_val_2:
                            fill = PatternFill(start_color=p_col_2, end_color=p_col_2, fill_type='solid')
                            cell.number_format = ';;;'
                        elif p_col_length >= 3 and cell_v is not None and cell_v <= (-1) * p_val_3:
                            fill = PatternFill(start_color=p_col_3, end_color=p_col_3, fill_type='solid')
                            cell.number_format = ';;;'
                        elif p_col_length >= 4 and cell_v is not None and cell_v <= (-1) * p_val_4:
                            fill = PatternFill(start_color=p_col_4, end_color=p_col_4, fill_type='solid')
                            cell.number_format = ';;;'
                        elif p_col_length >= 5 and cell_v is not None and cell_v <= (-1) * p_val_5:
                            fill = PatternFill(start_color=p_col_5, end_color=p_col_5, fill_type='solid')
                            cell.number_format = ';;;'
                        elif cell_v is not None and cell_v < 0:
                            fill = PatternFill(start_color=p_col_gtz, end_color=p_col_gtz, fill_type='solid')
                            cell.number_format = ';;;'
                        elif cell_v is not None and cell_v == 0:
                            fill = PatternFill(start_color=b_col_eqz, end_color=b_col_eqz, fill_type='solid')
                            cell.number_format = ';;;'
                        if cell_v is not None:
                            cell.fill = fill

            for row in ws.iter_rows(max_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')




        increase_progress(1)

    except Exception as e:
        error_message = traceback.format_exc() 
        print(error_message)
        tk.messagebox.showerror("Unknown Error in Peptide Difference Plot", error_message)
        if delete_faulty_sheets == True:
            for sheet_name in function_worksheet_names:
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
        
        
def r_chiclet():
    function_worksheet_names = []
    try:
        ws = wb.create_sheet("Chiclets")
        function_worksheet_names.append("Chiclets")
        plot_number = 0
        for state in states:
            asterisk_in_state = False
            if plot_number == 0:
                plot_start = 1
                plot_end = 50
            if plot_number > 0:
                plot_start = ws.max_column + 3
                plot_end = ws.max_column + 50
            ws.cell(row=1, column=plot_start, value=state)
            ws.cell(row=2, column=plot_start, value = "Sequence")
            ws.cell(row=2, column=plot_start+1, value = "Start")
            ws.cell(row=2, column=plot_start+2, value = "End")
            tpnum = 0
            for timepoint in s_timepoints[state]:
                if timepoint == 0:
                    tpnum = tpnum + 1
                    continue
                ws.cell(row=2, column=plot_start+3+tpnum, value = s_timepoints[state][tpnum])
                tpnum = tpnum + 1

            protein = state.split("~")[0]
            sorted_peptides = sorted(peplist[state], key=lambda x: (int(pro_peptide_starts.get((x[0], x[1]), [0])[0]), len(x[1])))

            pepnum = 0
            for peptide in sorted_peptides:
                startvalues = pro_peptide_starts.get((protein, peptide), None)
                startvalue= int(startvalues[0])
                endvalues = pro_peptide_ends.get((protein, peptide), None)
                endvalue = int(endvalues[0])

                ws.cell(row=3+pepnum, column=plot_start, value = peptide)
                ws.cell(row=3+pepnum, column=plot_start+1, value = startvalue)
                ws.cell(row=3+pepnum, column=plot_start+2, value = endvalue)
                tnum = 0
                for timepoint in s_timepoints[state]:
                    if timepoint == 0:
                        tnum = tnum + 1
                        continue
                    ws.cell(row=3+pepnum, column=plot_start+3+tnum, value=statedic_of_pepdic_cor[state][peptide][tnum][0])
                    tnum = tnum + 1
                try:
                    if noD_dic_states[state][peptide] == True:
                        ws.cell(row=3+pepnum, column=plot_start+3+tnum, value="*")
                        asterisk_in_state = True
                except:
                    pass
                pepnum = pepnum + 1

            if asterisk_in_state == True:
                ws.cell(row=7+pepnum, column=plot_start+3+tnum, value="* = no maxD for peptide")
                ws.cell(row=8+pepnum, column=plot_start+3+tnum, value="average back exchange used")
            ws.column_dimensions[get_column_letter(plot_start)].width = 30
            plot_number = plot_number + 1






        white_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type='solid')
        for row in ws.rows:
            for cell in row:
                cell.fill = white_fill
        for row in ws.iter_rows(max_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')

        for col in ws.iter_cols():
            for cell in col:
                if cell.value in ["Sequence", "Start", "End"]:
                    break
                else:
                    if cell.row >= 3:
                        cell_v = cell.value
                        if cell_v != "*" and cell_v != "* = no maxD for peptide" and cell_v != "average back exchange used":
                            if uptake_color_length >= 1 and cell_v is not None and cell_v >= uptake_val_1:
                                fill = PatternFill(start_color=uptake_col_1, end_color=uptake_col_1, fill_type='solid')
                                cell.number_format = ';;;'
                            elif uptake_color_length >= 2 and cell_v is not None and cell_v >= uptake_val_2:
                                fill = PatternFill(start_color=uptake_col_2, end_color=uptake_col_2, fill_type='solid')
                                cell.number_format = ';;;'
                            elif uptake_color_length >= 3 and cell_v is not None and cell_v >= uptake_val_3:
                                fill = PatternFill(start_color=uptake_col_3, end_color=uptake_col_3, fill_type='solid')
                                cell.number_format = ';;;'
                            elif uptake_color_length >= 4 and cell_v is not None and cell_v >= uptake_val_4:
                                fill = PatternFill(start_color=uptake_col_4, end_color=uptake_col_4, fill_type='solid')
                                cell.number_format = ';;;'
                            elif uptake_color_length >= 5 and cell_v is not None and cell_v >= uptake_val_5:
                                fill = PatternFill(start_color=uptake_col_5, end_color=uptake_col_5, fill_type='solid')
                                cell.number_format = ';;;'
                            elif uptake_color_length >= 6 and cell_v is not None and cell_v >= uptake_val_6:
                                fill = PatternFill(start_color=uptake_col_6, end_color=uptake_col_6, fill_type='solid')
                                cell.number_format = ';;;'
                            elif uptake_color_length >= 7 and cell_v is not None and cell_v >= uptake_val_7:
                                fill = PatternFill(start_color=uptake_col_7, end_color=uptake_col_7, fill_type='solid')
                                cell.number_format = ';;;'
                            elif uptake_color_length >= 8 and cell_v is not None and cell_v >= uptake_val_8:
                                fill = PatternFill(start_color=uptake_col_8, end_color=uptake_col_8, fill_type='solid')
                                cell.number_format = ';;;'
                            elif uptake_color_length >= 9 and cell_v is not None and cell_v >= uptake_val_9:
                                fill = PatternFill(start_color=uptake_col_9, end_color=uptake_col_9, fill_type='solid')
                                cell.number_format = ';;;'
                            elif cell_v is not None and cell_v > 0.0:
                                fill = PatternFill(start_color=uptake_gtz_key, end_color=uptake_gtz_key, fill_type='solid')
                                cell.number_format = ';;;'
                            elif cell_v == 0:
                                fill = PatternFill(start_color=uptake_eqz_key, end_color=uptake_eqz_key, fill_type='solid')
                                cell.number_format = ';;;'
                            elif cell_v == -99999:
                                fill = PatternFill(start_color=uptake_abs_key, end_color=uptake_abs_key, fill_type='solid')
                                cell.number_format = ';;;'
                            elif cell_v is not None and cell_v < 0.0:
                                fill = PatternFill(start_color=uptake_ltz_key, end_color=uptake_ltz_key, fill_type='solid')
                                cell.number_format = ';;;'
                            if cell_v is not None:
                                cell.fill = fill
        r_coincident_chiclet()
        increase_progress(0.33)
    except Exception as e:
        error_message = traceback.format_exc() 
        print(error_message)
        tk.messagebox.showerror("Unknown Error in Chiclet Plots", error_message)
        if delete_faulty_sheets == True:
            for sheet_name in function_worksheet_names:
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
    

def r_coincident_chiclet():
    function_worksheet_names = []
    try:
        ws = wb.create_sheet("Coincident Chiclets")
        function_worksheet_names.append("Coincident Chiclets")
        plot_number = 0
        for state in states:
            asterisk_in_state = False
            if plot_number == 0:
                plot_start = 1
                plot_end = 50
            if plot_number > 0:
                plot_start = ws.max_column + 3
                plot_end = ws.max_column + 50
            ws.cell(row=1, column=plot_start, value=state)
            ws.cell(row=2, column=plot_start, value = "Sequence")
            ws.cell(row=2, column=plot_start+1, value = "Start")
            ws.cell(row=2, column=plot_start+2, value = "End")
            tpnum = 0
            for timepoint in s_timepoints[state]:
                if timepoint == 0:
                    tpnum = tpnum + 1
                    continue
                ws.cell(row=2, column=plot_start+3+tpnum, value = s_timepoints[state][tpnum])
                tpnum = tpnum + 1

            protein = state.split("~")[0]
            sorted_peptides = sorted(coincident_protein_peptides[protein], key=lambda x: (int(pro_peptide_starts.get((x[0], x[1]), [0])[0]), len(x[1])))

            pepnum = 0
            for peptide in sorted_peptides:
                startvalues = pro_peptide_starts.get((protein, peptide), None)
                startvalue= int(startvalues[0]) 
                endvalues = pro_peptide_ends.get((protein, peptide), None)
                endvalue = int(endvalues[0])

                ws.cell(row=3+pepnum, column=plot_start, value = peptide)
                ws.cell(row=3+pepnum, column=plot_start+1, value = startvalue)
                ws.cell(row=3+pepnum, column=plot_start+2, value = endvalue)
                tnum = 0
                for timepoint in s_timepoints[state]:
                    if timepoint == 0:
                        tnum = tnum + 1
                        continue
                    ws.cell(row=3+pepnum, column=plot_start+3+tnum, value=statedic_of_pepdic_cor[state][peptide][tnum][0])
                    tnum = tnum + 1
                try:
                    if noD_dic_states[state][peptide] == True:
                        ws.cell(row=3+pepnum, column=plot_start+3+tnum, value="*")
                        asterisk_in_state = True
                except:
                    pass
                pepnum = pepnum + 1

            if asterisk_in_state == True:
                ws.cell(row=7+pepnum, column=plot_start+3+tnum, value="* = no maxD for peptide")
                ws.cell(row=8+pepnum, column=plot_start+3+tnum, value="average back exchange used")
            ws.column_dimensions[get_column_letter(plot_start)].width = 30
            plot_number = plot_number + 1






        white_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type='solid')
        for row in ws.rows:
            for cell in row:
                cell.fill = white_fill
        for row in ws.iter_rows(max_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')

        for col in ws.iter_cols():
            for cell in col:
                if cell.value in ["Sequence", "Start", "End"]:
                    break
                else:
                    if cell.row >= 3:
                        cell_v = cell.value
                        if cell_v != "*" and cell_v != "* = no maxD for peptide" and cell_v != "average back exchange used":
                            if uptake_color_length >= 1 and cell_v is not None and cell_v >= uptake_val_1:
                                fill = PatternFill(start_color=uptake_col_1, end_color=uptake_col_1, fill_type='solid')
                                cell.number_format = ';;;'
                            elif uptake_color_length >= 2 and cell_v is not None and cell_v >= uptake_val_2:
                                fill = PatternFill(start_color=uptake_col_2, end_color=uptake_col_2, fill_type='solid')
                                cell.number_format = ';;;'
                            elif uptake_color_length >= 3 and cell_v is not None and cell_v >= uptake_val_3:
                                fill = PatternFill(start_color=uptake_col_3, end_color=uptake_col_3, fill_type='solid')
                                cell.number_format = ';;;'
                            elif uptake_color_length >= 4 and cell_v is not None and cell_v >= uptake_val_4:
                                fill = PatternFill(start_color=uptake_col_4, end_color=uptake_col_4, fill_type='solid')
                                cell.number_format = ';;;'
                            elif uptake_color_length >= 5 and cell_v is not None and cell_v >= uptake_val_5:
                                fill = PatternFill(start_color=uptake_col_5, end_color=uptake_col_5, fill_type='solid')
                                cell.number_format = ';;;'
                            elif uptake_color_length >= 6 and cell_v is not None and cell_v >= uptake_val_6:
                                fill = PatternFill(start_color=uptake_col_6, end_color=uptake_col_6, fill_type='solid')
                                cell.number_format = ';;;'
                            elif uptake_color_length >= 7 and cell_v is not None and cell_v >= uptake_val_7:
                                fill = PatternFill(start_color=uptake_col_7, end_color=uptake_col_7, fill_type='solid')
                                cell.number_format = ';;;'
                            elif uptake_color_length >= 8 and cell_v is not None and cell_v >= uptake_val_8:
                                fill = PatternFill(start_color=uptake_col_8, end_color=uptake_col_8, fill_type='solid')
                                cell.number_format = ';;;'
                            elif uptake_color_length >= 9 and cell_v is not None and cell_v >= uptake_val_9:
                                fill = PatternFill(start_color=uptake_col_9, end_color=uptake_col_9, fill_type='solid')
                                cell.number_format = ';;;'
                            elif cell_v is not None and cell_v > 0.0:
                                fill = PatternFill(start_color=uptake_gtz_key, end_color=uptake_gtz_key, fill_type='solid')
                                cell.number_format = ';;;'
                            elif cell_v == 0:
                                fill = PatternFill(start_color=uptake_eqz_key, end_color=uptake_eqz_key, fill_type='solid')
                                cell.number_format = ';;;'
                            elif cell_v == -99999:
                                fill = PatternFill(start_color=uptake_abs_key, end_color=uptake_abs_key, fill_type='solid')
                                cell.number_format = ';;;'
                            elif cell_v is not None and cell_v < 0.0:
                                fill = PatternFill(start_color=uptake_ltz_key, end_color=uptake_ltz_key, fill_type='solid')
                                cell.number_format = ';;;'
                            if cell_v is not None:
                                cell.fill = fill
    except Exception as e:
        error_message = traceback.format_exc() 
        print(error_message)
        tk.messagebox.showerror("Unknown Error in Coincident Chiclets", error_message)
        if delete_faulty_sheets == True:
            for sheet_name in function_worksheet_names:
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]

def r_chicdif():
    function_worksheet_names = []
    try:
        ws = wb.create_sheet("Chiclet Differences")
        function_worksheet_names.append("Chiclet Differences")
        plot_number = 0
        for stt, pair in new_dic_of_dif_list.items():
            asterisk_in_state = False
            asterisk_in_start = False
            first = pair[0]
            second = pair[1]
            protein_one = first.split("~")[0]
            protein_two = second.split("~")[0]
            sorted_peptides_first = sorted(peplist[first], key=lambda x: (int(pro_peptide_starts.get((x[0], x[1]), [0])[0]), len(x[1])))
            sorted_peptides_second = sorted(peplist[second], key=lambda x: (int(pro_peptide_starts.get((x[0], x[1]), [0])[0]), len(x[1])))
            difname = f"{stt}"
            if plot_number == 0:
                plot_start = 1
                plot_end = 50
            if plot_number > 0:
                plot_start = ws.max_column + 3
                plot_end = ws.max_column + 50
            ws.cell(row=1, column=plot_start, value=difname)
            ws.cell(row=2, column=plot_start, value = "Sequence")
            ws.cell(row=2, column=plot_start+1, value = "Start")
            ws.cell(row=2, column=plot_start+2, value = "End")
            tpnum = 0
            for timepoint in s_timepoints[first]:
                if timepoint in s_timepoints[second]:
                    if timepoint == 0:
                        tpnum = tpnum + 1
                        continue
                    ws.cell(row=2, column=plot_start+3+tpnum, value = s_timepoints[first][tpnum])
                    tpnum = tpnum + 1



            pepset = set(peplist[first] + peplist[second])
            pepset_peptide_starts = {}
            for peptide in pepset:
                pepset_peptide_starts[peptide] = pro_peptide_starts.get((protein_one, peptide))
                if pepset_peptide_starts[peptide] == None:
                    pepset_peptide_starts[peptide] = pro_peptide_starts.get((protein_two, peptide))
            all_sorted_peptides = sorted(pepset, key=lambda p: (int(pepset_peptide_starts.get(p, [0])[0]), len(p)))



            pepnum = 0
            for peptide in all_sorted_peptides:
                if peptide in peplist[first] and peptide in peplist[second]:
                    startvalues = pro_peptide_starts.get((protein_one, peptide), None)
                    startvalue= int(startvalues[0])
                    endvalues = pro_peptide_ends.get((protein_one, peptide), None)
                    endvalue = int(endvalues[0])

                    startvalues2 = pro_peptide_starts.get((protein_two, peptide), None)
                    startvalue2 = int(startvalues2[0])





                    ws.cell(row=3+pepnum, column=plot_start, value = peptide)
                    if startvalue != startvalue2:
                        ws.cell(row=3+pepnum, column=plot_start+1, value = (str(startvalue)+"*"))
                        asterisk_in_start = True
                    else:
                        ws.cell(row=3+pepnum, column=plot_start+1, value = startvalue)
                    if startvalue != startvalue2:
                        ws.cell(row=3+pepnum, column=plot_start+2, value = (str(endvalue)+"*"))
                        asterisk_in_start = True
                    else:
                        ws.cell(row=3+pepnum, column=plot_start+2, value = endvalue)
                    tnum = 0
                    for timepoint in s_timepoints[first]:
                        if timepoint in s_timepoints[second]:
                            if timepoint == 0:
                                tnum = tnum + 1
                                continue 
                            if exp_bt_on_c and maxD_rfu_dif_on_c == True:
                                if statedic_of_pepdic_cor[first][peptide][tnum][0] != -99999 and statedic_of_pepdic_cor[second][peptide][tnum][0] != -99999:
                                    ws.cell(row=3+pepnum, column=plot_start+3+tnum, value=statedic_of_pepdic_cor[first][peptide][tnum][0] - statedic_of_pepdic_cor[second][peptide][tnum][0])
                                    tnum = tnum + 1
                                else:
                                    ws.cell(row=3+pepnum, column=plot_start+3+tnum, value = -99999)
                                    tnum = tnum + 1

                            if exp_bt_on_c and maxD_Da_dif_on_c == True:
                                if statedic_of_pepdic_cor[first][peptide][tnum][0] != -99999 and statedic_of_pepdic_cor[second][peptide][tnum][0] != -99999:
                                    max_theo = get_max_theo(peptide)
                                    ws.cell(row=3+pepnum, column=plot_start+3+tnum, value=max_theo*(statedic_of_pepdic_cor[first][peptide][tnum][0] - statedic_of_pepdic_cor[second][peptide][tnum][0]))
                                    tnum = tnum + 1
                                else:
                                    ws.cell(row=3+pepnum, column=plot_start+3+tnum, value = -99999)
                                    tnum = tnum + 1

                            if theo_bt_on_c and back_exchange == 0:
                                if statedic_of_pepdic_raw2[first][peptide][tnum][0] != -99999 and statedic_of_pepdic_raw2[second][peptide][tnum][0] != -99999:
                                    ws.cell(row=3+pepnum, column=plot_start+3+tnum, value=statedic_of_pepdic_raw2[first][peptide][tnum][0] - statedic_of_pepdic_raw2[second][peptide][tnum][0])
                                    tnum = tnum + 1
                                else:
                                    ws.cell(row=3+pepnum, column=plot_start+3+tnum, value = -99999)
                                    tnum = tnum + 1

                            if theo_bt_on_c and back_exchange != 0:
                                if statedic_of_pepdic_cor[first][peptide][tnum][0] != -99999 and statedic_of_pepdic_cor[second][peptide][tnum][0] != -99999:
                                    max_theo = get_max_theo(peptide)
                                    ws.cell(row=3+pepnum, column=plot_start+3+tnum, value=max_theo*(statedic_of_pepdic_cor[first][peptide][tnum][0] - statedic_of_pepdic_cor[second][peptide][tnum][0]))
                                    tnum = tnum + 1
                                else:
                                    ws.cell(row=3+pepnum, column=plot_start+3+tnum, value = -99999)
                                    tnum = tnum + 1

                    ch1 = False
                    ch2 = False
                    try:
                        if noD_dic_states[first][peptide] == True:
                            ch1 = True
                    except:
                        pass
                    try:
                        if noD_dic_states[second][peptide] == True:
                            ch2 = True
                    except:
                        pass
                    if ch1 == True or ch2 == True:
                        ws.cell(row=3+pepnum, column=plot_start+3+tnum, value="*")
                        asterisk_in_state = True

                    pepnum = pepnum + 1

                elif white_var.get() == 1:
                    startvalues = pro_peptide_starts.get((protein_one, peptide), None)
                    if startvalues == None:
                        startvalues = pro_peptide_starts.get((protein_two, peptide), None)
                    startvalue= int(startvalues[0])
                    endvalues = pro_peptide_ends.get((protein_one, peptide), None)
                    if endvalues == None:
                        endvalues = pro_peptide_ends.get((protein_two, peptide), None)
                    endvalue = int(endvalues[0])

                    ws.cell(row=3+pepnum, column=plot_start, value = peptide)
                    ws.cell(row=3+pepnum, column=plot_start+1, value = startvalue)
                    ws.cell(row=3+pepnum, column=plot_start+2, value = endvalue)
                    tnum = 0
                    for timepoint in s_timepoints[first]:
                        if timepoint in s_timepoints[second]:
                            if timepoint == 0:
                                tnum = tnum + 1
                                continue
                            ws.cell(row=3+pepnum, column=plot_start+3+tnum, value = None)
                            tnum = tnum + 1

                    pepnum += 1


            if asterisk_in_state == True:
                ws.cell(row=7+pepnum, column=plot_start+3+tnum, value="* = no maxD for peptide")
                ws.cell(row=8+pepnum, column=plot_start+3+tnum, value="average back exchange used")

            if asterisk_in_start == True:
                ws.cell(row=5+pepnum, column=plot_start+1, value="* = peptide sequence found at different positions in states")


            ws.column_dimensions[get_column_letter(plot_start)].width = 30
            plot_number = plot_number + 1




        white_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type='solid')
        for row in ws.rows:
            for cell in row:
                if cell.value != "*":
                    cell.fill = white_fill
        for row in ws.iter_rows(max_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')

        for col in ws.iter_cols():
            for cell in col:
                if cell.value in ["Sequence", "Start", "End"]:
                    break
                else:
                    if cell.row >= 3:
                        cell_v = cell.value
                        if cell_v != "*" and cell_v != "* = no maxD for peptide" and cell_v != "average back exchange used":
                            if cell_v == -99999:
                                fill = PatternFill(start_color=b_col_abs, end_color=b_col_abs, fill_type='solid')
                                cell.number_format = ';;;'
                            elif d_col_length >= 1 and cell_v is not None and cell_v >= d_val_1:
                                fill = PatternFill(start_color=d_col_1, end_color=d_col_1, fill_type='solid')
                                cell.number_format = ';;;'
                            elif d_col_length >= 2 and cell_v is not None and cell_v >= d_val_2:
                                fill = PatternFill(start_color=d_col_2, end_color=d_col_2, fill_type='solid')
                                cell.number_format = ';;;'
                            elif d_col_length >= 3 and cell_v is not None and cell_v >= d_val_3:
                                fill = PatternFill(start_color=d_col_3, end_color=d_col_3, fill_type='solid')
                                cell.number_format = ';;;'
                            elif d_col_length >= 4 and cell_v is not None and cell_v >= d_val_4:
                                fill = PatternFill(start_color=d_col_4, end_color=d_col_4, fill_type='solid')
                                cell.number_format = ';;;'
                            elif d_col_length >= 5 and cell_v is not None and cell_v >= d_val_5:
                                fill = PatternFill(start_color=d_col_5, end_color=d_col_5, fill_type='solid')
                                cell.number_format = ';;;'
                            elif cell_v is not None and cell_v > 0:
                                fill = PatternFill(start_color=d_col_gtz, end_color=d_col_gtz, fill_type='solid')
                                cell.number_format = ';;;'
                            elif p_col_length >= 1 and cell_v is not None and cell_v <= (-1) * p_val_1:
                                fill = PatternFill(start_color=p_col_1, end_color=p_col_1, fill_type='solid')
                                cell.number_format = ';;;'
                            elif p_col_length >= 2 and cell_v is not None and cell_v <= (-1) * p_val_2:
                                fill = PatternFill(start_color=p_col_2, end_color=p_col_2, fill_type='solid')
                                cell.number_format = ';;;'
                            elif p_col_length >= 3 and cell_v is not None and cell_v <= (-1) * p_val_3:
                                fill = PatternFill(start_color=p_col_3, end_color=p_col_3, fill_type='solid')
                                cell.number_format = ';;;'
                            elif p_col_length >= 4 and cell_v is not None and cell_v <= (-1) * p_val_4:
                                fill = PatternFill(start_color=p_col_4, end_color=p_col_4, fill_type='solid')
                                cell.number_format = ';;;'
                            elif p_col_length >= 5 and cell_v is not None and cell_v <= (-1) * p_val_5:
                                fill = PatternFill(start_color=p_col_5, end_color=p_col_5, fill_type='solid')
                                cell.number_format = ';;;'
                            elif cell_v is not None and cell_v < 0:
                                fill = PatternFill(start_color=p_col_gtz, end_color=p_col_gtz, fill_type='solid')
                                cell.number_format = ';;;'
                            elif cell_v is not None and cell_v == 0:
                                fill = PatternFill(start_color=b_col_eqz, end_color=b_col_eqz, fill_type='solid')
                                cell.number_format = ';;;'
                            if cell_v is not None:
                                cell.fill = fill


        for row in ws.iter_rows():
            for cell in row:
                if cell.value == -99999:
                    cell.value = None

        r_coincident_chicdif()
        increase_progress(0.33)
        
    except Exception as e:
        error_message = traceback.format_exc() 
        print(error_message)
        tk.messagebox.showerror("Unknown Error in Chiclet Differences", error_message) 
        if delete_faulty_sheets == True:
            for sheet_name in function_worksheet_names:
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
    
def r_coincident_chicdif():
    function_worksheet_names = []
    try:
        ws = wb.create_sheet("Coincident Chic Dif")
        function_worksheet_names.append("Coincident Chic Dif")
        plot_number = 0
        for stt, pair in new_dic_of_dif_list.items():
            asterisk_in_state = False
            asterisk_in_start = False
            first = pair[0]
            second = pair[1]
            protein_one = first.split("~")[0]
            protein_two = second.split("~")[0]
            sorted_peptides_first = sorted(coincident_protein_peptides[protein_one], key=lambda x: (int(pro_peptide_starts.get((x[0], x[1]), [0])[0]), len(x[1])))
            sorted_peptides_second = sorted(coincident_protein_peptides[protein_two], key=lambda x: (int(pro_peptide_starts.get((x[0], x[1]), [0])[0]), len(x[1])))
            difname = f"{stt}"
            if plot_number == 0:
                plot_start = 1
                plot_end = 50
            if plot_number > 0:
                plot_start = ws.max_column + 3
                plot_end = ws.max_column + 50
            ws.cell(row=1, column=plot_start, value=difname)
            ws.cell(row=2, column=plot_start, value = "Sequence")
            ws.cell(row=2, column=plot_start+1, value = "Start")
            ws.cell(row=2, column=plot_start+2, value = "End")
            tpnum = 0
            for timepoint in s_timepoints[first]:
                if timepoint in s_timepoints[second]:
                    if timepoint == 0:
                        tpnum = tpnum + 1
                        continue
                    ws.cell(row=2, column=plot_start+3+tpnum, value = s_timepoints[first][tpnum])
                    tpnum = tpnum + 1



            pepset = set(coincident_protein_peptides[protein_one] + coincident_protein_peptides[protein_two])
            pepset_peptide_starts = {}
            for peptide in pepset:
                pepset_peptide_starts[peptide] = pro_peptide_starts.get((protein_one, peptide))
                if pepset_peptide_starts[peptide] == None:
                    pepset_peptide_starts[peptide] = pro_peptide_starts.get((protein_two, peptide))
            all_sorted_peptides = sorted(pepset, key=lambda p: (int(pepset_peptide_starts.get(p, [0])[0]), len(p)))



            pepnum = 0
            for peptide in all_sorted_peptides:
                if peptide in peplist[first] and peptide in peplist[second]:
                    startvalues = pro_peptide_starts.get((protein_one, peptide), None)
                    startvalue= int(startvalues[0])
                    endvalues = pro_peptide_ends.get((protein_one, peptide), None)
                    endvalue = int(endvalues[0])

                    startvalues2 = pro_peptide_starts.get((protein_two, peptide), None)
                    startvalue2 = int(startvalues2[0])





                    ws.cell(row=3+pepnum, column=plot_start, value = peptide)
                    if startvalue != startvalue2:
                        ws.cell(row=3+pepnum, column=plot_start+1, value = (str(startvalue)+"*"))
                        asterisk_in_start = True
                    else:
                        ws.cell(row=3+pepnum, column=plot_start+1, value = startvalue)
                    if startvalue != startvalue2:
                        ws.cell(row=3+pepnum, column=plot_start+2, value = (str(endvalue)+"*"))
                        asterisk_in_start = True
                    else:
                        ws.cell(row=3+pepnum, column=plot_start+2, value = endvalue)
                    tnum = 0
                    for timepoint in s_timepoints[first]:
                        if timepoint in s_timepoints[second]:
                            if timepoint == 0:
                                tnum = tnum + 1
                                continue 
                            if exp_bt_on_c and maxD_rfu_dif_on_c == True:
                                if statedic_of_pepdic_cor[first][peptide][tnum][0] != -99999 and statedic_of_pepdic_cor[second][peptide][tnum][0] != -99999:
                                    ws.cell(row=3+pepnum, column=plot_start+3+tnum, value=statedic_of_pepdic_cor[first][peptide][tnum][0] - statedic_of_pepdic_cor[second][peptide][tnum][0])
                                    tnum = tnum + 1
                                else:
                                    ws.cell(row=3+pepnum, column=plot_start+3+tnum, value = -99999)
                                    tnum = tnum + 1

                            if exp_bt_on_c and maxD_Da_dif_on_c == True:
                                if statedic_of_pepdic_cor[first][peptide][tnum][0] != -99999 and statedic_of_pepdic_cor[second][peptide][tnum][0] != -99999:
                                    max_theo = get_max_theo(peptide)
                                    ws.cell(row=3+pepnum, column=plot_start+3+tnum, value=max_theo*(statedic_of_pepdic_cor[first][peptide][tnum][0] - statedic_of_pepdic_cor[second][peptide][tnum][0]))
                                    tnum = tnum + 1
                                else:
                                    ws.cell(row=3+pepnum, column=plot_start+3+tnum, value = -99999)
                                    tnum = tnum + 1

                            if theo_bt_on_c and back_exchange == 0:
                                if statedic_of_pepdic_raw2[first][peptide][tnum][0] != -99999 and statedic_of_pepdic_raw2[second][peptide][tnum][0] != -99999:
                                    ws.cell(row=3+pepnum, column=plot_start+3+tnum, value=statedic_of_pepdic_raw2[first][peptide][tnum][0] - statedic_of_pepdic_raw2[second][peptide][tnum][0])
                                    tnum = tnum + 1
                                else:
                                    ws.cell(row=3+pepnum, column=plot_start+3+tnum, value = -99999)
                                    tnum = tnum + 1

                            if theo_bt_on_c and back_exchange != 0:
                                if statedic_of_pepdic_cor[first][peptide][tnum][0] != -99999 and statedic_of_pepdic_cor[second][peptide][tnum][0] != -99999:
                                    max_theo = get_max_theo(peptide)
                                    ws.cell(row=3+pepnum, column=plot_start+3+tnum, value=max_theo*(statedic_of_pepdic_cor[first][peptide][tnum][0] - statedic_of_pepdic_cor[second][peptide][tnum][0]))
                                    tnum = tnum + 1
                                else:
                                    ws.cell(row=3+pepnum, column=plot_start+3+tnum, value = -99999)
                                    tnum = tnum + 1

                    ch1 = False
                    ch2 = False
                    try:
                        if noD_dic_states[first][peptide] == True:
                            ch1 = True
                    except:
                        pass
                    try:
                        if noD_dic_states[second][peptide] == True:
                            ch2 = True
                    except:
                        pass
                    if ch1 == True or ch2 == True:
                        ws.cell(row=3+pepnum, column=plot_start+3+tnum, value="*")
                        asterisk_in_state = True

                    pepnum = pepnum + 1

                elif white_var.get() == 1:
                    startvalues = pro_peptide_starts.get((protein_one, peptide), None)
                    if startvalues == None:
                        startvalues = pro_peptide_starts.get((protein_two, peptide), None)
                    startvalue= int(startvalues[0])
                    endvalues = pro_peptide_ends.get((protein_one, peptide), None)
                    if endvalues == None:
                        endvalues = pro_peptide_ends.get((protein_two, peptide), None)
                    endvalue = int(endvalues[0])

                    ws.cell(row=3+pepnum, column=plot_start, value = peptide)
                    ws.cell(row=3+pepnum, column=plot_start+1, value = startvalue)
                    ws.cell(row=3+pepnum, column=plot_start+2, value = endvalue)
                    tnum = 0
                    for timepoint in s_timepoints[first]:
                        if timepoint in s_timepoints[second]:
                            if timepoint == 0:
                                tnum = tnum + 1
                                continue
                            ws.cell(row=3+pepnum, column=plot_start+3+tnum, value = None)
                            tnum = tnum + 1

                    pepnum += 1


            if asterisk_in_state == True:
                ws.cell(row=7+pepnum, column=plot_start+3+tnum, value="* = no maxD for peptide")
                ws.cell(row=8+pepnum, column=plot_start+3+tnum, value="average back exchange used")

            if asterisk_in_start == True:
                ws.cell(row=5+pepnum, column=plot_start+1, value="* = peptide sequence found at different positions in states")


            ws.column_dimensions[get_column_letter(plot_start)].width = 30
            plot_number = plot_number + 1




        white_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type='solid')
        for row in ws.rows:
            for cell in row:
                if cell.value != "*":
                    cell.fill = white_fill
        for row in ws.iter_rows(max_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')

        for col in ws.iter_cols():
            for cell in col:
                if cell.value in ["Sequence", "Start", "End"]:
                    break
                else:
                    if cell.row >= 3:
                        cell_v = cell.value
                        if cell_v != "*" and cell_v != "* = no maxD for peptide" and cell_v != "average back exchange used":
                            if cell_v == -99999:
                                fill = PatternFill(start_color=b_col_abs, end_color=b_col_abs, fill_type='solid')
                                cell.number_format = ';;;'
                            elif d_col_length >= 1 and cell_v is not None and cell_v >= d_val_1:
                                fill = PatternFill(start_color=d_col_1, end_color=d_col_1, fill_type='solid')
                                cell.number_format = ';;;'
                            elif d_col_length >= 2 and cell_v is not None and cell_v >= d_val_2:
                                fill = PatternFill(start_color=d_col_2, end_color=d_col_2, fill_type='solid')
                                cell.number_format = ';;;'
                            elif d_col_length >= 3 and cell_v is not None and cell_v >= d_val_3:
                                fill = PatternFill(start_color=d_col_3, end_color=d_col_3, fill_type='solid')
                                cell.number_format = ';;;'
                            elif d_col_length >= 4 and cell_v is not None and cell_v >= d_val_4:
                                fill = PatternFill(start_color=d_col_4, end_color=d_col_4, fill_type='solid')
                                cell.number_format = ';;;'
                            elif d_col_length >= 5 and cell_v is not None and cell_v >= d_val_5:
                                fill = PatternFill(start_color=d_col_5, end_color=d_col_5, fill_type='solid')
                                cell.number_format = ';;;'
                            elif cell_v is not None and cell_v > 0:
                                fill = PatternFill(start_color=d_col_gtz, end_color=d_col_gtz, fill_type='solid')
                                cell.number_format = ';;;'
                            elif p_col_length >= 1 and cell_v is not None and cell_v <= (-1) * p_val_1:
                                fill = PatternFill(start_color=p_col_1, end_color=p_col_1, fill_type='solid')
                                cell.number_format = ';;;'
                            elif p_col_length >= 2 and cell_v is not None and cell_v <= (-1) * p_val_2:
                                fill = PatternFill(start_color=p_col_2, end_color=p_col_2, fill_type='solid')
                                cell.number_format = ';;;'
                            elif p_col_length >= 3 and cell_v is not None and cell_v <= (-1) * p_val_3:
                                fill = PatternFill(start_color=p_col_3, end_color=p_col_3, fill_type='solid')
                                cell.number_format = ';;;'
                            elif p_col_length >= 4 and cell_v is not None and cell_v <= (-1) * p_val_4:
                                fill = PatternFill(start_color=p_col_4, end_color=p_col_4, fill_type='solid')
                                cell.number_format = ';;;'
                            elif p_col_length >= 5 and cell_v is not None and cell_v <= (-1) * p_val_5:
                                fill = PatternFill(start_color=p_col_5, end_color=p_col_5, fill_type='solid')
                                cell.number_format = ';;;'
                            elif cell_v is not None and cell_v < 0:
                                fill = PatternFill(start_color=p_col_gtz, end_color=p_col_gtz, fill_type='solid')
                                cell.number_format = ';;;'
                            elif cell_v is not None and cell_v == 0:
                                fill = PatternFill(start_color=b_col_eqz, end_color=b_col_eqz, fill_type='solid')
                                cell.number_format = ';;;'
                            if cell_v is not None:
                                cell.fill = fill


        for row in ws.iter_rows():
            for cell in row:
                if cell.value == -99999:
                    cell.value = None
    
    except Exception as e:
        error_message = traceback.format_exc() 
        print(error_message)
        tk.messagebox.showerror("Unknown Error in Coincident Chiclet Differences", error_message)
        if delete_faulty_sheets == True:
            for sheet_name in function_worksheet_names:
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]

def r_condpeps():
    function_worksheet_names = []
    try:
        if reduce_states_var.get() == 0:
            states_to_look_in = statedic_of_pepdic_cor
        if reduce_states_var.get() == 1:
            states_to_look_in = order_state_dic.values()
            states_to_look_in = [x for x in states_to_look_in if x != False]

        whitefont = Font(color="FFFFFFFF")
        for state in states_to_look_in:
            asterisk_in_state = False
            highest_row = 1
            protein = state.split("~")[0]
            sorted_peptides = sorted(peplist[state], key=lambda x: (int(pro_peptide_starts.get((x[0], x[1]), [0])[0]), len(x[1])))
            ws_title = (f"{state}" + "_cond")[-30:]
            ws = wb.create_sheet(title=ws_title)
            function_worksheet_names.append(ws_title)
            cell_reference_list = list()
            ws.append(["Timepoint"])
            ws.append([" "] + seqlist_dic[state] + [" "])
            timepoint_number = 0
            for timepoint in s_timepoints[state]:
                if timepoint_number == 0:
                    startrow = 3
                    endrow = 250
                if timepoint_number != 0:
                    startrow = ws.max_row +5
                    endrow = ws.max_row + 250
                for peptide in sorted_peptides:
                    startvalues = pro_peptide_starts.get((protein, peptide), None)
                    startvalue= int(startvalues[0]) - seq_start[state]
                    endvalues = pro_peptide_ends.get((protein, peptide), None)
                    endvalue = int(endvalues[0]) - seq_start[state]
                    peptide_length = len(peptide)
                    Cuptake = None
                    Cuptake_SD = None

                    try:
                        for up, tp in statedic_of_pepdic_cor[state][peptide]:
                            if tp == timepoint:
                                Cuptake = up
                    except:
                        pass
                    try:
                        for sd, tp in statedic_of_sddic_cor[state][peptide]:
                            if tp == timepoint:
                                Cuptake_SD = sd
                    except:
                        Cuptake_SD = -99999
                        
                    if Cuptake_SD != -99999:
                        Cuptake_SD = Cuptake_SD / np.sqrt(float(SE_enter.get()))
                        

                    if Cuptake is not None:
                        for row in ws.iter_rows(min_row=startrow, max_row=startrow):
                            row[0].value = timepoint
                        for i, row in enumerate(ws.iter_rows(min_row=startrow,max_row=endrow), start=startrow):
                            if i > highest_row:
                                highest_row = i
                            cells = row[startvalue + 1:endvalue + 2]
                            if all(cell.value is None for cell in cells):
                                row[startvalue + 1].value = Cuptake
                                row[startvalue + 1].border = Border(top=Side(border_style='thin', color='FF000000'),
                                         bottom=Side(border_style='thin', color='FF000000'),
                                         left=Side(border_style='thin', color='FF000000'))
                                row[endvalue+1].value = Cuptake
                                row[endvalue+1].border = Border(top=Side(border_style='thin', color='FF000000'),
                                         bottom=Side(border_style='thin', color='FF000000'),
                                         right=Side(border_style='thin', color='FF000000'))
                                for cell in row[startvalue + 2:endvalue+1]:
                                    cell.value = Cuptake
                                    cell.border = Border(top=Side(border_style='thin', color='FF000000'),
                                         bottom=Side(border_style='thin', color='FF000000'))
                                middle = int((startvalue + 1 + endvalue + 1)/2)

                                if sd_checkvar.get() == 0:
                                    if len(peptide) >= 5:
                                        row[middle-1].value = round(Cuptake * 100, 1)
                                    else:
                                        row[middle].value = round(Cuptake * 100, 1)
                                else:
                                    if len(peptide) >= 5:
                                        if Cuptake_SD != -99999 and Cuptake_SD != -99999 and Cuptake_SD != 0 and Cuptake_SD != "-99999" and Cuptake_SD != "0":
                                            row[middle-1].value = str(round(Cuptake * 100)) + " " + "\u00B1" + str(round(Cuptake_SD * 100))
                                        else:
                                            row[middle-1].value = round(Cuptake * 100)
                                    else:
                                        if Cuptake_SD != -99999 and Cuptake_SD != -99999 and Cuptake_SD != 0 and Cuptake_SD != "-99999" and Cuptake_SD != "0":
                                            row[middle].value = str(round(Cuptake * 100)) + " " + "\u00B1" + str(round(Cuptake_SD * 100))
                                        else:
                                            row[middle].value = round(Cuptake * 100)



                                if Cuptake*100 >= 100:
                                    b = 14
                                else:
                                    b = 16


                                if len(peptide) >= 5:
                                    c = 1
                                else:
                                    c = 0
                                    if Cuptake*100 >= 100:
                                        b = 8
                                    else:
                                        b = 10

                                if uptake_color_length >= 1 and Cuptake > uptake_val_1:
                                    fill = PatternFill(start_color=uptake_col_1, end_color=uptake_col_1, fill_type='solid')
                                    font = Font(color=uptake_text_1, size=b)
                                elif uptake_color_length >= 2 and Cuptake > uptake_val_2:
                                    fill = PatternFill(start_color=uptake_col_2, end_color=uptake_col_2, fill_type='solid')
                                    font = Font(color=uptake_text_2, size=b)
                                elif uptake_color_length >= 3 and Cuptake > uptake_val_3:
                                    fill = PatternFill(start_color=uptake_col_3, end_color=uptake_col_3, fill_type='solid')
                                    font = Font(color=uptake_text_3, size=b)
                                elif uptake_color_length >= 4 and Cuptake > uptake_val_4:
                                    fill = PatternFill(start_color=uptake_col_4, end_color=uptake_col_4, fill_type='solid')
                                    font = Font(color=uptake_text_4, size=b)
                                elif uptake_color_length >= 5 and Cuptake > uptake_val_5:
                                    fill = PatternFill(start_color=uptake_col_5, end_color=uptake_col_5, fill_type='solid')
                                    font = Font(color=uptake_text_5, size=b)
                                elif uptake_color_length >= 6 and Cuptake > uptake_val_6:
                                    fill = PatternFill(start_color=uptake_col_6, end_color=uptake_col_6, fill_type='solid')
                                    font = Font(color=uptake_text_6, size=b)
                                elif uptake_color_length >= 7 and Cuptake > uptake_val_7:
                                    fill = PatternFill(start_color=uptake_col_7, end_color=uptake_col_7, fill_type='solid')
                                    font = Font(color=uptake_text_7, size=b)
                                elif uptake_color_length >= 8 and Cuptake > uptake_val_8:
                                    fill = PatternFill(start_color=uptake_col_8, end_color=uptake_col_8, fill_type='solid')
                                    font = Font(color=uptake_text_8, size=b)
                                elif uptake_color_length >= 9 and Cuptake > uptake_val_9:
                                    fill = PatternFill(start_color=uptake_col_9, end_color=uptake_col_9, fill_type='solid')
                                    font = Font(color=uptake_text_9, size=b)
                                elif Cuptake > 0.0:
                                    fill = PatternFill(start_color=uptake_gtz_key, end_color=uptake_gtz_key, fill_type='solid')
                                    font = Font(color=uptake_gtz_text, size=b)
                                elif Cuptake == 0:
                                    fill = PatternFill(start_color=uptake_eqz_key, end_color=uptake_eqz_key, fill_type='solid')
                                    row[middle-c].number_format = ';;;'
                                    font = Font(color=uptake_eqz_text, size=b)
                                elif Cuptake == -99999:
                                    fill = PatternFill(start_color=uptake_abs_key, end_color=uptake_abs_key, fill_type='solid')
                                    row[middle-c].number_format = ';;;'
                                    font = Font(color=uptake_abs_text, size=b)
                                elif Cuptake < 0.0:
                                    fill = PatternFill(start_color=uptake_ltz_key, end_color=uptake_ltz_key, fill_type='solid')
                                    font = Font(color=uptake_ltz_text, size=b)

                                row[middle-c].fill = fill
                                row[middle-c].font = font


                                if sd_checkvar.get() == 0:
                                    if len(peptide) >= 5:
                                        ws.merge_cells(start_row=row[middle-1].row, start_column=row[middle-1].column, end_row=row[middle+1].row, end_column=row[middle+1].column)
                                        middle_cell_reference = row[middle-1].coordinate
                                        cell_reference_list.append(middle_cell_reference)
                                        if row[middle-1].number_format != ';;;':
                                            row[middle-1].number_format = "0.0"
                                    else:
                                        ws.merge_cells(start_row=row[middle].row, start_column=row[middle].column, end_row=row[middle+1].row, end_column=row[middle+1].column)
                                        middle_cell_reference = row[middle].coordinate
                                        cell_reference_list.append(middle_cell_reference)
                                        if row[middle].number_format != ';;;':
                                            row[middle].number_format = "0.0"

                                else:   
                                    if len(peptide) >= 6:
                                        ws.merge_cells(start_row=row[middle-1].row, start_column=row[middle-1].column, end_row=row[middle+2].row, end_column=row[middle+2].column)
                                        middle_cell_reference = row[middle-1].coordinate
                                        cell_reference_list.append(middle_cell_reference)

                                    elif len(peptide) >= 5:
                                        ws.merge_cells(start_row=row[middle-1].row, start_column=row[middle-1].column, end_row=row[middle+1].row, end_column=row[middle+1].column)
                                        middle_cell_reference = row[middle-1].coordinate
                                        cell_reference_list.append(middle_cell_reference)

                                    elif len(peptide) <= 4:
                                        ws.merge_cells(start_row=row[middle].row, start_column=row[middle].column, end_row=row[middle+1].row, end_column=row[middle+1].column)
                                        middle_cell_reference = row[middle].coordinate
                                        cell_reference_list.append(middle_cell_reference)


                                row[middle-c].alignment = Alignment(horizontal='center', vertical='center')

                                try:
                                    if peptide in noD_dic_states[state]:
                                        if Cuptake != 0 and Cuptake != -99999:
                                            row[startvalue+1].value = "*"
                                            asterisk_in_state = True
                                            #row[startvalue+1].alignment = Alignment(textRotation=90, vertical='center')
                                            if uptake_color_length >= 1 and Cuptake > uptake_val_1:
                                                fill = PatternFill(start_color=uptake_col_1, end_color=uptake_col_1, fill_type='solid')
                                                font = Font(color=uptake_text_1, size=12)
                                            elif uptake_color_length >= 2 and Cuptake > uptake_val_2:
                                                fill = PatternFill(start_color=uptake_col_2, end_color=uptake_col_2, fill_type='solid')
                                                font = Font(color=uptake_text_2, size=12)
                                            elif uptake_color_length >= 3 and Cuptake > uptake_val_3:
                                                fill = PatternFill(start_color=uptake_col_3, end_color=uptake_col_3, fill_type='solid')
                                                font = Font(color=uptake_text_3, size=12)
                                            elif uptake_color_length >= 4 and Cuptake > uptake_val_4:
                                                fill = PatternFill(start_color=uptake_col_4, end_color=uptake_col_4, fill_type='solid')
                                                font = Font(color=uptake_text_4, size=12)
                                            elif uptake_color_length >= 5 and Cuptake > uptake_val_5:
                                                fill = PatternFill(start_color=uptake_col_5, end_color=uptake_col_5, fill_type='solid')
                                                font = Font(color=uptake_text_5, size=12)
                                            elif uptake_color_length >= 6 and Cuptake > uptake_val_6:
                                                fill = PatternFill(start_color=uptake_col_6, end_color=uptake_col_6, fill_type='solid')
                                                font = Font(color=uptake_text_6, size=12)
                                            elif uptake_color_length >= 7 and Cuptake > uptake_val_7:
                                                fill = PatternFill(start_color=uptake_col_7, end_color=uptake_col_7, fill_type='solid')
                                                font = Font(color=uptake_text_7, size=12)
                                            elif uptake_color_length >= 8 and Cuptake > uptake_val_8:
                                                fill = PatternFill(start_color=uptake_col_8, end_color=uptake_col_8, fill_type='solid')
                                                font = Font(color=uptake_text_8, size=12)
                                            elif uptake_color_length >= 9 and Cuptake > uptake_val_9:
                                                fill = PatternFill(start_color=uptake_col_9, end_color=uptake_col_9, fill_type='solid')
                                                font = Font(color=uptake_text_9, size=12)
                                            elif Cuptake > 0.0:
                                                fill = PatternFill(start_color=uptake_gtz_key, end_color=uptake_gtz_key, fill_type='solid')
                                                font = Font(color=uptake_gtz_text, size=12)
                                            elif Cuptake == 0:
                                                fill = PatternFill(start_color=uptake_eqz_key, end_color=uptake_eqz_key, fill_type='solid')
                                                font = Font(color=uptake_eqz_text, size=12)
                                                cell.number_format = ';;;'
                                            elif Cuptake == -99999:
                                                fill = PatternFill(start_color=uptake_abs_key, end_color=uptake_abs_key, fill_type='solid')
                                                font = Font(color=uptake_abs_text, size=12)
                                                cell.number_format = ';;;'
                                            elif Cuptake < 0.0:
                                                fill = PatternFill(start_color=uptake_ltz_key, end_color=uptake_ltz_key, fill_type='solid')
                                                font = Font(color=uptake_ltz_text, size=12)
                                            row[startvalue+1].fill = fill
                                            row[startvalue+1].font = font
                                except:
                                    pass



                                break
                            else:
                                continue



                timepoint_number = timepoint_number + 1



            for row in ws.iter_rows(min_row=2, max_row=2):
                for cell in row:
                    cell.font = courier_new_style

            if asterisk_in_state == True:
                for row in ws.iter_rows(min_row=highest_row+2, max_row=highest_row+2):
                    row[1].value = "* = no maxD for peptide, average back exchange used"


            white_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type='solid')
            for row in ws.rows:
                for cell in row:
                    if cell.value == None:
                        cell.fill = white_fill
            for row in ws.iter_rows(min_row=2, max_row=2):
                for cell in row:
                    cell.fill = white_fill




            for i, column in enumerate(ws.columns):
                if i == 0:
                    continue
    #            if con_pep_width_enter.get() == "2.5":
                ws.column_dimensions[column[0].column_letter].width = "2.504"
    #            else:
    #                ws.column_dimensions[column[0].column_letter].width = con_pep_width_enter.get()
    #




            for row in ws.iter_rows(min_row=3):
                for cell in row[1:]:
                    cell_v = cell.value
                    if cell.coordinate not in cell_reference_list and cell_v != "*" and cell_v != "* = no maxD for peptide, average back exchange used":
                        if uptake_color_length >= 1 and cell_v is not None and cell_v > uptake_val_1:
                            fill = PatternFill(start_color=uptake_col_1, end_color=uptake_col_1, fill_type='solid')
                            cell.number_format = ';;;'
                        elif uptake_color_length >= 2 and cell_v is not None and cell_v > uptake_val_2:
                            fill = PatternFill(start_color=uptake_col_2, end_color=uptake_col_2, fill_type='solid')
                            cell.number_format = ';;;'
                        elif uptake_color_length >= 3 and cell_v is not None and cell_v > uptake_val_3:
                            fill = PatternFill(start_color=uptake_col_3, end_color=uptake_col_3, fill_type='solid')
                            cell.number_format = ';;;'
                        elif uptake_color_length >= 4 and cell_v is not None and cell_v > uptake_val_4:
                            fill = PatternFill(start_color=uptake_col_4, end_color=uptake_col_4, fill_type='solid')
                            cell.number_format = ';;;'
                        elif uptake_color_length >= 5 and cell_v is not None and cell_v > uptake_val_5:
                            fill = PatternFill(start_color=uptake_col_5, end_color=uptake_col_5, fill_type='solid')
                            cell.number_format = ';;;'
                        elif uptake_color_length >= 6 and cell_v is not None and cell_v > uptake_val_6:
                            fill = PatternFill(start_color=uptake_col_6, end_color=uptake_col_6, fill_type='solid')
                            cell.number_format = ';;;'
                        elif uptake_color_length >= 7 and cell_v is not None and cell_v > uptake_val_7:
                            fill = PatternFill(start_color=uptake_col_7, end_color=uptake_col_7, fill_type='solid')
                            cell.number_format = ';;;'
                        elif uptake_color_length >= 8 and cell_v is not None and cell_v > uptake_val_8:
                            fill = PatternFill(start_color=uptake_col_8, end_color=uptake_col_8, fill_type='solid')
                            cell.number_format = ';;;'
                        elif uptake_color_length >= 9 and cell_v is not None and cell_v > uptake_val_9:
                            fill = PatternFill(start_color=uptake_col_9, end_color=uptake_col_9, fill_type='solid')
                            cell.number_format = ';;;'
                        elif cell_v is not None and cell_v > 0.0:
                            fill = PatternFill(start_color=uptake_gtz_key, end_color=uptake_gtz_key, fill_type='solid')
                            cell.number_format = ';;;'
                        elif cell_v == 0:
                            fill = PatternFill(start_color=uptake_eqz_key, end_color=uptake_eqz_key, fill_type='solid')
                            cell.number_format = ';;;'
                        elif cell_v == -99999:
                            fill = PatternFill(start_color=uptake_abs_key, end_color=uptake_abs_key, fill_type='solid')
                            cell.number_format = ';;;'
                        elif cell_v is not None and cell_v < 0.0:
                            fill = PatternFill(start_color=uptake_ltz_key, end_color=uptake_ltz_key, fill_type='solid')
                            cell.number_format = ';;;'
                        if cell_v is not None:
                            cell.fill = fill

            increase_progress(1)


            for row in ws.iter_rows(min_row=1, max_row=1):
                num = seq_start[state]
                for cell in row:
                    if cell.column >= 2 and cell.column < ws.max_column:
                        cell.value = num
                        num = num+1


            for row in ws.iter_rows(max_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')


    except Exception as e:
        error_message = traceback.format_exc() 
        print(error_message)
        tk.messagebox.showerror("Unknown Error in Condensed Peptide Map", error_message)
        if delete_faulty_sheets == True:
            for sheet_name in function_worksheet_names:
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
                
def r_difcond():
    function_worksheet_titles = []
    try:
        array_col = 1
        array = np.zeros((77,53))
        whitefont = Font(color="FFFFFFFF")
        for stt, pair in  new_dic_of_dif_list.items():
            asterisk_in_state = False
            highest_row = 1
            first = pair[0]
            second = pair[1]
            protein_one = first.split("~")[0]
            protein_two = second.split("~")[0]
            sorted_peptides_first = sorted(peplist[first], key=lambda x: (int(pro_peptide_starts.get((x[0], x[1]), [0])[0]), len(x[1])))
            sorted_peptides_second = sorted(peplist[second], key=lambda x: (int(pro_peptide_starts.get((x[0], x[1]), [0])[0]), len(x[1])))
            difname = f"{stt}"
            ws_title = (f"{difname}" + "_cond")
            ws = wb.create_sheet(title=ws_title)
            function_worksheet_titles.append(ws_title)
            cell_reference_list = list()
            ws.append(["Timepoint"])
            ws.append([" "] + seqlist_dic[first] + [" "])
            timepoint_number = 0
            for timepoint in s_timepoints[first]:
                if timepoint in s_timepoints[second]:
                    if timepoint_number == 0:
                        startrow = 3
                        endrow = 250
                    if timepoint_number != 0:
                        startrow = ws.max_row +5
                        endrow = ws.max_row + 250
                    for peptide in sorted_peptides_first:
                        if peptide in sorted_peptides_second:
                            startvalues = pro_peptide_starts.get((protein_one, peptide), None)
                            startvalue= int(startvalues[0]) - seq_start[first]
                            endvalues = pro_peptide_ends.get((protein_two, peptide), None)
                            endvalue = int(endvalues[0]) - seq_start[first]
                            peptide_length = len(peptide)
                            diftake = None
                            if exp_bt_on_c == True and maxD_rfu_dif_on_c == True:
                                up1 = None
                                up2 = None
                                diftake = None

                                try:
                                    for up, tp in statedic_of_pepdic_cor[first][peptide]:
                                        if tp == timepoint:
                                            up1 = up
                                    for up, tp in statedic_of_pepdic_cor[second][peptide]:
                                        if tp == timepoint:
                                            up2 = up
                                    if up1 is not None and up2 is not None and up1 != -99999 and up2 != -99999:
                                        diftake = up1 - up2
                                    elif up1 is not None and up2 is not None:
                                        diftake = -99999

                                except:
                                    pass

                                try:
                                    SD1 = None
                                    SD2 = None
                                    diftake_SD = None
                                    print("\n\n")
                                    for sd, tp in statedic_of_sddic_cor[first][peptide]:
                                        if tp == timepoint:
                                            SD1 = sd
                                            print(SD1)
                                    for sd, tp in statedic_of_sddic_cor[second][peptide]:
                                        if tp == timepoint:
                                            SD2 = sd
                                            print(SD2)
                                    if SD1 is not None and SD2 is not None and SD1 != -99999 and SD2 != -99999:
                                        SDs = np.array([SD1, SD2])
                                        diftake_SD = np.sqrt(np.sum(SDs ** 2))
                                    elif SD1 is not None and SD2 is not None:
                                        diftake_SD = -99999
                                except:
                                    diftake_SD = -99999
                            
                            if exp_bt_on_c == True and maxD_Da_dif_on_c == True:
                                up1 = None
                                up2 = None
                                diftake = None
                                print("\n\n")

                                try:
                                    for up, tp in statedic_of_pepdic_cor[first][peptide]:
                                        if tp == timepoint:
                                            up1 = up
                                    for up, tp in statedic_of_pepdic_cor[second][peptide]:
                                        if tp == timepoint:
                                            up2 = up
                                    if up1 is not None and up2 is not None and up1 != -99999 and up2 != -99999:
                                        max_theo = get_max_theo(peptide)
                                        diftake = max_theo * (up1 - up2)
                                    elif up1 is not None and up2 is not None:
                                        diftake = -99999

                                except:
                                    pass

                                try:
                                    SD1 = None
                                    SD2 = None
                                    diftake_SD = None
                                    for sd, tp in statedic_of_sddic_cor[first][peptide]:
                                        if tp == timepoint:
                                            SD1 = sd
                                    for sd, tp in statedic_of_sddic_cor[second][peptide]:
                                        if tp == timepoint:
                                            SD2 = sd
                                    if SD1 is not None and SD2 is not None and SD1 != -99999 and SD2 != -99999:
                                        SDs = np.array([SD1, SD2])
                                        diftake_SD = np.sqrt(np.sum(SDs ** 2))
                                        diftake_SD = diftake_SD * max_theo
                                    elif SD1 is not None and SD2 is not None:
                                        diftake_SD = -99999
                                except:
                                    diftake_SD = -99999

                            if theo_bt_on_c == True and back_exchange == 0:
                                up1 = None
                                up2 = None
                                diftake = None
                                try:
                                    for up, tp in statedic_of_pepdic_raw2[first][peptide]:
                                        if tp == timepoint:
                                            up1 = up
                                    for up, tp in statedic_of_pepdic_raw2[second][peptide]:
                                        if tp == timepoint:
                                            up2 = up
                                    if up1 is not None and up2 is not None and up1 != -99999 and up2 != -99999:
                                        diftake = up1 - up2
                                    elif up1 is not None and up2 is not None:
                                        diftake = -99999
                                except:
                                    pass

                                try:
                                    SD1 = None
                                    SD2 = None
                                    diftake_SD = None
                                    for sd, tp in statedic_of_sddic_raw2[first][peptide]:
                                        if tp == timepoint:
                                            SD1 = sd
                                    for sd, tp in statedic_of_sddic_raw2[second][peptide]:
                                        if tp == timepoint:
                                            SD2 = sd
                                    if SD1 is not None and SD2 is not None and SD1 != -99999 and SD2 != -99999:
                                        SDs = np.array([SD1, SD2])
                                        diftake_SD = np.sqrt(np.sum(SDs ** 2))
                                    elif SD1 is not None and SD2 is not None:
                                        diftake_SD = -99999
                                except:
                                    diftake_SD = -99999
                                
                            if theo_bt_on_c == True and back_exchange != 0:
                                up1 = None
                                up2 = None
                                diftake = None
                                try:
                                    for up, tp in statedic_of_pepdic_cor[first][peptide]:
                                        if tp == timepoint:
                                            up1 = up
                                    for up, tp in statedic_of_pepdic_cor[second][peptide]:
                                        if tp == timepoint:
                                            up2 = up
                                    max_theo = get_max_theo(peptide)
                                    if up1 is not None and up2 is not None and up1 != -99999 and up2 != -99999:
                                        diftake = max_theo * (up1 - up2)
                                    elif up1 is not None and up2 is not None:
                                        diftake = -99999
                                except:
                                    pass

                                try:
                                    SD1 = None
                                    SD2 = None
                                    diftake_SD = None
                                    for sd, tp in statedic_of_sddic_cor[first][peptide]:
                                        if tp == timepoint:
                                            SD1 = sd
                                    for sd, tp in statedic_of_sddic_cor[second][peptide]:
                                        if tp == timepoint:
                                            SD2 = sd
                                    max_theo = get_max_theo(peptide)
                                    if SD1 is not None and SD2 is not None and SD1 != -99999 and SD2 != -99999:
                                        SDs = np.array([SD1, SD2])
                                        diftake_SD = np.sqrt(np.sum(SDs ** 2))
                                        diftake_SD = diftake_SD * max_theo
                                    elif SD1 is not None and SD2 is not None:
                                        diftake_SD = -99999
                                except:
                                    diftake_SD = -99999

                            if diftake_SD != -99999:
                                diftake_SD = diftake_SD / np.sqrt(float(SE_enter.get()))
                                
                            if diftake is not None:
                                for row in ws.iter_rows(min_row=startrow, max_row=startrow):
                                    row[0].value = timepoint
                                for i, row in enumerate(ws.iter_rows(min_row=startrow,max_row=endrow), start=startrow):
                                    if i > highest_row:
                                        highest_row = i
                                    cells = row[startvalue + 1:endvalue + 2]
                                    if all(cell.value is None for cell in cells):
                                        row[startvalue + 1].value = diftake
                                        row[startvalue + 1].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                 bottom=Side(border_style='thin', color='FF000000'),
                                                 left=Side(border_style='thin', color='FF000000'))
                                        row[endvalue+1].value = diftake
                                        row[endvalue+1].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                 bottom=Side(border_style='thin', color='FF000000'),
                                                 right=Side(border_style='thin', color='FF000000'))
                                        for cell in row[startvalue + 2:endvalue+1]:
                                            cell.value = diftake
                                            cell.border = Border(top=Side(border_style='thin', color='FF000000'),
                                                 bottom=Side(border_style='thin', color='FF000000'))
                                        middle = int((startvalue + 1 + endvalue + 1)/2)



                                        if sd_checkvar.get() == 0:
                                            if len(peptide) >= 5:
                                                c = 1
                                            else:
                                                c = 0

                                            if exp_bt_on_c == True and maxD_rfu_dif_on_c == True:
                                                row[middle-c].value = round(diftake * 100, 1)
                                                row[middle-c].number_format = "0.0"


                                            if theo_bt_on_c == True or maxD_Da_dif_on_c == True:
                                                row[middle-c].value = round(diftake, 2)
                                                row[middle-c].number_format = "0.00"
                                        else: 
                                            if len(peptide) > 6:
                                                if diftake_SD != -99999 and diftake_SD != 0 and diftake_SD != "-99999" and diftake_SD != "0":
                                                    if exp_bt_on_c == True and maxD_rfu_dif_on_c == True:
                                                        if (str(diftake).startswith("-") and len(str(round(diftake * 100))) == 2) or len(str(round(diftake * 100))) == 1:
                                                            row[middle-2].value = str(round(diftake * 100)) + " " + "\u00B1" + str(round(diftake_SD * 100))
                                                        else:
                                                            row[middle-2].value = str(round(diftake * 100)) + " " + "\u00B1" + str(round(diftake_SD * 100))
                                                    else:
                                                        if (str(diftake).startswith("-") and len(str(round(diftake))) == 2) or len(str(round(diftake))) == 1:
                                                            row[middle-2].value = str(round(diftake, 1)) + " " + "\u00B1" + str(round(diftake_SD, 1))
                                                        else:
                                                            row[middle-2].value = str(round(diftake)) + " " + "\u00B1" + str(round(diftake_SD))
                                                else:
                                                    if exp_bt_on_c == True and maxD_rfu_dif_on_c == True:
                                                        row[middle-2].value = round(diftake * 100, 1)
                                                        row[middle-2].number_format = "0.0"
                                                    else:
                                                        row[middle-2].value = round(diftake, 1)
                                                        row[middle-2].number_format = "0.0"

                                            elif len(peptide) in [5, 6]:
                                                if diftake_SD != -99999 and diftake_SD != 0 and diftake_SD != "-99999" and diftake_SD != "0":
                                                    if exp_bt_on_c == True and maxD_rfu_dif_on_c == True:
                                                        row[middle-1].value = str(round(diftake * 100)) + " " + "\u00B1" + str(round(diftake_SD * 100))
                                                    else:
                                                        if (str(diftake).startswith("-") and len(str(round(diftake))) == 2) or len(str(round(diftake))) == 1:
                                                            row[middle-1].value = str(round(diftake, 1)) + " " + "\u00B1" + str(round(diftake_SD, 1))
                                                        else:
                                                            row[middle-1].value = str(round(diftake, 1)) + " " + "\u00B1" + str(round(diftake_SD, 1))
                                                else:
                                                    if exp_bt_on_c == True and maxD_rfu_dif_on_c == True:
                                                        row[middle-1].value = round(diftake * 100, 1)
                                                        row[middle-1].number_format = "0.0"
                                                    else:
                                                        row[middle-1].value = round(diftake, 1)
                                                        row[middle-1].number_format = "0.0"

                                            elif len(peptide) == 4:
                                                if exp_bt_on_c == True and maxD_rfu_dif_on_c == True:
                                                    row[middle].value = round(diftake * 100)
                                                    row[middle].number_format = "0.0"
                                                else:
                                                    row[middle].value = round(diftake, 1)
                                                    row[middle].number_format = "0.0"







                                        if sd_checkvar.get() == 0:
                                            if len(peptide) == 4:
                                                c = 0
                                                b = 10
                                            else:
                                                c = 1
                                                b = 14
                                        else:
                                            if len(peptide) == 4:
                                                c = 0
                                                b = 10
                                            elif len(peptide) in [5, 6]:
                                                c = 1
                                                b = 14
                                            else:
                                                c = 2
                                                b = 14

                                        if diftake == -99999:
                                            fill = PatternFill(start_color=b_col_abs, end_color=b_col_abs, fill_type='solid')
                                            font = font = Font(color=b_text_abs, size=b)
                                            row[middle-c].number_format = ';;;'
                                        elif p_col_length >= 1 and diftake >= d_val_1:
                                            fill = PatternFill(start_color=d_col_1, end_color=d_col_1, fill_type='solid')
                                            font = Font(color=d_text_1, size=b)
                                        elif p_col_length >= 2 and diftake >= d_val_2:
                                            fill = PatternFill(start_color=d_col_2, end_color=d_col_2, fill_type='solid')
                                            font = Font(color=d_text_2, size=b)
                                        elif p_col_length >= 3 and diftake >= d_val_3:
                                            fill = PatternFill(start_color=d_col_3, end_color=d_col_3, fill_type='solid')
                                            font = Font(color=d_text_3, size=b)
                                        elif p_col_length >= 4 and diftake >= d_val_4:
                                            fill = PatternFill(start_color=d_col_4, end_color=d_col_4, fill_type='solid')
                                            font = Font(color=d_text_4, size=b)
                                        elif p_col_length >= 5 and diftake >= d_val_5:
                                            fill = PatternFill(start_color=d_col_5, end_color=d_col_5, fill_type='solid')
                                            font = Font(color=d_text_5, size=b)
                                        elif diftake > 0:
                                            fill = PatternFill(start_color=d_col_gtz, end_color=d_col_gtz, fill_type='solid')
                                            font = Font(color=d_text_gtz, size=b)
                                            if insig_dif_chk.get() == 0:
                                                row[middle-c].number_format = ';;;'
                                        elif diftake == 0:
                                            fill = PatternFill(start_color=b_col_eqz, end_color=b_col_eqz, fill_type='solid')
                                            font = Font(color=b_text_eqz, size=b)
                                            row[middle-c].number_format = ';;;'
                                        elif d_col_length >= 1 and diftake <= (-1) * p_val_1:
                                            fill = PatternFill(start_color=p_col_1, end_color=p_col_1, fill_type='solid')
                                            font = Font(color=p_text_1, size=b)
                                        elif d_col_length >= 2 and diftake <= (-1) * p_val_2:
                                            fill = PatternFill(start_color=p_col_2, end_color=p_col_2, fill_type='solid')
                                            font = Font(color=p_text_2, size=b)
                                        elif d_col_length >= 3 and diftake <= (-1) * p_val_3:
                                            fill = PatternFill(start_color=p_col_3, end_color=p_col_3, fill_type='solid')
                                            font = Font(color=p_text_3, size=b)
                                        elif d_col_length >= 4 and diftake <= (-1) * p_val_4:
                                            fill = PatternFill(start_color=p_col_4, end_color=p_col_4, fill_type='solid')
                                            font = Font(color=p_text_4, size=b)
                                        elif d_col_length >= 5 and diftake <= (-1) * p_val_5:
                                            fill = PatternFill(start_color=p_col_5, end_color=p_col_5, fill_type='solid')
                                            font = Font(color=p_text_5, size=b)
                                        elif diftake < 0:
                                            fill = PatternFill(start_color=p_col_gtz, end_color=p_col_gtz, fill_type='solid')
                                            font = Font(color=p_text_gtz, size=b)
                                            if insig_dif_chk.get() == 0:
                                                row[middle-c].number_format = ';;;'
                                        else:
                                            print(diftake)

                                        row[middle-c].fill = fill
                                        row[middle-c].font = font




                                        if sd_checkvar.get() == 0:
    #                                        if exp_bt_on_c == True and maxD_rfu_dif_on_c == True:
    #                                            ws.merge_cells(start_row=row[middle-1].row, start_column=row[middle-1].column, end_row=row[middle+2].row, end_column=row[middle+2].column)
    #                                            middle_cell_reference = row[middle-1].coordinate
    #                                            cell_reference_list.append(middle_cell_reference)
    #
    #                                        else:
    #                                            ws.merge_cells(start_row=row[middle-1].row, start_column=row[middle-1].column, end_row=row[middle+2].row, end_column=row[middle+2].column)
    #                                            middle_cell_reference = row[middle-1].coordinate
    #                                            cell_reference_list.append(middle_cell_reference)


                                            if len(peptide) >= 5:
                                                ws.merge_cells(start_row=row[middle-1].row, start_column=row[middle-1].column, end_row=row[middle+1].row, end_column=row[middle+1].column)
                                                middle_cell_reference = row[middle-1].coordinate
                                                cell_reference_list.append(middle_cell_reference)
                                            else:
                                                ws.merge_cells(start_row=row[middle].row, start_column=row[middle].column, end_row=row[middle+1].row, end_column=row[middle+1].column)
                                                middle_cell_reference = row[middle].coordinate
                                                cell_reference_list.append(middle_cell_reference)


                                        if sd_checkvar.get() == 1:
                                            if len(peptide) == 6:
                                                ws.merge_cells(start_row=row[middle-1].row, start_column=row[middle-1].column, end_row=row[middle+2].row, end_column=row[middle+2].column)
                                                middle_cell_reference = row[middle-1].coordinate
                                            elif len(peptide) == 5:
                                                ws.merge_cells(start_row=row[middle-1].row, start_column=row[middle-1].column, end_row=row[middle+1].row, end_column=row[middle+1].column)
                                                middle_cell_reference = row[middle-1].coordinate
                                            elif len(peptide) == 4:
                                                ws.merge_cells(start_row=row[middle].row, start_column=row[middle].column, end_row=row[middle+1].row, end_column=row[middle+1].column)
                                                middle_cell_reference = row[middle].coordinate
                                            elif len(peptide) <= 10:
                                                ws.merge_cells(start_row=row[middle-2].row, start_column=row[middle-2].column, end_row=row[middle+1].row, end_column=row[middle+1].column)
                                                middle_cell_reference = row[middle-2].coordinate
                                            else:
                                                ws.merge_cells(start_row=row[middle-2].row, start_column=row[middle-2].column, end_row=row[middle+2].row, end_column=row[middle+2].column)
                                                middle_cell_reference = row[middle-2].coordinate

                                            cell_reference_list.append(middle_cell_reference)


                                        row[middle-c].alignment = Alignment(horizontal='center', vertical='center')

                                        ch1 = False
                                        ch2 = False
                                        try:
                                            if noD_dic_states[first][peptide] == True:
                                                ch1 = True
                                        except:
                                            pass
                                        try:
                                            if noD_dic_states[second][peptide] == True:
                                                ch2 = True
                                        except:
                                            pass
                                        if ch1 == True or ch2 == True:
                                            if diftake != 0 and diftake != -99999:
                                                row[startvalue+1].value = "*"
                                                asterisk_in_state = True
                                                #row[startvalue+1].alignment = Alignment(textRotation=90, vertical='center', horizontal='left')


                                                if d_col_length >= 1 and diftake >= d_val_1:
                                                    fill = PatternFill(start_color=d_col_1, end_color=d_col_1, fill_type='solid')
                                                    font = Font(color=d_text_1, size=12)
                                                elif d_col_length >= 2 and diftake >= d_val_2:
                                                    fill = PatternFill(start_color=d_col_2, end_color=d_col_2, fill_type='solid')
                                                    font = Font(color=d_text_2, size=12)
                                                elif d_col_length >= 3 and diftake >= d_val_3:
                                                    fill = PatternFill(start_color=d_col_3, end_color=d_col_3, fill_type='solid')
                                                    font = Font(color=d_text_3, size=12)
                                                elif d_col_length >= 4 and diftake >= d_val_4:
                                                    fill = PatternFill(start_color=d_col_4, end_color=d_col_4, fill_type='solid')
                                                    font = Font(color=d_text_4, size=12)
                                                elif d_col_length >= 5 and diftake >= d_val_5:
                                                    fill = PatternFill(start_color=d_col_5, end_color=d_col_5, fill_type='solid')
                                                    font = Font(color=d_text_5, size=12)
                                                elif diftake > 0:
                                                    fill = PatternFill(start_color=d_col_gtz, end_color=d_col_gtz, fill_type='solid')
                                                    font = Font(color=d_text_gtz, size=12)
                                                elif p_col_length >= 1 and diftake <= (-1) * p_val_1:
                                                    fill = PatternFill(start_color=p_col_1, end_color=p_col_1, fill_type='solid')
                                                    font = Font(color=p_text_1, size=12)
                                                elif p_col_length >= 2 and diftake <= (-1) * p_val_2:
                                                    fill = PatternFill(start_color=p_col_2, end_color=p_col_2, fill_type='solid')
                                                    font = Font(color=p_text_2, size=12)
                                                elif p_col_length >= 3 and diftake <= (-1) * p_val_3:
                                                    fill = PatternFill(start_color=p_col_3, end_color=p_col_3, fill_type='solid')
                                                    font = Font(color=p_text_3, size=12)
                                                elif p_col_length >= 4 and diftake <= (-1) * p_val_4:
                                                    fill = PatternFill(start_color=p_col_4, end_color=p_col_4, fill_type='solid')
                                                    font = Font(color=p_text_4, size=12)
                                                elif p_col_length >= 5 and diftake <= (-1) * p_val_5:
                                                    fill = PatternFill(start_color=p_col_5, end_color=p_col_5, fill_type='solid')
                                                    font = Font(color=p_text_5, size=12)
                                                elif diftake < 0:
                                                    fill = PatternFill(start_color=p_col_gtz, end_color=p_col_gtz, fill_type='solid')
                                                    font = Font(color=p_text_gtz, size=12)
                                                row[startvalue+1].fill = fill
                                                row[startvalue+1].font = font






                                        break
                                    else:
                                        continue



                        timepoint_number = timepoint_number + 1


            for row in ws.iter_rows(min_row=2, max_row=2):
                for cell in row:
                    cell.font = courier_new_style

            if asterisk_in_state == True:
                for row in ws.iter_rows(min_row=highest_row+2, max_row=highest_row+2):
                    row[1].value = "* = no maxD for peptide, average back exchange used"

            white_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type='solid')
            for row in ws.rows:
                for cell in row:
                    if cell.value == None:
                        cell.fill = white_fill
            for row in ws.iter_rows(min_row=2, max_row=2):
                for cell in row:
                    cell.fill = white_fill




            for i, column in enumerate(ws.columns):
                if i == 0:
                    continue
    #            if con_pep_width_enter.get() == "2.5":
                ws.column_dimensions[column[0].column_letter].width = "2.504"
    #            else:
    #                ws.column_dimensions[column[0].column_letter].width = con_pep_width_enter.get()


            for row in ws.iter_rows(min_row=3):
                for cell in row[1:]:
                    cell_v = cell.value
                    if cell.coordinate not in cell_reference_list and cell_v != "*" and cell_v != "* = no maxD for peptide, average back exchange used":
                        if cell_v == -99999:
                            fill = PatternFill(start_color=b_col_abs, end_color=b_col_abs, fill_type='solid')
                            cell.number_format = ';;;'
                        elif d_col_length >= 1 and cell_v is not None and cell_v >= d_val_1:
                            fill = PatternFill(start_color=d_col_1, end_color=d_col_1, fill_type='solid')
                            cell.number_format = ';;;'
                        elif d_col_length >= 2 and cell_v is not None and cell_v >= d_val_2:
                            fill = PatternFill(start_color=d_col_2, end_color=d_col_2, fill_type='solid')
                            cell.number_format = ';;;'
                        elif d_col_length >= 3 and cell_v is not None and cell_v >= d_val_3:
                            fill = PatternFill(start_color=d_col_3, end_color=d_col_3, fill_type='solid')
                            cell.number_format = ';;;'
                        elif d_col_length >= 4 and cell_v is not None and cell_v >= d_val_4:
                            fill = PatternFill(start_color=d_col_4, end_color=d_col_4, fill_type='solid')
                            cell.number_format = ';;;'
                        elif d_col_length >= 5 and cell_v is not None and cell_v >= d_val_5:
                            fill = PatternFill(start_color=d_col_5, end_color=d_col_5, fill_type='solid')
                            cell.number_format = ';;;'
                        elif cell_v is not None and cell_v > 0:
                            fill = PatternFill(start_color=d_col_gtz, end_color=d_col_gtz, fill_type='solid')
                            cell.number_format = ';;;'
                        elif p_col_length >=1 and cell_v is not None and cell_v <= (-1) * p_val_1:
                            fill = PatternFill(start_color=p_col_1, end_color=p_col_1, fill_type='solid')
                            cell.number_format = ';;;'
                        elif p_col_length >=2 and cell_v is not None and cell_v <= (-1) * p_val_2:
                            fill = PatternFill(start_color=p_col_2, end_color=p_col_2, fill_type='solid')
                            cell.number_format = ';;;'
                        elif p_col_length >=3 and cell_v is not None and cell_v <= (-1) * p_val_3:
                            fill = PatternFill(start_color=p_col_3, end_color=p_col_3, fill_type='solid')
                            cell.number_format = ';;;'
                        elif p_col_length >=4 and cell_v is not None and cell_v <= (-1) * p_val_4:
                            fill = PatternFill(start_color=p_col_4, end_color=p_col_4, fill_type='solid')
                            cell.number_format = ';;;'
                        elif p_col_length >=5 and cell_v is not None and cell_v <= (-1) * p_val_5:
                            fill = PatternFill(start_color=p_col_5, end_color=p_col_5, fill_type='solid')
                            cell.number_format = ';;;'
                        elif cell_v is not None and cell_v < 0:
                            fill = PatternFill(start_color=p_col_gtz, end_color=p_col_gtz, fill_type='solid')
                            cell.number_format = ';;;'
                        elif cell_v is not None and cell_v == 0:
                            fill = PatternFill(start_color=b_col_eqz, end_color=b_col_eqz, fill_type='solid')
                            cell.number_format = ';;;'
                        if cell.value is not None:
                            cell.fill = fill


            increase_progress(1)


            for row in ws.iter_rows(min_row=1, max_row=1):
                num = seq_start[first]
                for cell in row:
                    if cell.column >= 2 and cell.column < ws.max_column:
                        cell.value = num
                        num = num+1


            for row in ws.iter_rows(max_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')



            #df = pd.DataFrame(array)
    #        df.to_excel("your_file.xlsx", engine='openpyxl', index=False)
    
    except Exception as e:
        error_message = traceback.format_exc() 
        print(error_message)
        tk.messagebox.showerror("Unknown Error in Condensed Difference Map", error_message)
        if delete_faulty_sheets == True:
            for sheet_name in function_worksheet_names:
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]

def r_heat_map():
    try:
        for sheet_name in wb.sheetnames:
            if sheet_name.endswith("_dif"):
                sheet = wb[sheet_name]
                tp_starts = []

                linear_map_multiplier = 0.5/future_linear_map_multiplier

                for i, row in enumerate(sheet.iter_rows(values_only=True)):
                    if row[0] != None and row[0] != "0" and row[0] != 0:
                        tp_starts.append((i+1, row[0]))
                start_of_sheet = sheet_name.removesuffix("_dif")
                for j, (i, timepoint) in enumerate(tp_starts[2:], start=2):
                    ws_title = (start_of_sheet + "|" + str(timepoint)[0:5] + "#$")
                    ws = wb.create_sheet(title=ws_title)
                    for row_cells in sheet.iter_rows(min_row=1, max_row=1):
                        ws.append([cell.value for cell in row_cells])
                    for k, row_cells in enumerate(sheet.iter_rows(min_row=i, max_row=(tp_starts[j+1][0] - 1) if j+1 < len(tp_starts) else None)):
                        ws.append([(cell.value) for cell in row_cells])
                    ws.delete_cols(1)
                    ws.delete_rows(k)
                    for l, col in enumerate(ws.iter_cols(values_only=True)):
                        if col[0] is None:
                            continue
                        if all(cell is None or cell == -99999 for cell in col[1:]):
                            ws.cell(row=1, column=l+1, value=0)
                        else:
                            ws.cell(row=1, column=l+1, value=1)                    
                    for row in ws.iter_rows(min_row=2):
                        for p, cell in enumerate(row):
                            if cell.value is not None and cell.value != -99999:
                                try:
                                    cell.value = (cell.value * linear_map_multiplier)
                                except:
                                    if cell.value == "*":
                                        cell.value = (row[p+1].value * linear_map_multiplier)
                                    else:
                                        cell.value = None


        window_extend = 10
        class MyModel(nn.Module):
            def __init__(self, window_extend):
                super(MyModel, self).__init__()
                self.conv1 = nn.Conv2d(2, 32, kernel_size=5, padding=2)
                self.bn1 = nn.BatchNorm2d(32)
                self.pool = nn.MaxPool2d(2, 2)
                self.dropout1 = nn.Dropout(0.2)
                self.fc1 = nn.Linear(32 * 13 * 10, 64)
                self.dropout2 = nn.Dropout(0.1)
                self.fc2 = nn.Linear(64, 32)
                self.dropout3 = nn.Dropout(0.1)
                self.fc3 = nn.Linear(32, 16)
                self.fc4 = nn.Linear(16, 6)

            def forward(self, x):
                x = self.pool(Fu.relu(self.bn1(self.conv1(x))))
                x = x.reshape(x.size(0), -1)
                x = self.dropout1(x)
                x = Fu.relu(self.fc1(x))
                x = self.dropout2(x)
                x = Fu.relu(self.fc2(x))
                x = self.dropout3(x)
                x = Fu.relu(self.fc3(x))
                x = self.fc4(x)
                return x
    #    model = tf.keras.models.load_model("model_1")
        model = MyModel(window_extend)
        model.load_state_dict(torch.load('best_model_reshape.pth', map_location=torch.device('cpu')))
        print("\n\n")
        def label_peptides(matrix):
            last_value = 0
            peptide_length = 0
            indices_of_peptides = []
            for row_index, row in enumerate(matrix[2:, :, 0], start=2):
                for col_index, value in enumerate(row):
                    if value == 0:
                        last_value = 0
                        continue

                    if value != last_value:
                        #this is the start of a new peptide so we need to deal with the last peptide before updating any values
                        if peptide_length != 0: #we have no previous data to deal with the first time this loop is ran
                            indices_of_peptides.append((peptide_length, indices_of_peptide))
                        #here we reset values
                        indices_of_peptide = []  
                        peptide_length = 1
                        indices_of_peptide.append((row_index, col_index))
                    if value == last_value:
                        peptide_length += 1
                        indices_of_peptide.append((row_index, col_index))

                    last_value = value

            if peptide_length != 0:
                indices_of_peptides.append((peptide_length, indices_of_peptide)) #do the last peptide


            for length, indices in indices_of_peptides:
                for i, (row, col) in enumerate(indices):
                    if i == 0:
                        matrix[row, col, 0] = 0
                    if i == 1:
                        if n_min_two == True:
                            matrix[row, col, 0] = 0
            return matrix


        def trim_matrix(new_data_matrix):
            m = window_extend
            for k in range(0, window_extend+1):
                current_values = new_data_matrix[2:, m+k, 0]
                if k != 0:
                    if all(current_values == 0):
                        new_data_matrix[1:, m+k:, 0] = 0
                        break
                    lvs = []
                    cvs = []
                    for j, cv in enumerate(current_values):
                        if cv != 0:
                            cvs.append(cv)
                            lvs.append(last_values[j])
                    if all(cv != lv for cv, lv in zip(cvs, lvs)):
                        new_data_matrix[1:, m+k:, 0] = 0
                        break

                if k != window_extend - 1:
                    last_values = current_values


            m = window_extend
            for k in range(0, window_extend+1):
                current_values = new_data_matrix[2:, m-k, 0]
                if k != 0:
                    if all(current_values == 0):
                        new_data_matrix[1:, :m+1-k, 0] = 0
                        break
                    lvs = []
                    cvs = []
                    for j, cv in enumerate(current_values):
                        if cv != 0:
                            cvs.append(cv)
                            lvs.append(last_values[j])
                    if all(cv != lv for cv, lv in zip(cvs, lvs)):
                        new_data_matrix[1:, :m+1-k, 0] = 0
                        break

                if k != window_extend - 1:
                    last_values = current_values
            return new_data_matrix

        first_X_data = True
        def make_X_data(title, list1, list2, nexttitle):
            global first_X_data
            max_i = 0
            xls = pd.ExcelFile(title)
            for i, sheet_name in enumerate(xls.sheet_names):
                if not sheet_name.endswith("#$"):
                    continue
                df = pd.read_excel(xls, sheet_name, header=None)
                last_column_name = int(df.columns[-1])
                res = last_column_name + 2
                list2.append(res)
                matrix = np.full((27, res, 2), -0, dtype = float)
                if not np.issubdtype(df.values.dtype, np.number):
                    raise ValueError("Non-numeric data found in Excel sheet.")
                matrix[1:df.shape[0]+1, :df.shape[1], 0] = np.where(np.isnan(df.values), matrix[1:df.shape[0]+1, :df.shape[1], 0], df.values)
                matrix[matrix[:, :, 0] == -99999, 0] = 0
                matrix = label_peptides(matrix)
                matrix[2:, :, 1] = np.where(matrix[2:, :, 0] != 0, 1, 0) #this line gives bool values to the second layer
                list1.append((matrix, sheet_name.split("|")[0]))

    #        with pd.ExcelWriter(nexttitle, engine='openpyxl') as writer:
    #            for idx, (data_matrix, name) in enumerate(list1):
    #                df_to_save = pd.DataFrame(data_matrix[:,:,0], dtype = float)
    #                df_to_save.to_excel(writer, sheet_name=f'Sheet_{idx}', index=False, header=False)

            statename_dic = {}
            X = []
            X_complement = []
            for i, (data_matrix, statename) in enumerate(list1):
                if not statename in statename_dic.keys():
                    statename_dic[statename] = []
                statename_dic[statename].append(i)
                res = list2[i]
                x = 0
                while x < res:
                    new_data_matrix = data_matrix.copy()
                    new_data_matrix[0, :, 0] = 0  # Setting all values in the first row to 0
                    new_data_matrix[0, x, 0] = 1  # Setting the specific index in the first row to 1

                    pad_before = max(0, window_extend - x)
                    pad_after = max(0, window_extend - (res-x-1))
                    if pad_before > 0:
                        new_data_matrix = new_data_matrix[:, :x+window_extend+1, :]
                        new_data_matrix = np.pad(new_data_matrix, ((0, 0), (pad_before, 0), (0,0)), mode='constant', constant_values=0)
                    elif pad_after > 0:
                        new_data_matrix = new_data_matrix[:, x-window_extend:, :]
                        new_data_matrix = np.pad(new_data_matrix, ((0, 0), (0, pad_after), (0,0)), mode='constant', constant_values=0)
                    else:
                        new_data_matrix = new_data_matrix[:, x-window_extend:x+window_extend+1, :]
                    if new_data_matrix.shape != (27, ((window_extend*2) + 1), 2):
                        print(new_data_matrix.shape)


                    new_data_matrix = trim_matrix(new_data_matrix) 



                    X.append(new_data_matrix)
                    X_complement.append(i)
                    max_i = i
                    x += 1

            return np.array(X), X_complement, max_i, statename_dic

        temp_file_path_linearmap = 'temp_excel_file_linearmap.xlsx'
        wb.save(temp_file_path_linearmap)
        atexit.register(os.remove, temp_file_path_linearmap)

        test_data, test_res = [], []
        try:
            X_data, X_complement, max_i, statename_dic = make_X_data('temp_excel_file_linearmap.xlsx', test_data, test_res, 'Test Output Data.xlsx')
        except IndexError as e: 
            tk.messagebox.showerror("Localized Difference Plot Error", "No Data Found. Please make sure difference requests contain the same proteins.")
            return
        except Exception as e:
            tk.messagebox.showerror("Localized Difference Plot Error", f"An Unexpected Error Occured: {e}")
            return

        lm_X_data_dic = {}

        j = 0
        while j <= max_i:
            lm_X_data_dic[j] = []
            j += 1
        for i, new_data_matrix in enumerate(X_data):
            iteration = X_complement[i]
            lm_X_data_dic[iteration].append(new_data_matrix)
        for statename, i_list in statename_dic.items():
            ws_title = statename + "_predicts"
            ws_title = ws_title
            ws = wb.create_sheet(title = ws_title)

            for iteration, X_data in lm_X_data_dic.items():
                if iteration in i_list:
                    X_test= np.array(X_data)

                    shuffles = 10
                    augmented_predicted_labels = []
                    for i in range(shuffles):
                        if i == 0:
                            X_data = torch.tensor(X_test).float().permute(0, 3, 1, 2)
                        if i > 0:
                            X_test = shuffle_x_rows(X_test)
                            X_data = torch.tensor(X_test).float().permute(0, 3, 1, 2)
                        with torch.no_grad():
                            predictions = model(X_data)
                        predicted_labels = torch.argmax(predictions, dim=1).cpu().numpy()
                        augmented_predicted_labels.append(predicted_labels)
                    transposed = list(zip(*augmented_predicted_labels))
                    predicted_labels = [mode(group) for group in transposed]

    #                X_data = torch.tensor(X_data).float().permute(0, 3, 1, 2)
    #                with torch.no_grad():
    #                    predictions = model(X_data)
    #                predicted_labels = torch.argmax(predictions, dim=1).cpu().numpy()

                    all_predicted_labels_lengths[statename] = len(predicted_labels)


                    #predicted_labels_list = predicted_labels.tolist()[:-1]
                    predicted_labels_list = predicted_labels[:-1]

                    if proline_checkvar.get() == 1:
                        predicted_labels_list = set_prolines_to_none(predicted_labels_list, statename)


                    ws.append(predicted_labels_list)



        for sheet_name in wb.sheetnames:
            if sheet_name.endswith("#$"):
                ws = wb[sheet_name]
                wb.remove(ws)
                
    except Exception as e:
        error_message = traceback.format_exc() 
        print(error_message)
        tk.messagebox.showerror("Unknown Error in Linear Map Creation", error_message)

all_predicted_labels_lengths = {}
            
        
def set_prolines_to_none(predicted_labels_list, statename):
    if statename in new_dic_of_dif_list.keys():
        difpair = new_dic_of_dif_list[statename]
        first_dif = difpair[0]
        protein = first_dif.split("~")[0]
        for peptide in peplist[first_dif]:
            if "P" in peptide:
                startvalues = pro_peptide_starts.get((protein, peptide), None)
                startvalue= int(startvalues[0])
                proline_indexes = [i for i, n in enumerate(peptide) if n == "P"]
                for index in proline_indexes:
                    if predicted_labels_list[startvalue+index-seq_start[first_dif]] != 3:
                        predicted_labels_list[startvalue+index-seq_start[first_dif]] = 0      
    else:
        print("Couldn't find difpair")
        
    return predicted_labels_list

#def fix_n_min_two_linmap(predicted_labels_list, statename):
#    if statename in new_dic_of_dif_list.keys():
#        difpair = new_dic_of_dif_list[statename]
#        first_dif = difpair[0]
#        protein = first_dif.split("~")[0]
#        peptide_indices = []
#        for peptide in peplist[first_dif]:
#            startvalues = pro_peptide_starts.get((protein, peptide), None)
#            startvalue= int(startvalues[0])
#            for i in range(len(peptide)):
#                if startvalue+i not in peptide_indices:
#                    peptide_indices.append(startvalue + i)
#        print(peptide_indices)
#        print(len(peptide_indices))
#        print(len(predicted_labels_list))
#        print(predicted_labels_list)
#        for peptide_index in peptide_indices:
#            if predicted_labels_list[peptide_index-1] == 3:
#                try:
#                    predicted_labels_list[peptide_index-1] = 9
#                except:
#                    print(peptide_index-1)
#    
#    else:
#        print("Couldn't find difpair")
#        
#    return predicted_labels_list
    
def shuffle_rows(duplicate):
    indices = np.arange(2, 27)
    np.random.shuffle(indices)
    duplicate[2:, :, :] = duplicate[indices, :, :]
    return duplicate

def shuffle_x_rows(X_test):
    augmented_X = []
    for i in range(len(X_test)):
        duplicate = X_test[i].copy()
        duplicate = shuffle_rows(duplicate)
        augmented_X.append(duplicate)   
    return np.array(augmented_X) 

def r_uptake_plots():
    try:
        global a_horizontal, a_vertical
        last_filled_position = (0, 0)
        title_fontsize = 8 
        pdf_pages = PdfPages('uptake_plots.pdf')
        if a_vertical is True:
            fig, axes = plt.subplots(8, 6, figsize=(8.5, 11))
        if a_horizontal is True:
            fig, axes = plt.subplots(6, 8, figsize=(11, 8.5))
        fig.text(0.5, 0.04, f"{x_enter.get()}", ha='center', va='center', fontsize=12)
        fig.text(0.04, 0.5, f"{y_enter.get()}", ha='center', va='center', rotation='vertical', fontsize=12)
        plt.subplots_adjust(wspace=0.4, hspace=0.6)
        fig_idx = 0
        page_count = 1





        all_peptides = []      
        for state in order_state_dic.values():
            if state != False and state != "False":
                protein = state.split("~")[0]
                for peptide in peplist[state]:
                    if (protein, peptide) in all_peptides:
                        continue
                    all_peptides.append((protein, peptide))  

        seg_proteins = True
        if seg_proteins == False:             
            sorted_all_peptides = sorted(all_peptides, key=lambda x: (int(pro_peptide_starts.get((x[0], x[1]), [0])[0]), len(x[1])))
        if seg_proteins == True:
            sorted_all_peptides = sorted(
                all_peptides,
                key=lambda x: (
                    x[0],  # Sort primarily by protein name
                    int(pro_peptide_starts.get((x[0], x[1]), [0])[0]),  # Secondary sort by start position
                    len(x[1])  # Tertiary sort by peptide length
                )
            )

        new_sorted_all_peptides = sorted_all_peptides  

        max_protein_tps = {}
        for this_protein in protein_states.keys():
            max_state_tps = []
            for state in statedic_of_pepdic_raw2.keys():
                if not state.split("~")[0] == this_protein:
                    continue
                small_tuplist = []
                for any_peptide, tuplist in statedic_of_pepdic_raw2[state].items():
                    for each_tuple in tuplist:
                        small_tuplist.append(each_tuple[1])

                if show_last == False:
                    max_tp_state = sorted(set(small_tuplist), reverse=True)[1]
                else:
                    max_tp_state = max(small_tuplist)
                max_state_tps.append(max_tp_state)

            max_protein_tp = max(max_state_tps)
            max_protein_tps[this_protein] = max_protein_tp



        last_protein = None
        idx_increment = 0
        empty_protein_boxes = list()
        for idx, (protein, peptide) in enumerate(new_sorted_all_peptides):
            working_idx = idx + idx_increment
            if protein != last_protein:
                if a_vertical is True:
                    row = working_idx % 48 // 6
                    col = working_idx % 48 % 6
                if a_horizontal is True:
                    row = working_idx % 48 // 8
                    col = working_idx % 48 % 8
                empty_protein_box = (row, col)
                empty_protein_boxes.append(empty_protein_box)
                if last_protein != None:
                    if (working_idx+1) % 48 == 0:
                        for (row, col) in empty_protein_boxes:
                            axes[row, col].set_xticks([])
                            axes[row, col].set_yticks([])
                        empty_protein_boxes = []
                        pdf_pages.savefig(fig)
                        plt.close(fig)
                        if idx < len(new_sorted_all_peptides) - 1:
                            if a_vertical is True:
                                fig, axes = plt.subplots(8, 6, figsize=(8.5, 11))
                            if a_horizontal is True:
                                fig, axes = plt.subplots(6, 8, figsize=(11, 8.5))
                            fig.text(0.5, 0.04, f"{x_enter.get()}", ha='center', va='center', fontsize=12)
                            fig.text(0.04, 0.5, f"{y_enter.get()}", ha='center', va='center', rotation='vertical', fontsize=12)
                            plt.subplots_adjust(wspace=0.4, hspace=0.6)
                            page_count += 1

                idx_increment += 1


            working_idx = idx + idx_increment
            last_protein = protein


            if a_vertical is True:
                row = working_idx % 48 // 6
                col = working_idx % 48 % 6
            if a_horizontal is True:
                row = working_idx % 48 // 8
                col = working_idx % 48 % 8
            ax = axes[row, col]


            max_theo = get_max_theo(peptide)
            ax.set_xscale('log')
            ax.xaxis.set_major_locator(ticker.LogLocator(base=10.0, numticks=10))
            ax.xaxis.set_minor_locator(ticker.LogLocator(base=10.0, subs='auto', numticks=10))





            if calculate_theoretical_back_exchange is True:
                old_max_theo = max_theo
                try:
                    max_theo = int(np.ceil(max_theo))
                except:
                    pass

            if correction is True:
                max_theo += 2

            if max_theo <= 7:
                step = 1
            elif max_theo == 8 or max_theo == 10:
                step = 2
            elif max_theo == 9:
                step = 3
            elif max_theo in [11, 13, 14]:
                step = 2
            elif max_theo in [12, 15]:
                step = 3
            elif max_theo in [16, 17, 19, 20]:
                step = 4
            elif max_theo == 18:
                step = 6
            elif max_theo == 21:
                step = 7
            elif max_theo in [22, 23, 24]:
                step = 4
            elif max_theo >= 25:
                step = 5
            y_ticks = list(range(0, max_theo + 1, step))

            if max_theo % step > 1 and max_theo < 25:
                y_ticks.append(max_theo)
            if max_theo % step > 2 and max_theo >= 25:
                y_ticks.append(max_theo)
            ax.set_yticks(y_ticks)

            if correction is True:
                max_theo -= 2

            if calculate_theoretical_back_exchange is True:
                max_theo = old_max_theo


            startvalues = pro_peptide_starts.get((protein, peptide), None)
            startvalue= int(startvalues[0])
            endvalues = pro_peptide_ends.get((protein, peptide), None)
            endvalue = int(endvalues[0])
            if len(peptide) < 16:
                ax.set_title(f'$^{{{startvalue}}} {peptide} ^{{{endvalue}}}$', fontsize=5)
            else:
                new_peptide_name = peptide[:3] + "..." + peptide[-3:]
                ax.set_title(f'$^{{{startvalue}}} {new_peptide_name} ^{{{endvalue}}}$', fontsize=5)

            for state in statedic_of_pepdic_raw2:
                if not state.split("~")[0] == protein:
                    continue


                if show_last is False:
                    small_tuplist = []
                    for key, value in statedic_of_pepdic_raw2.items():
                        if not key.split("~")[0] == protein:
                            continue
                        for any_peptide, tuplist in value.items():
                            for each_tuple in tuplist:
                                small_tuplist.append(each_tuple[1])
                    max_tp_here = max(small_tuplist)

                up_list = []
                tp_list = []
                sd_list = []
                if correction is False:
                    ax.set_ylim(0, max_theo)
                    if peptide in statedic_of_pepdic_raw2[state]:
                        for up, tp in statedic_of_pepdic_raw2[state][peptide]:
                            up_list.append(up)
                            tp_list.append(tp)
                        for sd, tp in statedic_of_sddic_raw2[state][peptide]:
                            sd_list.append(sd)

                if correction is True:
                    ax.set_ylim(0, max_theo + 2)
                    if peptide in statedic_of_pepdic_cor[state]:
                        for up, tp in statedic_of_pepdic_cor[state][peptide]:
                            up_list.append(up)
                            tp_list.append(tp)
                        for sd, tp in statedic_of_sddic_cor[state][peptide]:
                            sd_list.append(sd)



                if tp_list != []:
                    if tp_list[0] == 0:
                        tp_list = tp_list[1:]
                        up_list = up_list[1:]
                        sd_list = sd_list[1:]
                    max_timepoint = max(tp_list)



                ax.tick_params(axis='x', labelsize=5)
                ax.tick_params(axis='y', labelsize=5)

                if (correction is False and peptide in statedic_of_pepdic_raw2[state]) or (correction is True and peptide in statedic_of_pepdic_cor[state]):
                    if tp_list != []:
                        if show_last is True:
                            pass
                        if show_last is False:
                            if tp_list[-1] == max_tp_here:
                                up_list = up_list[0:-1]
                                tp_list = tp_list[0:-1]
                                sd_list = sd_list[0:-1]

                        filtered_pairs = [(up, tp) for up, tp in zip(up_list, tp_list) if up != -99999]
                        if filtered_pairs:
                            up_list, tp_list = list(zip(*filtered_pairs))
                            up_list = list(up_list)
                            tp_list = list(tp_list)
                        else:
                            continue

                            
                        sd_list = [z if z != -99999 else np.nan for z in sd_list]
                
                        if correction is True:
                            up_list = [z * max_theo for z in up_list]
                            sd_list = [z * max_theo if z != np.nan else np.nan for z in sd_list]
                            
                        SE_list = [z / np.sqrt(float(SE_num_entry.get())) if z != np.nan else np.nan for z in sd_list]

                        for order, st in order_state_dic.items():
                            if st == state:
                                last_filled_position = (row, col)

                                if show_error_bars == False:
                                    ax.plot(tp_list, up_list, color=order_color_dic[order], linestyle=linestyle_in_use, linewidth = (linewidth_in_use/2))
                                else:
                                    try:
                                        ax.errorbar(tp_list, up_list, yerr=SE_list, color=order_color_dic[order], linestyle=linestyle_in_use, linewidth=(linewidth_in_use/2), capsize=(errorbar_capsize/2), elinewidth=(errorbar_linewidth/2), capthick=(errorbar_capthick/2))
                                    except:
                                        ax.plot(tp_list, up_list, color=order_color_dic[order], linestyle=linestyle_in_use, linewidth = (linewidth_in_use/2))

                                for x, y in zip(tp_list, up_list):
                                    ax.text(x, y, order_symbol_dic[order], color=order_color_dic[order], ha='center', va='center', fontsize=(order_size_dic[order]/3))


            end_of_page = False
            if a_vertical == True:
                if (row, col) == (7, 5):
                    end_of_page = True
            if a_horizontal == True:
                if (row, col) == (5, 7):
                    end_of_page = True
            if end_of_page == True and idx != 0:
                for (row, col) in empty_protein_boxes:
                    axes[row, col].set_xticks([])
                    axes[row, col].set_yticks([])
                empty_protein_boxes = []
                pdf_pages.savefig(fig)
                plt.close(fig)
                if idx < len(new_sorted_all_peptides) - 1:
                    if a_vertical is True:
                        fig, axes = plt.subplots(8, 6, figsize=(8.5, 11))
                    if a_horizontal is True:
                        fig, axes = plt.subplots(6, 8, figsize=(11, 8.5))
                    fig.text(0.5, 0.04, f"{x_enter.get()}", ha='center', va='center', fontsize=12)
                    fig.text(0.04, 0.5, f"{y_enter.get()}", ha='center', va='center', rotation='vertical', fontsize=12)
                    plt.subplots_adjust(wspace=0.4, hspace=0.6)
                    page_count += 1

        if a_vertical is True:
            max_row, max_col = 7, 5 
        if a_horizontal is True:
            max_row, max_col = 5, 7

        last_row, last_col = last_filled_position 

        for row in range(last_row, max_row + 1):  # +1 as range end is exclusive
            for col in range(last_col + 1 if row == last_row else 0, max_col + 1):  # +1 as range end is exclusive
                axes[row, col].axis('off')


        if (working_idx + 1) % 48 != 0 or idx == len(new_sorted_all_peptides) - 1:
            pdf_pages.savefig(fig)
        plt.close(fig)

        fig_legend = plt.figure(figsize=(8.5, 11))
        ax_legend = fig_legend.add_subplot(111)
        ax_legend.legend(handles=line_legend_entries, loc='center', frameon=True)
        ax_legend.axis('off')
        pdf_pages.savefig(fig_legend)
        plt.close(fig_legend)

        pdf_pages.close()

        increase_progress(2)





    except Exception as e:
        error_message = traceback.format_exc() 
        print(error_message)
        tk.messagebox.showerror("Unknown Error in Uptake Plots", error_message)          
       
                

def save_wb():
    global temp_file_path_excel2
    wb.remove(wb['Sheet'])
    
    def r_make_pretty_linearmap():
        for sheet_name in wb.sheetnames:
            if sheet_name.endswith("_colprdc"):
                sheet_to_remove = wb[sheet_name]
                wb.remove(sheet_to_remove)
        for sheet_name in wb.sheetnames:
            if sheet_name.endswith("_predicts"):
                source_sheet = wb[sheet_name]
                target_sheet_title = sheet_name.removesuffix("_predicts") + "_colprdc"
                target_sheet = wb.create_sheet(title=target_sheet_title)
                target_sheet.append([])
                for row in source_sheet.iter_rows():
                    row_data = [cell.value for cell in row]
                    row_data = [""] + row_data
                    target_sheet.append(row_data)
                    target_sheet.append([])
#                for row in target_sheet.iter_rows():
#                    if row[1].value is None:
#                        for cell in row:
#                            cell.border = Border()
#                            cell.fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type='solid')
#                        continue
#                    for i, cell in enumerate(row):
#                        if cell.value is None or cell.value == "":
#                            continue
#                        cell_v = cell.value
#                        if cell_v == 1:
#                            fill = PatternFill(start_color=f"{globals().get(f'p_col_{p_col_length}')}", end_color=f"{globals().get(f'p_col_{p_col_length}')}", fill_type='solid')
#                        if cell_v == 2:
#                            fill = PatternFill(start_color=f"{globals().get(f'p_col_{p_col_length-1}')}", end_color=f"{globals().get(f'p_col_{p_col_length-1}')}", fill_type='solid')
#                        if cell_v == 4:
#                            fill = PatternFill(start_color=f"{globals().get(f'd_col_{d_col_length}')}", end_color=f"{globals().get(f'd_col_{d_col_length}')}", fill_type='solid')
#                        if cell_v == 5:
#                            fill = PatternFill(start_color=f"{globals().get(f'd_col_{d_col_length-1}')}", end_color=f"{globals().get(f'd_col_{d_col_length-1}')}", fill_type='solid')
#                        if cell_v == 3:
#                            fill = PatternFill(start_color=b_col_abs, end_color=b_col_abs, fill_type='solid')
#                        if cell_v == 0:
#                            fill = PatternFill(start_color=d_col_gtz, end_color=d_col_gtz, fill_type='solid')
#                        cell.fill = fill
#                        
#                        if i == 0:
#                            pass
#                        
#                        elif i == 1:
#                            cell.border = Border(top=Side(border_style='thin', color='FF000000'),
#                                    bottom=Side(border_style='thin', color='FF000000'),
#                                    left=Side(border_style='thin', color='FF000000'))
#                        elif i == (len(row) - 1):
#                            cell.border = Border(top=Side(border_style='thin', color='FF000000'),
#                                    bottom=Side(border_style='thin', color='FF000000'),
#                                    right=Side(border_style='thin', color='FF000000'))
#                        else:
#                            cell.border = Border(top=Side(border_style='thin', color='FF000000'),
#                                    bottom=Side(border_style='thin', color='FF000000'))
#                        cell.number_format = ';;;'
           
        white_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type='solid')
        target_sheet_title = "localized differences"
        target_sheet = wb.create_sheet(title=target_sheet_title)
        target_sheet.append([])
        total_rows_used = 0
        last_row_index_list = []
        for sheet_name in wb.sheetnames:
            if sheet_name.endswith("_predicts"):
                source_sheet = wb[sheet_name]
                last_row_index = source_sheet.max_column
                last_row_index_list.append(last_row_index)
                
                dif_sheet_to_search = sheet_name.removesuffix("_predicts") + "_dif" 
                dif_sheet = wb[dif_sheet_to_search]
                dif_timepoints = list()
                for row in dif_sheet.iter_rows():
                    if row[0].value != "Timepoint" and row[0].value != 0 and row[0].value != None and row[0].value != "none" and row[0].value != " ":
                        dif_timepoints.append(row[0].value)
                        
                for row in dif_sheet.iter_rows(min_row=2, max_row=2):
                    row_sequence = [cell.value for cell in row][1:]
                for row in dif_sheet.iter_rows(min_row=1, max_row=1):
                    row_numbers = [cell.value for cell in row][1:]


                
                
                for row_index, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
                    for column_index, cell_value in enumerate(row, start=1):
                        # Get the cell in the target sheet
                        cell = target_sheet.cell(row=row_index + 2 + total_rows_used , column=column_index + 1)
                        cell.value = cell_value

                        # Color the cell based on its value
                        if cell_value == 1:
                            color = lcol1
                        elif cell_value == 2:
                            color = lcol2
                        elif cell_value == 3:
                            color = lcol3
                        elif cell_value == 4:
                            color = lcol4
                        elif cell_value == 5:
                            color = lcol5
                        elif cell_value == 0:
                            color = lcol0
                        elif cell_value == 6:
                            if lcol6 != False:
                                color = lcol6
                            else:
                                color = "000000"
                        elif cell_value == 7:
                            if lcol7 != False:
                                color = lcol7
                            else:
                                color = "000000"
                        elif cell_value == 8:
                            if lcol8 != False:
                                color = lcol8
                            else:
                                color = "000000"
                        elif cell_value == 9:
                            if lcol9 != False:
                                color = lcol9
                            else:
                                color = "000000"
                        else:
                            color = "000000"
                        
                        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        cell.fill = fill
                        cell.number_format = ';;;'
                        
                        if column_index + 1 == 2:
                            cell.border = Border(bottom=Side(border_style='thin', color='FF000000'),
                                    left=Side(border_style='thin', color='FF000000'),
                                    top=Side(border_style='thin', color='FF000000'))
                        elif column_index == last_row_index:
                            cell.border = Border(right=Side(border_style='thin', color='FF000000'),
                                    top=Side(border_style='thin', color='FF000000'),
                                    bottom=Side(border_style='thin', color='FF000000'))
                        else:
                            cell.border = Border(top=Side(border_style='thin', color='FF000000'),
                                    bottom=Side(border_style='thin', color='FF000000'))
                
               
                for row in target_sheet.iter_rows(min_row=total_rows_used+1, max_row=total_rows_used+1):
                    row[0].value = dif_sheet_to_search.removesuffix("_dif")
                    row[0].font = courier_new_style
                
                timepoint_number_increment = 0
                while timepoint_number_increment < len(dif_timepoints):
                    cell = target_sheet.cell(row=total_rows_used+3+timepoint_number_increment, column=1)
                    cell.value = dif_timepoints[timepoint_number_increment]
                    cell.font = courier_new_style
                    timepoint_number_increment += 1


                num_increment = 0
                while num_increment < len(row_numbers):
                    cell = target_sheet.cell(row=total_rows_used+1, column=2+num_increment)
                    cell.value = row_numbers[num_increment]
                    cell.font = size_5_courier_new_style
                    cell.fill = white_fill
                    num_increment += 1
                
                
                seq_increment = 0
                while seq_increment < len(row_sequence):
                    cell = target_sheet.cell(row=total_rows_used+2, column=2+seq_increment)
                    cell.value = row_sequence[seq_increment]
                    cell.font = courier_new_style
                    cell.fill = white_fill
                    seq_increment += 1
                        
                final_column = max(last_row_index_list)
                for col in range(2, final_column+2):
                    target_sheet.column_dimensions[get_column_letter(col)].width = 2.1
                    
                
                total_rows_used += 3
                total_rows_used += len(dif_timepoints)
                
        for sheet_name in wb.sheetnames:
            if sheet_name.endswith("_predicts"):
                ws = wb[sheet_name]
                wb.remove(ws)
            if sheet_name.endswith("_colprdc"):
                ws = wb[sheet_name]
                wb.remove(ws)
                

                        
                        

                        
                    
                
                    
    
    def get_user_title():
        #print("getting user title")
        if heatmap_bt_on:
            r_make_pretty_linearmap()
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            if sheet_name.endswith("_cond") or sheet_name.endswith("_dif"):
                if enumerate_residues_var.get() == 0:
                    if sheet_name.endswith("_cond"):
                        for i, column in enumerate(sheet.columns):
                            sheet.column_dimensions[column[0].column_letter].width = "2.94"
                        for i, row in enumerate(sheet.rows):
                            sheet.row_dimensions[row[0].row].height = "30"
                        for row in sheet.iter_rows():
                            for cell in row:
                                if cell.row >= 3 and cell.column >= 2:  
                                    if cell.font:
                                        new_font = Font(name=cell.font.name, size=16, color=cell.font.color)
                                        cell.font = new_font

                    for row in sheet.iter_rows(min_row=2, max_row=2):
                        sheet.row_dimensions[2].height = 20
                        for cell in row:
                            cell.font = Font(name="Courier New", size=20)
                    for row in sheet.iter_rows(min_row=1, max_row=1):
                        sheet.row_dimensions[1].height = 20
                        row[1].font = Font(name="Courier New", size=14)
                        for i, cell in enumerate(row[2:], start=2):
                            if cell.value == None:
                                continue
                            elif cell.value % 10 == 0:
                                try:
                                    row[i-1].value = cell.value
                                    sheet.merge_cells(start_row=cell.row, start_column=row[i-1].column, end_row=cell.row, end_column=row[i+1].column)
                                    row[i-1].alignment = Alignment(horizontal="center")
                                    row[i-1].font = Font(name="Courier New", size=14)  
                                except:
                                    print("Excepting during cell numbering cell merge")
                            else:
                                cell.value = None
                    for cell in sheet["A"]:
                        cell.font = Font(name="Courier New", size=14, bold=True)
                        cell.alignment = Alignment(horizontal="center") 
                        sheet.column_dimensions["A"].width = 15  
        
        global mapviewer_bt
        try:
            mapviewer_bt.destroy()
        except:
            pass
    
        wb_tit = filedialog.asksaveasfilename(filetypes=[("Excel Files", "*.xlsx")])
        if wb_tit:
            if not wb_tit.endswith(".xlsx"):
                wb_tit += ".xlsx"
            try:
                wb.save(wb_tit)
                tk.messagebox.showinfo("Save Workbook", f"The workbook has been saved as '{wb_tit}'.")
            except PermissionError as e:
                tk.messagebox.showerror("Save Workbook", f"PermissionError occurred: {e}\nPlease close this file or use another name")
        else:
            tk.messagebox.showwarning("Save Workbook", "No file path selected. The workbook was not saved.")

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet.page_setup.orientation = 'landscape'
        sheet.page_margins.left = 0
        sheet.page_margins.right = 0
        sheet.page_margins.top = 0.2
        if sheet_name.endswith("_cond"):
            for i, column in enumerate(sheet.columns):
                sheet.column_dimensions[column[0].column_letter].width = "2.504"
            for row in sheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.font = Font(size=6)
    
    temp_file_path_excel2 = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=True).name
    wb.save(temp_file_path_excel2)
        
    atexit.register(os.remove, temp_file_path_excel2)
    
    increase_progress(1)


    run_bt.config(state="normal")
    run_bt.config(relief="raised")

    global tit_bt
    tit_bt = tk.Button(window, text="Save Workbook", command=get_user_title)
    tit_bt.place(x=1290, y=260)
    

def save_pdf():
    increase_progress(1)
    global pdf_bt
    def get_pdf_title():
        pdf_tit = filedialog.asksaveasfilename(defaultextension=".pdf",
                                            filetypes=[("PDF files", "*.pdf")])
        existing_file_path = "uptake_plots.pdf"
        if pdf_tit:
            if not pdf_tit.endswith(".pdf"):
                pdf_tit += ".pdf"
            shutil.copy(existing_file_path, pdf_tit)
            tk.messagebox.showinfo("Save PDF", f"The PDF has been saved as '{pdf_tit}'.")
        else:
            tk.messagebox.showwarning("Save PDF", "No file path selected. The PDF was not saved.")
    
    global pdf_bt
    pdf_bt = tk.Button(window, text="Save Uptake Plots", command=get_pdf_title)
    pdf_bt.place(x=1285, y=290)
    
    run_bt.config(state="normal")
    run_bt.config(relief="raised")
    
    
    
def on_closing_mapviewer():
    global mapviewer_open
    mapviewer_open = False
    mapviewer.destroy()
    
def create_mapviewer_bt():
    global mapviewer_bt
    mapviewer_bt = tk.Button(window, text="Localized Difference Editor", command=open_mapviewer)
    mapviewer_bt.place(x=1260, y=230)
    
    run_bt.config(state="normal")
    run_bt.config(relief="raised")

mapviewer = None
mapviewer_open = False
def open_mapviewer():
    global state_dropdown, mapviewer, timepoint_dropdown, mapviewer_open
    if mapviewer_open:
        user_choice = tk.messagebox.askyesno("Localized Differences Editor", "Localized Differences Editor may already be open. Do you want to close and open a new window?", default='no')
        if user_choice:
            mapviewer_open = False
            mapviewer.destroy()
        else:
            mapviewer.lift()
            return
    mapviewer = tk.Toplevel(window)  # Create a new window for the popup menu
    mapviewer.geometry("1200x820")
    mapviewer.title("Localized Differences Editor")
    mapviewer_open = True
    mapviewer.protocol("WM_DELETE_WINDOW", on_closing_mapviewer)
    

    
    state_dropdown = ttk.Combobox(mapviewer, values=difference_titles, width=35)
    state_dropdown.set(difference_titles[0])
    state_dropdown.bind("<<ComboboxSelected>>", make_new_dropdowns)
    state_dropdown.bind("<<ComboboxSelected>>", create_pictures)
    state_dropdown.place(x=400, y=30)
    
    state = difference_titles[0]
    difference_tuple_index = difference_titles.index(state)
    difference_tuple = difference_states[difference_tuple_index]
    global common_elements
    common_elements = [x for x in s_timepoints[difference_tuple[0]] if x in s_timepoints[difference_tuple[1]] and x != 0]
    
    timepoint_dropdown = ttk.Combobox(mapviewer, values=common_elements, width=10)
    timepoint_dropdown.set(common_elements[0])
    timepoint_dropdown.bind("<<ComboboxSelected>>", create_pictures)
    timepoint_dropdown.place(x=700, y=30)
    
    
    create_pictures()
    
    
    
def make_new_dropdowns():
    global timepoint_dropdown
    state = state_dropdown.get()
    difference_tuple_index = difference_titles.index(state)
    difference_tuple = difference_states[difference_tuple_index]
    global common_elements
    common_elements = [x for x in s_timepoints[difference_tuple[0]] if x in s_timepoints[difference_tuple[1]] and x != 0]
    
    timepoint_dropdown = ttk.Combobox(mapviewer, values=common_elements, width=10)
    timepoint_dropdown.set(common_elements[0])
    timepoint_dropdown.bind("<<ComboboxSelected>>", create_pictures)
    timepoint_dropdown.place(x=700, y=30)
    

def create_pictures(event=None):
    current_state = state_dropdown.get()
    timepoint = float(timepoint_dropdown.get())
    sheet_to_search = current_state + "_cond"
    sheet_to_search = sheet_to_search
    ws = wb[sheet_to_search]
    
    
    
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        if row[0] == timepoint:
            timepoint_start = i
            break

    
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if row[0] is not None and row[0] != "":
            timepoint_1_index = i
            timepoint_1 = row[0]
            break
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if row[0] is not None and row[0] != timepoint_1 and row[0] != "":
            timepoint_2_index = i
            break
    

    if timepoint_1_index and timepoint_2_index:
        difference_in_timepoints = timepoint_2_index - timepoint_1_index
    else:
        print("Timepoints 1 and 2 not defined")
        

    
    def on_mouse_wheel(event):
        if event.delta > 0:
            move = -1
        elif event.delta < 0:
            move = 1
        else:
            move = 0
            
        h_canvas.xview_scroll(move, "units")
        m_canvas.xview_scroll(move, "units")
        v_canvas.xview_scroll(move, "units")
    
    h_canvas = tk.Canvas(mapviewer, bg="white")
    h_canvas.place(relx=0.1, rely=0.1, relwidth=0.8, relheight=0.15)
    h_frame = tk.Frame(h_canvas)
    h_canvas.create_window((0, 0), window=h_frame, anchor="nw")

    
    
    m_canvas = tk.Canvas(mapviewer, bg="white")
    m_canvas.place(relx=0.1, rely=0.28, relwidth=0.8, relheight=0.68)  # Adjust relheight as needed
    frame = tk.Frame(m_canvas)
    m_canvas.create_window((0, 0), window=frame, anchor="nw")
    
    v_canvas = tk.Canvas(mapviewer, bg="white")
    v_canvas.place(relx=0.1, rely=0.2, relwidth=0.8, relheight=0.1)
    v_frame = tk.Frame(v_canvas)
    v_canvas.create_window((0, 0), window=v_frame, anchor="nw")
    

    scrollbar = tk.Scrollbar(mapviewer, orient="horizontal")
    scrollbar.place(relx=0.5, rely=0.99, relwidth=0.8, anchor="s")

    # Configure both canvases to use the same horizontal scrollbar
    h_canvas.configure(xscrollcommand=scrollbar.set)
    m_canvas.configure(xscrollcommand=scrollbar.set)
    v_canvas.configure(xscrollcommand=scrollbar.set)

    # Set the scrollbar's command to control both canvases
    scrollbar.config(command=lambda *args: (h_canvas.xview(*args), m_canvas.xview(*args), v_canvas.xview(*args)))
    
    #mapviewer.bind("<MouseWheel>", on_mouse_wheel)



    
    temp_pdf_data_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    temp_pdf_data_file_path = temp_pdf_data_file.name
    temp_pdf_data_file.close()
    
    temp_pdf_header_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    temp_pdf_header_file_path = temp_pdf_header_file.name
    temp_pdf_header_file.close()
    
    
    book = xw.Book(temp_file_path_excel2)      
    sheet = book.sheets(sheet_to_search) 
    
    min_row = timepoint_start
    max_row = timepoint_start + difference_in_timepoints -1
    min_col = 2
    max_col = all_predicted_labels_lengths[current_state] + 1 
    if max_row - min_row > 23:
        row_dif = str((max_row - min_row)-27)
        tk.Label(mapviewer, text=f"{row_dif} additional row(s) could not fit in this interface. Please check excel sheet to see data in additional rows").place(x=350, y=7)
        max_row = min_row +23
    min_excel_cell = excel_cell(min_row, min_col)
    max_excel_cell = excel_cell(max_row, max_col)
    data_excel_range = f"{min_excel_cell}:{max_excel_cell}"

    
    sheet.range(data_excel_range).api.ExportAsFixedFormat(0, temp_pdf_data_file_path)
    
    
    min_row = 1
    max_row = 2
    min_col = 2
    max_col = all_predicted_labels_lengths[current_state] + 1 
    min_excel_cell = excel_cell(min_row, min_col)
    max_excel_cell = excel_cell(max_row, max_col)
    header_excel_range = f"{min_excel_cell}:{max_excel_cell}"
    
    sheet.range(header_excel_range).api.ExportAsFixedFormat(0, temp_pdf_header_file_path)
    
    book.close()
    
    
    ws = wb[current_state + "_predicts"]
    timepoint_index = common_elements.index(timepoint)
    for row in ws.iter_rows(min_row=timepoint_index + 1, max_row=timepoint_index + 1, values_only=True):
        all_predicts = list(row)
    num_cells_in_last_frame = len(all_predicts) % 54
    num_invisible_squares = 54 - num_cells_in_last_frame
    for _ in range(0, num_invisible_squares):
        all_predicts.append("x")
    
    
    global color_mapping
    new_items_list = [lcol0, lcol1, lcol2, lcol3, lcol4, lcol5, lcol6, lcol7, lcol8, lcol9]
    false_item_index_list = []
    for i, item in enumerate(new_items_list):
        if item == False:
            false_item_index_list.append(i)

    color_mapping = {}
    for i, item in enumerate(new_items_list):
        if i in false_item_index_list:
            color_mapping[i] = "#000000"
            continue
        color_mapping[i] = "#" + str(item)
    color_mapping["x"] = "#FFFFFF"



    color_indexes = []
    color_indexes_possible = [7, 6, 2, 1, 0, 4, 5, 8, 9, 3]
    for i, item in enumerate(color_indexes_possible):
        if item in false_item_index_list:
            continue
        if color_indexes_possible[i] == 3:
            continue
        color_indexes.append(color_indexes_possible[i])
    
    
    #fig, ax = plt.subplots(figsize=(len(color_indexes_possible), 2))
    fig, ax = plt.subplots(figsize=(6, 2))
    
    
        
    xpos = 0
    for n in color_indexes:
        color = color_mapping[n]
        square = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
        ax.plot([xpos+0.5, xpos+0.5], [1, 1.3], color='black', linewidth=1)
        ax.text(xpos+0.5, 1.35, str(n), ha='center', va='bottom', fontsize=12)
        xpos += 1
        ax.add_patch(square)

    xpos += 1
    color = color_mapping[3]
    square = patches.Rectangle((xpos, 0), 1, 1, linewidth=1, edgecolor='black', facecolor=color)
    ax.plot([xpos+0.5, xpos+0.5], [1, 1.3], color='black', linewidth=1)
    ax.text(xpos+0.5, 1.35, str(3), ha='center', va='bottom', fontsize=12)
    ax.add_patch(square)
    
    
    ax.spines['top'].set_visible(False)
    ax.spines['bottom'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.spines['right'].set_visible(False)

    ax.set_aspect('equal')
    ax.set_xticks([])
    ax.set_yticks([])
    
    
    def scale_figure(figure, scale_factor):
        old_size = figure.get_size_inches()
        new_size = (old_size[0] * scale_factor, old_size[1] * scale_factor)
        figure.set_size_inches(new_size)
        
        
    scale_factor = 0.3  
    scale_figure(fig, scale_factor)
    
   # fig.savefig('./RecentLegends/linear_map_scale.png', dpi=300)
    
    tk_bg_color_rgb = (240 / 255, 240 / 255, 240 / 255)
    fig.patch.set_facecolor(tk_bg_color_rgb)

    
    legend_canvas = FigureCanvasTkAgg(fig, master=mapviewer)
    legend_canvas_widget = legend_canvas.get_tk_widget()
    legend_canvas_widget.place(x=100, y=5)
    

    plt.close()
    
    
    fig2, ax2 = r_make_legend2(False)
    scale_factor = 0.4
    scale_figure(fig2, scale_factor)
    fig2.patch.set_facecolor(tk_bg_color_rgb)
                 
    legend2_canvas = FigureCanvasTkAgg(fig2, master=mapviewer)
    legend2_canvas_widget = legend2_canvas.get_tk_widget()
    legend2_canvas_widget.place(x=800, y=5)
    
    tk.Label(mapviewer, text="To edit labels, left click on the number of the residue you wish to edit and enter the intended value. Numbers can also be right clicked to paste the previously entered value.").place(x=130, y=60)
    
    tk.Label(mapviewer, text=f"Significance Cutoff used for Predictions = {future_linear_map_multiplier} (this can be changed by using a different color scheme)").place(x=270, y=5)
    

    
    tk.Label(mapviewer, text="Number Code:").place(x=15, y=5)
    
    tk.Label(mapviewer, text="0 - Insignificant", font=("Arial", 8)).place(x=12, y=25)
    tk.Label(mapviewer, text="Difference", font=("Arial", 8)).place(x=20, y=40)
    
    tk.Label(mapviewer, text="1 - Questionable", font=("Arial", 8)).place(x=12, y=60)
    tk.Label(mapviewer, text="Protection", font=("Arial", 8)).place(x=20, y=75)
    
    tk.Label(mapviewer, text="2 - Significant", font=("Arial", 8)).place(x=12, y=95)
    tk.Label(mapviewer, text="Protection", font=("Arial", 8)).place(x=20, y=110)
    
    tk.Label(mapviewer, text="3 - No Coverage", font=("Arial", 8)).place(x=12, y=130)
    
    tk.Label(mapviewer, text="4 - Questionable", font=("Arial", 8)).place(x=12, y=150)
    tk.Label(mapviewer, text="Deprotection", font=("Arial", 8)).place(x=20, y=165)
    
    tk.Label(mapviewer, text="5 - Significant", font=("Arial", 8)).place(x=12, y=185)
    tk.Label(mapviewer, text="Deprotection", font=("Arial", 8)).place(x=20, y=200)
    
    tk.Label(mapviewer, text="6/7/8/9 - Manual", font=("Arial", 8)).place(x=12, y=220)
    tk.Label(mapviewer, text="(Never predicted)", font=("Arial", 8)).place(x=12, y=235)
    
     
    

    

    
    pdf_data_document = fitz.open(temp_pdf_data_file_path)
    pdf_header_document = fitz.open(temp_pdf_header_file_path)
    for page_num, page in enumerate(pdf_data_document):
        page = pdf_data_document.load_page(page_num)
        page_pil = page.get_pixmap()
        page_image = Image.frombytes("RGB", [page_pil.width, page_pil.height], page_pil.samples)
        width, height = page_image.size
        page_image = page_image.resize((int(width*1.04), int(height*1.04)), Image.Resampling.LANCZOS)
        
        width, height = page_image.size
        left_margin = 0  # Adjust this value if needed
        top_margin = 0   # Adjust this value if needed
        right_margin = 12
        bottom_margin = 0  # Adjust this value if needed
        page_image = page_image.crop((left_margin, top_margin, width - right_margin, height - bottom_margin))
        
        page_image_tk = ImageTk.PhotoImage(page_image)
        map_image = tk.Label(frame, image=page_image_tk)
        map_image.grid(row=0, column=page_num)
        map_image.config(image=page_image_tk)
        map_image.image = page_image_tk
        
        page = pdf_header_document.load_page(page_num)
        page_pil = page.get_pixmap()
        page_header_image = Image.frombytes("RGB", [page_pil.width, page_pil.height], page_pil.samples)
        width, height = page_header_image.size
        page_header_image = page_header_image.resize((int(width*1.04), int(height*1.04)), Image.Resampling.LANCZOS)
        
        width, height = page_header_image.size
        left_margin = 0  # Adjust this value if needed
        top_margin = 0   # Adjust this value if needed
        right_margin = 12
        bottom_margin = 0  # Adjust this value if needed
        page_header_image = page_header_image.crop((left_margin, top_margin, width - right_margin, height - bottom_margin))
        
        page_header_image_tk = ImageTk.PhotoImage(page_header_image)
        map_header_image = tk.Label(h_frame, image=page_header_image_tk)
        map_header_image.grid(row=0, column=page_num)
        map_header_image.config(image=page_header_image_tk)
        map_header_image.image = page_header_image_tk
        
        
        if page_num == 0:
            global cell_sets, square_item_sets, square_canvas_sets
            cell_sets = []
            square_item_sets = []
            square_canvas_sets = []
            
        
        a_frame = tk.Frame(v_frame, width = width)
        if page_num == 0:
            a_frame.grid(row=1, column=page_num, padx=2)
        else:
            a_frame.grid(row=1, column=page_num, padx=2) #pre, padx = 9
            
        cells = [tk.Label(a_frame, text=value, padx=2) for value in all_predicts[0+(page_num*54):54+(page_num*54)]]
        square_canvas = tk.Canvas(a_frame, width=15 * len(cells), height=15)
        square_canvas.grid(row=1, column=0, columnspan=54, sticky="w")  # Position the canvas to the left of the cells
        square_items = [square_canvas.create_rectangle(i * 15, 0, (i + 1) * 15, 15, fill="green") for i in range(len(cells))]
        for i, cell in enumerate(cells):
            cell.grid(row=0, column=i, sticky="w")
        for i, cell in enumerate(cells):
            cell_value = all_predicts[i + (page_num * 54)]
            square_color = color_mapping.get(cell_value, "pink")
            square_canvas.itemconfig(square_items[i], fill=square_color)
            
            cell.bind("<Button-1>", lambda event, index=i, page_num=page_num: update_cell(event, index, page_num, ws, timepoint_index))
            cell.bind("<Button-3>", lambda event, index=i, page_num=page_num: copy_last_saved_value(event, index, page_num, ws, timepoint_index))# Right-click to copy last saved value
        cell_sets.append(cells)
        square_item_sets.append(square_items)
        square_canvas_sets.append(square_canvas)

        

    # Force the mapviewer window to redraw
    mapviewer.update()
    pdf_data_document.close()
    pdf_header_document.close()
    
    try:
        os.remove(temp_pdf_data_file_path)
        os.remove(temp_pdf_header_file_path)
    except:
        tk.messagebox.showerror("Excel Error", "Excel Error: Cannot access Excel. Please make sure your computer has access to Excel and try again.")
        return
    frame.update_idletasks()
    h_frame.update_idletasks()
    v_frame.update_idletasks()

    m_canvas.configure(scrollregion=m_canvas.bbox("all"))
    h_canvas.config(scrollregion=h_canvas.bbox("all"))
    v_canvas.config(scrollregion=v_canvas.bbox("all"))
    
#    def on_canvas_scroll(event):
#        h_canvas.xview_scroll(-1 * (event.delta // 120), "units")
#        m_canvas.xview_scroll(-1 * (event.delta // 120), "units")
#        v_canvas.xview_scroll(-1 * (event.delta // 120), "units")
#
#
#    h_canvas.bind("<MouseWheel>", on_canvas_scroll)
#    m_canvas.bind("<MouseWheel>", on_canvas_scroll)
#    v_canvas.bind("<MouseWheel>", on_canvas_scroll)
        

        
        
    
    # Create a button to retrieve values
    retrieve_button = tk.Button(mapviewer, text="Export to Pymol", command=lambda: export_to_pymol(ws, timepoint_index, current_state))
    retrieve_button.place(relx=0.9, rely=0.85)
    
#    save_linear_map_bt = tk.Button(mapviewer, text="Save Values", command=lambda: retrieve_values(ws, timepoint_index))
#    save_linear_map_bt.place(relx=0.9, rely=0.8)
    
def export_to_pymol(ws, timepoint_index, current_state):
    all_values = retrieve_values(ws, timepoint_index)
    if current_state in new_dic_of_dif_list.keys():
        difpair = new_dic_of_dif_list[current_state]
        first_dif = difpair[0]
        current_protein = first_dif.split("~")[0]
    pdb_file_path = filedialog.askopenfilename(title="Select a PDB File", filetypes=[("PDB Files", "*.pdb")])
    parser = PDB.PDBParser()
    structure = parser.get_structure("PDB_structure", pdb_file_path)
    chains = [chain.id for model in structure for chain in model]
    chain_dic = {}
    if len(chains) > 1:
        for chain in chains:
            chain_dic[chain] = False
            user_choice = tk.messagebox.askyesno(f"Color Chain? {chains}", f"{chain}")
            if user_choice:
                chain_dic[chain] = True
            else:
                chain_dic[chain] = False
    else:
        for chain in chains:
            chain_dic[chain] = True
    
    compiled_new_commands = []
    for chain_id, tf in chain_dic.items():
        if tf is False:
            continue
        pdb_sequence, first_residue_number = extract_sequence_and_first_residue_from_pdb(pdb_file_path, chain_id) 

        your_sequence = generate_best_fit_sequence(current_protein)



        alignments = align_sequences(pdb_sequence, your_sequence)[0]
        print(alignments)
        index_mapping = map_indices(alignments, first_residue_number)


        color_commands = []
        color_mapping2 = {}
        for value, hex_color in color_mapping.items():
            color_name = f"custom_color_{value}"
            rgb_color = hex_to_rgb(hex_color)
            color_command = f"set_color {color_name}, {rgb_color}"
            color_commands.append(color_command)
            color_mapping2[value] = color_name

        new_commands = generate_pymol_commands(index_mapping, all_values, color_mapping2, chain_id)
        if new_commands == False:
            return
        compiled_new_commands += new_commands
    
    commands = [f"load {pdb_file_path}"] + color_commands + [f"color {color_mapping2[3]}, polymer.protein"] + compiled_new_commands + ["hide (solvent)"]

    with open("recent_color_mapping.pml", "w") as file:
        for command in commands:
            file.write(command + "\n")

    try:
    #os.startfile("recent_color_mapping.pml")
        open_file("recent_color_mapping.pml")
    except:
        tk.messagebox.showerror("PyMol Error", "Error, could not find PyMOL. Please add PyMOL to PATH")
    
        
    
def hex_to_rgb(hex_color):
    hex_color = hex_color.lstrip('#')
    lv = len(hex_color)
    return tuple(int(hex_color[i:i + lv // 3], 16) / 255.0 for i in range(0, lv, lv // 3))

protein_pdb_dictionary = {}

def custom_warning_handler(message, category, filename, lineno, file=None, line=None):
    if issubclass(category, PDBConstructionWarning):
        print("ERROR: There is a gap in the .pdb file sequence. These residues cannot be colored")
    else:
        print(f"Standard Warning: {message}")
warnings.showwarning = custom_warning_handler



def extract_sequence_and_first_residue_from_pdb(pdb_file_path, chain_id):
    parser = PDBParser()
    structure = parser.get_structure('PDB_structure', pdb_file_path)
    for model in structure:
        for chain in model:
            if chain.id == chain_id:
                residues = [residue for residue in chain if residue.id[0] == ' ']
                if residues:
                    first_residue_number = residues[0].id[1]
                    sequence = ''.join([seq1(residue.resname) for residue in residues])
                    return sequence, first_residue_number
                else:
                    return '', None

def align_sequences(seq1, seq2):
    aligner = Align.PairwiseAligner()

    # Set the alignment method and scoring
    aligner.mode = 'global'  # Use 'local' for local alignment
    aligner.match_score = 1  # Score for identical characters
    aligner.mismatch_score = -1  # Penalty for non-identical characters
    aligner.open_gap_score = -0.5  # Penalty for opening a gap
    aligner.extend_gap_score = -0.1  # Penalty for extending a gap

    # Perform the alignment
    alignments = aligner.align(seq1, seq2)
    return alignments

def map_indices(alignments, first_residue_number):
    target_alignment = alignments[0]
    q_alignment = alignments[1]
    
    
    target_index = 0
    query_index = 0
    index_mapping = {}
    
    both_started = False
    for target_char, query_char in zip(target_alignment, q_alignment):        
        if target_char != '-':
            if query_char != '-':
                both_started = True
                # Both characters are not gaps
                index_mapping[query_index] = target_index + first_residue_number
            query_index += 1
            target_index += 1
        elif query_char != '-':
            # Only target character is a gap
            query_index += 1
            if both_started == True:
                target_index += 1
                
    return index_mapping





def generate_pymol_commands(mapping, all_values, color_mapping2, chain_id):
    new_commands = []
    for index, value in enumerate(all_values):
        if index in mapping:
            pdb_index = mapping[index]
            try:
                color = color_mapping2[value]
            except:
                tk.messagebox.showerror("Color Error", "At least one residue has been labeled with a dissalowed number. Please make sure all residues are labelled with a number available in the legend and try again.")
                mapviewer.focus_set()
                return False
            new_commands.append(f"color {color}, chain {chain_id} and resi {pdb_index}")
    return new_commands

        
        
def update_cell(event, cell_index, page_num, ws, timepoint_index):
    cells = cell_sets[page_num]
    cell_value = cells[cell_index].cget("text")
    
    def save_value(new_value, cells, page_num):
        global edit_cell_window, last_saved_value  # Declare global variables
        cells[cell_index].config(text=new_value)
        if edit_cell_window:
            edit_cell_window.destroy()  # Close the existing edit cell window
        edit_cell_window = None  # Reset the edit_cell_window reference
        last_saved_value = new_value  # Update the last saved value
        update_squares(cells, page_num)  # Update the squares when the cell value changes

    def edit_cell(event, cells, page_num):
        global edit_cell_window  # Declare global variables
        if edit_cell_window:
            edit_cell_window.destroy()  # Close the existing edit cell window
        edit_cell_window = tk.Toplevel()
        edit_cell_window.title("Edit Cell")
        edit_cell_window.geometry("200x100")
        
        new_value_entry = tk.Entry(edit_cell_window)
        new_value_entry.pack(pady=10)
        new_value_entry.focus_set()  # Set focus on the entry widget
        
        #save_button = tk.Button(edit_cell_window, text="Save", command=lambda: save_value(new_value_entry.get(), cells, page_num))
        save_button = tk.Button(edit_cell_window, text="Save", command=lambda: [save_value(new_value_entry.get(), cells, page_num), retrieve_values(ws, timepoint_index)])
        save_button.pack()
        tk.Label(edit_cell_window, text="(Or Press Enter)").pack()

        # Bind the "Enter" key to trigger the Save button
        edit_cell_window.bind("<Return>", lambda event: save_button.invoke())
    
    edit_cell(event, cells, page_num)


def retrieve_values(ws, timepoint_index):
    all_values = []
    for page_cells in cell_sets:
        current_values = [cell.cget("text") for cell in page_cells]
        all_values.extend(current_values)
        new_all_values = []
        for item in all_values:
            if item == "x":
                new_all_values.append(item)
            else:
                new_all_values.append(int(item))
        all_values = new_all_values
    x = 1
    while x > 0:
        if all_values[-1] == "x":
            all_values.pop()
        else:
            break
    

    for i, cell in enumerate(ws[timepoint_index+1]):
        cell.value = all_values[i]
    return all_values
        
            
#            
#    for row in ws.iter_rows(min_row=timepoint_index + 1, max_row=timepoint_index + 1, values_only=True):
#        row[i].value = None
#        
#        for i, val in enumerate(all_values):
#            row[i] = val

def copy_last_saved_value(event, cell_index, page_num, ws, timepoint_index):
    global last_saved_value  # Declare global variable
    cells = cell_sets[page_num]
    cells[cell_index].config(text=last_saved_value)
    update_squares(cells, page_num)  # Update the squares when copying the last saved value
    retrieve_values(ws, timepoint_index)
    


def update_squares(cells, page_num):
    square_canvas = square_canvas_sets[page_num]
    square_items = square_item_sets[page_num]
    for i, cell in enumerate(cells):
        square_color = color_mapping.get(int(cell.cget("text")), "pink")
        square_canvas.itemconfig(square_items[i], fill=square_color)



# Global variable to store the current edit cell window and the last saved value
edit_cell_window = None
last_saved_value = "1"



    
    
def excel_cell(row, col):
    """
    Convert row and column indices to Excel-style cell reference.
    :param row: Row index (1-based)
    :param col: Column index (1-based)
    :return: Excel-style cell reference (e.g., "A1", "B2", "C3", etc.)
    """
    col_letter = ""
    while col > 0:
        col, remainder = divmod(col - 1, 26)
        col_letter = chr(65 + remainder) + col_letter
    return col_letter + str(row)


def generate_best_fit_sequence(protein):
    protein_and_squiggle = protein + "~"
    for state in states:
        if state.startswith(protein_and_squiggle):
            new_sequence = {}
            for i, peptide in enumerate(peplist[state]):
                start = startvallist[state][i]
                for k, residue in enumerate(peptide, start=start):
                    new_sequence[k] = residue

#            min_num = startvallist[state][0]
#            max_num = startvallist[state][-1]
#
#            for i in range(min_num, max_num):
#                if i not in new_sequence.keys():
#                    new_sequence[i] = "-"


            # Sort the keys of the dictionary
            sorted_keys = sorted(new_sequence.keys())

            # Concatenate the values in the sorted order
            linear_sequence = ''.join(new_sequence[key] for key in sorted_keys)

            return linear_sequence
            
            
def temp_adjust_rate_calculation(k, tempK, activation_energy):
    #krc(T) = krc(293T)exp(-Ea[1/T - 1/293]/R)
    
    activation_energy = 4184*activation_energy ## convert to SI units
    
    k = k * np.exp(-1 *((activation_energy* ((1/tempK) - (1/293))) / 8.314))
        
    return k

def calculate_residue_rate(i, resL, resR, pH, tempK, log10_k_acid_ref, log10_k_base_ref, log10_k_water_ref, last):
    
    H = 10 ** (-1 * pH)
    pOH = 14 - pH
    OH = 10 ** (-1 * pOH)
    
    
    log10_k_acid = log10_k_acid_ref + resL.AR + resR.AL - pH    
    log10_k_base = log10_k_base_ref + resL.BR + resR.BL - pOH
    log10_k_water = log10_k_water_ref + resL.BR + resR.BL

    if i == 1:
        log10_k_acid += n_term.AR
        log10_k_base += n_term.BR
        log10_k_water += n_term.BR
        
    if last == True:
        log10_k_acid += c_term_COOH.AL
        log10_k_base += c_term_COOH.BL
        log10_k_water += c_term_COOH.BL
    
    k_acid = 10 ** log10_k_acid
    k_base = 10 ** log10_k_base
    k_water = 10 ** log10_k_water
    
    k_acid = temp_adjust_rate_calculation(k_acid, tempK, 14)
    k_base = temp_adjust_rate_calculation(k_base, tempK, 17)
    k_water = temp_adjust_rate_calculation(k_water, tempK, 19)
    
    k_rc = k_acid + k_base + k_water
    
    return k_rc

class residue:
    __slots__ = ("AL", "AR", "BL", "BR")
    AL: float
    AR: float
    BL: float
    BR: float
    
    def __init__(self, AL, AR, BL, BR):
        self.AL = AL
        self.AR = AR
        self.BL = BL
        self.BR = BR
        
ala = residue(AL = 0, AR = 0, BL = 0, BR = 0)
arg = residue(AL = -0.59, AR = -0.32, BL = 0.08, BR = 0.22)
asn = residue(AL = -0.58, AR = -0.13, BL = 0.49, BR = 0.32)
asp_COO = residue(AL = 0.9, AR = 0.58, BL = 0.1, BR = -0.18)
asp_COOH = residue(AL = -0.9, AR = -0.12, BL = 0.69, BR = 0.6)
cys = residue(AL=-0.54, AR = -0.46, BL = 0.62, BR = 0.55)
gly = residue(AL = -0.22, AR = 0.22, BL = -0.03, BR = 0.17)
gln = residue(AL = -0.47, AR = -0.27, BL = 0.06, BR = 0.2)
glu_COO = residue(AL = -0.9, AR = 0.31, BL = -0.11, BR = -0.15)
glu_COOH = residue(AL = -0.6, AR = -0.27, BL = 0.24, BR = 0.39)
his_H = residue(AL = -0.8, AR = -0.51, BL = 0.8, BR = 0.83)
ile = residue(AL = -0.91, AR = -0.59, BL = -0.73, BR = -0.23)
leu = residue(AL = -0.57, AR = -0.13, BL = -0.58, BR = -0.21)
lys = residue(AL = -0.56, AR = -0.29, BL = -0.04, BR = 0.12)
met = residue(AL = -0.64, AR = -0.28, BL = -0.01, BR = 0.11)
phe = residue(AL = -0.52, AR = -0.43, BL = -0.24, BR = 0.06)
pro_t = residue(AL = "NA", AR = -0.19, BL = "NA", BR = -0.24)
ser = residue(AL = -0.44, AR = -0.39, BL = 0.37, BR = 0.3)
thr = residue(AL = -0.79, AR = -0.47, BL = -0.07, BR = 0.2)
trp = residue(AL = -0.4, AR = -0.44, BL = -0.41, BR = -0.11)
tyr = residue(AL = -0.41, AR = -0.37, BL = -0.27, BR = 0.05)
val = residue(AL = -0.74, AR = -0.3, BL = -0.7, BR = -0.14)
n_term = residue(AL = "NA", AR = -1.32, BL = "NA", BR = 1.62)
c_term_COOH = residue(AL = 0.96, AR = "NA", BL = -1.8, BR = "NA")
c_term_COO = residue(AL = 0.05, AR = "NA", BL = "NA", BR = "NA")


amino_acid_dic = {
    "A": ala, "R": arg, "N": asn, "D": asp_COOH, "C": cys,
    "G": gly, "Q": gln, "E": glu_COOH, "H": his_H, "I": ile,
    "L": leu, "K": lys, "M": met, "F": phe, "P": pro_t,
    "S": ser, "T": thr, "W": trp, "Y": tyr, "V": val, "n_term": n_term
}

def calculate_peptide_deuterium_remaining(peptide, pH, tempK, time_in_H2O):
    peptide_residue_rates = []
    for i, residue in enumerate(peptide):
        if i == 0: 
            last = False
            resL = "n_term"
        else:
            resL = peptide[i-1]
        resR = peptide[i]


        if resR == "P":
            peptide_residue_rates.append(0)
            continue
        if resL == "n_term":
            peptide_residue_rates.append(0)
            continue

        resL = amino_acid_dic[resL]
        resR = amino_acid_dic[resR]

        if i == len(peptide) - 1:
            last = True

        residue_rate = calculate_residue_rate(i, resL, resR, pH, tempK, log10_k_acid_ref, log10_k_base_ref, log10_k_water_ref, last)
        peptide_residue_rates.append(residue_rate)

    peptide_residue_rates = np.array(peptide_residue_rates) ### in min^-1

    peptide_residue_deuterium = np.array([1 if residue != "P" else 0 for residue in peptide])
    peptide_residue_deuterium[0] = 0


    ## D(t) = Di * e^-kt
    deuterium_remaining = sum(np.array([peptide_residue_deuterium[i] * np.exp(-peptide_residue_rates[i] * time_in_H2O) for i in range(len(peptide))]))
    
    return(deuterium_remaining)
    


window.mainloop()
